from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from app.models import Assessment, CourseClass, Student, StudentScore
from app.services import get_clo_report, get_course_report


def import_marks_from_excel(session: Session, class_id: int, assessment_id: int, content: bytes) -> int:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    imported = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        student_no = str(row[0]).strip()
        raw_score = row[2] if len(row) > 2 else None
        if raw_score is None:
            continue
        student = session.exec(select(Student).where(Student.student_no == student_no)).first()
        if not student:
            continue
        existing = session.exec(
            select(StudentScore).where(
                StudentScore.assessment_id == assessment_id,
                StudentScore.student_id == student.id,
            )
        ).first()
        if existing:
            existing.score = float(raw_score)
            session.add(existing)
        else:
            session.add(StudentScore(assessment_id=assessment_id, student_id=student.id, score=float(raw_score)))
        imported += 1

    session.commit()
    return imported


def build_mark_template(session: Session, class_id: int, assessment_id: int) -> bytes:
    course_class = session.get(CourseClass, class_id)
    assessment = session.get(Assessment, assessment_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Marks"
    sheet.append(["Student ID", "Student Name", f"Score ({assessment.max_score})"])
    for enrollment in course_class.students:
        student = enrollment.student
        sheet.append([student.student_no, student.name_en, ""])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_course_report_excel(session: Session, class_id: int) -> bytes:
    report = get_course_report(session, class_id)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Course Report"
    summary.append(["Course", report["class"].course.title])
    summary.append(["Class", report["class"].name])
    summary.append([])
    summary.append(["Student ID", "Student Name", "Total", "Percent", "Status", "Grade"])
    for row in report["rows"]:
        summary.append([
            row["student"].student_no,
            row["student"].name_en,
            row["total"],
            row["percent"],
            row["status"],
            row["grade"],
        ])

    for clo in report["clos"]:
        clo_report = get_clo_report(session, class_id, clo.id)
        sheet = workbook.create_sheet(clo.code)
        sheet.append([f"{clo.code} Report", clo.description])
        sheet.append(["Pass", clo_report["pass_count"], "Fail", clo_report["fail_count"], "Achievement", clo_report["achievement"]])
        headers = ["Student ID", "Student Name"] + [item.name for item in clo_report["assessments"]] + ["Total", "Percent", "Status"]
        sheet.append(headers)
        for row in clo_report["rows"]:
            sheet.append(
                [row["student"].student_no, row["student"].name_en]
                + [mark["score"] for mark in row["marks"]]
                + [row["total"], row["percent"], row["status"]]
            )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
