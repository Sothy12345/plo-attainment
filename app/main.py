from collections import defaultdict

import csv
import math
import re
import smtplib
import sqlite3
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4
from typing import Annotated
from urllib.parse import quote
import os

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.exception_handlers import http_exception_handler as fastapi_http_exception_handler
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlmodel import Session, select
from sqlalchemy import inspect
from itsdangerous import BadSignature, URLSafeSerializer

from app.config import PROJECT_ROOT, settings as app_config
from app.database import create_db_and_tables, engine, get_session
from app.excel import build_course_report_excel, build_mark_template, import_marks_from_excel
from app.models import AcademicSemester, AcademicYear, Assessment, AuditLog, CLO, CLOPLOMapping, ClassStudent, ClassTeacher, Course, CourseClass, CoursePLOMapping, CourseTeacher, CohortOutcomeVersion, Document, Faculty, PEO, PEOPLOMapping, PLO, PLOTarget, PLOVersion, Program, Role, RoleDefinition, RolePermission, Student, StudentPromotionHistory, StudentScore, StudentSemesterEnrollment, SystemReport, SystemSetting, StudyPeriod, UserProgrammePreference, UserStudyPeriodPreference, Teacher, User
from app.permissions import ROLE_LABELS, can, can_access_program, scope_label, scoped_programs
from app.attainment import DEFAULT_TARGET as ATTAINMENT_DEFAULT_TARGET, cqi_report, faculty_attainment
from app.attainment import programme_plo_attainment
from app.faculty import faculty_overview
from app.security import hash_password, verify_password
from app.seed import ME_CURRICULUM, seed_data
from app.services import get_clo_report, get_course_report, get_plo_summary, get_program_report, normalized_mapping_weight

app = FastAPI(title="OBE PLO Attainment")
app.mount("/static", StaticFiles(directory=PROJECT_ROOT / "app" / "static"), name="static")
templates = Jinja2Templates(directory=PROJECT_ROOT / "app" / "templates")
templates.env.globals["role_labels"] = ROLE_LABELS
templates.env.globals["can"] = can
cookie_signer = URLSafeSerializer(app_config.secret_key, salt="obe-session")


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    if exc.status_code == 401:
        return redirect(f"/login?next={quote(str(request.url.path or '/dashboard'))}")
    return await fastapi_http_exception_handler(request, exc)

DEFAULT_SYSTEM_SETTINGS = {
    "system_name": "PLO Attainment System",
    "institution_name": "Institute of Technology",
    "address": "No. 123, Street 101, Sangkat Boeng Keng Kang I, Khan Chamkar Mon, Phnom Penh, Cambodia",
    "timezone": "(UTC+07:00) Bangkok, Hanoi, Jakarta",
    "date_format": "May 15, 2024 (MM/DD/YYYY)",
    "time_format": "12 Hour (10:30 AM)",
    "default_language": "English",
    "number_format": "1,234.56",
    "currency": "USD - US Dollar ($)",
    "multilingual_support": "Enabled",
    "academic_year": "2025-2026",
    "default_semester": "Semester 1",
    "passing_score": "50",
    "attainment_target": "70",
    "grading_scheme": "Percentage",
    "allow_manual_data_entry": "Enabled",
    "excellent_min": "80",
    "good_min": "70",
    "satisfactory_min": "50",
    "needs_improvement_max": "49",
    "email_notifications": "Enabled",
    "in_app_notifications": "Enabled",
    "sms_notifications": "Disabled",
    "notify_assessments": "Enabled",
    "notify_reports": "Enabled",
    "smtp_status": "Not tested",
    "smtp_host": "",
    "smtp_port": "587",
    "smtp_security": "STARTTLS",
    "smtp_last_test": "Never",
    "from_email": "",
    "email_template_footer": "PLO Attainment System",
    "session_timeout": "30 Minutes",
    "password_expiry": "90 Days",
    "two_factor_auth": "Disabled",
    "login_attempt_limit": "Enabled",
    "auto_backup": "Enabled",
    "backup_frequency": "Daily",
    "last_backup": "Never",
    "last_backup_file": "",
    "theme": "Light",
    "primary_color": "#2563EB",
    "compact_mode": "Disabled",
    "api_enabled": "Disabled",
    "webhook_url": "",
    "lms_integration": "Disabled",
    "audit_retention": "2 Years",
    "audit_export_format": "CSV",
    "audit_tracking": "Enabled",
    "maintenance_mode": "Disabled",
    "dashboard_tips": "Enabled",
    "version": "1.0.0",
}


def plo_sort_key(plo: PLO) -> int:
    digits = "".join(character for character in plo.code if character.isdigit())
    return int(digits or 0)


def cohort_code_parts(value: str | None) -> dict[str, str | int] | None:
    """Parse a full cohort-class code such as 21ME11Mb1."""
    match = re.fullmatch(
        r"(\d{2})([A-Za-z]+)([1-4])([1-2])([A-Za-z])([A-Za-z])(\d+)",
        str(value or "").strip(),
    )
    if not match:
        return None
    generation, programme, year, semester, shift, degree, group = match.groups()
    return {
        "generation": generation,
        "programme": programme.upper(),
        "year": int(year),
        "semester": int(semester),
        "shift": shift.upper(),
        "degree": degree.lower(),
        "group": group,
    }


def cohort_family_key(value: str | None) -> str | None:
    """Return the stable four-year cohort identity without year/semester."""
    parts = cohort_code_parts(value)
    if not parts:
        return None
    return "|".join(
        str(parts[key])
        for key in ("generation", "programme", "shift", "degree", "group")
    )


def outcome_cohort_key(value: str | None) -> str:
    """Return the one four-year identity used for an outcome-version assignment."""
    raw = str(value or "").strip()
    if not raw:
        return ""
    # New records store the already-normalized identity; old records may store
    # a semester class such as 21ME11Mb1.
    return raw if "|" in raw else (cohort_family_key(raw) or raw)


def outcome_cohort_label(value: str | None) -> str:
    """Create a readable four-year cohort label for the admin assignment list."""
    parts = cohort_code_parts(value)
    if not parts:
        raw = str(value or "").strip()
        if "|" in raw:
            generation, programme, shift, degree, group = raw.split("|", 4)
            parts = {
                "generation": generation,
                "programme": programme,
                "shift": shift,
                "degree": degree,
                "group": group,
            }
        else:
            return raw
    shift = {"M": "Morning", "A": "Afternoon", "E": "Evening"}.get(str(parts["shift"]), str(parts["shift"]))
    degree = {"b": "Bachelor", "m": "Master"}.get(str(parts["degree"]).lower(), str(parts["degree"]))
    return f"{parts['generation']}{parts['programme']} · {shift} · {degree} · Group {parts['group']} · Years 1–4"


@app.on_event("startup")
def on_startup() -> None:
    create_db_and_tables()
    with Session(engine) as session:
        seed_data(session)
        seed_study_periods(session)
        backfill_student_semester_enrollments(session)


def current_user(request: Request, session: Annotated[Session, Depends(get_session)]) -> User | None:
    session_cookie = request.cookies.get("obe_session")
    if not session_cookie:
        return None
    try:
        user_id = int(cookie_signer.loads(session_cookie))
    except (BadSignature, TypeError, ValueError):
        return None
    return session.get(User, user_id)


def require_user(user: Annotated[User | None, Depends(current_user)]) -> User:
    if not user:
        raise HTTPException(status_code=401)
    return user


@app.post("/study-period/select")
def set_global_study_period(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    study_period_id: int = Form(...),
    return_to: str = Form("/dashboard"),
) -> RedirectResponse:
    period = session.get(StudyPeriod, study_period_id)
    if not period or not period.is_active:
        raise HTTPException(status_code=400, detail="Invalid study period")

    preference = session.exec(
        select(UserStudyPeriodPreference).where(UserStudyPeriodPreference.user_id == user.id)
    ).first()
    if preference is None:
        preference = UserStudyPeriodPreference(user_id=user.id, study_period_id=period.id)
    else:
        preference.study_period_id = period.id
        preference.updated_at = datetime.utcnow()
    session.add(preference)
    session.commit()

    response = redirect(safe_return_path(return_to))
    response.set_cookie("study_period_id", str(period.id), max_age=60 * 60 * 24 * 365, samesite="lax")
    response.set_cookie("global_academic_year", period.academic_year, max_age=60 * 60 * 24 * 365, samesite="lax")
    response.set_cookie("global_semester", str(period.semester), max_age=60 * 60 * 24 * 365, samesite="lax")
    return response


def redirect(path: str) -> RedirectResponse:
    return RedirectResponse(path, status_code=303)


def status_badge(active: bool = True) -> str:
    return "Active" if active else "Inactive"


def admin_badge(text: str, tone: str = "blue") -> dict:
    return {"text": text, "tone": tone}


def semester_number(value: str | int | None) -> str:
    text = str(value or "").strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits[:1] if digits else "1"


def semester_label(value: str | int | None) -> str:
    return f"Semester {semester_number(value)}"


def sync_study_periods_from_academic_semesters(session: Session) -> list[StudyPeriod]:
    """Keep the global Study Period selector aligned with Admin semester records.

    StudyPeriod rows are updated in place instead of deleted so saved user
    preferences keep valid foreign keys. A removed or inactive semester simply
    makes its corresponding StudyPeriod unavailable in the sidebar.
    """
    semesters = list(
        session.exec(
            select(AcademicSemester).order_by(
                AcademicSemester.academic_year.desc(), AcademicSemester.code
            )
        ).all()
    )
    source_by_key: dict[tuple[str, int], AcademicSemester] = {}
    for semester in semesters:
        academic_year = str(semester.academic_year or "").strip()
        semester_no = int(semester_number(semester.name or semester.code))
        if not academic_year or semester_no not in {1, 2}:
            continue
        key = (academic_year, semester_no)
        current_source = source_by_key.get(key)
        if current_source is None or (semester.is_default and not current_source.is_default):
            source_by_key[key] = semester

    periods = list(session.exec(select(StudyPeriod)).all())
    period_by_key = {(item.academic_year, int(item.semester)): item for item in periods}
    previous_current_key = next(
        ((item.academic_year, int(item.semester)) for item in periods if item.is_current),
        None,
    )
    now = datetime.utcnow()
    for key, semester in source_by_key.items():
        period = period_by_key.get(key)
        if period is None:
            period = StudyPeriod(
                academic_year=key[0],
                semester=key[1],
                label=f"{key[0]} • Semester {key[1]}",
            )
            session.add(period)
            period_by_key[key] = period
        period.label = f"{key[0]} • Semester {key[1]}"
        period.start_date = semester.start_date or None
        period.end_date = semester.end_date or None
        period.is_active = bool(semester.is_active)
        period.updated_at = now
        session.add(period)

    # Preserve old rows for preferences, but hide periods no longer configured.
    for key, period in period_by_key.items():
        if key not in source_by_key:
            period.is_active = False
            period.is_current = False
            period.updated_at = now
            session.add(period)

    active_keys = [key for key, source in source_by_key.items() if source.is_active]
    default_key = next(
        (key for key, source in source_by_key.items() if source.is_active and source.is_default),
        None,
    )
    if default_key is None and previous_current_key in active_keys:
        default_key = previous_current_key
    if default_key is None and active_keys:
        default_year = session.exec(
            select(AcademicYear).where(AcademicYear.is_default == True)  # noqa: E712
        ).first()
        if default_year:
            default_key = next(
                (key for key in sorted(active_keys) if key[0] == default_year.name),
                None,
            )
    if default_key is None and active_keys:
        default_key = sorted(active_keys, key=lambda item: (item[0], -item[1]), reverse=True)[0]

    for key, period in period_by_key.items():
        period.is_current = bool(period.is_active and key == default_key)
        session.add(period)
    session.commit()
    return list(
        session.exec(
            select(StudyPeriod)
            .where(StudyPeriod.is_active == True)  # noqa: E712
            .order_by(StudyPeriod.academic_year.desc(), StudyPeriod.semester)
        ).all()
    )


def seed_study_periods(session: Session) -> None:
    """Initialize academic records, then derive Study Periods from semesters."""
    classes = list(session.exec(select(CourseClass)).all())
    ensure_academic_records(session, classes)
    sync_study_periods_from_academic_semesters(session)


def available_study_periods() -> list[StudyPeriod]:
    with Session(engine) as db:
        return db.exec(
            select(StudyPeriod)
            .where(StudyPeriod.is_active == True)  # noqa: E712
            .order_by(StudyPeriod.academic_year.desc(), StudyPeriod.semester)
        ).all()


def selected_study_period(request: Request, user: User | None = None) -> StudyPeriod | None:
    """Resolve DB preference first, then cookie, then the system current period."""
    with Session(engine) as db:
        period = None
        if user and user.id:
            preference = db.exec(
                select(UserStudyPeriodPreference).where(UserStudyPeriodPreference.user_id == user.id)
            ).first()
            if preference:
                period = db.get(StudyPeriod, preference.study_period_id)

        if period is None:
            cookie_id = request.cookies.get("study_period_id")
            if cookie_id and str(cookie_id).isdigit():
                period = db.get(StudyPeriod, int(cookie_id))

        if period is None or not period.is_active:
            period = db.exec(
                select(StudyPeriod)
                .where(StudyPeriod.is_current == True)  # noqa: E712
                .order_by(StudyPeriod.academic_year, StudyPeriod.semester)
            ).first()

        if period is None:
            period = db.exec(
                select(StudyPeriod)
                .where(StudyPeriod.is_active == True)  # noqa: E712
                .order_by(StudyPeriod.academic_year, StudyPeriod.semester)
            ).first()
        return period


def safe_return_path(value: str | None) -> str:
    path = str(value or "/dashboard")
    return path if path.startswith("/") and not path.startswith("//") else "/dashboard"


templates.env.globals["available_study_periods"] = available_study_periods
templates.env.globals["selected_study_period"] = selected_study_period


def dean_programme_options(user: User | None) -> list:
    """Programmes a Dean can switch between, for the /manager page switcher."""
    if not user or user.role != Role.DEAN or user.faculty_id is None:
        return []
    with Session(engine) as db:
        return list(
            db.exec(
                select(Program).where(Program.faculty_id == user.faculty_id).order_by(Program.code)
            ).all()
        )


def dean_active_programme(user: User | None):
    if not user or user.role != Role.DEAN:
        return None
    with Session(engine) as db:
        return manager_program(db, user)


def radar_sin(index: int, count: int) -> float:
    """X offset for radar axis `index` of `count`, starting at the top."""
    if not count:
        return 0.0
    return round(math.sin(2 * math.pi * index / count), 6)


def radar_cos(index: int, count: int) -> float:
    """Y offset for radar axis `index` of `count`, starting at the top."""
    if not count:
        return 0.0
    return round(math.cos(2 * math.pi * index / count), 6)


templates.env.globals["radar_sin"] = radar_sin
templates.env.globals["radar_cos"] = radar_cos
templates.env.globals["dean_programme_options"] = dean_programme_options
templates.env.globals["dean_active_programme"] = dean_active_programme


def is_full_class_code(code: str | None) -> bool:
    text = clean_class_code(code) if "clean_class_code" in globals() else "".join(str(code or "").split())
    # Expected pattern like 21ME11Mb1: generation + programme + year + semester + shift + degree + group.
    return len(text) >= 8 and any(ch.isdigit() for ch in text[-4:])


def normalize_curriculum_semester(value: str | int | None) -> str:
    return semester_number(value)


def format_datetimeish(value: object) -> str:
    if not value:
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%b %d, %Y %I:%M %p")
    return str(value)


templates.env.filters["datetime"] = format_datetimeish


def optional_int(value: str | int | None) -> int | None:
    """Parse an optional integer, treating unparseable input as absent.

    Most call sites pass raw query-string or form values, so a hand-edited URL
    like ?program_id=abc must fall back to the default rather than raise.
    """
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def normalized_import_header(value: object) -> str:
    return "".join(character for character in str(value or "").strip().lower() if character.isalnum())


def import_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_student_import_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    header_aliases = {
        "student_no": {"studentid", "studentno", "studentnumber", "id", "studentcode", "code"},
        "name_en": {"studentname", "name", "fullname", "englishname", "nameen"},
        "name_kh": {"khmername", "namekh", "khname"},
        "email": {"email", "emailaddress", "mail"},
    }

    def map_record(raw_record: dict[str, object]) -> dict[str, str]:
        normalized = {normalized_import_header(key): value for key, value in raw_record.items()}
        mapped: dict[str, str] = {}
        for target, aliases in header_aliases.items():
            mapped[target] = ""
            for alias in aliases:
                if alias in normalized:
                    mapped[target] = import_cell(normalized[alias])
                    break
        return mapped

    if suffix == "csv":
        records = csv.DictReader(StringIO(content.decode("utf-8-sig")))
        rows = [map_record(record) for record in records]
    elif suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            return []
        headers = [import_cell(header) for header in raw_rows[0]]
        rows = [map_record(dict(zip(headers, row))) for row in raw_rows[1:]]
        workbook.close()
    else:
        raise ValueError("Please upload an .xlsx or .csv student file.")

    valid_rows: list[dict[str, str]] = []
    for row in rows:
        if not row["student_no"] and not row["name_en"]:
            continue
        if not row["student_no"] or not row["name_en"]:
            continue
        valid_rows.append(row)
    return valid_rows


def normalize_user_scope(session: Session, role: Role, faculty_id: int | None, program_id: int | None) -> tuple[int | None, int | None]:
    if role == Role.SUPER_ADMIN:
        return None, None
    if role == Role.DEAN:
        return faculty_id, None
    if program_id:
        program = session.get(Program, program_id)
        if not program:
            raise HTTPException(status_code=400, detail="Programme scope not found")
        return program.faculty_id, program_id
    return faculty_id, None


def role_scope_is_missing(role: Role, faculty_id: int | None, program_id: int | None) -> bool:
    if role == Role.DEAN:
        return faculty_id is None
    if role in {Role.PROGRAM_MANAGER, Role.STUDENT}:
        return program_id is None
    return False


def ensure_academic_records(session: Session, classes: list[CourseClass]) -> tuple[list[AcademicYear], list[AcademicSemester]]:
    years = session.exec(select(AcademicYear).order_by(AcademicYear.name.desc())).all()
    if not years:
        class_years = sorted({item.academic_year for item in classes if item.academic_year}, reverse=True)
        if not class_years:
            class_years = ["2025-2026", "2024-2025"]
        for index, year_name in enumerate(class_years):
            start_year = year_name.split("-")[0]
            end_year = year_name.split("-")[-1]
            session.add(
                AcademicYear(
                    name=year_name,
                    start_date=f"Aug 01, {start_year}",
                    end_date=f"Jul 31, {end_year}",
                    is_active=index < 2,
                    is_default=index == 0,
                )
            )
        session.commit()
        years = session.exec(select(AcademicYear).order_by(AcademicYear.name.desc())).all()

    semesters = session.exec(select(AcademicSemester).order_by(AcademicSemester.academic_year.desc(), AcademicSemester.code)).all()
    if not semesters:
        pairs = sorted({(item.academic_year, item.semester) for item in classes if item.academic_year and item.semester}, reverse=True)
        if not pairs:
            pairs = [("2025-2026", "1"), ("2025-2026", "2")]
        for index, (year_name, semester) in enumerate(pairs):
            code = f"{year_name.replace('-', '')}-S{semester}"
            session.add(
                AcademicSemester(
                    name=f"Semester {semester}",
                    code=code,
                    academic_year=year_name,
                    start_date="Aug 01, 2025" if str(semester) == "1" else "Jan 02, 2026",
                    end_date="Dec 31, 2025" if str(semester) == "1" else "May 31, 2026",
                    is_active=index < 4,
                    is_default=index == 0,
                )
            )
        session.commit()
        semesters = session.exec(select(AcademicSemester).order_by(AcademicSemester.academic_year.desc(), AcademicSemester.code)).all()
    return years, semesters


def ensure_admin_support_records(session: Session, programs: list[Program], plos: list[PLO]) -> tuple[list[PLOTarget], list[SystemReport], list[AuditLog], dict[str, str]]:
    targets = session.exec(select(PLOTarget)).all()
    if not targets and plos:
        for plo in plos:
            session.add(
                PLOTarget(
                    program_id=plo.program_id,
                    plo_id=plo.id,
                    academic_year="2024-2025",
                    cohort="Cohort 2024",
                    target=70,
                )
            )
        session.commit()
        targets = session.exec(select(PLOTarget)).all()

    reports = session.exec(select(SystemReport)).all()
    if not reports:
        report_seed = [
            ("User Activity Report", "User Management", "Summary of user logins and activities within the system.", "PDF"),
            ("Role Permissions Report", "Role Management", "List of roles and their permissions.", "Excel"),
            ("Faculty Summary Report", "Faculty Management", "Summary of faculties and programmes.", "PDF"),
            ("Programme Report", "Programme Management", "List of programmes and total enrollments.", "Excel"),
            ("PLO Attainment Summary", "PLO Management", "Overall PLO attainment summary by programme.", "PDF"),
            ("PLO Target Achievement Report", "PLO Target Setup", "PLO target vs achievement analysis.", "Excel"),
        ]
        for name, category, description, report_format in report_seed:
            session.add(SystemReport(name=name, category=category, description=description, format=report_format))
        session.commit()
        reports = session.exec(select(SystemReport)).all()

    logs = session.exec(select(AuditLog)).all()
    if not logs:
        log_seed = [
            ("User Management", "CREATE", "Created a new user account.", "User: admin"),
            ("Role Management", "UPDATE", "Updated role scope information.", "Role: Dean"),
            ("PLO Management", "UPDATE", "Updated PLO description.", "PLO: PLO2"),
            ("Academic Year", "CREATE", "Created academic year 2024-2025.", "Academic Year: 2024-2025"),
            ("System Settings", "UPDATE", "Updated system general settings.", "General Settings"),
            ("PLO Target Setup", "DELETE", "Deleted PLO target setup.", "Target ID: TGT-045"),
            ("Programme Management", "UPDATE", "Updated programme details.", "Programme: BSME"),
            ("Cohort / Batch", "CREATE", "Created new cohort.", "Cohort: 2024-BME-01"),
            ("PEO Management", "UPDATE", "Updated PEO statement.", "PEO: PEO1"),
            ("User Management", "DELETE", "Deleted user account.", "User: test.user"),
        ]
        for index, (module, action, description, item_record) in enumerate(log_seed, 1):
            session.add(
                AuditLog(
                    date_time=f"May {15 if index < 6 else 14}, 2024 10:{20 + index:02d}:34 AM",
                    module=module,
                    action=action,
                    description=description,
                    item_record=item_record,
                    status="Failed" if index == 10 else "Success",
                )
            )
        session.commit()
        logs = session.exec(select(AuditLog)).all()

    settings = {item.key: item.value for item in session.exec(select(SystemSetting)).all()}
    if not settings:
        for key, value in DEFAULT_SYSTEM_SETTINGS.items():
            session.add(SystemSetting(key=key, value=value, category="general"))
        session.commit()
        settings = {item.key: item.value for item in session.exec(select(SystemSetting)).all()}
    else:
        missing_settings = {key: value for key, value in DEFAULT_SYSTEM_SETTINGS.items() if key not in settings}
        if missing_settings:
            for key, value in missing_settings.items():
                session.add(SystemSetting(key=key, value=value, category="general"))
            session.commit()
            settings = {item.key: item.value for item in session.exec(select(SystemSetting)).all()}

    return list(targets), list(reports), list(logs), settings


ROLE_PERMISSION_MODULES = [
    "Dashboard",
    "User Management",
    "Role Management",
    "Faculty Management",
    "Programme Management",
    "Student Management",
    "PLO Management",
    "PEO Management",
    "PEO-PLO Mapping",
    "PLO Target Setup",
    "Course Management",
    "Assessment Mapping",
    "Score Entry",
    "Reports",
    "Audit Logs",
    "System Settings",
]


def default_role_definitions() -> list[dict]:
    return [
        {"role_key": Role.SUPER_ADMIN.value, "role_name": "Admin", "role_code": "ADMIN", "description": "Full system administration access.", "abac_scope_type": "All", "menu_access": "admin:*"},
        {"role_key": Role.DEAN.value, "role_name": "Dean", "role_code": "DEAN", "description": "View faculty-level reports and programmes under assigned faculty.", "abac_scope_type": "Faculty", "menu_access": "dean:*"},
        {"role_key": Role.PROGRAM_MANAGER.value, "role_name": "Programme Manager", "role_code": "PROGRAM_MANAGER", "description": "Manage one assigned programme, curriculum, PLO/PEO, courses and promotion.", "abac_scope_type": "Programme", "menu_access": "manager:*"},
        {"role_key": Role.TEACHER.value, "role_name": "Teacher", "role_code": "TEACHER", "description": "Input marks and view reports for assigned courses only.", "abac_scope_type": "Own Data", "menu_access": "teacher:*"},
        {"role_key": Role.STUDENT.value, "role_name": "Student", "role_code": "STUDENT", "description": "View own courses, marks and reports only.", "abac_scope_type": "Own Data", "menu_access": "student:*"},
    ]


def default_permission_values(role_key: str | None, module: str) -> dict[str, bool]:
    if role_key == Role.SUPER_ADMIN.value:
        return {"can_view": True, "can_create": True, "can_edit": True, "can_delete": True, "can_export": True}
    if role_key == Role.DEAN.value:
        can_view = module in {"Dashboard", "Faculty Management", "Programme Management", "Student Management", "PLO Management", "PEO Management", "PEO-PLO Mapping", "PLO Target Setup", "Reports", "Audit Logs"}
        return {"can_view": can_view, "can_create": False, "can_edit": False, "can_delete": False, "can_export": can_view and module in {"Reports", "PLO Management", "PEO Management"}}
    if role_key == Role.PROGRAM_MANAGER.value:
        can_manage = module in {"Student Management", "PLO Management", "PEO Management", "PEO-PLO Mapping", "PLO Target Setup", "Course Management", "Assessment Mapping", "Reports"}
        return {"can_view": module not in {"System Settings"}, "can_create": can_manage, "can_edit": can_manage, "can_delete": can_manage and module not in {"Reports"}, "can_export": module in {"Reports", "PLO Management", "PEO Management", "Course Management"}}
    if role_key == Role.TEACHER.value:
        can_view = module in {"Dashboard", "Course Management", "Assessment Mapping", "Score Entry", "Reports"}
        return {"can_view": can_view, "can_create": module == "Score Entry", "can_edit": module == "Score Entry", "can_delete": False, "can_export": module == "Reports"}
    if role_key == Role.STUDENT.value:
        can_view = module in {"Dashboard", "Course Management", "Reports"}
        return {"can_view": can_view, "can_create": False, "can_edit": False, "can_delete": False, "can_export": False}
    return {"can_view": False, "can_create": False, "can_edit": False, "can_delete": False, "can_export": False}


def ensure_role_definitions(session: Session) -> list[RoleDefinition]:
    for spec in default_role_definitions():
        record = session.exec(select(RoleDefinition).where(RoleDefinition.role_code == spec["role_code"])).first()
        if not record:
            record = RoleDefinition(**spec, status="Active", is_system_role=True)
            session.add(record)
            session.commit()
            session.refresh(record)
        else:
            record.role_key = spec["role_key"]
            record.is_system_role = True
            session.add(record)
    session.commit()
    roles = list(session.exec(select(RoleDefinition).order_by(RoleDefinition.is_system_role.desc(), RoleDefinition.role_name)))
    for role_def in roles:
        for module in ROLE_PERMISSION_MODULES:
            permission = session.exec(select(RolePermission).where(RolePermission.role_definition_id == role_def.id, RolePermission.module == module)).first()
            if not permission:
                session.add(RolePermission(role_definition_id=role_def.id, module=module, **default_permission_values(role_def.role_key, module)))
    session.commit()
    return list(session.exec(select(RoleDefinition).order_by(RoleDefinition.is_system_role.desc(), RoleDefinition.role_name)))


def ensure_plo_versions(session: Session, programs: list[Program], academic_years: list[AcademicYear] | None = None) -> list[PLOVersion]:
    years = academic_years or list(session.exec(select(AcademicYear).order_by(AcademicYear.name.desc())))
    default_year = next((year for year in years if year.is_default), years[0] if years else None)
    for program in programs:
        if not program.id:
            continue
        version = session.exec(select(PLOVersion).where(PLOVersion.programme_id == program.id).order_by(PLOVersion.id)).first()
        if not version:
            version = PLOVersion(programme_id=program.id, version_name=f"{program.code} 2024-2025", effective_academic_year_id=default_year.id if default_year else None, status="Active")
            session.add(version)
            session.commit()
            session.refresh(version)
        for plo in session.exec(select(PLO).where(PLO.program_id == program.id)).all():
            if not plo.plo_version_id:
                plo.plo_version_id = version.id
                session.add(plo)
        for peo in session.exec(select(PEO).where(PEO.program_id == program.id)).all():
            if not peo.plo_version_id:
                peo.plo_version_id = version.id
                session.add(peo)
        for mapping in session.exec(select(PEOPLOMapping).join(PEO, PEOPLOMapping.peo_id == PEO.id).where(PEO.program_id == program.id)).all():
            if not mapping.program_id:
                mapping.program_id = program.id
            if not mapping.plo_version_id:
                mapping.plo_version_id = version.id
            if not mapping.contribution_percentage:
                mapping.contribution_percentage = stored_percent(getattr(mapping, "weight", 0))
            session.add(mapping)
    session.commit()
    return list(session.exec(select(PLOVersion).order_by(PLOVersion.programme_id, PLOVersion.version_name)))


def active_plo_version_for_program(session: Session, program_id: int | None) -> PLOVersion | None:
    if not program_id:
        return None
    return session.exec(select(PLOVersion).where(PLOVersion.programme_id == program_id, PLOVersion.status == "Active").order_by(PLOVersion.id.desc())).first() or session.exec(select(PLOVersion).where(PLOVersion.programme_id == program_id).order_by(PLOVersion.id.desc())).first()


def build_admin_management_page(section: str, session: Session, program_id: int | None = None, version_id: int | None = None) -> dict:
    users = list(session.exec(select(User).order_by(User.name)))
    faculties = list(session.exec(select(Faculty)))
    programs = list(session.exec(select(Program)))
    faculties_by_id = {faculty.id: faculty for faculty in faculties}
    programs_by_id = {program.id: program for program in programs}
    courses = sorted(session.exec(select(Course)).all(), key=lambda course: (course.curriculum_year or 99, course.curriculum_semester or "", course.code))
    classes = list(session.exec(select(CourseClass)))
    students = list(session.exec(select(Student).order_by(Student.student_no)))
    academic_years, academic_semesters = ensure_academic_records(session, classes)
    role_definitions = ensure_role_definitions(session)
    plo_versions = ensure_plo_versions(session, programs, academic_years)
    plo_versions_by_id = {version.id: version for version in plo_versions}
    plos = list(session.exec(select(PLO).order_by(PLO.code)))
    peos = list(session.exec(select(PEO).order_by(PEO.code)))
    plo_targets, system_reports, audit_logs, system_settings = ensure_admin_support_records(session, programs, plos)
    assessments = list(session.exec(select(Assessment)))
    courses_by_id = {course.id: course for course in courses if course.id}
    classes_by_id = {course_class.id: course_class for course_class in classes if course_class.id}
    teachers = list(session.exec(select(Teacher)))
    teachers_by_user_id = {teacher.user_id: teacher for teacher in teachers}
    teacher_course_ids: dict[int, set[int]] = {teacher.id: set() for teacher in teachers if teacher.id}
    teacher_class_ids: dict[int, set[int]] = {teacher.id: set() for teacher in teachers if teacher.id}
    for assignment in session.exec(select(CourseTeacher)).all():
        if assignment.teacher_id and assignment.course_id:
            teacher_course_ids.setdefault(assignment.teacher_id, set()).add(assignment.course_id)
    for assignment in session.exec(select(ClassTeacher)).all():
        if assignment.teacher_id and assignment.class_id:
            teacher_class_ids.setdefault(assignment.teacher_id, set()).add(assignment.class_id)

    def course_scope_label(course: Course | None) -> str:
        if not course:
            return "-"
        program = course.program
        faculty = program.faculty if program else None
        parts = [
            faculty.name if faculty else "No faculty",
            program.code if program else "No programme",
            f"{course.code} · {course.title}",
        ]
        return " · ".join(parts)

    def class_scope_label(course_class: CourseClass | None) -> str:
        if not course_class:
            return "-"
        course = course_class.course
        course_label = f"{course.code} · {course.title}" if course else "No course"
        program = course.program if course else None
        faculty = program.faculty if program else None
        parts = [
            faculty.name if faculty else "No faculty",
            program.code if program else "No programme",
            clean_class_code(course_class.name),
            course_label,
            f"{course_class.academic_year} · Semester {semester_number(course_class.semester)}",
        ]
        return " · ".join(str(part) for part in parts if part)

    teacher_faculty_options = [{"id": faculty.id, "label": faculty.name} for faculty in faculties if faculty.id]
    teacher_program_options = [
        {
            "id": program.id,
            "faculty_id": program.faculty_id,
            "faculty_label": program.faculty.name if program.faculty else "No faculty",
            "label": f"{program.code} · {program.name} ({program.faculty.name if program.faculty else 'No faculty'})",
        }
        for program in sorted(programs, key=lambda item: (item.faculty.name if item.faculty else "", item.code))
        if program.id
    ]
    user_records = []
    for item in users:
        teacher = teachers_by_user_id.get(item.id)
        assigned_course_ids = sorted(teacher_course_ids.get(teacher.id, set())) if teacher and teacher.id else []
        assigned_class_ids = sorted(teacher_class_ids.get(teacher.id, set())) if teacher and teacher.id else []
        assigned_course_labels = [course_scope_label(courses_by_id.get(course_id)) for course_id in assigned_course_ids]
        assigned_class_labels = [class_scope_label(classes_by_id.get(class_id)) for class_id in assigned_class_ids]
        assigned_program_ids = {
            course.program_id
            for course_id in assigned_course_ids
            for course in [courses_by_id.get(course_id)]
            if course and course.program_id
        }
        assigned_program_ids.update(
            course_class.course.program_id
            for class_id in assigned_class_ids
            for course_class in [classes_by_id.get(class_id)]
            if course_class and course_class.course and course_class.course.program_id
        )
        assigned_faculty_ids = {
            programs_by_id[program_id].faculty_id
            for program_id in assigned_program_ids
            if program_id in programs_by_id and programs_by_id[program_id].faculty_id
        }
        assigned_program_labels = [
            f"{programs_by_id[program_id].code} · {programs_by_id[program_id].name}"
            for program_id in sorted(assigned_program_ids)
            if program_id in programs_by_id
        ]
        assigned_faculty_labels = [
            faculties_by_id[faculty_id].name
            for faculty_id in sorted(assigned_faculty_ids)
            if faculty_id in faculties_by_id
        ]
        teaching_scope = f"{len(assigned_program_ids)} programmes · {len(assigned_class_ids)} classes" if teacher else "-"
        user_records.append(
            {
            "id": item.id,
            "name": item.name,
            "username": item.email.split("@")[0],
            "email": item.email,
            "role": item.role,
            "role_label": ROLE_LABELS[item.role],
            "role_tone": "blue"
            if item.role == Role.SUPER_ADMIN
            else "purple"
            if item.role == Role.DEAN
            else "green"
            if item.role == Role.PROGRAM_MANAGER
            else "orange"
            if item.role == Role.TEACHER
            else "cyan",
            "faculty_id": item.faculty_id,
            "program_id": item.program_id,
            "faculty": faculties_by_id[item.faculty_id].name if item.faculty_id in faculties_by_id else "All faculties",
            "program": programs_by_id[item.program_id].name if item.program_id in programs_by_id else ("All programmes" if item.role in [Role.SUPER_ADMIN, Role.DEAN] else "-"),
            "staff_no": teacher.staff_no if teacher else "",
            "teacher_course_ids": assigned_course_ids,
            "teacher_class_ids": assigned_class_ids,
            "teacher_faculty_ids": sorted(assigned_faculty_ids),
            "teacher_program_ids": sorted(assigned_program_ids),
            "teacher_faculties": assigned_faculty_labels,
            "teacher_programmes": assigned_program_labels,
            "teacher_courses": assigned_course_labels,
            "teacher_classes": assigned_class_labels,
            "teaching_scope": teaching_scope,
            "is_active": item.is_active,
            }
        )

    base_stats = {
        "blue": "bi-people",
        "green": "bi-shield-check",
        "orange": "bi-person-x",
        "purple": "bi-building",
    }

    faculty_codes = {
        "Faculty of Business Administration": "FBA",
        "Faculty of Science and Technology": "FST",
        "Faculty of Agriculture": "FA",
        "Faculty of Art, Humanity and Foreign Language": "FAHFL",
        "Faculty of Social Sciences": "FSS",
    }
    faculty_descriptions = {
        "Faculty of Business Administration": "Business, accounting, management, marketing, finance and banking programmes.",
        "Faculty of Science and Technology": "Computer science, mathematics, mechanical engineering, electronics and automation.",
        "Faculty of Agriculture": "Agronomy, animal science, veterinary and rural development programmes.",
        "Faculty of Art, Humanity and Foreign Language": "Humanity, arts and foreign language programmes.",
        "Faculty of Social Sciences": "Social science and public administration programmes.",
    }
    mock_faculties = [
        [
            str(index),
            admin_badge(faculty_codes.get(faculty.name, faculty.name[:4].upper())),
            faculty.name,
            faculty_descriptions.get(faculty.name, "Academic faculty."),
            len(faculty.programs),
            status_badge(),
            "actions",
        ]
        for index, faculty in enumerate(faculties, 1)
    ]
    faculty_records = [
        {
            "id": faculty.id,
            "code": faculty_codes.get(faculty.name, faculty.name[:4].upper()),
            "name": faculty.name,
            "description": faculty_descriptions.get(faculty.name, "Academic faculty."),
            "programmes": len(faculty.programs),
            "is_active": True,
        }
        for faculty in faculties
    ]
    program_records = [
        {
            "id": program.id,
            "code": program.code,
            "name": program.name,
            "degree": "Bachelor",
            "faculty_id": program.faculty_id,
            "faculty": program.faculty.name if program.faculty else "-",
            "students": 0,
            "is_active": True,
        }
        for program in programs
    ]
    academic_year_records = [
        {
            "id": item.id,
            "name": item.name,
            "start_date": item.start_date,
            "end_date": item.end_date,
            "is_active": item.is_active,
            "is_default": item.is_default,
            "created_by": item.created_by,
            "created_at": item.created_at,
        }
        for item in academic_years
    ]
    semester_records = [
        {
            "id": item.id,
            "name": item.name,
            "code": item.code,
            "academic_year": item.academic_year,
            "start_date": item.start_date,
            "end_date": item.end_date,
            "is_active": item.is_active,
            "is_default": item.is_default,
            "created_at": item.created_at,
        }
        for item in academic_semesters
    ]
    cohort_records = []
    grouped_classes: dict[tuple[str, str, str, str], dict] = {}
    for course_class in classes:
        semester_no = semester_number(course_class.semester)
        programme_name = course_class.course.program.name if course_class.course and course_class.course.program else "-"
        programme_code = course_class.course.program.code if course_class.course and course_class.course.program else ""
        class_code = clean_class_code(course_class.name)
        item = {
            "id": course_class.id,
            "code": class_code,
            "name": class_code,
            "course_id": course_class.course_id,
            "programme": programme_name,
            "programme_code": programme_code,
            "program_id": course_class.course.program_id if course_class.course else None,
            "course": f"{course_class.course.code} · {course_class.course.title}" if course_class.course else "-",
            "academic_year": course_class.academic_year,
            "semester": semester_no,
            "semester_label": semester_label(semester_no),
            "start_date": course_class.semester_start or "",
            "end_date": course_class.semester_end or "",
            "students": len(course_class.students),
            "is_active": True,
            "code_complete": is_full_class_code(class_code),
        }
        cohort_records.append(item)
        key = (class_code, programme_name, course_class.academic_year, semester_no)
        group = grouped_classes.setdefault(key, {
            "id": course_class.id,
            "code": class_code,
            "programme": programme_name,
            "programme_code": programme_code,
            "program_id": course_class.course.program_id if course_class.course else None,
            "academic_year": course_class.academic_year,
            "semester": semester_no,
            "semester_label": semester_label(semester_no),
            "course_count": 0,
            "students": 0,
            "courses": [],
            "code_complete": is_full_class_code(class_code),
        })
        group["course_count"] += 1
        group["students"] = max(group["students"], len(course_class.students))
        group["courses"].append(item)
    cohort_group_records = sorted(grouped_classes.values(), key=lambda item: (item["programme"], item["code"], item["academic_year"], item["semester"]))
    # Student intake uses one option for the whole four-year cohort. The eight
    # semester class codes remain available in Academic Structure and are
    # reached later through Student Promotion.
    intake_options_by_family: dict[tuple[int, str], dict] = {}
    for cohort in cohort_group_records:
        parts = cohort_code_parts(cohort["code"])
        family_key = cohort_family_key(cohort["code"])
        program_id_for_cohort = cohort.get("program_id")
        if not parts or not family_key or not program_id_for_cohort:
            continue
        key = (program_id_for_cohort, family_key)
        position = (int(parts["year"]) - 1) * 2 + int(parts["semester"])
        shift_label = {"M": "Morning", "A": "Afternoon", "E": "Evening"}.get(
            str(parts["shift"]), str(parts["shift"])
        )
        degree_label = {"b": "Bachelor", "m": "Master"}.get(
            str(parts["degree"]), str(parts["degree"]).upper()
        )
        option = intake_options_by_family.setdefault(
            key,
            {
                "id": cohort["id"],
                "program_id": program_id_for_cohort,
                "programme_code": cohort["programme_code"],
                "programme": cohort["programme"],
                "generation": parts["generation"],
                "group": parts["group"],
                "label": f"Generation {parts['generation']} · {cohort['programme_code']} · {shift_label} · {degree_label} · Group {parts['group']}",
                "intake_class_code": cohort["code"],
                "intake_academic_year": cohort["academic_year"],
                "intake_course_count": cohort["course_count"],
                "configured_semesters": set(),
                "position": position,
            },
        )
        option["configured_semesters"].add(position)
        # The representative is always the earliest configured class, normally
        # Year 1 Semester 1. Only this first period is enrolled at intake.
        if position < option["position"]:
            option["id"] = cohort["id"]
            option["intake_class_code"] = cohort["code"]
            option["intake_academic_year"] = cohort["academic_year"]
            option["intake_course_count"] = cohort["course_count"]
            option["position"] = position
    student_cohort_options = sorted(
        [item for item in intake_options_by_family.values() if item["position"] == 1],
        key=lambda item: (item["programme"], item["generation"], item["group"]),
    )
    semester_enrollments_by_student: dict[int, list[StudentSemesterEnrollment]] = {}
    for enrollment in session.exec(
        select(StudentSemesterEnrollment).order_by(
            StudentSemesterEnrollment.created_at,
            StudentSemesterEnrollment.id,
        )
    ).all():
        semester_enrollments_by_student.setdefault(enrollment.student_id, []).append(enrollment)

    student_records = []
    for student in students:
        course_enrollments = list(student.enrollments)
        study_enrollments = semester_enrollments_by_student.get(student.id, [])
        current_study_enrollment = max(
            study_enrollments,
            key=lambda item: (
                int(str(item.academic_year or "0").split("-")[0])
                if str(item.academic_year or "").split("-")[0].isdigit()
                else 0,
                int(normalize_semester(item.semester) or 0),
                item.id or 0,
            ),
        ) if study_enrollments else None
        primary_class = course_enrollments[-1].course_class if course_enrollments else None
        primary_course = primary_class.course if primary_class and primary_class.course else None
        primary_program = (
            session.get(Program, current_study_enrollment.program_id)
            if current_study_enrollment
            else primary_course.program if primary_course and primary_course.program else None
        )
        cohort_name = current_study_enrollment.cohort_name if current_study_enrollment else primary_class.name if primary_class else "-"
        study_year, study_semester = _cohort_year_semester(cohort_name) if cohort_name != "-" else (None, None)
        student_records.append(
            {
                "id": student.id,
                "student_no": student.student_no,
                "name_en": student.name_en,
                "name_kh": student.name_kh or "",
                "email": student.user.email if student.user else "",
                "user_id": student.user_id,
                "class_id": primary_class.id if primary_class else None,
                "cohort": cohort_name,
                "programme": primary_program.name if primary_program else "-",
                "faculty": primary_program.faculty.name if primary_program and primary_program.faculty else "-",
                "program_id": primary_program.id if primary_program else None,
                "study_year": study_year,
                "study_semester": study_semester,
                "study_level": f"Year {study_year} · Semester {study_semester}" if study_year and study_semester else "Not enrolled",
                "enrollments": len(study_enrollments),
                "course_enrollments": len(course_enrollments),
                "scores": len(student.scores),
                "is_active": student.user.is_active if student.user else True,
            }
        )
    plos_by_id = {plo.id: plo for plo in plos}
    peo_links_by_peo: dict[int, dict[int, float]] = {}
    for link in session.exec(select(PEOPLOMapping)).all():
        peo_links_by_peo.setdefault(link.peo_id, {})[link.plo_id] = stored_percent(link.contribution_percentage)
    selected_mapping_program_id = optional_int(None)
    selected_mapping_version_id = optional_int(None)
    peo_records = []
    for index, peo in enumerate(peos, 1):
        mapped_weights = peo_links_by_peo.get(peo.id, {})
        mapped_plo_ids = list(mapped_weights)
        mapped_plos = [plos_by_id[plo_id] for plo_id in mapped_plo_ids if plo_id in plos_by_id]
        weight_values = [weight for weight in mapped_weights.values() if weight > 0]
        result = round(sum(weight_values), 2) if weight_values else 0
        version = plo_versions_by_id.get(peo.plo_version_id)
        peo_records.append(
            {
                "id": peo.id,
                "program_id": peo.program_id,
                "plo_version_id": peo.plo_version_id,
                "version_name": version.version_name if version else "Default",
                "version_locked": version.is_locked if version else False,
                "code": peo.code,
                "description": peo.description,
                "programme": programs_by_id[peo.program_id].name if peo.program_id in programs_by_id else "-",
                "domain": "PEO",
                "status": peo.status,
                "remark": peo.remark,
                "is_active": peo.status == "Active",
                "created_at": format_datetimeish(peo.created_at),
                "mapped_plo_ids": mapped_plo_ids,
                "mapped_weights": mapped_weights,
                "mapped_plo_codes": [plo.code for plo in mapped_plos],
                "result": result,
                "result_status": "Complete" if abs(result - 100) < 0.01 else "Needs Review",
            }
        )
    plo_records = [
        {
            "id": plo.id,
            "program_id": plo.program_id,
            "plo_version_id": plo.plo_version_id,
            "version_name": plo_versions_by_id[plo.plo_version_id].version_name if plo.plo_version_id in plo_versions_by_id else "Default",
            "version_locked": plo_versions_by_id[plo.plo_version_id].is_locked if plo.plo_version_id in plo_versions_by_id else False,
            "code": plo.code,
            "description": plo.description,
            "programme": plo.program.name if plo.program else "-",
            "domain": plo.domain,
            "bloom_level": plo.bloom_level,
            "status": plo.status,
            "remark": plo.remark,
            "is_active": plo.status == "Active",
            "created_at": format_datetimeish(plo.created_at),
        }
        for index, plo in enumerate(plos, 1)
    ]
    active_peo_records = [record for record in peo_records if record["is_active"]]
    selected_mapping_program_id = active_peo_records[0]["program_id"] if active_peo_records else (programs[0].id if programs else None)
    selected_mapping_version_id = active_peo_records[0]["plo_version_id"] if active_peo_records else (
        active_plo_version_for_program(session, selected_mapping_program_id).id if selected_mapping_program_id else None
    )
    peo_mapping_peos = [
        record
        for record in active_peo_records
        if record["program_id"] == selected_mapping_program_id and record["plo_version_id"] == selected_mapping_version_id
    ]
    peo_mapping_plos = [
        plo
        for plo in plos
        if plo.program_id == selected_mapping_program_id and plo.plo_version_id == selected_mapping_version_id and plo.status == "Active"
    ]
    peo_mapping_version = plo_versions_by_id.get(selected_mapping_version_id)
    target_records = [
        {
            "id": target.id,
            "program_id": target.program_id,
            "plo_id": target.plo_id,
            "programme": programs_by_id[target.program_id].name if target.program_id in programs_by_id else "-",
            "academic_year": target.academic_year,
            "cohort": target.cohort,
            "plo_code": plos_by_id[target.plo_id].code if target.plo_id in plos_by_id else "-",
            "plo_description": plos_by_id[target.plo_id].description if target.plo_id in plos_by_id else "-",
            "target": target.target,
            "set_by": target.set_by,
            "updated_at": target.updated_at,
        }
        for target in plo_targets
    ]
    report_records = [
        {
            "id": report.id,
            "name": report.name,
            "category": report.category,
            "description": report.description,
            "created_by": report.created_by,
            "last_generated": report.last_generated,
            "format": report.format,
            "status": report.status,
        }
        for report in system_reports
    ]
    audit_records = [
        {
            "id": log.id,
            "date_time": log.date_time,
            "user_name": log.user_name,
            "module": log.module,
            "action": log.action,
            "description": log.description,
            "item_record": log.item_record,
            "ip_address": log.ip_address,
            "status": log.status,
        }
        for log in audit_logs
    ]
    role_permissions = {
        role_def.id: list(session.exec(select(RolePermission).where(RolePermission.role_definition_id == role_def.id).order_by(RolePermission.module)))
        for role_def in role_definitions
        if role_def.id
    }
    role_records = [
        {
            "id": role_def.id,
            "role": Role(role_def.role_key) if role_def.role_key in {role.value for role in Role} else None,
            "name": role_def.role_name,
            "code": role_def.role_code,
            "tone": "blue" if role_def.abac_scope_type == "All" else "purple" if role_def.abac_scope_type == "Faculty" else "green" if role_def.abac_scope_type == "Programme" else "cyan",
            "scope": role_def.abac_scope_type,
            "description": role_def.description,
            "attribute": "No scope attribute" if role_def.abac_scope_type == "All" else "faculty_id" if role_def.abac_scope_type == "Faculty" else "program_id" if role_def.abac_scope_type == "Programme" else "own record",
            "allowed": role_def.menu_access,
            "status": role_def.status,
            "is_system_role": role_def.is_system_role,
            "users": [item for item in user_records if item["role"].value == role_def.role_key] if role_def.role_key else [],
            "active_users": sum(1 for item in users if item.role.value == role_def.role_key and item.is_active) if role_def.role_key else 0,
            "permissions": role_permissions.get(role_def.id, []),
        }
        for role_def in role_definitions
    ]
    configs = {
        "users": {
            "kind": "users",
            "title": "User Management",
            "description": "Manage system users, their roles and access permissions.",
            "button": "Add New User",
            "button_url": "/admin/users#add-user",
            "list_title": "User List",
            "list_description": "View and manage all system users.",
            "stats": [
                ("Total Users", len(users), "bi-people", "blue"),
                ("Active Users", sum(1 for item in users if item.is_active), "bi-shield-check", "green"),
                ("Inactive Users", sum(1 for item in users if not item.is_active), "bi-person-x", "orange"),
                ("Roles", len(Role), "bi-people-fill", "purple"),
            ],
            "columns": ["No.", "Full Name", "Username", "Email", "Role", "Faculty", "Programme", "Teaching", "Status", "Actions"],
            "rows": [
                [
                    str(index),
                    user.name,
                    user.email.split("@")[0],
                    user.email,
                    admin_badge(ROLE_LABELS[user.role], "blue" if user.role == Role.SUPER_ADMIN else "green" if user.role == Role.PROGRAM_MANAGER else "orange" if user.role == Role.TEACHER else "cyan"),
                    faculties_by_id[user.faculty_id].name if user.faculty_id in faculties_by_id else "All faculties",
                    programs_by_id[user.program_id].name if user.program_id in programs_by_id else ("All programmes" if user.role in [Role.SUPER_ADMIN, Role.DEAN] else "-"),
                    status_badge(user.is_active),
                    "actions",
                ]
                for index, user in enumerate(users, 1)
            ],
            "showing": f"Showing 1 to {len(users)} of {len(users)} users",
            "user_records": user_records,
            "faculties": faculties,
            "programs": programs,
            "roles": list(Role),
            "student_cohort_options": student_cohort_options,
            "teacher_faculty_options": teacher_faculty_options,
            "teacher_program_options": teacher_program_options,
        },
        "roles": {
            "kind": "roles",
            "title": "Role Management",
            "description": "Manage role permissions with attribute-based access control by faculty and programme.",
            "button": "Add Role",
            "button_url": "/admin/roles#add-role",
            "list_title": "Role List",
            "list_description": "View and manage all user roles and their permissions.",
            "stats": [("Total Roles", len(role_records), "bi-people", "blue"), ("Active Roles", sum(1 for item in role_records if item["status"] == "Active"), "bi-shield-check", "green"), ("Inactive Roles", sum(1 for item in role_records if item["status"] != "Active"), "bi-person-x", "orange"), ("System Roles", sum(1 for item in role_records if item["is_system_role"]), "bi-wrench", "purple")],
            "columns": ["No.", "Role Name", "Role Code", "ABAC Scope", "Users", "Status", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(role_records)} of {len(role_records)} roles",
            "summary": [("Access Model", "RBAC + ABAC", "bi-diagram-3"), ("Permission Modules", len(ROLE_PERMISSION_MODULES), "bi-grid"), ("System Roles", sum(1 for item in role_records if item["is_system_role"]), "bi-shield-lock"), ("Scoped Data", "Faculty/Programme/Own", "bi-person-check")],
            "access_rules": [
                ("Role", "Attribute Used", "Allowed Data"),
                ("Super Admin", "No scope attribute", "All faculties, programmes, users, setup and reports."),
                ("Dean", "faculty_id", "Only reports and programmes where programme.faculty_id equals dean.faculty_id."),
                ("Programme Coordinator", "faculty_id + program_id", "Only one assigned programme, including PEO/PLO, courses, CLO assessment mapping and reports."),
                ("Teacher", "program_id + assigned class", "Only assigned classes for score input and CLO/course reports."),
                ("Student", "program_id + own student profile", "Only their own marks, CLO/PLO reports, documents and calendar."),
            ],
            "role_records": role_records,
            "permission_modules": ROLE_PERMISSION_MODULES,
            "users": users,
            "faculties": faculties,
            "programs": programs,
            "roles": list(Role),
        },
        "students": {
            "kind": "students",
            "title": "Student Management",
            "description": "Create each student account once, then progress the same student through the four-year programme.",
            "button": "Add New Student",
            "button_url": "/admin/students#add-student",
            "list_title": "Student List",
            "list_description": "View and manage all student records.",
            "stats": [
                ("Total Students", len(student_records), "bi-people", "blue"),
                ("Active Students", sum(1 for item in student_records if item["is_active"]), "bi-person-check", "green"),
                ("Inactive Students", sum(1 for item in student_records if not item["is_active"]), "bi-person-x", "orange"),
                ("Enrolled Cohorts", len({item["cohort"] for item in student_records if item["cohort"] != "-"}), "bi-journal-text", "purple"),
            ],
            "columns": ["No.", "Student ID", "Student Name", "Khmer Name", "Email", "Faculty", "Programme", "Current Cohort", "Study Progress", "Status", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(student_records)} of {len(student_records)} students",
            "about": "One Student and one User account are kept for all four years. Semester enrollment records are added during promotion; the student identity, login and historical scores are never duplicated.",
            "records": student_records,
            "classes": classes,
            "cohort_groups": cohort_group_records,
            "student_cohort_options": student_cohort_options,
            "programs": programs,
            "form_anchor": "add-student",
        },
        "faculties": {
            "kind": "faculties",
            "title": "Faculty Management",
            "description": "Manage faculties in the system.",
            "button": "Add New Faculty",
            "button_url": "/admin/faculties#add-faculty",
            "list_title": "Faculty List",
            "list_description": "View and manage all faculties.",
            "stats": [("Total Faculties", len(faculties), "bi-building", "blue"), ("Active Faculties", len(faculties), "bi-shield-check", "green"), ("Inactive Faculties", 0, "bi-person-x", "orange"), ("Total Programmes", len(programs), "bi-people-fill", "purple")],
            "columns": ["No.", "Faculty Code", "Faculty Name", "Description", "Total Programmes", "Status", "Actions"],
            "rows": mock_faculties,
            "showing": f"Showing 1 to {len(faculties)} of {len(faculties)} faculties",
            "about": "Faculties are the top-level academic entities in the system. Programmes are managed directly under each faculty.",
            "records": faculty_records,
            "form_anchor": "add-faculty",
        },
        "programmes": {
            "kind": "programmes",
            "title": "Programme Management",
            "description": "Manage academic programmes under faculties.",
            "button": "Add New Programme",
            "button_url": "/admin/programmes#add-programme",
            "list_title": "Programme List",
            "list_description": "View and manage all academic programmes.",
            "stats": [("Total Programmes", len(programs), "bi-mortarboard", "blue"), ("Active Programmes", len(programs), "bi-shield-check", "green"), ("Inactive Programmes", 0, "bi-person-x", "orange"), ("Total Faculties", len(faculties), "bi-building", "purple")],
            "columns": ["No.", "Programme Code", "Programme Name", "Degree Level", "Faculty", "Total Students", "Status", "Actions"],
            "rows": [[str(i), admin_badge(program.code), program.name, "Bachelor", program.faculty.name if program.faculty else "-", 0, status_badge(), "actions"] for i, program in enumerate(programs, 1)] or [],
            "showing": f"Showing 1 to {len(programs)} of {len(programs)} programmes",
            "about": "Programmes are academic offerings within faculties. You can manage programme details, duration, intake and related information.",
            "records": program_records,
            "faculties": faculties,
            "form_anchor": "add-programme",
        },
        "academic-years": {
            "kind": "academic_years",
            "title": "Academic Year",
            "description": "Manage academic years used in the system.",
            "button": "Add New Academic Year",
            "button_url": "/admin/academic-years#add-academic-year",
            "list_title": "Academic Year List",
            "list_description": "View and manage all academic years.",
            "stats": [("Total Academic Years", len(academic_year_records), "bi-calendar3", "blue"), ("Active Academic Years", sum(1 for item in academic_year_records if item["is_active"]), "bi-shield-check", "green"), ("Inactive Academic Years", sum(1 for item in academic_year_records if not item["is_active"]), "bi-wallet", "orange"), ("Default Academic Year", next((item["name"] for item in academic_year_records if item["is_default"]), "-"), "bi-bar-chart", "purple")],
            "columns": ["No.", "Academic Year", "Start Date", "End Date", "Status", "Default", "Created By", "Created At", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(academic_year_records)} of {len(academic_year_records)} academic years",
            "about": "Academic year is used to organize academic activities, courses, assessments and reports within a specific period.",
            "records": academic_year_records,
            "form_anchor": "add-academic-year",
        },
        "semesters": {
            "kind": "semesters",
            "title": "Semester",
            "description": "Manage semesters within academic years.",
            "button": "Add New Semester",
            "button_url": "/admin/semesters#add-semester",
            "list_title": "Semester List",
            "list_description": "View and manage all semesters.",
            "stats": [("Total Semesters", len(semester_records), "bi-calendar3", "blue"), ("Active Semesters", sum(1 for item in semester_records if item["is_active"]), "bi-shield-check", "green"), ("Inactive Semesters", sum(1 for item in semester_records if not item["is_active"]), "bi-wallet", "orange"), ("Default Semester", next((f'{item["name"]} ({item["academic_year"]})' for item in semester_records if item["is_default"]), "-"), "bi-calendar2-week", "purple")],
            "columns": ["No.", "Semester Name", "Semester Code", "Academic Year", "Start Date", "End Date", "Status", "Default", "Created At", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(semester_records)} of {len(semester_records)} semesters",
            "about": "Semesters are academic periods within an academic year used to organize teaching, learning, assessments and reporting.",
            "records": semester_records,
            "academic_years": academic_years,
            "form_anchor": "add-semester",
        },
        "cohorts": {
            "kind": "cohorts",
            "title": "Cohort / Batch Management",
            "description": "Manage student cohorts or batches in the system.",
            "button": "Add New Cohort / Batch",
            "button_url": "/admin/cohorts#add-cohort",
            "list_title": "Cohort / Batch List",
            "list_description": "View and manage all cohorts or batches.",
            "stats": [("Total Cohorts / Batches", len(cohort_records), "bi-people", "blue"), ("Active Cohorts / Batches", len(cohort_records), "bi-shield-check", "green"), ("Inactive Cohorts / Batches", 0, "bi-person-x", "orange"), ("Default Cohort / Batch", cohort_records[0]["name"] if cohort_records else "-", "bi-calendar3", "purple")],
            "columns": ["No.", "Cohort / Batch Code", "Cohort / Batch Name", "Programme", "Intake Year", "Start Year", "Expected Graduation Year", "Status", "Default", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(cohort_records)} of {len(cohort_records)} cohorts / batches",
            "about": "Cohorts or batches represent student groups based on their intake year and programme. They are used for tracking academic progress, PLO attainment, and reporting.",
            "records": cohort_records,
            "courses": courses,
            "programs": programs,
            "academic_years": academic_years,
            "form_anchor": "add-cohort",
        },
        "academic-structure": {
            "kind": "academic_structure",
            "title": "Academic Structure",
            "description": "Manage academic years, semesters, and class/cohort codes in one simple workflow.",
            "button": "Create New Class",
            "button_url": "/admin/cohorts#classes",
            "list_title": "Academic Structure",
            "list_description": "Use tabs to manage the foundation data used by student enrollment and promotion.",
            "stats": [
                ("Academic Years", len(academic_year_records), "bi-calendar3", "blue"),
                ("Semesters", len(semester_records), "bi-calendar2-week", "green"),
                ("Classes / Cohorts", len(cohort_group_records), "bi-people", "purple"),
                ("Active Year", next((item["name"] for item in academic_year_records if item["is_default"]), academic_year_records[0]["name"] if academic_year_records else "-"), "bi-check-circle", "orange"),
            ],
            "columns": [],
            "rows": [],
            "showing": "",
            "about": "Academic structure should be created in order: academic year, semester, then class/cohort code. Promotion uses class codes such as 21ME11Mb1 → 21ME12Mb1.",
            "academic_year_records": academic_year_records,
            "semester_records": semester_records,
            "cohort_records": cohort_records,
            "cohort_group_records": cohort_group_records,
            "academic_years": academic_years,
            "courses": courses,
            "programs": programs,
            "form_anchor": "classes",
        },

        "plos": {
            "kind": "plos",
            "title": "PLO Management",
            "description": "Manage Programme Learning Outcomes (PLOs) for all programmes.",
            "button": "Add New PLO",
            "button_url": "/admin/plos#add-plo",
            "list_title": "PLO List",
            "list_description": "View and manage all Programme Learning Outcomes.",
            "stats": [("Total PLOs", len(plo_records), "bi-bullseye", "blue"), ("Active PLOs", sum(1 for item in plo_records if item["is_active"]), "bi-shield-check", "green"), ("Inactive PLOs", sum(1 for item in plo_records if not item["is_active"]), "bi-x-circle", "orange"), ("Programmes", max(len(programs), 1), "bi-building", "purple")],
            "columns": ["No.", "PLO Version", "PLO Code", "PLO Description", "Programme", "Domain", "Bloom", "Status", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(plo_records)} of {len(plo_records)} PLOs",
            "about": "Programme Learning Outcomes (PLOs) describe what students are expected to know and be able to do by the time of graduation.",
            "records": plo_records,
            "programs": programs,
            "plo_versions": plo_versions,
            "academic_years": academic_years,
            "form_anchor": "add-plo",
        },
        "peos": {
            "kind": "peos",
            "title": "PEO Management",
            "description": "Manage Programme Educational Objectives (PEOs) for all programmes.",
            "button": "Add New PEO",
            "button_url": "/admin/peos#add-peo",
            "list_title": "PEO List",
            "list_description": "View and manage all Programme Educational Objectives.",
            "stats": [("Total PEOs", len(peo_records), "bi-bullseye", "blue"), ("Active PEOs", sum(1 for item in peo_records if item["is_active"]), "bi-shield-check", "green"), ("Inactive PEOs", sum(1 for item in peo_records if not item["is_active"]), "bi-x-circle", "orange"), ("Programmes", max(len(programs), 1), "bi-building", "purple")],
            "columns": ["No.", "PLO Version", "PEO Code", "PEO Description", "Programme", "Mapped PLOs", "Status", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(peo_records)} of {len(peo_records)} PEOs",
            "about": "Programme Educational Objectives (PEOs) are broad statements that describe career and professional accomplishments the programme prepares graduates to achieve.",
            "records": peo_records,
            "programs": programs,
            "plo_versions": plo_versions,
            "plos": plos,
            "mapping_program_id": selected_mapping_program_id,
            "mapping_version_id": selected_mapping_version_id,
            "mapping_version_locked": peo_mapping_version.is_locked if peo_mapping_version else False,
            "mapping_peos": peo_mapping_peos,
            "mapping_plos": peo_mapping_plos,
            "mapping_updated_at": format_datetimeish(
                max(
                    [
                        mapping.updated_at or mapping.created_at
                        for mapping in session.exec(select(PEOPLOMapping).where(PEOPLOMapping.plo_version_id == selected_mapping_version_id)).all()
                        if mapping.updated_at or mapping.created_at
                    ],
                    default=None,
                )
            ),
            "form_anchor": "add-peo",
        },
        "targets": {
            "kind": "targets",
            "title": "PLO Target Setup",
            "description": "Define target attainment levels for each PLO by programme and academic year.",
            "button": "Add Target Setup",
            "button_url": "/admin/targets#add-target",
            "list_title": "PLO Target List",
            "list_description": "View and manage target attainment by PLO.",
            "stats": [("Total Programmes", max(len(programs), 1), "bi-bullseye", "blue"), ("Academic Years", len(academic_years), "bi-calendar3", "green"), ("PLOs", len(plos), "bi-bullseye", "orange"), ("Total Target Setups", len(target_records), "bi-bar-chart", "purple")],
            "columns": ["No.", "Programme", "Academic Year", "Cohort / Batch", "PLO Code", "PLO Description", "Target Attainment Level (%)", "Set By", "Last Updated", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(target_records)} of {len(target_records)} target setups",
            "about": "Target attainment level is the expected minimum percentage of students that should achieve each PLO by the end of the academic year.",
            "records": target_records,
            "programs": programs,
            "plos": plos,
            "academic_years": academic_years,
            "classes": classes,
            "form_anchor": "add-target",
        },
        "reports": {
            "kind": "reports",
            "title": "System Reports",
            "description": "View, generate and export system reports for monitoring and analysis.",
            "button": "Generate Custom Report",
            "button_url": "/admin/reports#add-report",
            "list_title": "Reports List",
            "list_description": "View and manage all system reports.",
            "stats": [("Total Reports", len(report_records), "bi-file-earmark-bar-graph", "blue"), ("Generated Today", min(7, len(report_records)), "bi-shield-check", "green"), ("Scheduled Reports", sum(1 for item in report_records if item["status"] == "Scheduled"), "bi-file-earmark", "purple"), ("Total Downloads", 156, "bi-file-earmark-arrow-down", "cyan")],
            "columns": ["No.", "Report Name", "Category", "Description", "Created By", "Last Generated", "Format", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(report_records)} of {len(report_records)} reports",
            "records": report_records,
            "form_anchor": "add-report",
        },
        "audit-logs": {
            "kind": "audit_logs",
            "title": "Audit Logs",
            "description": "Track and review all important activities and changes performed in the system.",
            "button": "Export Logs",
            "button_url": "/admin/audit-logs/export",
            "list_title": "Audit Log List",
            "list_description": "View and manage audit events.",
            "stats": [("Total Logs", len(audit_records), "bi-file-text", "blue"), ("Today's Logs", min(126, len(audit_records)), "bi-person", "green"), ("Create Actions", sum(1 for item in audit_records if item["action"] == "CREATE"), "bi-pencil", "purple"), ("Delete Actions", sum(1 for item in audit_records if item["action"] == "DELETE"), "bi-trash", "red")],
            "columns": ["No.", "Date & Time", "User", "Module", "Action", "Description", "Item / Record", "IP Address", "Status", "Actions"],
            "rows": [],
            "showing": f"Showing 1 to {len(audit_records)} of {len(audit_records)} logs",
            "records": audit_records,
            "form_anchor": "add-audit-log",
        },
    }

    if section == "settings":
        return {
            "kind": "settings",
            "title": "System Settings",
            "description": "Manage and configure system preferences and parameters.",
            "button": "Save Changes",
            "button_url": "/admin/settings#settings-form",
            "stats": [("System Settings", len(system_settings), "bi-gear", "blue"), ("User Roles", len(Role), "bi-shield-check", "green"), ("Total Users", len(users), "bi-people", "purple"), ("Database Size", "256.4 MB", "bi-database", "orange"), ("Last Backup", system_settings.get("last_backup", "-"), "bi-clock", "cyan")],
            "settings": system_settings,
        }

    if section == "outcome-versions" and programs:
        selected_program = session.get(Program, program_id) if program_id else programs[0]
        if not selected_program:
            selected_program = programs[0]
        selected_version, _versions = selected_outcome_version(session, selected_program, version_id)
        version_data = programme_version_data(session, selected_program, selected_version)
        return {
            "kind": "outcome_versions",
            "title": "Outcome Version Management",
            "description": "Create, publish and assign protected PEO/PLO outcome versions for each four-year cohort.",
            "button": "Create Draft Version",
            "button_url": "#create-draft-version",
            "programs": programs,
            "selected_program": selected_program,
            "version_data": version_data,
            "stats": [
                ("Versions", len(version_data["versions"]), "bi-layers", "blue"),
                ("Published", sum(1 for item in version_data["versions"] if item.is_locked), "bi-lock-fill", "green"),
                ("Four-Year Cohorts", len(version_data["cohorts"]), "bi-people", "purple"),
                ("Assigned Cohorts", sum(version_data["cohort_counts"].values()), "bi-link-45deg", "orange"),
            ],
        }

    if section in {"plos", "peos", "targets"} and programs:
        selected_program = session.get(Program, program_id) if program_id else programs[0]
        if not selected_program:
            selected_program = programs[0]
        selected_version, selected_versions = selected_outcome_version(session, selected_program, version_id)
        selected_plo_records = [
            record for record in plo_records
            if record["program_id"] == selected_program.id and record["plo_version_id"] == selected_version.id
        ]
        selected_plos = [
            plo for plo in plos
            if plo.program_id == selected_program.id and plo.plo_version_id == selected_version.id
        ]
        page = configs[section]
        page["plo_versions"] = selected_versions
        page["selected_program"] = selected_program
        page["version_data"] = programme_version_data(session, selected_program, selected_version)
        page["plos"] = selected_plos
        if section == "plos":
            page["records"] = selected_plo_records
            page["button_url"] = f"/admin/plos?program_id={selected_program.id}&version_id={selected_version.id}#add-plo"
            page["stats"] = [
                ("Total PLOs", len(selected_plo_records), "bi-bullseye", "blue"),
                ("Active PLOs", sum(1 for item in selected_plo_records if item["is_active"]), "bi-shield-check", "green"),
                ("Inactive PLOs", sum(1 for item in selected_plo_records if not item["is_active"]), "bi-x-circle", "orange"),
                ("Version", selected_version.version_name, "bi-layers", "purple"),
            ]
            page["showing"] = f"Showing 1 to {len(selected_plo_records)} of {len(selected_plo_records)} PLOs"
        elif section == "peos":
            selected_records = [
                record for record in peo_records
                if record["program_id"] == selected_program.id and record["plo_version_id"] == selected_version.id
            ]
            mapping_plos = [plo for plo in selected_plos if plo.status == "Active"]
            page["records"] = selected_records
            page["mapping_program_id"] = selected_program.id
            page["mapping_version_id"] = selected_version.id
            page["mapping_version_locked"] = selected_version.is_locked
            page["mapping_peos"] = [record for record in selected_records if record["is_active"]]
            page["mapping_plos"] = mapping_plos
            current_mapping_rows = session.exec(select(PEOPLOMapping).where(PEOPLOMapping.plo_version_id == selected_version.id)).all()
            page["mapping_updated_at"] = format_datetimeish(
                max((item.updated_at or item.created_at for item in current_mapping_rows if item.updated_at or item.created_at), default=None)
            )
            page["button_url"] = f"/admin/peos?program_id={selected_program.id}&version_id={selected_version.id}&show_add=1#add-peo"
            page["stats"] = [
                ("Total PEOs", len(selected_records), "bi-bullseye", "blue"),
                ("Active PEOs", sum(1 for item in selected_records if item["is_active"]), "bi-shield-check", "green"),
                ("Inactive PEOs", sum(1 for item in selected_records if not item["is_active"]), "bi-x-circle", "orange"),
                ("Version", selected_version.version_name, "bi-layers", "purple"),
            ]
            page["showing"] = f"Showing 1 to {len(selected_records)} of {len(selected_records)} PEOs"
        else:
            selected_plo_ids = {plo.id for plo in selected_plos}
            selected_records = [
                record for record in target_records
                if record["program_id"] == selected_program.id and record["plo_id"] in selected_plo_ids
            ]
            page["records"] = selected_records
            page["button_url"] = f"/admin/targets?program_id={selected_program.id}&version_id={selected_version.id}#add-target"
            page["stats"] = [
                ("PLOs", len(selected_plos), "bi-bullseye", "blue"),
                ("Targets Set", len(selected_records), "bi-check-circle", "green"),
                ("Academic Years", len(academic_years), "bi-calendar3", "orange"),
                ("Version", selected_version.version_name, "bi-layers", "purple"),
            ]
            page["showing"] = f"Showing 1 to {len(selected_records)} of {len(selected_records)} target setups"
        return page

    return configs.get(section, configs["users"])


def normalize_semester(value: str | int | None) -> str:
    """Normalize semester values stored as int or text.

    Supports 1, 2, "1", "2", "Semester 1", and "Semester 2".
    """
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return text.replace("Semester", "").replace("semester", "").strip()


def course_semester_matches(course: Course, semester: str) -> bool:
    return normalize_semester(course.curriculum_semester) == normalize_semester(semester)


def next_academic_year_name(academic_year: str, from_semester: str, to_semester: str) -> str:
    """Return the target academic year for a promotion.

    Semester 1 -> 2 stays in the same academic year.
    Semester 2 -> 1 moves to the next academic year, e.g. 2025-2026 -> 2026-2027.
    """
    try:
        from_value = int(normalize_semester(from_semester))
        to_value = int(normalize_semester(to_semester))
    except ValueError:
        return academic_year
    if to_value >= from_value:
        return academic_year
    parts = (academic_year or "").split("-")
    if len(parts) == 2 and all(part.strip().isdigit() for part in parts):
        return f"{int(parts[0]) + 1}-{int(parts[1]) + 1}"
    return academic_year


def next_curriculum_year(source_year: int | None, from_semester: str, to_semester: str) -> int | None:
    if source_year is None:
        return None
    try:
        from_value = int(normalize_semester(from_semester))
        to_value = int(normalize_semester(to_semester))
    except ValueError:
        return source_year
    return source_year + 1 if to_value < from_value else source_year


def curriculum_year_from_classes(source_classes: list[CourseClass]) -> int | None:
    years = sorted(
        course_class.course.curriculum_year
        for course_class in source_classes
        if course_class.course and course_class.course.curriculum_year is not None
    )
    return years[0] if years else None


def unique_courses_by_code(courses: list[Course]) -> list[Course]:
    unique: dict[str, Course] = {}
    for course in courses:
        key = course.code.strip().upper()
        if key not in unique:
            unique[key] = course
    return list(unique.values())


def promotion_scoped_faculties_and_programs(session: Session, user: User) -> tuple[list[Faculty], list[Program]]:
    all_faculties = list(session.exec(select(Faculty).order_by(Faculty.name)))
    all_programs = list(session.exec(select(Program).order_by(Program.name)))
    if user.role == Role.SUPER_ADMIN:
        return all_faculties, all_programs
    programs = scoped_programs(user, all_programs)
    faculty_ids = {program.faculty_id for program in programs}
    return [faculty for faculty in all_faculties if faculty.id in faculty_ids], programs



def next_class_code_for_promotion(class_code: str) -> str:
    """Return the next semester class code for codes like 21ME11Mb1."""
    parts = parse_class_code(class_code)
    if not parts:
        return clean_class_code(class_code)
    year_no = int(parts["study_year"])
    semester_no = int(parts["semester_no"])
    if semester_no == 1:
        return class_code_for_semester(class_code, year_no, 2)
    return class_code_for_semester(class_code, year_no + 1, 1)


def class_code_semester(class_code: str, fallback: str = "1") -> str:
    parts = parse_class_code(class_code)
    return parts["semester_no"] if parts else semester_number(fallback)


def class_code_study_year(class_code: str) -> int | None:
    parts = parse_class_code(class_code)
    return int(parts["study_year"]) if parts else None


def database_table_exists(table_name: str) -> bool:
    """Return True when a database table exists. Keeps optional demo tables from causing 500 errors."""
    try:
        return table_name in inspect(engine).get_table_names()
    except Exception:
        return False


def build_student_promotion_page(request: Request, session: Session, user: User) -> dict:
    faculties, programs = promotion_scoped_faculties_and_programs(session, user)
    classes = list(session.exec(select(CourseClass).order_by(CourseClass.academic_year.desc(), CourseClass.name, CourseClass.semester)))
    academic_years, semesters = ensure_academic_records(session, classes)
    histories = list(session.exec(select(StudentPromotionHistory).order_by(StudentPromotionHistory.created_at.desc())))

    query = request.query_params
    scoped_program_ids = {program.id for program in programs if program.id}
    requested_faculty_id = optional_int(query.get("faculty_id"))
    requested_program_id = optional_int(query.get("program_id"))
    available_programs_for_requested_faculty = [
        program for program in programs if requested_faculty_id is None or program.faculty_id == requested_faculty_id
    ]

    if requested_program_id and any(program.id == requested_program_id for program in available_programs_for_requested_faculty):
        selected_program_id = requested_program_id
    else:
        # Default to the user's working programme. Falling back to the first
        # programme by name landed a Dean on whichever programme sorts first
        # (e.g. Computer Science), which has no classes and looked like a bug.
        working_program = (
            manager_program(session, user) if user.role in PROGRAMME_SCOPE_ROLES else None
        )
        if working_program and any(
            program.id == working_program.id for program in available_programs_for_requested_faculty
        ):
            selected_program_id = working_program.id
        else:
            selected_program_id = available_programs_for_requested_faculty[0].id if available_programs_for_requested_faculty else None

    selected_program = session.get(Program, selected_program_id) if selected_program_id else None
    if selected_program and selected_program.id not in scoped_program_ids:
        raise HTTPException(status_code=403)

    selected_faculty_id = requested_faculty_id or (selected_program.faculty_id if selected_program else (faculties[0].id if faculties else None))
    available_programs = [program for program in programs if selected_faculty_id is None or program.faculty_id == selected_faculty_id]

    program_classes = [
        course_class for course_class in classes
        if course_class.course and course_class.course.program_id == selected_program_id
    ]

    # The sidebar Study Period is the single period selector for this page.
    # Ignore the old `study_period` query parameter so the page cannot disagree
    # with the global period selected by the current user.
    global_period = selected_study_period(request, user)
    if global_period is not None:
        selected_academic_year = str(global_period.academic_year or "").strip()
        selected_from_semester = normalize_semester(global_period.semester)
    else:
        available_periods = sorted(
            {
                (str(item.academic_year or "").strip(), normalize_semester(item.semester))
                for item in program_classes
                if str(item.academic_year or "").strip()
            },
            key=lambda item: (item[0], item[1]),
            reverse=True,
        )
        selected_academic_year, selected_from_semester = (
            available_periods[0] if available_periods else ("2025-2026", "1")
        )
    # The chosen Study Period can be one where this programme has no enrolled
    # students yet (the system default is the newest year). Rather than render
    # an empty table, fall back to the most recent period that does have
    # students and tell the user on the page.
    programme_class_ids = {item.id for item in program_classes if item.id}
    populated_periods: set[tuple[str, str]] = set()
    if programme_class_ids:
        populated_ids = {
            row.class_id
            for row in session.exec(
                select(ClassStudent).where(ClassStudent.class_id.in_(programme_class_ids))
            ).all()
        }
        for item in program_classes:
            if item.id in populated_ids:
                populated_periods.add(
                    (str(item.academic_year or "").strip(), normalize_semester(item.semester))
                )

    period_fallback = None
    if populated_periods and (selected_academic_year, selected_from_semester) not in populated_periods:
        fallback_year, fallback_semester = sorted(populated_periods, reverse=True)[0]
        period_fallback = {
            "requested": f"{selected_academic_year} · Semester {selected_from_semester}",
            "used": f"{fallback_year} · Semester {fallback_semester}",
        }
        selected_academic_year, selected_from_semester = fallback_year, fallback_semester

    selected_period = f"{selected_academic_year}|{selected_from_semester}"

    period_classes = [
        item for item in program_classes
        if str(item.academic_year or "").strip() == selected_academic_year
        and normalize_semester(item.semester) == selected_from_semester
    ]

    class_groups: dict[str, dict] = {}
    for course_class in period_classes:
        class_code = clean_class_code(course_class.name)
        if not class_code:
            continue
        year, semester_no = _cohort_year_semester(class_code)
        group = class_groups.setdefault(class_code, {
            "class_code": class_code,
            "year": year,
            "semester": semester_no or selected_from_semester,
            "generation": _cohort_generation(class_code),
            "course_class_ids": [],
            "courses": {},
            "students": {},
        })
        if course_class.id:
            group["course_class_ids"].append(course_class.id)
        if course_class.course:
            group["courses"][course_class.course.id] = course_class.course

    all_ids = [cid for group in class_groups.values() for cid in group["course_class_ids"]]
    if all_ids:
        enrollments = session.exec(select(ClassStudent).where(ClassStudent.class_id.in_(all_ids))).all()
        class_by_id = {cid: code for code, group in class_groups.items() for cid in group["course_class_ids"]}
        for enrollment in enrollments:
            code = class_by_id.get(enrollment.class_id)
            if not code or not enrollment.student:
                continue
            row = class_groups[code]["students"].setdefault(enrollment.student_id, {
                "student": enrollment.student,
                "course_count": 0,
                "status": enrollment.status or "Active",
            })
            row["course_count"] += 1

    # Add a promotion progress status to every class card.  The status is
    # calculated from real target-class enrolments, not only from history rows,
    # so the card remains correct after a partial promotion or a retry.
    for group in class_groups.values():
        group["student_count"] = len(group["students"])
        group["course_count"] = len(group["courses"])
        group["target_class_code"] = next_class_code_for_promotion(group["class_code"])
        group["target_semester"] = class_code_semester(
            group["target_class_code"],
            "2" if normalize_semester(group["semester"]) == "1" else "1",
        )
        group["target_academic_year"] = next_academic_year_name(
            selected_academic_year,
            normalize_semester(group["semester"]),
            group["target_semester"],
        )

        matching_target_classes = [
            course_class for course_class in classes
            if course_class.course
            and course_class.course.program_id == selected_program_id
            and clean_class_code(course_class.name) == group["target_class_code"]
            and str(course_class.academic_year or "").strip() == group["target_academic_year"]
            and normalize_semester(course_class.semester) == normalize_semester(group["target_semester"])
        ]
        matching_target_ids = [item.id for item in matching_target_classes if item.id]
        promoted_student_ids: set[int] = set()
        if matching_target_ids:
            target_rows = session.exec(
                select(ClassStudent).where(ClassStudent.class_id.in_(matching_target_ids))
            ).all()
            promoted_student_ids = {row.student_id for row in target_rows}

        source_student_ids = set(group["students"].keys())
        group["promoted_count"] = len(source_student_ids & promoted_student_ids)
        group["remaining_count"] = max(group["student_count"] - group["promoted_count"], 0)

        if group["student_count"] == 0:
            group["promotion_status"] = "empty"
            group["promotion_status_label"] = "No students"
        elif group["promoted_count"] == 0:
            group["promotion_status"] = "pending"
            group["promotion_status_label"] = "Not promoted"
        elif group["promoted_count"] < group["student_count"]:
            group["promotion_status"] = "partial"
            group["promotion_status_label"] = "Partially promoted"
        else:
            group["promotion_status"] = "complete"
            group["promotion_status_label"] = "Promoted"

    classes_by_year = {year: [] for year in range(1, 5)}
    for group in class_groups.values():
        if group["year"] in classes_by_year:
            classes_by_year[group["year"]].append(group)
    for groups in classes_by_year.values():
        groups.sort(key=lambda item: item["class_code"])

    requested_class_code = clean_class_code(query.get("cohort"))
    selected_cohort = requested_class_code if requested_class_code in class_groups else ""
    if not selected_cohort:
        ordered_groups = [g for year in range(1, 5) for g in classes_by_year[year]]
        # Prefer a class that actually has students; the lowest study year can be
        # an empty shell class in this period, which made the page look broken.
        selected_cohort = next(
            (g["class_code"] for g in ordered_groups if g["student_count"]),
            next((g["class_code"] for g in ordered_groups), ""),
        )

    selected_to_semester = class_code_semester(next_class_code_for_promotion(selected_cohort), "2" if selected_from_semester == "1" else "1")
    target_cohort = next_class_code_for_promotion(selected_cohort) if selected_cohort else ""
    target_academic_year = next_academic_year_name(selected_academic_year, selected_from_semester, selected_to_semester)

    source_group = class_groups.get(selected_cohort)
    source_classes = [item for item in period_classes if clean_class_code(item.name) == selected_cohort]
    preview_by_student = dict(source_group["students"]) if source_group else {}

    target_classes = [
        course_class for course_class in classes
        if course_class.course
        and course_class.course.program_id == selected_program_id
        and clean_class_code(course_class.name) == target_cohort
        and str(course_class.academic_year or "").strip() == target_academic_year
        and normalize_semester(course_class.semester) == normalize_semester(selected_to_semester)
    ]
    target_courses = unique_courses_by_code([course_class.course for course_class in target_classes if course_class.course])

    if preview_by_student and target_classes:
        target_class_ids = [course_class.id for course_class in target_classes if course_class.id]
        target_enrollments = session.exec(select(ClassStudent).where(ClassStudent.class_id.in_(target_class_ids))).all()
        already_student_ids = {enrollment.student_id for enrollment in target_enrollments}
        for student_id in already_student_ids:
            if student_id in preview_by_student:
                preview_by_student[student_id]["already_target"] = True
        for student_id in preview_by_student:
            preview_by_student[student_id].setdefault("already_target", False)
    else:
        for student_id in preview_by_student:
            preview_by_student[student_id]["already_target"] = False

    # Grade-completion dashboard for the selected source class. Load course
    # records by foreign key instead of depending only on an already-loaded
    # relationship. This keeps the dashboard connected to the real CourseClass,
    # Course, CLO, Assessment, CourseTeacher/ClassTeacher and StudentScore data.
    source_courses_by_id: dict[int, Course] = {}
    for course_class in source_classes:
        course = session.get(Course, course_class.course_id) if course_class.course_id else None
        if course and course.id:
            source_courses_by_id[course.id] = course

    # Fallback for databases where classes were created before their course
    # relationships were fully populated. Use the programme curriculum for the
    # selected year and semester, but only when no class-linked courses exist.
    if not source_courses_by_id and selected_program_id:
        source_year = class_code_study_year(selected_cohort)
        curriculum_courses = session.exec(
            select(Course).where(Course.program_id == selected_program_id)
        ).all()
        for course in curriculum_courses:
            if (
                int(course.curriculum_year or 0) == int(source_year or 0)
                and normalize_semester(course.curriculum_semester) == normalize_semester(selected_from_semester)
                and course.id
            ):
                source_courses_by_id[course.id] = course

    source_courses = sorted(
        source_courses_by_id.values(),
        key=lambda course: (course.curriculum_year or 99, normalize_semester(course.curriculum_semester), course.code),
    )
    source_student_ids = set(preview_by_student.keys())
    assessment_rows = []
    for course in source_courses:
        course_assessments = teacher_course_assessments(session, course.id)
        assessment_rows.extend(course_assessments)

    assessment_ids = [item.id for item in assessment_rows if item.id]
    score_records = []
    if assessment_ids and source_student_ids and database_table_exists("studentscore"):
        score_records = list(session.exec(
            select(StudentScore).where(
                StudentScore.assessment_id.in_(assessment_ids),
                StudentScore.student_id.in_(list(source_student_ids)),
            )
        ).all())
    score_lookup = {(item.student_id, item.assessment_id): item for item in score_records}

    course_grade_summary = []
    course_assessment_map = {}
    for course in source_courses:
        assessments_for_course = teacher_course_assessments(session, course.id)
        course_assessment_map[course.id] = assessments_for_course
        expected = len(source_student_ids) * len(assessments_for_course)
        entered_records = [
            score_lookup[(student_id, assessment.id)]
            for student_id in source_student_ids
            for assessment in assessments_for_course
            if (student_id, assessment.id) in score_lookup
        ]
        entered = len(entered_records)
        locked = sum(1 for item in entered_records if bool(item.locked))
        if not assessments_for_course:
            grade_status = "no_assessment"
            grade_status_label = "No assessments"
        elif entered == 0:
            grade_status = "not_started"
            grade_status_label = "Not input"
        elif entered < expected:
            grade_status = "partial"
            grade_status_label = "In progress"
        elif locked < expected:
            grade_status = "complete_draft"
            grade_status_label = "Scores complete"
        else:
            grade_status = "submitted"
            grade_status_label = "Submitted"
        assigned_teachers: list[str] = []

        # The selected cohort's class-level assignment is authoritative. A
        # CourseTeacher row is programme-wide legacy data and must not make a
        # teacher appear in every cohort that offers the same course.
        source_class_ids = [item.id for item in source_classes if item.id and item.course_id == course.id]
        if source_class_ids and database_table_exists("classteacher"):
            for row in session.exec(select(ClassTeacher).where(ClassTeacher.class_id.in_(source_class_ids))).all():
                teacher = session.get(Teacher, row.teacher_id) if row.teacher_id else None
                if not teacher:
                    continue
                teacher_user = session.get(User, teacher.user_id) if teacher.user_id else None
                name = teacher_user.name if teacher_user else (teacher.staff_no or "Teacher")
                if name not in assigned_teachers:
                    assigned_teachers.append(name)

        # Compatibility fallback for old databases that only recorded a
        # course-level assignment. Never mix it with an explicit class-level
        # assignment for the selected cohort.
        if not assigned_teachers and database_table_exists("courseteacher"):
            for row in session.exec(select(CourseTeacher).where(CourseTeacher.course_id == course.id)).all():
                teacher = session.get(Teacher, row.teacher_id) if row.teacher_id else None
                if not teacher:
                    continue
                teacher_user = session.get(User, teacher.user_id) if teacher.user_id else None
                name = teacher_user.name if teacher_user else (teacher.staff_no or "Teacher")
                if name not in assigned_teachers:
                    assigned_teachers.append(name)

        # CLO, assessment, weight and maximum score come directly from the
        # course's Assessment Mapping (CLO -> Assessment). Score completion is
        # calculated only for students in the selected source cohort.
        assessment_details = []
        weight_total = 0.0
        weight_entered = 0.0
        weight_submitted = 0.0
        weight_locked = 0.0
        for assessment in assessments_for_course:
            assessment_scores = [
                score_lookup[(student_id, assessment.id)]
                for student_id in source_student_ids
                if (student_id, assessment.id) in score_lookup
            ]
            assessment_expected = len(source_student_ids)
            assessment_entered = len(assessment_scores)
            assessment_locked = sum(1 for score in assessment_scores if bool(score.locked))
            assessment_submitted = sum(
                1
                for score in assessment_scores
                if bool(score.locked)
                or score.submitted_at is not None
                or str(score.status or "").strip().lower() == "submitted"
            )
            assessment_weight = assessment_weight_percent(assessment)
            entered_ratio = assessment_entered / assessment_expected if assessment_expected else 0.0
            submitted_ratio = assessment_submitted / assessment_expected if assessment_expected else 0.0
            locked_ratio = assessment_locked / assessment_expected if assessment_expected else 0.0
            weighted_entered = assessment_weight * entered_ratio
            weighted_submitted = assessment_weight * submitted_ratio
            weighted_locked = assessment_weight * locked_ratio
            latest_score_at = max(
                (score.updated_at or score.created_at for score in assessment_scores),
                default=None,
            )

            if assessment_expected == 0:
                detail_status = "not_started"
                detail_status_label = "No students"
            elif assessment_entered == 0:
                detail_status = "not_started"
                detail_status_label = "Not input"
            elif assessment_entered < assessment_expected:
                detail_status = "partial"
                detail_status_label = "In progress"
            elif assessment_locked < assessment_expected:
                detail_status = "complete_draft"
                detail_status_label = "Scores complete"
            else:
                detail_status = "submitted"
                detail_status_label = "Submitted"

            clo = session.get(CLO, assessment.clo_id) if assessment.clo_id else None
            assessment_details.append({
                "id": assessment.id,
                "clo_code": clo.code if clo else "-",
                "name": assessment.name,
                "weight": assessment_weight,
                "max_score": float(assessment.max_score or 0),
                "expected": assessment_expected,
                "entered": assessment_entered,
                "submitted": assessment_submitted,
                "locked": assessment_locked,
                "weighted_expected": assessment_weight,
                "weighted_entered": round(weighted_entered, 2),
                "weighted_submitted": round(weighted_submitted, 2),
                "weighted_locked": round(weighted_locked, 2),
                "last_score_input_at": format_datetimeish(latest_score_at),
                "status": detail_status,
                "status_label": detail_status_label,
            })
            weight_total += assessment_weight
            weight_entered += weighted_entered
            weight_submitted += weighted_submitted
            weight_locked += weighted_locked

        course_grade_summary.append({
            "course": course,
            "teacher_names": ", ".join(name for name in assigned_teachers if name) or "Not assigned",
            "assessment_count": len(assessments_for_course),
            "assessment_details": assessment_details,
            "expected": expected,
            "entered": entered,
            "locked": locked,
            "student_count": len(source_student_ids),
            "weight_total": round(weight_total, 2),
            "weight_entered": round(weight_entered, 2),
            "weight_submitted": round(weight_submitted, 2),
            "weight_locked": round(weight_locked, 2),
            "status": grade_status,
            "status_label": grade_status_label,
        })

    for student_id, preview_row in preview_by_student.items():
        course_scores = []
        all_student_courses_complete = True
        for course in source_courses:
            assessments_for_course = course_assessment_map.get(course.id, [])
            entered_count = 0
            locked_count = 0
            weighted_total = 0.0
            for assessment in assessments_for_course:
                score = score_lookup.get((student_id, assessment.id))
                if not score:
                    continue
                entered_count += 1
                if score.locked:
                    locked_count += 1
                max_score = float(assessment.max_score or assessment_weight_percent(assessment) or 100)
                if max_score > 0:
                    weighted_total += (float(score.score) / max_score) * assessment_weight_percent(assessment)
            expected_count = len(assessments_for_course)
            if expected_count == 0:
                cell_status = "no_assessment"
                cell_label = "No assessment"
                all_student_courses_complete = False
            elif entered_count == 0:
                cell_status = "not_started"
                cell_label = "No score"
                all_student_courses_complete = False
            elif entered_count < expected_count:
                cell_status = "partial"
                cell_label = f"{entered_count}/{expected_count} input"
                all_student_courses_complete = False
            elif locked_count < expected_count:
                cell_status = "complete_draft"
                cell_label = "Draft"
                all_student_courses_complete = False
            else:
                cell_status = "submitted"
                cell_label = "Submitted"
            course_scores.append({
                "course_id": course.id,
                "course_code": course.code,
                "course_title": course.title,
                "score": round(weighted_total, 2) if entered_count else None,
                "entered": entered_count,
                "expected": expected_count,
                "locked": locked_count,
                "status": cell_status,
                "status_label": cell_label,
            })
        preview_row["course_scores"] = course_scores
        preview_row["scores_ready"] = all_student_courses_complete

    all_course_scores_ready = bool(source_courses) and bool(source_student_ids) and all(
        item["status"] == "submitted" for item in course_grade_summary
    )

    return {
        "title": "Promote Students",
        "description": "Promote students from the selected study-period class to the next class.",
        "faculties": faculties,
        "programs": available_programs,
        "academic_years": academic_years,
        "semesters": semesters,
        "selected_period": selected_period,
        "period_fallback": period_fallback,
        "selected_academic_year": selected_academic_year,
        "target_academic_year": target_academic_year,
        "selected_faculty_id": selected_faculty_id,
        "selected_program_id": selected_program_id,
        "selected_program": selected_program,
        "selected_cohort": selected_cohort,
        "target_cohort": target_cohort,
        "selected_from_semester": selected_from_semester,
        "selected_to_semester": selected_to_semester,
        "classes_by_year": classes_by_year,
        "source_classes": source_classes,
        "source_curriculum_year": class_code_study_year(selected_cohort),
        "target_curriculum_year": class_code_study_year(target_cohort),
        "target_courses": target_courses,
        "target_classes": target_classes,
        "source_courses": source_courses,
        "course_grade_summary": course_grade_summary,
        "all_course_scores_ready": all_course_scores_ready,
        "preview_rows": sorted(preview_by_student.values(), key=lambda row: row["student"].student_no),
        "history": histories,
        "has_program_classes": bool(program_classes),
        "has_source_classes": bool(source_classes),
    }


def get_or_create_target_class(session: Session, source_class: CourseClass | None, course: Course, cohort_name: str, academic_year: str, semester: str) -> CourseClass:
    existing = session.exec(
        select(CourseClass).where(
            CourseClass.course_id == course.id,
            CourseClass.name == cohort_name,
            CourseClass.academic_year == academic_year,
            CourseClass.semester == semester,
        )
    ).first()
    if existing:
        return existing
    target = CourseClass(
        course_id=course.id,
        name=cohort_name,
        academic_year=academic_year,
        semester=semester,
        semester_start=source_class.semester_end if source_class else None,
        semester_end=None,
    )
    session.add(target)
    session.flush()
    session.refresh(target)
    return target


def promote_selected_students(
    session: Session,
    user: User,
    academic_year: str,
    faculty_id: int | None,
    program_id: int,
    cohort_name: str,
    from_semester: str,
    to_semester: str,
    student_ids: list[int],
) -> tuple[int, int, int]:
    program = session.get(Program, program_id)
    scoped_program_ids = {program.id for program in promotion_scoped_faculties_and_programs(session, user)[1]}
    if not program or program.id not in scoped_program_ids:
        raise HTTPException(status_code=403)
    cohort_name = clean_class_code(cohort_name)
    if not cohort_name:
        return 0, 0, len(student_ids)

    # New cohort system: source cohort is a full class code; target cohort is the next class code.
    if parse_class_code(cohort_name):
        from_semester = class_code_semester(cohort_name, from_semester)
        target_cohort_name = next_class_code_for_promotion(cohort_name)
        to_semester = class_code_semester(target_cohort_name, to_semester)
    else:
        target_cohort_name = cohort_name

    if normalize_semester(from_semester) == normalize_semester(to_semester) and cohort_name == target_cohort_name:
        return 0, 0, len(student_ids)

    target_academic_year = next_academic_year_name(academic_year, from_semester, to_semester)

    all_classes = list(session.exec(select(CourseClass)).all())
    source_classes = [
        course_class
        for course_class in all_classes
        if course_class.course
        and course_class.course.program_id == program_id
        and course_class.name == cohort_name
        and course_class.academic_year == academic_year
        and normalize_semester(course_class.semester) == normalize_semester(from_semester)
    ]
    if not source_classes:
        return 0, 0, len(student_ids)

    # Server-side grade completion guard.  Do not allow a direct POST to
    # bypass the UI: every source course must have assessments and every
    # selected student must have a locked/submitted score for each assessment.
    source_courses_for_guard = unique_courses_by_code([item.course for item in source_classes if item.course])
    if not source_courses_for_guard or not student_ids:
        return 0, 0, len(student_ids)
    for source_course in source_courses_for_guard:
        guard_assessments = teacher_course_assessments(session, source_course.id)
        if not guard_assessments:
            return 0, 0, len(student_ids)
        guard_assessment_ids = [item.id for item in guard_assessments if item.id]
        guard_scores = list(session.exec(
            select(StudentScore).where(
                StudentScore.student_id.in_(student_ids),
                StudentScore.assessment_id.in_(guard_assessment_ids),
            )
        ).all())
        guard_lookup = {(item.student_id, item.assessment_id): item for item in guard_scores}
        for student_id in student_ids:
            for assessment in guard_assessments:
                score = guard_lookup.get((student_id, assessment.id))
                if not score or not bool(score.locked):
                    return 0, 0, len(student_ids)

    # Prefer target classes already created by /admin/cohorts. If missing, create them from curriculum courses.
    target_classes = [
        course_class
        for course_class in all_classes
        if course_class.course
        and course_class.course.program_id == program_id
        and course_class.name == target_cohort_name
        and course_class.academic_year == target_academic_year
        and normalize_semester(course_class.semester) == normalize_semester(to_semester)
    ]
    if not target_classes:
        target_year = class_code_study_year(target_cohort_name) or next_curriculum_year(curriculum_year_from_classes(source_classes), from_semester, to_semester)
        target_courses = [
            course
            for course in session.exec(select(Course).where(Course.program_id == program_id)).all()
            if course_semester_matches(course, to_semester) and (target_year is None or course.curriculum_year == target_year)
        ]
        target_courses = unique_courses_by_code(target_courses)
        target_classes = [get_or_create_target_class(session, source_classes[0], course, target_cohort_name, target_academic_year, to_semester) for course in target_courses]

    if not target_classes:
        return 0, 0, len(student_ids)

    try:
        promoted_student_ids: set[int] = set()
        skipped = 0
        now = datetime.utcnow()
        source_class_ids = [course_class.id for course_class in source_classes if course_class.id]

        for student_id in student_ids:
            existing_target_semester = session.exec(
                select(StudentSemesterEnrollment).where(
                    StudentSemesterEnrollment.student_id == student_id,
                    StudentSemesterEnrollment.program_id == program_id,
                    StudentSemesterEnrollment.cohort_name == target_cohort_name,
                    StudentSemesterEnrollment.academic_year == target_academic_year,
                    StudentSemesterEnrollment.semester == to_semester,
                )
            ).first()
            if existing_target_semester:
                skipped += 1
                continue

            student_source_enrollments = session.exec(
                select(ClassStudent).where(ClassStudent.student_id == student_id, ClassStudent.class_id.in_(source_class_ids))
            ).all()
            if not student_source_enrollments:
                skipped += 1
                continue

            source_semester_enrollment = session.exec(
                select(StudentSemesterEnrollment).where(
                    StudentSemesterEnrollment.student_id == student_id,
                    StudentSemesterEnrollment.program_id == program_id,
                    StudentSemesterEnrollment.cohort_name == cohort_name,
                    StudentSemesterEnrollment.academic_year == academic_year,
                    StudentSemesterEnrollment.semester == from_semester,
                )
            ).first()
            if not source_semester_enrollment:
                source_semester_enrollment = StudentSemesterEnrollment(
                    student_id=student_id,
                    program_id=program_id,
                    cohort_name=cohort_name,
                    academic_year=academic_year,
                    semester=from_semester,
                    status="Promoted",
                    promoted_by_user_id=user.id,
                    promoted_at=now,
                )
            else:
                source_semester_enrollment.status = "Promoted"
                source_semester_enrollment.promoted_by_user_id = user.id
                source_semester_enrollment.promoted_at = now
            session.add(source_semester_enrollment)
            session.flush()

            created_for_student = 0
            first_target_id = None
            for target_class in target_classes:
                existing = session.exec(select(ClassStudent).where(ClassStudent.class_id == target_class.id, ClassStudent.student_id == student_id)).first()
                if existing:
                    continue
                session.add(ClassStudent(class_id=target_class.id, student_id=student_id, status="Active"))
                created_for_student += 1
                first_target_id = target_class.id if first_target_id is None else first_target_id

            if created_for_student:
                session.add(
                    StudentSemesterEnrollment(
                        student_id=student_id,
                        program_id=program_id,
                        cohort_name=target_cohort_name,
                        academic_year=target_academic_year,
                        semester=to_semester,
                        status="Active",
                        promoted_from_id=source_semester_enrollment.id,
                        promoted_by_user_id=user.id,
                        promoted_at=now,
                    )
                )
                promoted_student_ids.add(student_id)
                for enrollment in student_source_enrollments:
                    enrollment.status = "Promoted"
                    enrollment.promoted_to_class_id = first_target_id
                    enrollment.promoted_at = now
                    enrollment.promoted_by_user_id = user.id
                    session.add(enrollment)
            else:
                skipped += 1

        promoted_count = len(promoted_student_ids)
        if promoted_count:
            session.add(
                StudentPromotionHistory(
                    promoted_by_user_id=user.id,
                    promoted_by_name=user.name,
                    academic_year=academic_year,
                    faculty_id=faculty_id,
                    program_id=program_id,
                    cohort_name=f"{cohort_name} → {target_cohort_name}",
                    from_semester=from_semester,
                    to_semester=to_semester,
                    student_count=promoted_count,
                    course_count=len(target_classes),
                    skipped_count=skipped,
                )
            )
            add_audit_record(
                session=session,
                user=user,
                module="Student Promotion",
                action="CREATE",
                description=f"Promoted {promoted_count} students from {cohort_name} to {target_cohort_name}.",
                item_record=f"{cohort_name} → {target_cohort_name}",
            )
        session.commit()
        return promoted_count, len(target_classes), skipped
    except Exception:
        session.rollback()
        raise


def faculty_code(faculty: Faculty | None) -> str:
    if not faculty:
        return "ALL"
    known_codes = {
        "Faculty of Business Administration": "FBA",
        "Faculty of Science and Technology": "FST",
        "Faculty of Agriculture": "FA",
        "Faculty of Art, Humanity and Foreign Language": "FAHFL",
        "Faculty of Social Sciences": "FSS",
    }
    return known_codes.get(faculty.name, "".join(word[0] for word in faculty.name.split()[:4]).upper())


def dean_programmes(session: Session, user: User) -> list[dict]:
    """Programme summary rows for the Dean, with real coordinators and attainment."""
    overview = dean_faculty_overview(session, user)
    return [
        {
            "id": row["program"].id,
            "code": row["code"],
            "name": row["name"],
            "level": row["level"],
            "faculty": row["program"].faculty.name if row["program"].faculty else "Faculty",
            "coordinator": row["coordinator"],
            "students": row["students"],
            "courses": row["courses"],
            "value": row["attainment"],
            "attainment": f"{row['attainment']}%" if row["attainment"] is not None else "No data",
            "status": row["status"],
        }
        for row in overview["programme_rows"]
    ]


def dean_stats(kind: str = "programmes", programme_count: int = 6, plo_count: int = 36) -> list[tuple]:
    stat_sets = {
        "dashboard": [("Total Programmes", programme_count, "bi-people-fill", "blue"), ("Faculty Members", 48, "bi-person-fill", "green"), ("Assessments Conducted", 124, "bi-file-earmark-bar-graph", "purple"), ("PLOs Monitored", plo_count, "bi-bullseye", "orange"), ("Overall PLO Attainment", "78.4%", "bi-graph-up-arrow", "cyan")],
        "profile": [("Total Programmes", programme_count, "bi-mortarboard", "blue"), ("Active Programmes", programme_count, "bi-check-circle", "green"), ("Total Faculty Members", 48, "bi-people-fill", "purple"), ("Total Students", "1,872", "bi-people", "orange"), ("Average PLO Attainment", "78.4%", "bi-graph-up-arrow", "cyan")],
        "members": [("Total Faculty Members", 48, "bi-people-fill", "blue"), ("Full-time Faculty", 38, "bi-person-fill", "green"), ("Part-time Faculty", 10, "bi-person", "purple"), ("Average PLO Attainment", "78.4%", "bi-mortarboard", "orange"), ("Reports This Year", 24, "bi-graph-up-arrow", "cyan")],
        "reports": [("Overall PLO Attainment", "78.4%", "bi-graph-up-arrow", "blue"), ("PLOs On Track", max(0, plo_count - 2), "bi-bullseye", "green"), ("PLOs At Risk", 2, "bi-exclamation-triangle", "orange"), ("PLOs Below Target", 0, "bi-arrow-down", "red"), ("Reports Generated", 24, "bi-file-earmark-bar-graph", "purple")],
        "programmes": [("Total Programmes", programme_count, "bi-grid-fill", "blue"), ("Active Programmes", programme_count, "bi-check-circle", "green"), ("Inactive Programmes", 0, "bi-pause-circle", "purple"), ("Average PLO Attainment", "78.4%", "bi-bar-chart", "orange"), ("Programme Reports", 24, "bi-file-earmark-text", "cyan")],
        "mapping": [("Total PLOs", 7, "bi-diagram-3", "blue"), ("Fully mapped", 6, "bi-check-circle", "green"), ("Partially Mapped", 1, "bi-diagram-2", "orange"), ("Not Mapped", 0, "bi-x-circle", "red"), ("Last Updated", "May 14, 2025", "bi-file-earmark-bar-graph", "purple")],
        "targets": [("Total PLOs", 7, "bi-bullseye", "purple"), ("Targets Set", "7 (100%)", "bi-bullseye", "green"), ("Average Target", "70.0%", "bi-graph-up-arrow", "orange"), ("Last Updated", "May 14, 2025", "bi-calendar3", "blue"), ("Target Based On", "Historical + Benchmark", "bi-file-earmark-text", "purple")],
        "announcements": [("Total Announcements", 24, "bi-megaphone", "blue"), ("Published", 18, "bi-check-circle", "green"), ("Scheduled", 4, "bi-clock", "orange"), ("Draft", 2, "bi-clipboard", "purple"), ("Expired", 5, "bi-x-circle", "red")],
        "documents": [("Total Documents", 156, "bi-folder", "blue"), ("Published", 112, "bi-file-earmark-check", "green"), ("Draft", 18, "bi-clock", "orange"), ("Archived", 21, "bi-archive", "purple"), ("Expired", 5, "bi-file-earmark-x", "red")],
        "calendar": [("Total Events", 42, "bi-calendar3", "blue"), ("Upcoming Events", 12, "bi-calendar-check", "green"), ("Meetings", 8, "bi-people", "orange"), ("Deadlines", 10, "bi-clipboard2-check", "red"), ("This Month", 18, "bi-calendar2-week", "purple")],
        "audit": [("Total Activities", "3,842", "bi-list-ul", "blue"), ("Users", 54, "bi-person-fill", "green"), ("Today's Activities", 42, "bi-calendar2-check", "purple"), ("Critical Events", 7, "bi-activity", "orange"), ("Success Rate", "99.8%", "bi-shield-check", "cyan")],
    }
    return stat_sets.get(kind, stat_sets["programmes"])


def percent_label(value: float | None) -> str:
    return "No data" if value is None else f"{value}%"


def dean_attainment_page(session: Session, user: User, focus: str, program_id: int | None) -> dict:
    """Faculty-wide PLO/PEO attainment computed from real student scores."""
    programs = scoped_programs(user, session.exec(select(Program).order_by(Program.code)).all())

    def version_for(program: Program) -> PLOVersion:
        selected, _versions = selected_outcome_version(session, program, None)
        return selected

    summary = faculty_attainment(session, programs, version_for)
    entries = summary["programmes"]
    selected_entry = next(
        (item for item in entries if item["program"].id == program_id),
        next((item for item in entries if item[focus]["has_data"]), entries[0] if entries else None),
    )

    if focus == "peo":
        average = summary["peo_average"]
        achieved = sum(item["peo"]["achieved"] for item in entries)
        measured = sum(item["peo"]["measured_count"] for item in entries)
        stats = [
            ("Faculty Average PEO Attainment", percent_label(average), "bi-bullseye", "blue"),
            ("Programmes Reporting", f"{summary['peo_reporting_count']} / {summary['programme_count']}", "bi-mortarboard", "purple"),
            ("PEOs Achieved", achieved, "bi-check-circle", "green"),
            ("PEOs Below Target", max(0, measured - achieved), "bi-exclamation-triangle", "orange"),
            ("Total PEOs", summary["peo_total"], "bi-list-check", "cyan"),
        ]
        title = "PEO Performance & Attainment"
        description = "PEO attainment per programme and the faculty average, rolled up from PLO attainment."
    else:
        average = summary["plo_average"]
        stats = [
            ("Faculty Average PLO Attainment", percent_label(average), "bi-bullseye", "blue"),
            ("Programmes Reporting", f"{summary['plo_reporting_count']} / {summary['programme_count']}", "bi-mortarboard", "purple"),
            ("PLOs On Track", summary["on_track"], "bi-check-circle", "green"),
            ("PLOs At Risk", summary["at_risk"], "bi-exclamation-triangle", "orange"),
            ("PLOs Below Target", summary["below_target"], "bi-arrow-down", "red"),
        ]
        title = "PLO Performance & Attainment"
        description = "PLO attainment per programme and the faculty average, computed from entered student scores."

    return {
        "kind": "attainment",
        "focus": focus,
        "title": title,
        "description": description,
        "stats": stats,
        "summary": summary,
        "faculty_average": average,
        "selected_entry": selected_entry,
        "target": ATTAINMENT_DEFAULT_TARGET,
    }


def dean_faculty_overview(session: Session, user: User) -> dict:
    """Real faculty data for the Dean pages, on the default outcome version."""

    def version_for(program: Program) -> PLOVersion:
        selected, _versions = selected_outcome_version(session, program, None)
        return selected

    return faculty_overview(session, user, version_for)


def dean_overview_stats(kind: str, overview: dict) -> list[tuple]:
    """Metric cards built from real counts instead of fixed sample numbers."""
    counts = overview["counts"]
    summary = overview["summary"]
    plo_average = percent_label(summary["plo_average"])
    peo_average = percent_label(summary["peo_average"])
    reporting = f"{summary['plo_reporting_count']} / {summary['programme_count']}"
    measured = [row["attainment"] for row in overview["assessments"] if row["attainment"] is not None]
    assessment_average = round(sum(measured) / len(measured), 1) if measured else None
    stat_sets = {
        "profile": [
            ("Programmes", counts["programmes"], "bi-mortarboard", "blue"),
            ("Staff Members", counts["members"], "bi-people-fill", "green"),
            ("Students", counts["students"], "bi-person", "purple"),
            ("Courses", counts["courses"], "bi-journal-text", "orange"),
            ("Average PLO Attainment", plo_average, "bi-graph-up-arrow", "cyan"),
        ],
        "members": [
            ("Total Staff", counts["members"], "bi-people-fill", "blue"),
            ("Lecturers", counts["lecturers"], "bi-person-fill", "green"),
            ("Programme Coordinators", counts["coordinators"], "bi-person-badge", "purple"),
            ("Courses Covered", counts["courses"], "bi-journal-text", "orange"),
            ("Programmes", counts["programmes"], "bi-mortarboard", "cyan"),
        ],
        "reports": [
            ("Average PLO Attainment", plo_average, "bi-graph-up-arrow", "blue"),
            ("Average PEO Attainment", peo_average, "bi-bullseye", "cyan"),
            ("PLOs On Track", summary["on_track"], "bi-check-circle", "green"),
            ("PLOs At Risk", summary["at_risk"], "bi-exclamation-triangle", "orange"),
            ("Programmes Reporting", reporting, "bi-mortarboard", "purple"),
        ],
        "assessment": [
            ("Total Assessments", counts["assessments"], "bi-clipboard-check", "blue"),
            ("With Scores Entered", counts["assessments_scored"], "bi-check-circle", "green"),
            ("Awaiting Scores", counts["assessments_pending"], "bi-clock", "orange"),
            ("Courses", counts["courses"], "bi-journal-text", "purple"),
            ("Average Score", percent_label(assessment_average), "bi-bar-chart", "cyan"),
        ],
        "programmes": [
            ("Programmes", counts["programmes"], "bi-grid-fill", "blue"),
            ("Reporting Attainment", reporting, "bi-check-circle", "green"),
            ("Total PLOs", counts["plos"], "bi-bullseye", "purple"),
            ("Total PEOs", counts["peos"], "bi-diagram-3", "orange"),
            ("Average PLO Attainment", plo_average, "bi-graph-up-arrow", "cyan"),
        ],
    }
    return stat_sets.get(kind, stat_sets["programmes"])


def dean_overview_programmes(overview: dict) -> list[dict]:
    """Flatten the faculty overview into the shape the dashboard tables expect."""
    return [
        {
            "id": row["program"].id,
            "code": row["code"],
            "name": row["name"],
            "coordinator": row["coordinator"],
            "students": row["students"],
            "courses": row["courses"],
            "value": row["attainment"],
            "attainment": f"{row['attainment']}%" if row["attainment"] is not None else "No data",
            "status": row["status"],
        }
        for row in overview["programme_rows"]
    ]


# Distinct colours per programme for the bar and radar charts.
PROGRAMME_COLOURS = [
    "#2563eb", "#16a34a", "#f59e0b", "#a855f7",
    "#0891b2", "#dc2626", "#db2777", "#65a30d",
]


def _radar_data(entries: list[dict], key: str) -> dict:
    """Radar axes and one polygon per programme for the given outcome type."""
    codes: list[str] = []
    for entry in entries:
        for row in entry[key]["rows"]:
            if row["code"] not in codes:
                codes.append(row["code"])
    codes.sort(key=outcome_code_sort_key)

    # Every programme in the faculty appears, so the Dean always sees the full
    # list. Programmes with nothing measured carry has_data=False; the chart
    # lists them in the legend but plots no polygon, because drawing them at
    # zero would claim their students scored zero.
    series = []
    for entry in entries:
        by_code = {row["code"]: row["attainment"] for row in entry[key]["rows"]}
        target_by_code = {row["code"]: row["target"] for row in entry[key]["rows"]}
        points = [by_code.get(code) for code in codes]
        # Targets are per outcome, so the chart ring follows each PLO/PEO's own
        # target rather than a single faculty-wide number.
        targets = [target_by_code.get(code) for code in codes]
        measured_targets = [value for value in targets if value is not None]
        series.append(
            {
                "label": entry["code"],
                "name": entry["name"],
                "colour": entry["colour"],
                "points": points,
                "targets": targets,
                "mean_target": round(sum(measured_targets) / len(measured_targets), 1) if measured_targets else None,
                "has_data": any(value is not None for value in points),
            }
        )
    # Averaged across the programmes that actually measured each outcome, so a
    # programme with no scores does not drag the faculty average toward zero.
    averages = []
    for index in range(len(codes)):
        values = [item["points"][index] for item in series if item["points"][index] is not None]
        averages.append(round(sum(values) / len(values), 1) if values else None)
    target_averages = []
    for index in range(len(codes)):
        values = [item["targets"][index] for item in series if item["targets"][index] is not None]
        target_averages.append(round(sum(values) / len(values), 1) if values else None)
    # A programme that defines no targets of its own shows the faculty target on
    # each axis, so every card carries the same goal line and the shapes stay
    # comparable. Only the plotted values differ (0 where nothing is measured).
    for item in series:
        item["targets"] = [
            value if value is not None else target_averages[index]
            for index, value in enumerate(item["targets"])
        ]
        own = [value for value in item["targets"] if value is not None]
        item["mean_target"] = round(sum(own) / len(own), 1) if own else None

    measured_targets = [value for value in target_averages if value is not None]
    average_series = {
        "label": "Faculty average",
        "name": "Faculty average",
        "colour": "#0b2355",
        "points": averages,
        "targets": target_averages,
        "mean_target": round(sum(measured_targets) / len(measured_targets), 1) if measured_targets else None,
        "has_data": any(value is not None for value in averages),
    }
    return {
        "codes": codes,
        "series": series,
        "averages": averages,
        "target_averages": target_averages,
        "average_series": average_series,
    }


def dean_attainment_detail_page(session: Session, user: User, focus: str) -> dict:
    """Faculty attainment page: charts, comparison table and CQI actions.

    Always covers every programme in the Dean's own faculty, and carries both
    the PLO and PEO profiles so either page shows the full outcome picture.
    """
    programmes = faculty_programmes(session, user)
    default_target = system_attainment_target(session)

    def version_for(program: Program) -> PLOVersion:
        selected, _versions = selected_outcome_version(session, program, None)
        return selected

    summary = faculty_attainment(session, programmes, version_for, None, default_target)
    entries = summary["programmes"]
    for index, entry in enumerate(entries):
        entry["colour"] = PROGRAMME_COLOURS[index % len(PROGRAMME_COLOURS)]

    key = "peo" if focus == "peo" else "plo"
    other_key = "plo" if key == "peo" else "peo"
    average = summary["peo_average"] if focus == "peo" else summary["plo_average"]
    other_average = summary["plo_average"] if focus == "peo" else summary["peo_average"]

    measured = [item for item in entries if item[key]["overall"] is not None]
    highest = max(measured, key=lambda item: item[key]["overall"], default=None)
    lowest = min(measured, key=lambda item: item[key]["overall"], default=None)

    radar = {"plo": _radar_data(entries, "plo"), "peo": _radar_data(entries, "peo")}

    comparison = []
    for entry in entries:
        for row in entry[key]["rows"]:
            gap = None if row["attainment"] is None else round(row["attainment"] - (average or 0), 1)
            comparison.append(
                {
                    "programme_code": entry["code"],
                    "programme_name": entry["name"],
                    "colour": entry["colour"],
                    "code": row["code"],
                    "description": row["description"],
                    "attainment": row["attainment"],
                    "target": row["target"],
                    "status": "Achieved" if row["attainment"] is not None and row["attainment"] >= row["target"] else ("No Data" if row["attainment"] is None else "Below Target"),
                    "vs_faculty": gap,
                    "cqi_required": row["attainment"] is None or row["attainment"] < row["target"],
                }
            )

    below_target = sum(1 for row in comparison if row["status"] == "Below Target")
    unit = "PEO" if focus == "peo" else "PLO"
    other_unit = "PLO" if unit == "PEO" else "PEO"

    return {
        "kind": "attainment_detail",
        "focus": focus,
        "unit": unit,
        "other_unit": other_unit,
        "other_key": other_key,
        "title": f"{unit} Attainment",
        "description": f"{unit} attainment for every programme in your faculty, with programme comparison and CQI actions.",
        "stats": [
            ("Programmes", len(entries), "bi-mortarboard", "blue"),
            (f"Faculty Average {unit}", percent_label(average), "bi-graph-up-arrow", "cyan"),
            ("Highest Programme", f"{highest['code']} · {highest[key]['overall']}%" if highest else "No data", "bi-arrow-up-circle", "green"),
            ("Lowest Programme", f"{lowest['code']} · {lowest[key]['overall']}%" if lowest else "No data", "bi-arrow-down-circle", "orange"),
            (f"{unit}s Below Target", below_target, "bi-exclamation-triangle", "red"),
        ],
        "summary": summary,
        "entries": entries,
        "faculty_average": average,
        "other_average": other_average,
        "default_target": default_target,
        "radar": radar,
        "comparison": comparison,
        "cqi": cqi_report(entries, focus),
        "has_any_data": any(entry[key]["has_data"] for entry in entries),
        "has_other_data": any(entry[other_key]["has_data"] for entry in entries),
    }


def outcome_code_sort_key(code: str) -> tuple:
    digits = "".join(character for character in str(code) if character.isdigit())
    return (int(digits) if digits else 0, str(code))


def system_attainment_target(session: Session) -> float:
    setting = session.exec(
        select(SystemSetting).where(SystemSetting.key == "attainment_target")
    ).first()
    try:
        return float(str(setting.value).strip()) if setting and setting.value else ATTAINMENT_DEFAULT_TARGET
    except (TypeError, ValueError):
        return ATTAINMENT_DEFAULT_TARGET


def build_dean_page(section: str, session: Session, user: User, program_id: int | None = None) -> dict:
    if section == "plo-performance":
        page = dean_attainment_page(session, user, "plo", program_id)
    elif section == "peo-performance":
        page = dean_attainment_page(session, user, "peo", program_id)
    else:
        page = None
    if page is not None:
        faculty = session.get(Faculty, user.faculty_id) if user.faculty_id else None
        programmes = dean_programmes(session, user)
        page.update(
            {
                "faculty_name": faculty.name if faculty else "All Faculties",
                "faculty_code": faculty_code(faculty),
                "programmes": programmes,
                "programme_count": len(programmes),
            }
        )
        return page
    faculty = session.get(Faculty, user.faculty_id) if user.faculty_id else None
    # One pass over the faculty data; the programme rows are derived from it.
    overview = dean_faculty_overview(session, user)
    programmes = dean_overview_programmes(overview)
    programme_count = len(programmes)
    pages = {
        "profile": {"kind": "profile", "title": "Faculty Profile", "description": f"{faculty.name if faculty else 'Assigned faculty'} profile and academic portfolio.", "stats": dean_overview_stats("profile", overview), "overview": overview},
        "members": {"kind": "members", "title": "Faculty Members", "description": "Staff attached to this faculty and the programmes under it.", "stats": dean_overview_stats("members", overview), "overview": overview},
        "faculty-reports": {"kind": "reports", "title": "Faculty Level Reports", "description": "Faculty-level PLO attainment computed from entered student scores.", "stats": dean_overview_stats("reports", overview), "overview": overview},
        "assessment-reports": {"kind": "assessment", "title": "Assessment Reports", "description": "Assessment score coverage and attainment across the faculty.", "stats": dean_overview_stats("assessment", overview), "overview": overview},
        "programmes": {"kind": "programmes", "title": "Programmes", "description": "Programmes under your faculty, with their real PLO attainment.", "stats": dean_overview_stats("programmes", overview), "overview": overview},
        "announcements": {"kind": "list", "title": "Announcements", "description": "Create and manage faculty announcements.", "stats": dean_stats("announcements"), "button": "Create Announcement"},
        "documents": document_page_data(session, user, "/dean/documents"),
        "calendar": {"kind": "calendar", "title": "Calendar", "description": "Faculty academic calendar and events.", "stats": dean_stats("calendar")},
        "audit-logs": {"kind": "audit", "title": "Audit Logs", "description": "Track important faculty activities and changes.", "stats": dean_stats("audit")},
    }
    page = pages.get(section, pages["programmes"])
    page.update(
        {
            "faculty_name": faculty.name if faculty else "All Faculties",
            "faculty_code": faculty_code(faculty),
            "programmes": programmes,
            "programme_count": programme_count,
        }
    )
    return page


PROGRAMME_SCOPE_ROLES = {Role.PROGRAM_MANAGER, Role.DEAN}


def require_programme_scope(user: User) -> None:
    """Guard the /manager pages, which Deans share with Programme Managers."""
    if user.role not in PROGRAMME_SCOPE_ROLES:
        raise HTTPException(status_code=403)


def faculty_programmes(session: Session, user: User) -> list[Program]:
    """Every programme a Dean may work on, ordered for the switcher."""
    if user.faculty_id is None:
        return []
    return list(
        session.exec(
            select(Program).where(Program.faculty_id == user.faculty_id).order_by(Program.code)
        ).all()
    )


def active_programme_preference(session: Session, user: User) -> int | None:
    if not user.id:
        return None
    preference = session.exec(
        select(UserProgrammePreference).where(UserProgrammePreference.user_id == user.id)
    ).first()
    return preference.program_id if preference else None


def set_active_programme(session: Session, user: User, program: Program) -> None:
    """Remember the programme a Dean switched to, so it survives navigation."""
    if not user.id or program.id is None:
        return
    preference = session.exec(
        select(UserProgrammePreference).where(UserProgrammePreference.user_id == user.id)
    ).first()
    if preference:
        preference.program_id = program.id
        preference.updated_at = datetime.utcnow()
    else:
        preference = UserProgrammePreference(user_id=user.id, program_id=program.id)
    session.add(preference)
    session.commit()


def manager_program(session: Session, user: User) -> Program:
    """The programme the /manager pages act on for this user.

    Programme Managers always get their own programme. Deans get whichever
    programme in their faculty they last switched to, defaulting to the first.
    """
    if user.role == Role.DEAN:
        programmes = faculty_programmes(session, user)
        if not programmes:
            raise HTTPException(status_code=404, detail="No programmes exist in this faculty")
        preferred_id = active_programme_preference(session, user)
        return next((item for item in programmes if item.id == preferred_id), programmes[0])
    program = session.get(Program, user.program_id) if user.program_id else None
    if not program:
        program = session.exec(select(Program).where(Program.code == "ME")).first()
    if not program:
        program = session.exec(select(Program)).first()
    if not program:
        raise HTTPException(status_code=404)
    return program


def ensure_programme_outcome_versions(session: Session, program: Program) -> list[PLOVersion]:
    """Create a safe legacy version and attach old unversioned PEO/PLO rows."""
    versions = list(session.exec(select(PLOVersion).where(PLOVersion.programme_id == program.id).order_by(PLOVersion.id)).all())
    if not versions:
        legacy = PLOVersion(programme_id=program.id, version_name="V1", status="Active", is_locked=False)
        session.add(legacy)
        session.flush()
        versions = [legacy]
    default_version = versions[0]
    changed = False
    for plo in session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == None)).all():
        plo.plo_version_id = default_version.id
        session.add(plo); changed = True
    for peo in session.exec(select(PEO).where(PEO.program_id == program.id, PEO.plo_version_id == None)).all():
        peo.plo_version_id = default_version.id
        session.add(peo); changed = True
    for link in session.exec(select(PEOPLOMapping).where(PEOPLOMapping.program_id == program.id, PEOPLOMapping.plo_version_id == None)).all():
        link.plo_version_id = default_version.id
        session.add(link); changed = True
    if changed:
        session.commit()
    return list(session.exec(select(PLOVersion).where(PLOVersion.programme_id == program.id).order_by(PLOVersion.id.desc())).all())


def selected_outcome_version(session: Session, program: Program, version_id: int | None = None) -> tuple[PLOVersion, list[PLOVersion]]:
    versions = ensure_programme_outcome_versions(session, program)
    selected = session.get(PLOVersion, version_id) if version_id else None
    if not selected or selected.programme_id != program.id:
        selected = next((item for item in versions if item.status == "Active"), versions[0])
    return selected, versions


def editable_outcome_version(session: Session, program: Program, version_id: int | None) -> PLOVersion:
    """Return the requested programme version only when CQI work may still edit it."""
    version = session.get(PLOVersion, version_id) if version_id else None
    if not version or version.programme_id != program.id:
        raise HTTPException(status_code=400, detail="A valid outcome version is required.")
    if version.is_locked or version.status in {"Published", "Retired"}:
        raise HTTPException(status_code=409, detail="Published outcome versions are read-only. Copy it to a new Draft version for CQI changes.")
    return version


def programme_version_data(session: Session, program: Program, selected: PLOVersion) -> dict:
    versions = ensure_programme_outcome_versions(session, program)
    enrollments = session.exec(select(StudentSemesterEnrollment).where(StudentSemesterEnrollment.program_id == program.id)).all()
    classes = [
        item for item in session.exec(select(CourseClass)).all()
        if item.course and item.course.program_id == program.id
    ]
    cohort_samples: dict[str, str] = {}
    for raw_name in [
        *(str(item.cohort_name).strip() for item in enrollments if item.cohort_name),
        *(str(item.name).strip() for item in classes if item.name),
    ]:
        family = cohort_family_key(raw_name)
        if family:
            cohort_samples.setdefault(family, raw_name)
    assignments = session.exec(select(CohortOutcomeVersion).where(CohortOutcomeVersion.programme_id == program.id)).all()
    assignment_by_cohort: dict[str, CohortOutcomeVersion] = {}
    for item in assignments:
        key = outcome_cohort_key(item.cohort_name)
        if key:
            assignment_by_cohort.setdefault(key, item)
            cohort_samples.setdefault(key, item.cohort_name)
    counts = {version.id: 0 for version in versions}
    for item in assignment_by_cohort.values():
        counts[item.outcome_version_id] = counts.get(item.outcome_version_id, 0) + 1
    cohorts = [
        {"key": key, "label": outcome_cohort_label(sample)}
        for key, sample in cohort_samples.items()
        if key
    ]
    cohorts.sort(key=lambda item: item["label"])
    return {
        "versions": versions,
        "selected": selected,
        "cohorts": cohorts,
        "cohort_names": sorted({sample for sample in cohort_samples.values() if sample}),
        "assignments": assignment_by_cohort,
        "cohort_counts": counts,
    }


def manager_plos(session: Session, user: User, version_id: int | None = None) -> list[dict]:
    """PLOs for the active programme with attainment from real student scores.

    `value` stays numeric (0 when nothing has been assessed) so existing bar
    widths keep working; check `has_data` before presenting it as a result.
    """
    program = manager_program(session, user)
    selected, _versions = selected_outcome_version(session, program, version_id)
    result = programme_plo_attainment(session, program, selected.id)
    return [
        {
            "id": row["plo"].id,
            "code": row["code"],
            "description": row["description"],
            "value": row["attainment"] if row["attainment"] is not None else 0,
            "attainment": row["attainment"],
            "target": row["target"],
            "students_assessed": row["students_assessed"],
            "students_meeting_target": row["students_meeting_target"],
            "mapped_clos": row["mapped_clos"],
            "has_data": row["has_data"],
            "status": row["status"],
        }
        for row in result["rows"]
    ]


def manager_plo_management_data(session: Session, user: User, version_id: int | None = None) -> dict:
    program = manager_program(session, user)
    selected_version, versions = selected_outcome_version(session, program, version_id)
    rows = sorted(
        session.exec(
            select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == selected_version.id)
        ).all(),
        key=plo_sort_key,
    )
    return {
        "program": program,
        "versions": versions,
        "selected_version": selected_version,
        "rows": rows,
        "editable": not selected_version.is_locked and selected_version.status not in {"Published", "Retired"},
        "domains": ["Knowledge", "Skills", "Attitude"],
        "bloom_levels": ["C1", "C2", "C3", "C4", "C5", "C6", "P1", "P2", "P3", "P4", "P5", "A1", "A2", "A3", "A4", "A5"],
    }


def course_mapping_percent(mapping: CoursePLOMapping | None) -> int:
    if not mapping:
        return 0
    if mapping.level > 3:
        return max(0, min(100, mapping.level))
    return {0: 0, 1: 30, 2: 60, 3: 100}.get(mapping.level, 0)


def clamp_percent(value: str | int | float | None) -> float:
    try:
        percent = float(value if value not in (None, "") else 0)
    except (TypeError, ValueError):
        percent = 0
    return max(0, min(100, percent))


def stored_percent(value: float | int | None) -> float:
    if not value:
        return 0
    if 0 < value <= 1:
        return round(float(value) * 100, 2)
    return round(clamp_percent(value), 2)


def assessment_weight_percent(assessment: Assessment) -> float:
    if assessment.weight == 1 and assessment.max_score != 1:
        return round(clamp_percent(assessment.max_score), 2)
    return stored_percent(assessment.weight)


def course_teacher_names(session: Session, course_id: int | None) -> str:
    if not course_id:
        return "Not assigned"
    assignments = session.exec(select(CourseTeacher).where(CourseTeacher.course_id == course_id)).all()
    names = [assignment.teacher.user.name for assignment in assignments if assignment.teacher and assignment.teacher.user]
    return ", ".join(names) if names else "Not assigned"


def manager_courses(
    session: Session,
    user: User | None,
    plos: list[dict] | None = None,
    program: Program | None = None,
) -> list[dict]:
    """Return programme course mappings for either Manager or scoped Admin views."""
    program = program or (manager_program(session, user) if user else None)
    if not program:
        raise HTTPException(status_code=404, detail="Programme not found")
    if plos is None:
        if not user:
            raise HTTPException(status_code=400, detail="PLO version is required")
        plos = manager_plos(session, user)
    # Programme management must use only base curriculum courses. Cohort course
    # copies have the same course code but separate ids and assignments; mixing
    # them here made Manager show "Not assigned" while Teacher used the assigned
    # base course with the same visible code.
    raw_courses = sorted(
        session.exec(
            select(Course).where(
                Course.program_id == program.id,
                Course.cohort_id == None,  # noqa: E711
            )
        ).all(),
        key=lambda item: (item.curriculum_year or 99, item.curriculum_semester or "", item.code, item.id or 0),
    )
    course_by_code: dict[str, Course] = {}
    for course in raw_courses:
        current = course_by_code.get(course.code)
        if not current or (len(course.clos), course.id or 0) > (len(current.clos), current.id or 0):
            course_by_code[course.code] = course
    if program.code == "ME":
        official_codes = [code for _year, _semester, code, _title, _credits in ME_CURRICULUM]
        official_code_set = set(official_codes)
        courses = [course_by_code[code] for code in official_codes if code in course_by_code]
        # Always show every custom course. Course management must not depend on
        # whether a Course-PLO mapping has already been created.
        courses.extend(
            sorted(
                [
                    course
                    for code, course in course_by_code.items()
                    if code not in official_code_set
                ],
                key=lambda item: (item.curriculum_year or 99, item.curriculum_semester or "", item.code),
            )
        )
    else:
        courses = sorted(course_by_code.values(), key=lambda item: (item.curriculum_year or 99, item.curriculum_semester or "", item.code))
    course_rows = []
    for course in courses:
        # Course-level PLO weights are derived from the sum of all CLO-PLO
        # mappings for this course. This keeps /manager/programme-mapping
        # consistent with /manager/course-assessment-setup.
        stored_mappings = session.exec(
            select(CoursePLOMapping).where(CoursePLOMapping.course_id == course.id)
        ).all()
        mapping_by_plo_id = {mapping.plo_id: mapping for mapping in stored_mappings}
        totals_by_plo_id: dict[int, float] = {int(plo["id"]): 0.0 for plo in plos}

        clo_records = session.exec(select(CLO).where(CLO.course_id == course.id)).all()
        for clo in clo_records:
            clo_mappings = session.exec(
                select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)
            ).all()
            for clo_mapping in clo_mappings:
                if clo_mapping.plo_id in totals_by_plo_id:
                    totals_by_plo_id[clo_mapping.plo_id] += stored_percent(clo_mapping.weight)

        levels = []
        changed = False
        for plo in plos:
            plo_id = int(plo["id"])
            percent = round(totals_by_plo_id.get(plo_id, 0.0), 2)
            symbol = f"{percent:g}%" if percent else ""
            levels.append({"level": percent, "percent": percent, "symbol": symbol})

            # Synchronize the derived value into CoursePLOMapping so reports and
            # any other pages that read this table do not use stale values.
            mapping = mapping_by_plo_id.get(plo_id)
            if mapping:
                if round(float(mapping.level or 0), 2) != percent or (mapping.symbol or "") != symbol:
                    mapping.level = percent
                    mapping.symbol = symbol
                    session.add(mapping)
                    changed = True
            else:
                session.add(
                    CoursePLOMapping(
                        course_id=course.id,
                        plo_id=plo_id,
                        level=percent,
                        symbol=symbol,
                    )
                )
                changed = True

        if changed:
            session.commit()

        mapped_count = sum(1 for item in levels if item["percent"] > 0)
        course_rows.append(
            {
                "id": course.id,
                "code": course.code,
                "title": course.title,
                "year": course.curriculum_year or "-",
                "semester": course.curriculum_semester or "-",
                "credits": course.credits,
                "lecturer": course_teacher_names(session, course.id),
                "clos": len(course.clos),
                "mapping": "Mapped" if mapped_count >= max(1, len(plos) // 2) else "Partial" if mapped_count else "Not Mapped",
                "levels": levels,
            }
        )
    return course_rows


def manager_selected_course_data(session: Session, user: User, plos: list[dict], course_id: int | None = None) -> dict:
    program = manager_program(session, user)
    courses = session.exec(
        select(Course)
        .where(
            Course.program_id == program.id,
            Course.cohort_id == None,  # noqa: E711
        )
        .order_by(Course.curriculum_year, Course.curriculum_semester, Course.code)
    ).all()
    selected = session.get(Course, course_id) if course_id else None
    if not selected or selected.program_id != program.id or selected.cohort_id is not None:
        selected = courses[0] if courses else None
    if not selected:
        return {"selected": None, "courses": [], "clos": [], "levels": [], "mapping_by_plo": {}}
    mapping_by_plo = {
        mapping.plo_id: mapping
        for mapping in session.exec(select(CoursePLOMapping).where(CoursePLOMapping.course_id == selected.id)).all()
    }

    # Course-level PLO weights are derived from the real CLO-PLO mappings.
    # Example: CLO1->PLO2 50% + CLO2->PLO2 15% = Course PLO2 65%.
    clo_records = session.exec(select(CLO).where(CLO.course_id == selected.id).order_by(CLO.code)).all()
    course_plo_totals: dict[int, float] = {int(plo["id"]): 0.0 for plo in plos}
    for clo_item in clo_records:
        for clo_mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo_item.id)).all():
            if clo_mapping.plo_id in course_plo_totals:
                course_plo_totals[clo_mapping.plo_id] += stored_percent(clo_mapping.weight)

    levels = []
    mappings_changed = False
    for plo in plos:
        plo_id = int(plo["id"])
        percent = round(course_plo_totals.get(plo_id, 0.0), 2)
        levels.append({"plo": plo, "level": percent, "percent": percent, "symbol": f"{percent:g}%" if percent else ""})

        # Keep CoursePLOMapping synchronized so reports and other pages use the
        # same derived values instead of stale manually entered values.
        mapping = mapping_by_plo.get(plo_id)
        symbol = f"{percent:g}%" if percent else ""
        if mapping:
            if round(float(mapping.level or 0), 2) != percent or (mapping.symbol or "") != symbol:
                mapping.level = percent
                mapping.symbol = symbol
                session.add(mapping)
                mappings_changed = True
        else:
            mapping = CoursePLOMapping(course_id=selected.id, plo_id=plo_id, level=percent, symbol=symbol)
            session.add(mapping)
            mapping_by_plo[plo_id] = mapping
            mappings_changed = True

    if mappings_changed:
        session.commit()

    clos = []
    assessment_rows = []
    for clo in clo_records:
        clo_mappings = {
            mapping.plo_id: stored_percent(mapping.weight)
            for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all()
        }
        assessments = session.exec(select(Assessment).where(Assessment.clo_id == clo.id).order_by(Assessment.id)).all()
        weights = [clo_mappings.get(plo["id"], 0) for plo in plos]
        assessment_items = [
            {
                "id": assessment.id,
                "name": assessment.name,
                "max_score": assessment.max_score,
                "weight": assessment_weight_percent(assessment),
                "description": assessment.description or "",
            }
            for assessment in assessments
        ]
        plo_total = round(sum(weights), 2)
        clo_target = round(max(weights, default=0), 2)
        assessment_total = round(sum(item["weight"] for item in assessment_items), 2)
        row = {
            "clo": clo,
            "weights": weights,
            "assessments": assessment_items,
            "plo_total": plo_total,
            "clo_target": clo_target,
            "assessment_total": assessment_total,
            "is_balanced": abs(clo_target - assessment_total) < 0.01,
        }
        clos.append(row)
        for assessment in row["assessments"]:
            assessment_rows.append({"clo": clo, "weights": row["weights"], "assessment": assessment})
    return {
        "selected": selected,
        "courses": courses,
        "teacher_names": course_teacher_names(session, selected.id),
        "clos": clos,
        "assessment_rows": assessment_rows,
        "levels": levels,
        "mapping_by_plo": mapping_by_plo,
    }


def manager_teacher_assignment_data(
    session: Session,
    user: User,
    course_id: int | None = None,
    study_period: StudyPeriod | None = None,
) -> dict:
    program = manager_program(session, user)
    courses = session.exec(
        select(Course)
        .where(
            Course.program_id == program.id,
            Course.cohort_id == None,  # noqa: E711
        )
        .order_by(Course.curriculum_year, Course.curriculum_semester, Course.code)
    ).all()
    selected = session.get(Course, course_id) if course_id else None
    if not selected or selected.program_id != program.id or selected.cohort_id is not None:
        selected = courses[0] if courses else None
    teachers = session.exec(select(Teacher).join(User).where(User.role == Role.TEACHER).order_by(User.name)).all()
    classes: list[CourseClass] = []
    assignments: list[ClassTeacher] = []
    assigned_ids: set[int] = set()
    if selected:
        classes = [
            item
            for item in session.exec(select(CourseClass).where(CourseClass.course_id == selected.id)).all()
            if class_matches_study_period(item, study_period)
        ]
        classes.sort(key=lambda item: item.name)
        class_ids = [item.id for item in classes if item.id]
        assignments = list(
            session.exec(select(ClassTeacher).where(ClassTeacher.class_id.in_(class_ids))).all()
        ) if class_ids else []
        assignments.sort(key=lambda item: (item.course_class.name if item.course_class else "", item.teacher.user.name if item.teacher and item.teacher.user else ""))
        assigned_ids = {assignment.teacher_id for assignment in assignments}
    return {
        "courses": courses,
        "selected": selected,
        "teachers": teachers,
        "classes": classes,
        "assignments": assignments,
        "assigned_ids": assigned_ids,
        "study_period": study_period,
    }


def peo_page_data(session: Session, user: User, version_id: int | None = None, read_only: bool = False) -> dict:
    # Deans resolve through manager_program() too, so the PEO page follows the
    # programme they switched to rather than always the first in the faculty.
    program = manager_program(session, user) if user.role in PROGRAMME_SCOPE_ROLES else None
    if user.role == Role.SUPER_ADMIN:
        program = session.exec(select(Program).where(Program.code == "ME")).first() or session.exec(select(Program)).first()
    if not program:
        raise HTTPException(status_code=404)
    selected_version, _versions = selected_outcome_version(session, program, version_id)
    plos = sorted(session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == selected_version.id)).all(), key=plo_sort_key)
    peos = list(session.exec(select(PEO).where(PEO.program_id == program.id, PEO.plo_version_id == selected_version.id).order_by(PEO.code)).all())
    links = session.exec(select(PEOPLOMapping).where(PEOPLOMapping.plo_version_id == selected_version.id)).all()
    link_weights = {(link.peo_id, link.plo_id): stored_percent(link.contribution_percentage) for link in links}
    rows = []
    for peo in peos:
        weight_map = {plo.id: link_weights.get((peo.id, plo.id), 0) for plo in plos if plo.id is not None}
        mapped_plos = [{"plo": plo, "weight": weight_map.get(plo.id, 0)} for plo in plos if weight_map.get(plo.id, 0) > 0]
        weights = [item["weight"] for item in mapped_plos]
        result = round(sum(weights) / len(weights), 2) if weights else 0
        rows.append({"peo": peo, "mapped_plos": mapped_plos, "weight_map": weight_map, "result": result, "status": "Achieved" if result >= 70 else "Not Achieved"})
    return {"program": program, "plos": plos, "rows": rows, "read_only": read_only}


def manager_stats(kind: str, plos: list[dict] | None = None, courses: list[dict] | None = None, peos: list | None = None) -> list[tuple]:
    plo_count = len(plos or [])
    course_count = len(courses or [])
    mapped_count = sum(1 for course in (courses or []) if course.get("mapping") == "Mapped")
    peo_count = len(peos or [])
    stat_sets = {
        "performance": [("Overall PLO Attainment", "70.0%", "bi-bullseye", "purple"), ("PLOs Achieved Target", "4 / 7", "bi-pie-chart", "blue"), ("Highest Attainment", "PLO7", "bi-graph-up-arrow", "cyan"), ("Lowest Attainment", "PLO6", "bi-arrow-counterclockwise", "red"), ("Last Updated", "May 14, 2025", "bi-calendar3", "orange"), ("Data Source", "18", "bi-file-earmark-bar-graph", "purple")],
        "mapping": [("Total PLOs", plo_count, "bi-diagram-3", "purple"), ("Fully Mapped", mapped_count, "bi-check-circle", "green"), ("Partially Mapped", max(course_count - mapped_count, 0), "bi-share", "orange"), ("Not Mapped", sum(1 for course in (courses or []) if course.get("mapping") == "Not Mapped"), "bi-x-circle", "red"), ("Last Updated", "ME Specification", "bi-file-earmark-arrow-up", "purple")],
        "outcome_versions": [("Programme", "Scoped", "bi-diagram-3", "blue"), ("PLOs", plo_count, "bi-bullseye", "purple"), ("Courses", course_count, "bi-journal-text", "orange"), ("Source", "Database", "bi-database", "green")],
        "peos": [("Total PEOs", peo_count, "bi-bullseye", "blue"), ("Mapped PEOs", peo_count, "bi-diagram-3", "green"), ("Programme PLOs", plo_count, "bi-list-check", "purple"), ("Average Result", "78.0%", "bi-graph-up-arrow", "orange")],
        "plos": [("Total PLOs", plo_count, "bi-bullseye", "blue"), ("Active PLOs", sum(1 for item in (plos or []) if item.get("status") == "Active"), "bi-check-circle", "green"), ("Courses", course_count, "bi-journal-text", "purple"), ("Versioned", "Yes", "bi-layers", "orange")],
        "targets": [("Total PLOs", 7, "bi-bullseye", "purple"), ("Targets Set", "7 / 7", "bi-check-circle", "green"), ("Average Target", "70.0%", "bi-graph-up-arrow", "blue"), ("Above Current Baseline", 4, "bi-arrow-up-right", "cyan"), ("Needs Attention", 2, "bi-exclamation-triangle", "orange"), ("Last Updated", "May 14, 2025", "bi-calendar3", "purple")],
        "courses": [("Total Courses", course_count, "bi-journal-text", "purple"), ("Active Courses", course_count, "bi-check-circle", "green"), ("This Semester", sum(1 for course in (courses or []) if str(course.get("semester")) == "1"), "bi-calendar3", "blue"), ("Total Credits", sum(float(course.get("credits") or 0) for course in (courses or [])), "bi-layers", "orange"), ("Courses Mapped to PLOs", mapped_count, "bi-diagram-3", "cyan"), ("Last Updated", "ME Specification", "bi-calendar3", "purple")],
        "assessments": [("Total Assessments", 18, "bi-calendar-check", "purple"), ("This Semester", 8, "bi-calendar3", "blue"), ("Scheduled", 5, "bi-clock", "orange"), ("Completed", 9, "bi-check-circle", "green"), ("Courses Assessed", 12, "bi-book", "cyan"), ("Last Updated", "May 14, 2025", "bi-calendar3", "purple")],
        "documents": [("Total Documents", 156, "bi-file-earmark-text", "blue"), ("Published", 112, "bi-check-circle", "green"), ("Draft", 18, "bi-pencil", "orange"), ("Archived", 21, "bi-archive", "purple"), ("Templates", 15, "bi-file-earmark-richtext", "blue"), ("Last Updated", "May 14, 2025", "bi-calendar3", "blue")],
        "announcements": [("Total Announcements", 24, "bi-megaphone", "purple"), ("Published", 16, "bi-check-circle", "green"), ("Scheduled", 4, "bi-calendar-event", "orange"), ("Drafts", 3, "bi-pencil", "orange"), ("Archived", 1, "bi-archive", "purple"), ("Last Updated", "May 14, 2025", "bi-calendar3", "blue")],
        "reports": [("Total Reports", 24, "bi-file-earmark-text", "blue"), ("Generated This Semester", 8, "bi-calendar-check", "green"), ("Scheduled Reports", 5, "bi-clock", "purple"), ("Pending Review", 3, "bi-hourglass-split", "orange"), ("Exports", 42, "bi-download", "cyan"), ("Last Generated", "May 14, 2025", "bi-calendar3", "blue")],
        "users": [("Total Users", 28, "bi-people", "blue"), ("Active Users", 24, "bi-person-check", "green"), ("Lecturers", 16, "bi-mortarboard", "purple"), ("Assessors", 5, "bi-people-fill", "orange"), ("Programme Staff", 4, "bi-briefcase", "cyan"), ("Pending Invites", 3, "bi-envelope", "orange")],
        "audit": [("Total Activities", 245, "bi-list-task", "blue"), ("Users", 28, "bi-people", "green"), ("Today", 10, "bi-calendar-check", "purple"), ("Create Actions", 52, "bi-plus-circle", "green"), ("Success Rate", "99.8%", "bi-shield-check", "cyan")],
    }
    return stat_sets.get(kind, [])


def manager_support_data(session: Session, user: User, program: Program, courses: list[dict]) -> dict:
    course_ids = [int(course["id"]) for course in courses if course.get("id")]
    course_rows = session.exec(select(Course).where(Course.id.in_(course_ids))).all() if course_ids else []
    clo_ids = [clo.id for course in course_rows for clo in course.clos if clo.id]
    assessments = session.exec(select(Assessment).where(Assessment.clo_id.in_(clo_ids)).order_by(Assessment.id.desc())).all() if clo_ids else []
    clos_by_id = {clo.id: clo for course in course_rows for clo in course.clos if clo.id}
    courses_by_id = {course.id: course for course in course_rows if course.id}
    users = session.exec(select(User).where(User.program_id == program.id).order_by(User.role, User.name)).all()
    reports = session.exec(select(SystemReport).order_by(SystemReport.id.desc())).all()
    settings = {item.key: item.value for item in session.exec(select(SystemSetting)).all()}
    now_label = format_datetimeish(datetime.utcnow())
    calendar_rows = []
    for assessment in assessments[:20]:
        clo = clos_by_id.get(assessment.clo_id)
        course = courses_by_id.get(clo.course_id) if clo else None
        calendar_rows.append(
            {
                "title": assessment.name,
                "category": "Assessment",
                "owner": course.code if course else "-",
                "status": "Configured",
                "updated_at": f"Assessment #{assessment.id}",
            }
        )
    return {
        "calendar": calendar_rows,
        "announcements": [
            {"title": "PLO target review", "category": "Academic", "owner": user.name, "status": "Published", "updated_at": now_label},
            {"title": "Assessment score entry reminder", "category": "Assessment", "owner": user.name, "status": "Published", "updated_at": now_label},
        ],
        "reports": [
            {"title": report.name, "category": report.category, "owner": report.created_by, "status": report.status, "updated_at": report.last_generated}
            for report in reports[:20]
        ],
        "users": [
            {
                "title": item.name,
                "category": ROLE_LABELS.get(item.role, item.role.value),
                "owner": item.email,
                "status": "Active" if item.is_active else "Inactive",
                "updated_at": program.code,
            }
            for item in users
        ],
        "settings": [
            {"title": key.replace("_", " ").title(), "category": "Setting", "owner": program.code, "status": value, "updated_at": "Current"}
            for key, value in sorted(settings.items())
            if key in {"system_name", "institution_name", "academic_year", "default_semester", "attainment_target", "passing_score", "email_notifications", "in_app_notifications"}
        ],
        "settings_values": settings,
        "activities": [
            {"title": report.name, "category": "Report", "owner": report.created_by, "status": report.status, "updated_at": report.last_generated}
            for report in reports[:5]
        ],
    }


def manager_plo_target_data(session: Session, user: User, target_id: int | None = None, version_id: int | None = None) -> dict:
    """Load programme PLO target records for the manager page."""
    program = manager_program(session, user)
    selected_version, _versions = selected_outcome_version(session, program, version_id)
    plos = list(
        session.exec(
            select(PLO)
            .where(PLO.program_id == program.id, PLO.plo_version_id == selected_version.id)
            .order_by(PLO.code)
        ).all()
    )
    plo_by_id = {plo.id: plo for plo in plos if plo.id is not None}
    targets = list(
        session.exec(
            select(PLOTarget)
            .where(PLOTarget.program_id == program.id)
            .order_by(PLOTarget.academic_year.desc(), PLOTarget.cohort, PLOTarget.id)
        ).all()
    )
    rows = [
        {"target": target, "plo": plo_by_id.get(target.plo_id)}
        for target in targets
        if plo_by_id.get(target.plo_id) is not None
    ]

    study_period_records = list(
        session.exec(
            select(StudyPeriod).order_by(StudyPeriod.academic_year.desc(), StudyPeriod.semester)
        ).all()
    )
    academic_years = sorted(
        {str(item.academic_year).strip() for item in study_period_records if item.academic_year}
        | {str(item.academic_year).strip() for item in targets if item.academic_year},
        reverse=True,
    )
    cohort_values = {str(item.cohort).strip() for item in targets if item.cohort}
    cohort_values.update(
        str(item.cohort_name).strip()
        for item in session.exec(
            select(StudentSemesterEnrollment).where(StudentSemesterEnrollment.program_id == program.id)
        ).all()
        if item.cohort_name
    )
    cohorts = sorted(value for value in cohort_values if value)

    selected = None
    if target_id:
        target = session.get(PLOTarget, target_id)
        if target and target.program_id == program.id and target.plo_id in plo_by_id:
            selected = {"target": target, "plo": plo_by_id[target.plo_id]}

    return {
        "program": program,
        "plos": plos,
        "rows": rows,
        "academic_years": academic_years,
        "cohorts": cohorts,
        "selected": selected,
    }


def build_manager_page(
    section: str,
    session: Session,
    user: User,
    course_id: int | None = None,
    version_id: int | None = None,
    study_period: StudyPeriod | None = None,
) -> dict:
    kind_map = {
        "outcome-versions": "outcome_versions",
        "plo-performance": "performance",
        "programme-mapping": "mapping",
        "plo-targets": "targets",
        "plo-target-setup": "targets",
        "peos": "peos",
        "peo-management": "peos",
        "plo-management": "plos",
        "courses": "courses",
        "course-create": "course_create",
        "course-mapping": "course_mapping",
        "assign-teachers": "assign_teachers",
        "documents": "documents",
        "calendar": "calendar",
        "announcements": "announcements",
        "reports": "reports",
        "users": "users",
        "settings": "settings",
        "course-assessment-setup": "course_setup",
    }
    titles = {
        "outcome-versions": ("Outcome Version Management", "Create CQI draft outcome versions and manage approved programme outcomes."),
        "plo-performance": ("PLO Performance", "Monitor and evaluate the attainment of Programme Learning Outcomes (PLOs)."),
        "programme-mapping": ("Programme Mapping", "Define and manage the mapping between courses and Programme Learning Outcomes (PLOs)."),
        "plo-targets": ("PLO Target Setup", "Set and review target attainment levels for Programme Learning Outcomes (PLOs)."),
        "plo-target-setup": ("PLO Target Setup", "Set and review target attainment levels for Programme Learning Outcomes (PLOs)."),
        "peos": ("PEO Management", "Create PEOs, map them to PLOs, and view PEO results for your programme."),
        "peo-management": ("PEO Management", "Create PEOs, map them to PLOs, and view PEO results for your programme."),
        "plo-management": ("PLO Management", "Create, edit, and archive PLOs for your programme outcome version."),
        "courses": ("Courses", "Manage courses, lecturers, CLO coverage, and curriculum alignment for your programme."),
        "course-create": ("Create Course", "Create or update course information for the selected programme."),
        "course-mapping": ("Course Mapping", "Manage course-to-PLO mapping for the selected programme."),
        "assign-teachers": ("Assign Teachers", "Assign one or more teachers to each course in your programme."),
        "assessments": ("Assessments", "Manage assessment plans, methods, schedules, and attainment tracking for your programme."),
        "documents": ("Documents", "Manage programme documents, templates, evidence files, and supporting materials."),
        "calendar": ("Calendar", "View and manage all activities, assessments, deadlines and events."),
        "announcements": ("Announcements", "Manage programme announcements, notices, reminders, and communication updates."),
        "reports": ("Reports", "Generate, review, and export programme-level reports and attainment summaries."),
        "users": ("Users", "Manage lecturers, assessors, staff, and programme-related user access."),
        "audit-logs": ("Audit Logs", "Track all system activities and changes."),
        "settings": ("Settings", "Manage system settings and preferences."),
        "course-assessment-setup": ("Assessment Mapping", "Define course information, CLOs, assessment weights, and mapping to PLOs."),
    }
    title, description = titles[section]
    kind = kind_map[section]
    program = manager_program(session, user)
    selected_version, _versions = selected_outcome_version(session, program, version_id)
    plos = manager_plos(session, user, selected_version.id)
    courses = manager_courses(session, user, plos)
    support_data = manager_support_data(session, user, program, courses)
    peo_data = peo_page_data(session, user, selected_version.id) if kind == "peos" else None
    plo_management_data = manager_plo_management_data(session, user, selected_version.id) if kind == "plos" else None
    stats = manager_stats(kind, plos, courses, peo_data["rows"] if peo_data else None)
    if kind in {"calendar", "announcements", "reports", "users", "settings"}:
        support_rows = support_data.get(kind, [])
        active_count = sum(1 for row in support_rows if row.get("status") in {"Active", "Published", "Ready", "Configured"})
        stats = [
            ("Records", len(support_rows), "bi-list-check", "blue"),
            ("Active", active_count, "bi-check-circle", "green"),
            ("Programme", program.code, "bi-diagram-3", "purple"),
            ("Source", "Database", "bi-database", "orange"),
        ]
    target_data = manager_plo_target_data(session, user, course_id, selected_version.id) if kind == "targets" else None
    course_data = manager_selected_course_data(session, user, plos, course_id) if kind in {"courses", "course_create", "course_mapping", "course_setup"} else None
    assignment_data = manager_teacher_assignment_data(session, user, course_id, study_period) if kind in {"assign_teachers", "course_mapping"} else None
    page = {
        "kind": kind,
        "title": title,
        "description": description,
        "stats": stats,
        "plos": plos,
        "courses": courses,
        "program": program,
        "version_data": programme_version_data(session, program, selected_version),
        "support_data": support_data,
    }
    if peo_data:
        page["peo_data"] = peo_data
    if plo_management_data:
        page["plo_management_data"] = plo_management_data
    if target_data:
        page["target_data"] = target_data
    if course_data:
        # The Courses page opens in Create mode unless a course_id was explicitly
        # selected from the Edit action. Other course pages keep their default
        # selected-course behavior.
        if kind == "courses" and course_id is None:
            course_data["selected"] = None
        page["course_data"] = course_data
    if assignment_data:
        page["assignment_data"] = assignment_data
    return page


def teacher_profile_for_user(session: Session, user: User) -> Teacher | None:
    return session.exec(select(Teacher).where(Teacher.user_id == user.id)).first()


def class_matches_study_period(course_class: CourseClass, study_period: StudyPeriod | None) -> bool:
    if study_period is None:
        return True
    return (
        str(course_class.academic_year or "").strip() == str(study_period.academic_year or "").strip()
        and normalize_semester(course_class.semester) == normalize_semester(study_period.semester)
    )


def teacher_assigned_classes(session: Session, user: User, study_period: StudyPeriod | None) -> list[CourseClass]:
    """Return only classes this teacher may access in the selected Study Period.

    ClassTeacher is authoritative. Once this teacher has any explicit class
    assignment, only exact assignments in the selected period are returned.
    CourseTeacher is a compatibility fallback only for an old database where
    the teacher has no class-level assignments at all.
    """
    teacher = teacher_profile_for_user(session, user)
    if not teacher or not teacher.id:
        return []

    period_classes = [
        item
        for item in session.exec(select(CourseClass)).all()
        if item.id and class_matches_study_period(item, study_period)
    ]
    if not period_classes:
        return []

    # Once a teacher has at least one class-level assignment, ClassTeacher is
    # authoritative for every Study Period. Checking only the selected period
    # caused a teacher with no assignment in that period to fall back to the
    # legacy CourseTeacher table. A course offered to two cohorts was then
    # expanded into two false classes (for example, 7 courses became 14).
    explicit_assignments = list(
        session.exec(select(ClassTeacher).where(ClassTeacher.teacher_id == teacher.id)).all()
    )
    if explicit_assignments:
        period_class_ids = {item.id for item in period_classes if item.id}
        explicit_class_ids = {
            assignment.class_id
            for assignment in explicit_assignments
            if assignment.class_id in period_class_ids
        }
        assigned = [
            course_class for course_class in period_classes
            if course_class.id in explicit_class_ids
        ]
    else:
        legacy_course_ids = {
            assignment.course_id
            for assignment in session.exec(
                select(CourseTeacher).where(CourseTeacher.teacher_id == teacher.id)
            ).all()
        }
        assigned = [
            course_class
            for course_class in period_classes
            if course_class.course_id in legacy_course_ids
            and is_full_class_code(course_class.name)
        ]
    return sorted(
        assigned,
        key=lambda item: (
            item.academic_year,
            normalize_semester(item.semester),
            item.name,
            item.course.code if item.course else "",
        ),
    )


def teacher_assigned_courses(session: Session, user: User, study_period: StudyPeriod | None = None) -> list[Course]:
    teacher = teacher_profile_for_user(session, user)
    if not teacher or not teacher.id:
        return []
    courses_by_id: dict[int, Course] = {}
    if study_period is not None:
        for course_class in teacher_assigned_classes(session, user, study_period):
            if course_class.course and course_class.course.id:
                courses_by_id[course_class.course.id] = course_class.course
    else:
        for assignment in session.exec(select(CourseTeacher).where(CourseTeacher.teacher_id == teacher.id)).all():
            if assignment.course and assignment.course.id:
                courses_by_id[assignment.course.id] = assignment.course
        for assignment in session.exec(select(ClassTeacher).where(ClassTeacher.teacher_id == teacher.id)).all():
            if assignment.course_class and assignment.course_class.course and assignment.course_class.course.id:
                courses_by_id[assignment.course_class.course.id] = assignment.course_class.course
    return sorted(
        courses_by_id.values(),
        key=lambda course: (course.curriculum_year or 99, course.curriculum_semester or "", course.code),
    )


def course_student_count(session: Session, course_id: int | None) -> int:
    if not course_id:
        return 0
    classes = session.exec(select(CourseClass).where(CourseClass.course_id == course_id)).all()
    return sum(len(course_class.students) for course_class in classes)


def course_assessment_count(session: Session, course_id: int | None) -> int:
    if not course_id:
        return 0
    clos = session.exec(select(CLO).where(CLO.course_id == course_id)).all()
    assessment_count = 0
    for clo in clos:
        assessment_count += len(clo.assessments)
    return assessment_count


def teacher_courses(session: Session, user: User, study_period: StudyPeriod | None) -> list[dict]:
    assigned_classes = teacher_assigned_classes(session, user, study_period)
    icons = ["bi-database", "bi-diagram-3", "bi-code-slash", "bi-book", "bi-cpu"]
    tones = ["blue", "green", "orange", "purple", "cyan"]
    records = []
    for index, course_class in enumerate(assigned_classes):
        course = course_class.course
        if not course or not course.id:
            continue
        student_rows = teacher_course_student_rows(session, course.id, course_class.id)
        assessment_rows = teacher_course_assessments(session, course.id)
        student_ids = {item["id"] for item in student_rows}
        assessment_ids = [item.id for item in assessment_rows if item.id]
        score_count = len(
            session.exec(
                select(StudentScore).where(
                    StudentScore.student_id.in_(student_ids),
                    StudentScore.assessment_id.in_(assessment_ids),
                )
            ).all()
        ) if student_ids and assessment_ids else 0
        students = len(student_rows)
        assessments = len(assessment_rows)
        expected_scores = students * assessments
        pending = max(expected_scores - score_count, 0)
        attainment = round(
            sum(float(item["percent"]) for item in student_rows) / students,
            2,
        ) if students else 0
        records.append(
            {
                "id": course.id,
                "class_id": course_class.id,
                "class_name": course_class.name,
                "academic_year": course_class.academic_year,
                "study_semester": normalize_semester(course_class.semester),
                "code": course.code,
                "title": course.title,
                "icon": icons[index % len(icons)],
                "students": students,
                "pending": pending,
                "submitted": score_count,
                "assessment_count": assessments,
                "attainment": f"{attainment}%",
                "tone": tones[index % len(tones)],
                "credits": course.credits,
                "semester": course.curriculum_semester or "-",
            }
        )
    return records


def teacher_stats(kind: str) -> list[tuple]:
    stat_sets = {
        "courses": [("Total Courses", 3, "bi-book", "blue"), ("Total Students", 96, "bi-people-fill", "green"), ("Total Assessments", 15, "bi-clipboard-check", "orange"), ("Pending Score Input", 28, "bi-clipboard-data", "purple")],
        "assessments": [("Total Assessments", 5, "bi-clipboard-check", "blue"), ("Pending Scores", 18, "bi-hourglass-split", "orange"), ("Students", 56, "bi-people", "purple"), ("PLO Attainment", "82%", "bi-graph-up-arrow", "green")],
        "students": [("Total Students", 56, "bi-people-fill", "blue"), ("Active Students", 54, "bi-person-check", "green"), ("Inactive Students", 2, "bi-person-x", "orange"), ("Class Average", "76.45", "bi-mortarboard", "purple")],
        "reports": [("Total Students", 56, "bi-people-fill", "blue"), ("Total Assessments", 5, "bi-clipboard-check", "green"), ("PLOs Mapped", 11, "bi-diagram-3", "purple"), ("Overall Attainment", "76.45%", "bi-file-earmark-bar-graph", "orange"), ("PLOs Achieved", "6 / 11", "bi-check2-square", "cyan")],
        "announcements": [("Total Announcements", 12, "bi-megaphone", "blue"), ("Unread", 5, "bi-send", "green"), ("Pinned", 3, "bi-pin-angle", "orange"), ("This Week", 2, "bi-calendar3", "purple")],
        "documents": [("Total Documents", 48, "bi-folder", "blue"), ("Total Size", "2.45 GB", "bi-file-earmark", "green"), ("Downloads", 256, "bi-download", "purple"), ("Recently Added", 8, "bi-eye", "orange"), ("Shared Documents", 12, "bi-people", "cyan")],
    }
    return stat_sets.get(kind, [])


def teacher_stats_for_courses(courses: list[dict]) -> list[tuple]:
    return [
        ("Assigned Classes", len(courses), "bi-book", "blue"),
        ("Class Enrolments", sum(int(course.get("students") or 0) for course in courses), "bi-people-fill", "green"),
        ("Total Assessments", sum(int(course.get("assessment_count") or 0) for course in courses), "bi-clipboard-check", "orange"),
        ("Pending Score Input", sum(int(course.get("pending") or 0) for course in courses), "bi-clipboard-data", "purple"),
    ]


def teacher_course_classes(
    session: Session,
    course_id: int | None,
    user: User | None = None,
    study_period: StudyPeriod | None = None,
) -> list[CourseClass]:
    if not course_id:
        return []
    if user is not None:
        classes = [
            item
            for item in teacher_assigned_classes(session, user, study_period)
            if item.course_id == course_id
        ]
    else:
        classes = list(session.exec(select(CourseClass).where(CourseClass.course_id == course_id)).all())
    return sorted(
        classes,
        key=lambda item: (item.academic_year, normalize_semester(item.semester), item.name),
    )


def teacher_student_plo_averages(
    session: Session,
    course_id: int | None,
    student_ids: list[int],
    plo_ids: list[int],
) -> dict:
    """Average raw score per PLO for each student, from this course's own CLOs.

    A student's CLO score is the marks they earned on that CLO's assessments;
    each PLO is those CLO scores weighted by the CLO-PLO mapping. Values are
    marks, not percentages. Students with no marks for a PLO get None rather
    than 0, so an unmarked assessment does not read as a zero score.
    """
    plo_ids = [pid for pid in plo_ids if pid is not None]
    empty = {
        "per_student": {sid: {pid: None for pid in plo_ids} for sid in student_ids},
        "student_total": {sid: None for sid in student_ids},
        "plo_average": {pid: None for pid in plo_ids},
        "total": None,
    }
    if not course_id or not student_ids or not plo_ids:
        return empty

    clos = list(session.exec(select(CLO).where(CLO.course_id == course_id)).all())
    clo_ids = [clo.id for clo in clos if clo.id is not None]
    if not clo_ids:
        return empty

    assessments = list(session.exec(select(Assessment).where(Assessment.clo_id.in_(clo_ids))).all())
    assessment_clo = {item.id: item.clo_id for item in assessments if item.id is not None}
    assessment_ids = [item.id for item in assessments if item.id is not None]
    scores = (
        list(session.exec(select(StudentScore).where(StudentScore.assessment_id.in_(assessment_ids))).all())
        if assessment_ids
        else []
    )
    totals: dict[tuple[int, int], float] = defaultdict(float)
    seen: set[tuple[int, int]] = set()
    for row in scores:
        clo_id = assessment_clo.get(row.assessment_id)
        if clo_id is None or row.student_id not in set(student_ids):
            continue
        totals[(clo_id, row.student_id)] += float(row.score or 0)
        seen.add((clo_id, row.student_id))

    # Raw marks earned per CLO, keyed the same way as before.
    clo_score: dict[int, dict[int, float]] = defaultdict(dict)
    for (clo_id, student_id) in seen:
        clo_score[clo_id][student_id] = totals[(clo_id, student_id)]

    mappings = list(session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id.in_(clo_ids))).all())
    by_plo: dict[int, list] = defaultdict(list)
    for mapping in mappings:
        if mapping.plo_id in plo_ids:
            by_plo[mapping.plo_id].append(mapping)

    per_student: dict[int, dict[int, float | None]] = {sid: {} for sid in student_ids}
    for plo_id in plo_ids:
        for student_id in student_ids:
            weighted = 0.0
            weight_total = 0.0
            for mapping in by_plo.get(plo_id, []):
                weight = normalized_mapping_weight(float(mapping.weight or 0))
                if weight <= 0:
                    continue
                value = clo_score.get(mapping.clo_id, {}).get(student_id)
                if value is None:
                    continue
                weighted += value * weight
                weight_total += weight
            per_student[student_id][plo_id] = round(weighted / weight_total, 1) if weight_total > 0 else None

    # Total score is the sum of the student's PLO scores.
    student_total = {}
    for student_id in student_ids:
        values = [v for v in per_student[student_id].values() if v is not None]
        student_total[student_id] = round(sum(values), 1) if values else None

    plo_average = {}
    for plo_id in plo_ids:
        values = [per_student[sid][plo_id] for sid in student_ids if per_student[sid][plo_id] is not None]
        plo_average[plo_id] = round(sum(values) / len(values), 1) if values else None

    # Keeps the class-average row internally consistent: its total is the sum of
    # the per-PLO averages shown in the same row.
    measured = [v for v in plo_average.values() if v is not None]
    return {
        "per_student": per_student,
        "student_total": student_total,
        "plo_average": plo_average,
        "total": round(sum(measured), 1) if measured else None,
    }


def teacher_course_student_rows(session: Session, course_id: int | None, class_id: int | None = None) -> list[dict]:
    if not course_id:
        return []
    rows_by_student: dict[int, dict] = {}
    classes = [
        course_class
        for course_class in teacher_course_classes(session, course_id)
        if not class_id or course_class.id == class_id
    ]
    assessments = teacher_course_assessments(session, course_id)
    score_lookup = {
        (score.student_id, score.assessment_id): score
        for score in session.exec(select(StudentScore)).all()
    }
    for course_class in classes:
        for enrollment in session.exec(select(ClassStudent).where(ClassStudent.class_id == course_class.id)).all():
            student = enrollment.student
            if not student or not student.id or student.id in rows_by_student:
                continue
            scores = []
            score_by_assessment = {}
            weighted_total = 0.0
            raw_total = 0.0
            any_locked = False
            for assessment in assessments:
                score_record = score_lookup.get((student.id, assessment.id))
                value = score_record.score if score_record else None
                scores.append(value)
                score_by_assessment[assessment.id] = value
                if score_record and score_record.locked:
                    any_locked = True
                if value is not None:
                    max_score = float(assessment.max_score or assessment_weight_percent(assessment) or 100)
                    if max_score > 0:
                        weighted_total += (float(value) / max_score) * assessment_weight_percent(assessment)
                    raw_total += float(value)
            percent = round(weighted_total, 2)
            rows_by_student[student.id] = {
                "id": student.id,
                "student_no": student.student_no,
                "name": student.name_en,
                "email": student.user.email if student.user else "",
                "status": enrollment.status or "Active",
                "class_name": course_class.name,
                "class_id": course_class.id,
                "scores": scores,
                "score_by_assessment": score_by_assessment,
                "total": round(weighted_total, 2),
                "raw_total": round(raw_total, 2),
                "percent": percent,
                "locked": any_locked,
            }
    return sorted(rows_by_student.values(), key=lambda item: item["student_no"])


def selected_teacher_class(
    classes: list[CourseClass],
    selected_class_id: int | None,
    academic_year: str | None = None,
    semester: str | None = None,
) -> CourseClass | None:
    if selected_class_id:
        for course_class in classes:
            if course_class.id == selected_class_id:
                return course_class
    if academic_year or semester:
        for course_class in classes:
            if academic_year and course_class.academic_year != academic_year:
                continue
            if semester and course_class.semester != semester:
                continue
            return course_class
    return classes[0] if classes else None


def teacher_course_assessments(session: Session, course_id: int | None) -> list[Assessment]:
    if not course_id:
        return []
    return list(session.exec(select(Assessment).join(CLO).where(CLO.course_id == course_id).order_by(CLO.code, Assessment.id)).all())


def assessment_type(name: str) -> str:
    lowered = name.lower()
    if "quiz" in lowered or "test" in lowered:
        return "Quiz"
    if "exam" in lowered or "midterm" in lowered or "final" in lowered:
        return "Exam"
    if "lab" in lowered or "practice" in lowered:
        return "Lab"
    if "project" in lowered:
        return "Project"
    if "presentation" in lowered:
        return "Presentation"
    return "Assignment"


def teacher_assessment_rows(session: Session, course_id: int | None, students: list[dict] | None = None) -> list[dict]:
    assessments = teacher_course_assessments(session, course_id)
    students = students if students is not None else teacher_course_student_rows(session, course_id)
    student_ids = {student["id"] for student in students}
    rows = []
    for assessment in assessments:
        score_records = [
            score
            for score in session.exec(select(StudentScore).where(StudentScore.assessment_id == assessment.id)).all()
            if score.student_id in student_ids
        ]
        score_count = len(score_records)
        pending = max(len(student_ids) - score_count, 0)
        locked = any(score.locked for score in score_records)
        rows.append(
            {
                "id": assessment.id,
                "name": assessment.name,
                "description": assessment.description or assessment.clo.description if assessment.clo else "",
                "type": assessment_type(assessment.name),
                "weight": assessment_weight_percent(assessment),
                "max_score": assessment.max_score,
                "status": "Submitted" if locked else "Completed" if pending == 0 and student_ids else "Pending" if score_count == 0 else "In Progress",
                "submitted": score_count,
                "pending": pending,
                "locked": locked,
                "clo_code": assessment.clo.code if assessment.clo else "",
            }
        )
    return rows


def assessment_weight_total(assessments: list[dict]) -> float:
    return round(sum(float(assessment.get("weight") or 0) for assessment in assessments), 2)


def course_scores_locked(session: Session, course_id: int | None, student_ids: set[int] | None = None) -> bool:
    if not course_id:
        return False
    assessment_ids = [assessment.id for assessment in teacher_course_assessments(session, course_id) if assessment.id]
    if not assessment_ids:
        return False
    statement = select(StudentScore).where(StudentScore.assessment_id.in_(assessment_ids), StudentScore.locked == True)  # noqa: E712
    if student_ids:
        statement = statement.where(StudentScore.student_id.in_(student_ids))
    return session.exec(statement).first() is not None


def score_unlock_page_data(session: Session, user: User, study_period: StudyPeriod | None = None) -> dict:
    if user.role == Role.PROGRAM_MANAGER:
        program = manager_program(session, user)
        courses = session.exec(select(Course).where(Course.program_id == program.id).order_by(Course.code)).all()
    elif user.role == Role.DEAN and user.faculty_id:
        program_ids = [program.id for program in session.exec(select(Program).where(Program.faculty_id == user.faculty_id)).all() if program.id]
        courses = session.exec(select(Course).where(Course.program_id.in_(program_ids)).order_by(Course.code)).all() if program_ids else []
    elif can(user, "manage_system"):
        courses = session.exec(select(Course).order_by(Course.code)).all()
    else:
        raise HTTPException(status_code=403)
    program_ids = {course.program_id for course in courses if course.program_id}
    programs = {program.id: program for program in session.exec(select(Program).where(Program.id.in_(program_ids))).all()} if program_ids else {}
    rows = []
    total_locked = 0
    for course in courses:
        assessment_ids = [assessment.id for clo in course.clos for assessment in clo.assessments if assessment.id]
        locked_scores = session.exec(
            select(StudentScore).where(StudentScore.assessment_id.in_(assessment_ids), StudentScore.locked == True)  # noqa: E712
        ).all() if assessment_ids else []
        if not locked_scores:
            continue
        locked_student_ids = {score.student_id for score in locked_scores}
        course_classes = session.exec(select(CourseClass).where(CourseClass.course_id == course.id)).all()
        if study_period:
            period_classes = [item for item in course_classes if class_matches_study_period(item, study_period)]
            if period_classes:
                course_classes = period_classes
        best_class = None
        best_match_count = 0
        for course_class in course_classes:
            if course_class.id is None:
                continue
            class_student_ids = {
                enrollment.student_id
                for enrollment in session.exec(select(ClassStudent).where(ClassStudent.class_id == course_class.id)).all()
            }
            match_count = len(locked_student_ids & class_student_ids)
            if match_count > best_match_count:
                best_match_count = match_count
                best_class = course_class
        total_locked += len(locked_scores)
        rows.append({
            "course": course,
            "class": best_class,
            "program": programs.get(course.program_id),
            "assessments": len(assessment_ids),
            "locked_scores": len(locked_scores),
            "students": len(locked_student_ids),
        })
    return {
        "kind": "score_unlock",
        "title": "Score Unlock",
        "description": "Unlock submitted final scores so teachers can edit and resubmit corrected results.",
        "stats": [
            ("Locked Courses", len(rows), "bi-lock-fill", "orange"),
            ("Locked Scores", total_locked, "bi-clipboard-x", "red"),
            ("Courses in Scope", len(courses), "bi-book", "blue"),
            ("Role", ROLE_LABELS.get(user.role, user.role.value), "bi-shield-check", "green"),
        ],
        "rows": rows,
    }


def teacher_mapping_rows(session: Session, course_id: int | None, plos: list[dict]) -> list[dict]:
    if not course_id:
        return []
    rows = []
    for clo in session.exec(select(CLO).where(CLO.course_id == course_id).order_by(CLO.code)).all():
        mappings = {mapping.plo_id: stored_percent(mapping.weight) for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all()}
        assessments = session.exec(select(Assessment).where(Assessment.clo_id == clo.id).order_by(Assessment.id)).all()
        weights = [mappings.get(plo["id"], 0) for plo in plos]
        rows.append(
            {
                "clo": clo,
                "weights": weights,
                "plo_total": round(sum(weights), 2),
                "assessments": [
                    {
                        "id": assessment.id,
                        "name": assessment.name,
                        "weight": assessment_weight_percent(assessment),
                        "max_score": assessment.max_score,
                    }
                    for assessment in assessments
                ],
            }
        )
    return rows


def teacher_plo_rows(
    session: Session,
    user: User,
    course: Course | None,
    course_class: CourseClass | None,
    students: list[dict],
) -> list[dict]:
    """Calculate PLO attainment from this class's real scores and mappings."""
    program = course.program if course and course.program else manager_program(session, user)
    version_id: int | None = None
    if course_class and course_class.name:
        cohort_key = outcome_cohort_key(course_class.name)
        version_link = next(
            (
                item for item in session.exec(
                    select(CohortOutcomeVersion).where(CohortOutcomeVersion.programme_id == program.id)
                ).all()
                if outcome_cohort_key(item.cohort_name) == cohort_key
            ),
            None,
        )
        version_id = version_link.outcome_version_id if version_link else None
    version, _versions = selected_outcome_version(session, program, version_id)
    plos = sorted(
        session.exec(
            select(PLO).where(
                PLO.program_id == program.id,
                PLO.plo_version_id == version.id,
                PLO.status == "Active",
            )
        ).all(),
        key=plo_sort_key,
    )
    if not course or not course.id:
        return [{"id": plo.id, "code": plo.code, "description": plo.description, "value": 0.0} for plo in plos]

    assessments = teacher_course_assessments(session, course.id)
    assessment_by_id = {item.id: item for item in assessments if item.id}
    student_ids = {item["id"] for item in students}
    score_rows = list(
        session.exec(
            select(StudentScore).where(
                StudentScore.student_id.in_(student_ids),
                StudentScore.assessment_id.in_(list(assessment_by_id)),
            )
        ).all()
    ) if student_ids and assessment_by_id else []
    mappings_by_clo: dict[int, list[CLOPLOMapping]] = {}
    clo_ids = {item.clo_id for item in assessments if item.clo_id}
    if clo_ids:
        for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id.in_(clo_ids))).all():
            mappings_by_clo.setdefault(mapping.clo_id, []).append(mapping)

    numerator = {plo.id: 0.0 for plo in plos}
    denominator = {plo.id: 0.0 for plo in plos}
    valid_plo_ids = set(numerator)
    for score in score_rows:
        assessment = assessment_by_id.get(score.assessment_id)
        if not assessment:
            continue
        max_score = float(assessment.max_score or 0)
        if max_score <= 0:
            continue
        score_percent = max(0.0, min(100.0, (float(score.score) / max_score) * 100))
        assessment_weight = assessment_weight_percent(assessment)
        for mapping in mappings_by_clo.get(assessment.clo_id, []):
            if mapping.plo_id not in valid_plo_ids:
                continue
            combined_weight = assessment_weight * stored_percent(mapping.weight)
            numerator[mapping.plo_id] += score_percent * combined_weight
            denominator[mapping.plo_id] += combined_weight

    return [
        {
            "id": plo.id,
            "code": plo.code,
            "description": plo.description,
            "value": round(numerator[plo.id] / denominator[plo.id], 2) if denominator[plo.id] else 0.0,
        }
        for plo in plos
    ]


def teacher_clo_rows(session: Session, course: Course | None, students: list[dict]) -> list[dict]:
    """Calculate CLO attainment from this class's real assessment scores."""
    if not course or not course.id:
        return []
    clos = list(session.exec(select(CLO).where(CLO.course_id == course.id).order_by(CLO.code)).all())
    if not clos:
        return []
    assessments = teacher_course_assessments(session, course.id)
    assessments_by_clo: dict[int, list[Assessment]] = {}
    assessment_by_id: dict[int, Assessment] = {}
    for assessment in assessments:
        if assessment.id is None:
            continue
        assessment_by_id[assessment.id] = assessment
        assessments_by_clo.setdefault(assessment.clo_id, []).append(assessment)

    student_ids = {item["id"] for item in students}
    assessment_ids = list(assessment_by_id)
    score_rows = (
        list(
            session.exec(
                select(StudentScore).where(
                    StudentScore.student_id.in_(student_ids),
                    StudentScore.assessment_id.in_(assessment_ids),
                )
            ).all()
        )
        if student_ids and assessment_ids
        else []
    )

    numerator = {clo.id: 0.0 for clo in clos if clo.id is not None}
    denominator = {clo.id: 0.0 for clo in clos if clo.id is not None}
    for score in score_rows:
        assessment = assessment_by_id.get(score.assessment_id)
        if not assessment:
            continue
        max_score = float(assessment.max_score or 0)
        if max_score <= 0:
            continue
        clo_id = assessment.clo_id
        score_percent = max(0.0, min(100.0, (float(score.score) / max_score) * 100))
        weight = assessment_weight_percent(assessment)
        numerator[clo_id] = numerator.get(clo_id, 0.0) + (score_percent * weight)
        denominator[clo_id] = denominator.get(clo_id, 0.0) + weight

    return [
        {
            "id": clo.id,
            "code": clo.code,
            "description": clo.description,
            "value": round(numerator.get(clo.id, 0.0) / denominator.get(clo.id, 0.0), 2)
            if clo.id is not None and denominator.get(clo.id, 0.0)
            else 0.0,
            "assessment_count": len(assessments_by_clo.get(clo.id, [])) if clo.id is not None else 0,
        }
        for clo in clos
    ]


def course_score_count(session: Session, course_id: int | None) -> int:
    if not course_id:
        return 0
    assessment_ids = [assessment.id for assessment in teacher_course_assessments(session, course_id) if assessment.id]
    if not assessment_ids:
        return 0
    return len(session.exec(select(StudentScore).where(StudentScore.assessment_id.in_(assessment_ids))).all())


def course_average_percent(session: Session, course_id: int | None) -> float:
    assessments = teacher_course_assessments(session, course_id)
    if not assessments:
        return 0
    values = []
    for assessment in assessments:
        max_score = float(assessment.max_score or assessment_weight_percent(assessment) or 100)
        if max_score <= 0:
            continue
        for score in session.exec(select(StudentScore).where(StudentScore.assessment_id == assessment.id)).all():
            values.append((float(score.score) / max_score) * 100)
    return sum(values) / len(values) if values else 0


def selected_teacher_course(
    session: Session,
    user: User,
    selected_course_id: int | None,
    study_period: StudyPeriod | None = None,
) -> Course | None:
    courses = teacher_assigned_courses(session, user, study_period)
    if selected_course_id:
        for course in courses:
            if course.id == selected_course_id:
                return course
        return None
    return courses[0] if courses else None


def selected_teacher_assessment(assessments: list[dict], selected_assessment_id: int | None) -> dict | None:
    if selected_assessment_id:
        for assessment in assessments:
            if assessment["id"] == selected_assessment_id:
                return assessment
    return assessments[0] if assessments else None


def teacher_page_stats(
    kind: str,
    selected_course: Course | None,
    students: list[dict],
    assessments: list[dict],
    plos: list[dict],
    clos: list[dict] | None = None,
) -> list[tuple]:
    student_count = len(students)
    assessment_count = len(assessments)
    completed = len([item for item in assessments if item["status"] == "Completed"])
    pending = sum(int(item["pending"]) for item in assessments)
    average = round(sum(student["percent"] for student in students) / student_count, 2) if student_count else 0
    report_outcomes = clos if kind == "course_reports" and clos is not None else plos
    achieved_outcomes = len([item for item in report_outcomes if float(item.get("value") or 0) >= 70])
    if kind in {"outcome_versions", "course_mapping", "plo_target_view"}:
        return [
            ("Assigned Course", selected_course.code if selected_course else "-", "bi-journal-text", "blue"),
            ("Students", student_count, "bi-people", "green"),
            ("Assessments", assessment_count, "bi-clipboard-check", "purple"),
            ("PLOs Visible", len(plos), "bi-bullseye", "orange"),
        ]
    if kind == "assessments":
        return [
            ("Total Assessments", assessment_count, "bi-clipboard-check", "blue"),
            ("Pending Scores", pending, "bi-hourglass-split", "orange"),
            ("Students", student_count, "bi-people", "purple"),
            ("Course Average", f"{average}%", "bi-graph-up-arrow", "green"),
        ]
    if kind == "students":
        active = len([student for student in students if student["status"].lower() == "active"])
        return [
            ("Total Students", student_count, "bi-people-fill", "blue"),
            ("Active Students", active, "bi-person-check", "green"),
            ("Inactive Students", max(student_count - active, 0), "bi-person-x", "orange"),
            ("Class Average", f"{average}%", "bi-mortarboard", "purple"),
        ]
    if kind in {"plo_reports", "course_reports"}:
        outcome_label = "CLOs" if kind == "course_reports" else "PLOs"
        return [
            ("Total Students", student_count, "bi-people-fill", "blue"),
            ("Total Assessments", assessment_count, "bi-clipboard-check", "green"),
            (f"{outcome_label} Mapped", len(report_outcomes), "bi-diagram-3", "purple"),
            ("Overall Attainment", f"{average}%", "bi-file-earmark-bar-graph", "orange"),
            (f"{outcome_label} Achieved", f"{achieved_outcomes} / {len(report_outcomes)}", "bi-check2-square", "cyan"),
        ]
    return teacher_stats(kind)


def teacher_outcome_scope_data(
    session: Session,
    user: User,
    selected_course: Course | None,
    version_id: int | None = None,
) -> dict:
    """Version/PEO/PLO data visible to a teacher for the selected assigned course."""
    program = selected_course.program if selected_course and selected_course.program else None
    if not program:
        return {"program": None, "version_data": None, "peos": [], "plos": [], "targets": [], "editable": False}
    selected_version, _versions = selected_outcome_version(session, program, version_id)
    peos = list(
        session.exec(
            select(PEO)
            .where(PEO.program_id == program.id, PEO.plo_version_id == selected_version.id)
            .order_by(PEO.code)
        ).all()
    )
    plos = sorted(
        session.exec(
            select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == selected_version.id)
        ).all(),
        key=plo_sort_key,
    )
    plo_ids = [plo.id for plo in plos if plo.id is not None]
    targets = (
        session.exec(
            select(PLOTarget)
            .where(PLOTarget.program_id == program.id, PLOTarget.plo_id.in_(plo_ids))
            .order_by(PLOTarget.academic_year.desc(), PLOTarget.cohort, PLOTarget.id)
        ).all()
        if plo_ids
        else []
    )
    return {
        "program": program,
        "version_data": programme_version_data(session, program, selected_version),
        "peos": peos,
        "plos": plos,
        "targets": targets,
        "editable": not selected_version.is_locked and selected_version.status not in {"Published", "Retired"},
    }


def build_teacher_page(
    section: str,
    session: Session,
    user: User,
    selected_course_id: int | None = None,
    selected_assessment_id: int | None = None,
    selected_class_id: int | None = None,
    study_period: StudyPeriod | None = None,
    version_id: int | None = None,
) -> dict:
    kind_map = {
        "outcome-versions": "outcome_versions",
        "course-mapping": "course_mapping",
        "plo-target-view": "plo_target_view",
        "plo-target-setup": "plo_target_view",
        "courses": "courses",
        "assessments": "assessments",
        "enter-scores": "scores",
        "students": "students",
        "plo-reports": "plo_reports",
        "course-reports": "course_reports",
        "announcements": "announcements",
        "calendar": "calendar",
        "documents": "documents",
    }
    titles = {
        "outcome-versions": ("Outcome Version Management", "View outcome versions for the programme connected to your assigned course."),
        "course-mapping": ("Course Mapping", "Edit CLO-PLO mapping only for your assigned course."),
        "plo-target-view": ("PLO Target View", "View PLO targets for the programme connected to your assigned course."),
        "plo-target-setup": ("PLO Target View", "View PLO targets for the programme connected to your assigned course."),
        "courses": ("My Courses", "View and manage the courses you are teaching."),
        "assessments": ("Assessments", "View assessment details and enter scores for your students."),
        "enter-scores": ("Enter Scores", "Select a class, view course mapping, and enter all assessment scores for all students."),
        "students": ("Students", "View and manage students enrolled in this class."),
        "plo-reports": ("PLO Report", "View PLO attainment results for this course based on assessment scores."),
        "course-reports": ("CLO Report", "View comprehensive CLO attainment results for this course."),
        "announcements": ("Announcements", "View important updates and announcements."),
        "calendar": ("Calendar", "View your schedule, important dates, and course activities."),
        "documents": ("Documents", "Access and manage course documents and learning materials."),
    }
    title, description = titles[section]
    kind = kind_map[section]
    courses = teacher_courses(session, user, study_period)
    assigned_classes = teacher_assigned_classes(session, user, study_period)
    selected_class = next(
        (item for item in assigned_classes if selected_class_id and item.id == selected_class_id),
        None,
    )
    selected_course = selected_class.course if selected_class and selected_class.course else selected_teacher_course(
        session, user, selected_course_id, study_period
    )
    classes = teacher_course_classes(
        session,
        selected_course.id if selected_course else None,
        user,
        study_period,
    )
    if selected_class is None or selected_class not in classes:
        selected_class = selected_teacher_class(classes, selected_class_id)
    students = teacher_course_student_rows(session, selected_course.id if selected_course else None, selected_class.id if selected_class else None)
    plos = teacher_plo_rows(session, user, selected_course, selected_class, students)
    # Per-student averages for each PLO, shown on the student list.
    student_plo = teacher_student_plo_averages(
        session,
        selected_course.id if selected_course else None,
        [student["id"] for student in students],
        [plo["id"] for plo in plos],
    )
    clos = teacher_clo_rows(session, selected_course, students)
    assessments = teacher_assessment_rows(session, selected_course.id if selected_course else None, students)
    selected_assessment = selected_teacher_assessment(assessments, selected_assessment_id)
    mapping_rows = teacher_mapping_rows(session, selected_course.id if selected_course else None, plos)
    outcome_scope = teacher_outcome_scope_data(session, user, selected_course, version_id)
    course_mapping_rows = []
    if selected_course and outcome_scope.get("plos"):
        version_plos = outcome_scope["plos"]
        version_plo_ids = {plo.id for plo in version_plos if plo.id is not None}
        for clo in sorted(selected_course.clos, key=lambda item: item.code):
            weights = {
                mapping.plo_id: stored_percent(mapping.weight)
                for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all()
                if mapping.plo_id in version_plo_ids
            }
            course_mapping_rows.append(
                {
                    "clo": clo,
                    "weights": [weights.get(plo.id, 0) for plo in version_plos],
                    "total": sum(weights.values()),
                }
            )
    if selected_course and kind in {"assessments", "scores", "students", "plo_reports", "course_reports", "course_mapping", "plo_target_view"}:
        title = f"{title} - {selected_course.code} {selected_course.title}"
    stats = teacher_stats_for_courses(courses) if kind in {"courses", "announcements", "calendar", "documents"} else teacher_page_stats(kind, selected_course, students, assessments, plos, clos)
    return {
        "kind": kind,
        "title": title,
        "description": description,
        "stats": stats,
        "plos": plos,
        "clos": clos,
        "courses": courses,
        "study_period": study_period,
        "selected_course": selected_course,
        "classes": classes,
        "selected_class": selected_class,
        "students": students,
        "assessments": assessments,
        "selected_assessment": selected_assessment,
        "mapping_rows": mapping_rows,
        "outcome_scope": outcome_scope,
        "course_mapping_rows": course_mapping_rows,
        "assessment_weight_total": assessment_weight_total(assessments),
        "scores_locked": course_scores_locked(session, selected_course.id if selected_course else None, {student["id"] for student in students}),
        "score_error": "",
        "score_saved": False,
        "student_plo": student_plo,
    }

def build_manager_score_entry_data(
    session: Session,
    user: User,
    course_id: int | None = None,
    class_id: int | None = None,
    assessment_id: int | None = None,
    study_period: StudyPeriod | None = None,
) -> dict:
    """Build data for manager score entry page, similar to teacher page but with full programme scope."""
    program = manager_program(session, user)

    # Get all curriculum courses in the programme
    all_courses = session.exec(
        select(Course)
        .where(Course.program_id == program.id, Course.cohort_id == None)
        .order_by(Course.curriculum_year, Course.curriculum_semester, Course.code)
    ).all()

    # Get classes for each course, filtered by study period
    course_class_map: dict[int, list[CourseClass]] = {}
    class_groups: dict[tuple[str, str, str], dict] = {}
    for course in all_courses:
        classes = session.exec(select(CourseClass).where(CourseClass.course_id == course.id)).all()
        if study_period:
            classes = [c for c in classes if class_matches_study_period(c, study_period)]
        if classes:
            course_class_map[course.id] = classes
        for course_class in classes:
            key = (course_class.name, course_class.academic_year, str(course_class.semester))
            group = class_groups.setdefault(
                key,
                {
                    "id": course_class.id,
                    "name": course_class.name,
                    "academic_year": course_class.academic_year,
                    "semester": course_class.semester,
                    "courses": [],
                    "student_ids": set(),
                },
            )
            group["courses"].append(
                {
                    "id": course.id,
                    "class_id": course_class.id,
                    "code": course.code,
                    "title": course.title,
                    "credits": course.credits,
                    "year": course.curriculum_year,
                    "semester": course.curriculum_semester,
                }
            )
            if course_class.id is not None:
                for enrollment in session.exec(select(ClassStudent).where(ClassStudent.class_id == course_class.id)).all():
                    group["student_ids"].add(enrollment.student_id)

    selected_course = None
    selected_class = None

    if course_id:
        selected_course = session.get(Course, course_id)
        if selected_course and selected_course.program_id != program.id:
            selected_course = None

    # If no course selected or invalid, pick first course with classes
    if not selected_course and course_class_map:
        first_course_id = next(iter(course_class_map))
        selected_course = session.get(Course, first_course_id)
        if selected_course and course_class_map.get(selected_course.id):
            selected_class = course_class_map[selected_course.id][0]

    # If class_id provided, try to select that class for the selected course
    if class_id and selected_course:
        cls = session.get(CourseClass, class_id)
        if cls and cls.course_id == selected_course.id and class_matches_study_period(cls, study_period):
            selected_class = cls

    # If still no selected_class, pick first class for selected course
    if selected_course and not selected_class and course_class_map.get(selected_course.id):
        selected_class = course_class_map[selected_course.id][0]

    selected_class_group_key = None
    if selected_class:
        selected_class_group_key = (selected_class.name, selected_class.academic_year, str(selected_class.semester))

    formatted_class_groups = []
    for key, group in sorted(class_groups.items(), key=lambda item: (item[0][1], item[0][2], item[0][0])):
        courses_for_group = sorted(group["courses"], key=lambda item: (item.get("year") or 0, str(item.get("semester") or ""), item["code"]))
        formatted_class_groups.append(
            {
                "id": group["id"],
                "name": group["name"],
                "academic_year": group["academic_year"],
                "semester": group["semester"],
                "courses": courses_for_group,
                "students": len(group["student_ids"]),
                "active": selected_class_group_key == key,
            }
        )

    # Load data for selected course/class
    students = []
    assessments = []
    mapping_rows = []
    plos_objs = []
    assessment_rows = []
    score_locked = False
    teacher_names = ""
    total_assessments = 0
    pending = 0
    submitted = 0
    average = 0.0
    plo_values = []
    course_plo_data = []

    if selected_course and selected_class:
        # Students
        students = teacher_course_student_rows(session, selected_course.id, selected_class.id)
        # Assessments (Assessment objects)
        assessments = teacher_course_assessments(session, selected_course.id)
        # PLOs for the programme version
        version, _ = selected_outcome_version(session, program, None)
        plos_objs = session.exec(
            select(PLO)
            .where(PLO.program_id == program.id, PLO.plo_version_id == version.id, PLO.status == "Active")
            .order_by(PLO.code)
        ).all()
        plos_objs = sorted(plos_objs, key=plo_sort_key)
        # Mapping rows (CLO-PLO weights)
        mapping_rows = teacher_mapping_rows(
            session,
            selected_course.id,
            [{"id": p.id, "code": p.code} for p in plos_objs]
        )
        # Assessment rows for gradebook
        assessment_rows = teacher_assessment_rows(session, selected_course.id, students)
        # Score locked status
        score_locked = course_scores_locked(session, selected_course.id, {s["id"] for s in students})
        # Teacher names
        teacher_names = course_teacher_names(session, selected_course.id)
        # Stats
        total_assessments = len(assessments)
        pending = sum(1 for a in assessment_rows if a["pending"] > 0)
        submitted = sum(a["submitted"] for a in assessment_rows)
        average = course_average_percent(session, selected_course.id)
        # PLO values for charts
        plo_values = teacher_plo_rows(session, user, selected_course, selected_class, students)

    # Build stats for the score entry section
    stats = teacher_page_stats("assessments", selected_course, students, assessment_rows, plo_values) if selected_course else []

    # Get list of classes for dropdown (for selected course)
    class_options = []
    if selected_course:
        classes_for_course = session.exec(select(CourseClass).where(CourseClass.course_id == selected_course.id)).all()
        if study_period:
            classes_for_course = [c for c in classes_for_course if class_matches_study_period(c, study_period)]
        class_options = [
            {"id": c.id, "name": c.name, "academic_year": c.academic_year, "semester": c.semester}
            for c in classes_for_course
        ]

    return {
        "kind": "manager_score_entry",
        "title": "Manager Score Entry",
        "description": "Enter and manage scores for courses in your programme.",
        "stats": stats,
        "courses": [{"id": c.id, "code": c.code, "title": c.title} for c in all_courses],
        "class_groups": formatted_class_groups,
        "selected_course": selected_course,
        "selected_class": selected_class,
        "students": students,
        "assessments": assessment_rows,  # list of dicts from teacher_assessment_rows
        "plos": plos_objs,  # list of PLO objects
        "mapping_rows": mapping_rows,
        "assessment_weight_total": assessment_weight_total(assessment_rows),
        "scores_locked": score_locked,
        "teacher_names": teacher_names,
        "total_assessments": total_assessments,
        "pending": pending,
        "submitted": submitted,
        "average": average,
        "study_period": study_period,
        "classes": class_options,
        "plo_values": plo_values,  # list of dicts with id, code, description, value
        "course_plo_data": course_plo_data,
    }







def student_stats(kind: str) -> list[tuple]:
    stat_sets = {
        "dashboard": [("Enrolled Courses", 5, "bi-people", "blue"), ("Completed Assessments", 12, "bi-clipboard-check", "green"), ("Overall PLO Attainment", "76.45%", "bi-bullseye", "purple"), ("PLOs >= 70%", "6 / 11", "bi-graph-up-arrow", "orange"), ("PLOs < 70%", "5 / 11", "bi-graph-down-arrow", "cyan"), ("Total Assessments", 5, "bi-file-earmark-text", "blue")],
        "courses": [("Total Courses", 29, "bi-book", "blue"), ("Completed", 11, "bi-check-circle", "green"), ("In Progress", 7, "bi-hourglass-split", "orange"), ("Pending", 11, "bi-clock", "purple")],
        "assessments": [("Total Assessments", 5, "bi-file-earmark-check", "blue"), ("Completed Assessments", 4, "bi-check-circle", "green"), ("Pending Assessments", 1, "bi-hourglass-split", "orange"), ("Overall CLO Attainment", "76.45%", "bi-bullseye", "purple"), ("PLOs >= 70%", "6 / 11", "bi-graph-up-arrow", "blue")],
        "clo": [("Overall CLO Attainment", "76.45%", "bi-bullseye", "blue"), ("CLOs Achieved", "3 / 5", "bi-check-circle", "green"), ("CLOs Not Achieved", "2 / 5", "bi-graph-up-arrow", "orange"), ("Total Assessments", 5, "bi-file-earmark-text", "purple"), ("Your Overall Status", "Good", "bi-people", "blue")],
        "reports": [("Overall PLO Attainment", "76.45%", "bi-bullseye", "blue"), ("PLOs >= 70%", "6 / 11", "bi-check-circle", "green"), ("PLOs < 70%", "5 / 11", "bi-graph-up-arrow", "orange"), ("Total Courses", 45, "bi-file-earmark-text", "purple"), ("Total Students", 56, "bi-people", "blue")],
        "course_reports": [("Overall PLO Attainment", "78.60%", "bi-bullseye", "blue"), ("PLOs >= 70%", "8 / 11", "bi-check-circle", "green"), ("PLOs < 70%", "3 / 11", "bi-graph-up-arrow", "orange"), ("Assessments", 6, "bi-file-earmark-text", "purple"), ("Students", 56, "bi-people", "blue")],
        "announcements": [("Total Announcements", 18, "bi-megaphone", "blue"), ("Unread", 6, "bi-envelope", "orange"), ("Pinned", 3, "bi-pin-angle", "purple"), ("This Week", 4, "bi-calendar3", "green")],
        "documents": [("All Documents", 28, "bi-folder", "blue"), ("Guidelines", 7, "bi-file-earmark-check", "green"), ("Academic", 9, "bi-mortarboard", "orange"), ("Assessment", 6, "bi-clipboard", "purple"), ("Reports", 4, "bi-graph-up-arrow", "orange"), ("Other", 2, "bi-folder", "blue")],
    }
    return stat_sets.get(kind, [])


def student_profile(session: Session, user: User) -> Student | None:
    return session.exec(select(Student).where(Student.user_id == user.id)).first()
# ===== Helper: get student portal data (reuse existing logic) =====
def student_portal_data_for_student(session: Session, student: Student) -> dict:
    """
    Returns the same portal data structure as student_portal_data but for a given student.
    """
    if not student or not student.id:
        return {
            "student": None,
            "courses": [],
            "assessments": [],
            "clo_values": [],
            "plo_values": [],
            "course_plo_data": [],
            "events": [],
            "announcements": [],
            "documents": [],
            "overall": 0.0,
            "target": 70.0,
        }

    # ---- 1. Fetch student's class enrollments, courses, CLOs, assessments, scores ----
    class_links = list(
        session.exec(select(ClassStudent).where(ClassStudent.student_id == student.id)).all()
    )
    class_ids = [link.class_id for link in class_links if link.class_id]
    classes = list(session.exec(select(CourseClass).where(CourseClass.id.in_(class_ids))).all()) if class_ids else []
    class_by_course: dict[int, CourseClass] = {}
    for course_class in classes:
        class_by_course.setdefault(course_class.course_id, course_class)

    enrolled_course_ids = list(class_by_course.keys())
    courses = list(session.exec(select(Course).where(Course.id.in_(enrolled_course_ids)).order_by(Course.code)).all()) if enrolled_course_ids else []
    course_ids = [course.id for course in courses if course.id]

    clos = list(session.exec(select(CLO).where(CLO.course_id.in_(course_ids)).order_by(CLO.code)).all()) if course_ids else []
    clo_ids = [clo.id for clo in clos if clo.id]

    assessments = list(session.exec(select(Assessment).where(Assessment.clo_id.in_(clo_ids)).order_by(Assessment.id)).all()) if clo_ids else []
    assessment_ids = [assessment.id for assessment in assessments if assessment.id]

    scores = list(
        session.exec(select(StudentScore).where(StudentScore.student_id == student.id, StudentScore.assessment_id.in_(assessment_ids))).all()
    ) if assessment_ids else []
    scores_by_assessment = {score.assessment_id: score for score in scores}

    course_by_id = {course.id: course for course in courses if course.id}
    clo_by_id = {clo.id: clo for clo in clos if clo.id}
    assessments_by_course: dict[int, list[Assessment]] = {}
    for assessment in assessments:
        clo = clo_by_id.get(assessment.clo_id)
        if clo:
            assessments_by_course.setdefault(clo.course_id, []).append(assessment)

    # ---- 2. Build course_rows (with total score) ----
    course_rows = []
    for course in courses:
        course_assessments = assessments_by_course.get(course.id, [])
        course_clos = [clo for clo in clos if clo.course_id == course.id]
        course_class = class_by_course.get(course.id)
        instructors = []
        if course_class:
            for assignment in getattr(course_class, "teachers", []) or []:
                teacher = getattr(assignment, "teacher", None)
                teacher_user = getattr(teacher, "user", None) if teacher else None
                teacher_name = (
                    getattr(teacher_user, "name", None)
                    or getattr(teacher, "name", None)
                    or getattr(teacher, "name_en", None)
                    or getattr(teacher, "staff_no", None)
                )
                if teacher_name:
                    instructors.append(str(teacher_name))
        entered = sum(1 for assessment in course_assessments if assessment.id in scores_by_assessment)
        total = len(course_assessments)
        locked = sum(1 for assessment in course_assessments if (scores_by_assessment.get(assessment.id).locked if scores_by_assessment.get(assessment.id) else False))
        progress = round((entered / total) * 100, 2) if total else 0.0
        status = "Completed" if total and locked == total else "In Progress" if entered else "Pending"
        latest = max((score.updated_at for score in scores if score.assessment_id in {item.id for item in course_assessments} and score.updated_at), default=None)

        total_score = 0.0
        total_possible = 0.0
        for assessment in course_assessments:
            if assessment.max_score:
                total_possible += assessment.max_score
            score_obj = scores_by_assessment.get(assessment.id)
            if score_obj:
                total_score += float(score_obj.score)
        score_percent = round((total_score / total_possible) * 100, 2) if total_possible > 0 else 0.0

        course_rows.append({
            "id": course.id,
            "code": course.code,
            "title": course.title,
            "credits": course.credits,
            "year": course.curriculum_year,
            "class_name": class_by_course.get(course.id).name if course.id in class_by_course else "-",
            "academic_year": class_by_course.get(course.id).academic_year if course.id in class_by_course else "-",
            "semester": class_by_course.get(course.id).semester if course.id in class_by_course else "-",
            "assessments": total,
            "entered": entered,
            "progress": progress,
            "status": status,
            "last_activity": format_datetimeish(latest) if latest else "No score yet",
            "total_score": round(total_score, 2),
            "total_possible": round(total_possible, 2),
            "score_percent": score_percent,
        })

    # ---- 3. Build assessment_rows and clo_scores ----
    assessment_rows = []
    clo_scores: dict[int, list[float]] = {}
    for assessment in assessments:
        clo = clo_by_id.get(assessment.clo_id)
        course = course_by_id.get(clo.course_id) if clo else None
        score = scores_by_assessment.get(assessment.id)
        percent = round((float(score.score) / float(assessment.max_score or 100)) * 100, 2) if score and assessment.max_score else None
        if percent is not None and clo and clo.id:
            clo_scores.setdefault(clo.id, []).append(percent)
        assessment_rows.append({
            "id": assessment.id,
            "name": assessment.name,
            "type": assessment.description or "Assessment",
            "course": course,
            "clo": clo,
            "weight": assessment_weight_percent(assessment),
            "max_score": assessment.max_score,
            "score": score.score if score else None,
            "percent": percent,
            "status": "Submitted" if score and score.locked else "Completed" if score else "Pending",
        })

    # ---- 4. CLO attainment summary ----
    clo_values = []
    for clo in clos:
        values = clo_scores.get(clo.id, [])
        percent = round(sum(values) / len(values), 2) if values else 0.0
        clo_values.append((clo.code, clo.description, percent, "Achieved" if percent >= 70 else "Not Achieved"))

    # ---- 5. PLO data (sorted numerically) ----
    plo_ids = []
    if courses:
        program_id = courses[0].program_id
        selected_version, _versions = selected_outcome_version(session, session.get(Program, program_id))
        plos = list(session.exec(select(PLO).where(
            PLO.program_id == program_id,
            PLO.plo_version_id == selected_version.id,
            PLO.status == "Active"
        ).order_by(PLO.code)).all()) if selected_version else []
        plos = sorted(plos, key=plo_sort_key)
        plo_ids = [plo.id for plo in plos if plo.id]
    else:
        plos = []

    mappings = list(session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id.in_(clo_ids), CLOPLOMapping.plo_id.in_(plo_ids))).all()) if clo_ids and plo_ids else []
    clo_percent = {clo.id: (round(sum(clo_scores.get(clo.id, [])) / len(clo_scores.get(clo.id, [])), 2) if clo_scores.get(clo.id, []) else 0.0) for clo in clos if clo.id}
    colors = ["#2f7dff", "#22c55e", "#f8b41c", "#8059d6", "#26b8a6", "#ff8a3d", "#48b8e8", "#ec5a7a", "#78c850", "#a7c7e7", "#d6d6d6"]

    plo_values = []
    for index, plo in enumerate(plos):
        related = [mapping for mapping in mappings if mapping.plo_id == plo.id and clo_percent.get(mapping.clo_id, 0) > 0]
        if related:
            total_weight = sum(float(item.weight or 0) for item in related) or len(related)
            percent = sum(clo_percent.get(item.clo_id, 0) * (float(item.weight or 1) / total_weight) for item in related)
        else:
            percent = 0.0
        plo_values.append((plo.code, round(percent, 2), colors[index % len(colors)]))

    overall_values = [value for _code, value, _color in plo_values if value > 0]
    overall = round(sum(overall_values) / len(overall_values), 2) if overall_values else 0.0

    # ---- 6. Per-course PLO spider data ----
    course_plo_data = []
    for course in courses:
        course_clo_ids = [clo.id for clo in clos if clo.course_id == course.id]
        course_assessments = [assess for assess in assessments if assess.clo_id in course_clo_ids]
        course_scores = [score for score in scores if score.assessment_id in [a.id for a in course_assessments]]

        course_clo_scores = {}
        for clo in clos:
            if clo.course_id != course.id:
                continue
            clo_scores_list = []
            for assess in course_assessments:
                if assess.clo_id == clo.id:
                    score_obj = scores_by_assessment.get(assess.id)
                    if score_obj and assess.max_score:
                        percent = (float(score_obj.score) / float(assess.max_score)) * 100
                        clo_scores_list.append(percent)
            if clo_scores_list:
                avg = sum(clo_scores_list) / len(clo_scores_list)
            else:
                avg = 0.0
            course_clo_scores[clo.id] = avg

        course_mappings = [m for m in mappings if m.clo_id in course_clo_ids]
        course_plo_vals = []
        for idx, plo in enumerate(plos):
            related = [m for m in course_mappings if m.plo_id == plo.id and course_clo_scores.get(m.clo_id, 0) > 0]
            if related:
                total_weight = sum(float(m.weight or 0) for m in related) or len(related)
                percent = sum(course_clo_scores.get(m.clo_id, 0) * (float(m.weight or 1) / total_weight) for m in related)
            else:
                percent = 0.0
            course_plo_vals.append((plo.code, round(percent, 2), colors[idx % len(colors)]))

        course_overall = round(sum([v for _, v, _ in course_plo_vals if v > 0]) / len([v for _, v, _ in course_plo_vals if v > 0]), 2) if any(v > 0 for _, v, _ in course_plo_vals) else 0.0
        course_plo_data.append({
            "course_id": course.id,
            "course_code": course.code,
            "course_title": course.title,
            "plo_values": course_plo_vals,
            "overall": course_overall,
        })

    # ---- 7. Events, announcements, documents ----
    events = [
        {"title": row["name"], "date": row["course"].code if row["course"] else "-", "type": row["type"], "course": row["course"].title if row["course"] else "-"}
        for row in assessment_rows[:8]
    ]

    return {
        "student": student,
        "courses": course_rows,
        "assessments": assessment_rows,
        "clo_values": clo_values,
        "plo_values": plo_values,
        "course_plo_data": course_plo_data,
        "events": events,
        "announcements": [
            {"title": "Assessment score update", "body": "New submitted scores are reflected in your attainment reports.", "course": "All Courses", "type": "Academic", "priority": "Medium", "status": "Read"},
            {"title": "Study period reminder", "body": "Use the Study Period selector to view another semester.", "course": "Portal", "type": "General", "priority": "Low", "status": "Read"},
        ],
        "documents": [
            {"name": "PLO Handbook (Student Guide)", "category": "Guidelines", "type": "PDF", "size": "-", "uploaded_by": "Admin"},
            {"name": "Assessment Rubric", "category": "Assessment", "type": "PDF", "size": "-", "uploaded_by": "Teacher"},
        ],
        "overall": overall,
        "target": 70.0,
    }

def radar_chart_data(items: list[dict], radius: float = 42.0, center: float = 50.0) -> dict:
    total = len(items)
    if not total:
        return {"items": [], "points": "", "rings": [], "axes": []}
    chart_items = []
    points = []
    axes = []
    for index, item in enumerate(items):
        angle = (2 * math.pi * index / total) - (math.pi / 2)
        value = max(0.0, min(100.0, float(item.get("value") or 0)))
        value_radius = radius * (value / 100.0)
        x = center + math.cos(angle) * value_radius
        y = center + math.sin(angle) * value_radius
        label_x = center + math.cos(angle) * (radius + 7)
        label_y = center + math.sin(angle) * (radius + 7)
        axis_x = center + math.cos(angle) * radius
        axis_y = center + math.sin(angle) * radius
        points.append(f"{x:.2f},{y:.2f}")
        axes.append({"x": f"{axis_x:.2f}", "y": f"{axis_y:.2f}"})
        chart_items.append({**item, "label_x": f"{label_x:.2f}", "label_y": f"{label_y:.2f}"})
    rings = []
    for level in (25, 50, 75, 100):
        level_radius = radius * (level / 100.0)
        ring_points = []
        for index in range(total):
            angle = (2 * math.pi * index / total) - (math.pi / 2)
            ring_points.append(f"{center + math.cos(angle) * level_radius:.2f},{center + math.sin(angle) * level_radius:.2f}")
        rings.append({"level": level, "points": " ".join(ring_points)})
    return {"items": chart_items, "points": " ".join(points), "rings": rings, "axes": axes}


def student_portal_data(session: Session, user: User) -> dict:
    student = student_profile(session, user)
    if not student or not student.id:
        return {
            "student": None,
            "courses": [],
            "assessments": [],
            "clo_values": [],
            "plo_values": [],
            "course_plo_data": [],
            "events": [],
            "announcements": [],
            "documents": [],
            "course_sheets": [],
            "semester_spider": [],
            "year_spider": [],
            "semester_spider_chart": radar_chart_data([]),
            "year_spider_chart": radar_chart_data([]),
            "overall": 0.0,
            "target": 70.0,
        }

    # ---- 1. Fetch student's class enrollments, courses, CLOs, assessments, scores ----
    class_links = list(
        session.exec(select(ClassStudent).where(ClassStudent.student_id == student.id)).all()
    )
    class_ids = [link.class_id for link in class_links if link.class_id]
    classes = list(session.exec(select(CourseClass).where(CourseClass.id.in_(class_ids))).all()) if class_ids else []
    class_by_course: dict[int, CourseClass] = {}
    for course_class in classes:
        class_by_course.setdefault(course_class.course_id, course_class)

    enrolled_course_ids = list(class_by_course.keys())
    courses = list(session.exec(select(Course).where(Course.id.in_(enrolled_course_ids)).order_by(Course.code)).all()) if enrolled_course_ids else []
    course_ids = [course.id for course in courses if course.id]

    clos = list(session.exec(select(CLO).where(CLO.course_id.in_(course_ids)).order_by(CLO.code)).all()) if course_ids else []
    clo_ids = [clo.id for clo in clos if clo.id]

    assessments = list(session.exec(select(Assessment).where(Assessment.clo_id.in_(clo_ids)).order_by(Assessment.id)).all()) if clo_ids else []
    assessment_ids = [assessment.id for assessment in assessments if assessment.id]

    scores = list(
        session.exec(select(StudentScore).where(StudentScore.student_id == student.id, StudentScore.assessment_id.in_(assessment_ids))).all()
    ) if assessment_ids else []
    scores_by_assessment = {score.assessment_id: score for score in scores}

    course_by_id = {course.id: course for course in courses if course.id}
    clo_by_id = {clo.id: clo for clo in clos if clo.id}
    assessments_by_course: dict[int, list[Assessment]] = {}
    for assessment in assessments:
        clo = clo_by_id.get(assessment.clo_id)
        if clo:
            assessments_by_course.setdefault(clo.course_id, []).append(assessment)

    # ---- 2. Build course_rows (with total score) ----
    course_rows = []
    for course in courses:
        course_assessments = assessments_by_course.get(course.id, [])
        course_clos = [clo for clo in clos if clo.course_id == course.id]
        course_class = class_by_course.get(course.id)
        instructors = []
        if course_class:
            for assignment in getattr(course_class, "teachers", []) or []:
                teacher = getattr(assignment, "teacher", None)
                teacher_user = getattr(teacher, "user", None) if teacher else None
                teacher_name = (
                    getattr(teacher_user, "name", None)
                    or getattr(teacher, "name", None)
                    or getattr(teacher, "name_en", None)
                    or getattr(teacher, "staff_no", None)
                )
                if teacher_name:
                    instructors.append(str(teacher_name))
        entered = sum(1 for assessment in course_assessments if assessment.id in scores_by_assessment)
        total = len(course_assessments)
        locked = sum(1 for assessment in course_assessments if (scores_by_assessment.get(assessment.id).locked if scores_by_assessment.get(assessment.id) else False))
        progress = round((entered / total) * 100, 2) if total else 0.0
        status = "Completed" if total and locked == total else "In Progress" if entered else "Pending"
        latest = max((score.updated_at for score in scores if score.assessment_id in {item.id for item in course_assessments} and score.updated_at), default=None)

        # ---- NEW: total score for this course ----
        total_score = 0.0
        total_possible = 0.0
        for assessment in course_assessments:
            if assessment.max_score:
                total_possible += assessment.max_score
            score_obj = scores_by_assessment.get(assessment.id)
            if score_obj:
                total_score += float(score_obj.score)
        score_percent = round((total_score / total_possible) * 100, 2) if total_possible > 0 else 0.0

        course_rows.append({
            "id": course.id,
            "code": course.code,
            "title": course.title,
            "credits": course.credits,
            "year": course.curriculum_year,
            "class_name": course_class.name if course_class else "-",
            "academic_year": course_class.academic_year if course_class else "-",
            "semester": course_class.semester if course_class else "-",
            "instructor": ", ".join(dict.fromkeys(instructors)) or "-",
            "clos": len(course_clos),
            "assessments": total,
            "entered": entered,
            "progress": progress,
            "status": status,
            "last_activity": format_datetimeish(latest) if latest else "No score yet",
            # ---- NEW fields ----
            "total_score": round(total_score, 2),
            "total_possible": round(total_possible, 2),
            "score_percent": score_percent,
        })

    # ---- 3. Build assessment_rows and clo_scores ----
    assessment_rows = []
    clo_scores: dict[int, list[float]] = {}
    for assessment in assessments:
        clo = clo_by_id.get(assessment.clo_id)
        course = course_by_id.get(clo.course_id) if clo else None
        score = scores_by_assessment.get(assessment.id)
        percent = round((float(score.score) / float(assessment.max_score or 100)) * 100, 2) if score and assessment.max_score else None
        if percent is not None and clo and clo.id:
            clo_scores.setdefault(clo.id, []).append(percent)
        assessment_rows.append({
            "id": assessment.id,
            "name": assessment.name,
            "type": assessment.description or "Assessment",
            "course": course,
            "clo": clo,
            "weight": assessment_weight_percent(assessment),
            "max_score": assessment.max_score,
            "score": score.score if score else None,
            "percent": percent,
            "status": "Submitted" if score and score.locked else "Completed" if score else "Pending",
        })

    # ---- 4. CLO attainment summary ----
    clo_values = []
    for clo in clos:
        values = clo_scores.get(clo.id, [])
        percent = round(sum(values) / len(values), 2) if values else 0.0
        clo_values.append((clo.code, clo.description, percent, "Achieved" if percent >= 70 else "Not Achieved"))

    # ---- 5. PLO data (sorted numerically) ----
    plo_ids = []
    if courses:
        program_id = courses[0].program_id
        selected_version, _versions = selected_outcome_version(session, session.get(Program, program_id))
        plos = list(session.exec(select(PLO).where(
            PLO.program_id == program_id,
            PLO.plo_version_id == selected_version.id,
            PLO.status == "Active"
        ).order_by(PLO.code)).all()) if selected_version else []
        # ---- KEY FIX: sort PLOs numerically (PLO1, PLO2, ... PLO10) ----
        plos = sorted(plos, key=plo_sort_key)
        plo_ids = [plo.id for plo in plos if plo.id]
    else:
        plos = []

    mappings = list(session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id.in_(clo_ids), CLOPLOMapping.plo_id.in_(plo_ids))).all()) if clo_ids and plo_ids else []
    clo_percent = {clo.id: (round(sum(clo_scores.get(clo.id, [])) / len(clo_scores.get(clo.id, [])), 2) if clo_scores.get(clo.id, []) else 0.0) for clo in clos if clo.id}
    colors = ["#2f7dff", "#22c55e", "#f8b41c", "#8059d6", "#26b8a6", "#ff8a3d", "#48b8e8", "#ec5a7a", "#78c850", "#a7c7e7", "#d6d6d6"]

    # Overall PLO values (already in numeric order)
    plo_values = []
    for index, plo in enumerate(plos):
        related = [mapping for mapping in mappings if mapping.plo_id == plo.id and clo_percent.get(mapping.clo_id, 0) > 0]
        if related:
            total_weight = sum(float(item.weight or 0) for item in related) or len(related)
            percent = sum(clo_percent.get(item.clo_id, 0) * (float(item.weight or 1) / total_weight) for item in related)
        else:
            percent = 0.0
        plo_values.append((plo.code, round(percent, 2), colors[index % len(colors)]))

    overall_values = [value for _code, value, _color in plo_values if value > 0]
    overall = round(sum(overall_values) / len(overall_values), 2) if overall_values else 0.0

    # ---- 6. Per-course PLO spider data (also uses sorted plos) ----
    course_plo_data = []
    for course in courses:
        course_clo_ids = [clo.id for clo in clos if clo.course_id == course.id]
        course_assessments = [assess for assess in assessments if assess.clo_id in course_clo_ids]
        course_scores = [score for score in scores if score.assessment_id in [a.id for a in course_assessments]]

        course_clo_scores = {}
        for clo in clos:
            if clo.course_id != course.id:
                continue
            clo_scores_list = []
            for assess in course_assessments:
                if assess.clo_id == clo.id:
                    score_obj = scores_by_assessment.get(assess.id)
                    if score_obj and assess.max_score:
                        percent = (float(score_obj.score) / float(assess.max_score)) * 100
                        clo_scores_list.append(percent)
            if clo_scores_list:
                avg = sum(clo_scores_list) / len(clo_scores_list)
            else:
                avg = 0.0
            course_clo_scores[clo.id] = avg

        course_mappings = [m for m in mappings if m.clo_id in course_clo_ids]
        course_plo_vals = []
        for idx, plo in enumerate(plos):
            related = [m for m in course_mappings if m.plo_id == plo.id and course_clo_scores.get(m.clo_id, 0) > 0]
            if related:
                total_weight = sum(float(m.weight or 0) for m in related) or len(related)
                percent = sum(course_clo_scores.get(m.clo_id, 0) * (float(m.weight or 1) / total_weight) for m in related)
            else:
                percent = 0.0
            course_plo_vals.append((plo.code, round(percent, 2), colors[idx % len(colors)]))

        course_overall = round(sum([v for _, v, _ in course_plo_vals if v > 0]) / len([v for _, v, _ in course_plo_vals if v > 0]), 2) if any(v > 0 for _, v, _ in course_plo_vals) else 0.0
        course_plo_data.append({
            "course_id": course.id,
            "course_code": course.code,
            "course_title": course.title,
            "plo_values": course_plo_vals,
            "overall": course_overall,
        })

    def period_sort_value(raw_semester: object) -> int:
        normalized = normalize_semester(raw_semester)
        try:
            return int(normalized or 0)
        except (TypeError, ValueError):
            return 0

    course_rows = sorted(
        course_rows,
        key=lambda row: (
            str(row.get("academic_year") or ""),
            period_sort_value(row.get("semester")),
            str(row.get("code") or ""),
        ),
    )
    sheet_lookup: dict[tuple[str, int], dict] = {}
    for row in course_rows:
        academic_year = str(row.get("academic_year") or "-")
        semester_number = period_sort_value(row.get("semester"))
        sheet = sheet_lookup.setdefault(
            (academic_year, semester_number),
            {
                "academic_year": academic_year,
                "semester": semester_number if semester_number else row.get("semester") or "-",
                "label": f"{academic_year} · Semester {semester_number if semester_number else row.get('semester') or '-'}",
                "courses": [],
                "average": 0.0,
                "credits": 0.0,
            },
        )
        sheet["courses"].append(row)
        sheet["credits"] += float(row.get("credits") or 0)
    course_sheets = []
    for _key, sheet in sorted(sheet_lookup.items(), key=lambda item: item[0]):
        measured = [float(course.get("score_percent") or 0) for course in sheet["courses"] if float(course.get("score_percent") or 0) > 0]
        sheet["average"] = round(sum(measured) / len(measured), 2) if measured else 0.0
        sheet["credits"] = round(sheet["credits"], 2)
        course_sheets.append(sheet)

    semester_spider = []
    for semester_number in range(1, 9):
        period_courses = [
            row for row in course_rows
            if period_sort_value(row.get("semester")) == semester_number
        ]
        values = [float(row.get("score_percent") or 0) for row in period_courses if float(row.get("score_percent") or 0) > 0]
        semester_spider.append(
            {
                "code": f"S{semester_number}",
                "label": f"Semester {semester_number}",
                "value": round(sum(values) / len(values), 2) if values else 0.0,
                "courses": len(period_courses),
            }
        )
    year_spider = []
    for year_number in range(1, 5):
        period_courses = [
            row for row in course_rows
            if row.get("year") == year_number
            or ((period_sort_value(row.get("semester")) - 1) // 2 + 1 == year_number if period_sort_value(row.get("semester")) else False)
        ]
        values = [float(row.get("score_percent") or 0) for row in period_courses if float(row.get("score_percent") or 0) > 0]
        year_spider.append(
            {
                "code": f"Y{year_number}",
                "label": f"Year {year_number}",
                "value": round(sum(values) / len(values), 2) if values else 0.0,
                "courses": len(period_courses),
            }
        )
    semester_spider_chart = radar_chart_data(semester_spider)
    year_spider_chart = radar_chart_data(year_spider)

    # ---- 7. Events, announcements, documents ----
    events = [
        {"title": row["name"], "date": row["course"].code if row["course"] else "-", "type": row["type"], "course": row["course"].title if row["course"] else "-"}
        for row in assessment_rows[:8]
    ]

    return {
        "student": student,
        "courses": course_rows,
        "assessments": assessment_rows,
        "clo_values": clo_values,
        "plo_values": plo_values,
        "course_plo_data": course_plo_data,
        "course_sheets": course_sheets,
        "semester_spider": semester_spider,
        "year_spider": year_spider,
        "semester_spider_chart": semester_spider_chart,
        "year_spider_chart": year_spider_chart,
        "events": events,
        "announcements": [
            {"title": "Assessment score update", "body": "New submitted scores are reflected in your attainment reports.", "course": "All Courses", "type": "Academic", "priority": "Medium", "status": "Read"},
            {"title": "Study period reminder", "body": "Use the Study Period selector to view another semester.", "course": "Portal", "type": "General", "priority": "Low", "status": "Read"},
        ],
        "documents": [
            {"name": "PLO Handbook (Student Guide)", "category": "Guidelines", "type": "PDF", "size": "-", "uploaded_by": "Admin"},
            {"name": "Assessment Rubric", "category": "Assessment", "type": "PDF", "size": "-", "uploaded_by": "Teacher"},
        ],
        "overall": overall,
        "target": 70.0,
    }


###################
def build_student_page(section: str, session: Session, user: User) -> dict:
    kind_map = {
        "dashboard": "dashboard",
        "courses": "courses",
        "assessments": "assessments",
        "clo-attainment": "clo",
        "my-scores": "student_my_scores",          # <-- added
        "plo-overview": "plo_report",
        "course-reports": "course_reports",
        "announcements": "announcements",
        "calendar": "calendar",
        "documents": "documents",
    }
    titles = {
        "dashboard": ("Student Dashboard", "Here is what is happening with your courses and PLO attainment."),
        "courses": ("My Courses", "View and track all courses you are enrolled in."),
        "assessments": ("Assessment & CLO Attainment", "View all assessments and CLO attainment for the selected course."),
        "clo-attainment": ("CLO Attainment", "View CLO attainment results for the selected course."),
        "my-scores": ("My Scores & PLO Graph", "View your assessment scores and PLO attainment spider graph."),   # <-- added
        "plo-overview": ("PLO Attainment Report", "View PLO attainment performance based on your selection."),
        "course-reports": ("Course Report", "View PLO attainment performance for each course."),
        "announcements": ("Announcements", "View important course updates and academic announcements."),
        "calendar": ("Calendar", "View your academic schedule, important dates, and events."),
        "documents": ("Documents", "Access and download important documents and guidelines."),
    }
    title, description = titles[section]
    kind = kind_map[section]
    data = student_portal_data(session, user)
    attained = sum(1 for _code, value, _color in data["plo_values"] if value >= data["target"])
    not_attained = max(len(data["plo_values"]) - attained, 0)
    completed_assessments = sum(1 for row in data["assessments"] if row["score"] is not None)
    stats = [
        ("Enrolled Courses", len(data["courses"]), "bi-book", "blue"),
        ("Completed Assessments", completed_assessments, "bi-clipboard-check", "green"),
        ("Overall PLO Attainment", f"{data['overall']}%", "bi-bullseye", "purple"),
        ("PLOs >= 70%", f"{attained} / {len(data['plo_values'])}", "bi-graph-up-arrow", "orange"),
        ("PLOs < 70%", not_attained, "bi-graph-down-arrow", "red"),
        ("Total Assessments", len(data["assessments"]), "bi-file-earmark-text", "cyan"),
    ]
    page = {"kind": kind, "title": title, "description": description, "stats": stats}
    page.update(data)
    page["plo_attained"] = attained
    page["plo_not_attained"] = not_attained
    return page
    
    
def build_curriculum_years(courses: list[Course], classes: list[CourseClass]) -> list[dict]:
    classes_by_course: dict[int, list[CourseClass]] = {}
    for course_class in classes:
        classes_by_course.setdefault(course_class.course_id, []).append(course_class)

    years = sorted({course.curriculum_year for course in courses if course.curriculum_year is not None})
    curriculum_years = []
    for year in years:
        year_courses = [course for course in courses if course.curriculum_year == year]
        year_courses.sort(key=lambda course: (course.curriculum_semester or "", course.code))
        semester_counts: dict[str, int] = {}
        for course in year_courses:
            semester_counts[course.curriculum_semester or "-"] = semester_counts.get(course.curriculum_semester or "-", 0) + 1

        seen_semesters: set[str] = set()
        rows = []
        for index, course in enumerate(year_courses):
            semester = course.curriculum_semester or "-"
            rows.append(
                {
                    "course": course,
                    "classes": classes_by_course.get(course.id, []),
                    "show_year": index == 0,
                    "year_rowspan": len(year_courses),
                    "show_semester": semester not in seen_semesters,
                    "semester": semester,
                    "semester_rowspan": semester_counts[semester],
                }
            )
            seen_semesters.add(semester)
        curriculum_years.append({"year": year, "rows": rows})
    return curriculum_years


@app.get("/", response_class=HTMLResponse)
def home(request: Request, user: Annotated[User | None, Depends(current_user)]):
    if user:
        return redirect("/dashboard")
    return templates.TemplateResponse("login.html", {"request": request})


@app.get("/login", response_class=HTMLResponse)
def login_form(request: Request, user: Annotated[User | None, Depends(current_user)]):
    if user:
        next_path = safe_return_path(request.query_params.get("next"))
        return redirect(next_path)
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login")
def login(
    response: Response,
    session: Annotated[Session, Depends(get_session)],
    email: str = Form(...),
    password: str = Form(...),
    next: str = Form("/dashboard"),
):
    user = session.exec(select(User).where(User.email == email)).first()
    if not user or not verify_password(password, user.password_hash):
        return RedirectResponse("/login?error=1", status_code=303)
    response = redirect(safe_return_path(next))
    response.set_cookie(
        "obe_session",
        cookie_signer.dumps(user.id),
        httponly=True,
        secure=bool(os.getenv("VERCEL")),
        samesite="lax",
        max_age=60 * 60 * 12,
    )
    return response


@app.post("/logout")
def logout():
    response = redirect("/")
    response.delete_cookie("obe_session")
    return response


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    all_programs = list(session.exec(select(Program).order_by(Program.code)))
    selected_program_id = optional_int(request.query_params.get("program_id"))
    if user.role == Role.SUPER_ADMIN:
        dashboard_program_options = all_programs
        selected_program = session.get(Program, selected_program_id) if selected_program_id else None
        if selected_program is None:
            selected_program_id = None
        programs = [selected_program] if selected_program else all_programs
    else:
        programs = scoped_programs(user, all_programs)
        if user.role == Role.TEACHER and not programs:
            assigned_program_ids = {course.program_id for course in teacher_assigned_courses(session, user)}
            programs = [program for program in all_programs if program.id in assigned_program_ids]
        dashboard_program_options = programs
        selected_program_id = programs[0].id if len(programs) == 1 else None

    scoped_program_ids = {program.id for program in programs if program.id is not None}
    courses = sorted(
        [
            course
            for course in session.exec(select(Course)).all()
            if course.program_id in scoped_program_ids
        ],
        key=lambda course: (course.curriculum_year or 99, course.curriculum_semester or "", course.code),
    )
    course_ids = {course.id for course in courses if course.id is not None}

    dashboard_period = selected_study_period(request, user)
    # Deans get faculty-wide figures computed from real scores.
    dean_overview = dean_faculty_overview(session, user) if user.role == Role.DEAN else None
    period_academic_year = str(dashboard_period.academic_year).strip() if dashboard_period else ""
    period_semester = normalize_semester(dashboard_period.semester) if dashboard_period else ""

    raw_classes = [
        course_class
        for course_class in session.exec(select(CourseClass)).all()
        if course_class.course_id in course_ids
        and (
            not dashboard_period
            or (
                str(course_class.academic_year or "").strip() == period_academic_year
                and normalize_semester(course_class.semester) == period_semester
            )
        )
    ]
    classes_by_generation: dict[tuple[int, str], CourseClass] = {}
    for course_class in raw_classes:
        course = session.get(Course, course_class.course_id)
        key = (course.program_id if course else 0, " ".join(course_class.name.split()).lower())
        classes_by_generation.setdefault(key, course_class)
    classes = sorted(
        classes_by_generation.values(),
        key=lambda item: (item.academic_year, normalize_semester(item.semester), item.name),
    )

    period_student_ids: set[int] = set()
    semester_enrollments = list(session.exec(select(StudentSemesterEnrollment)).all())
    for enrollment in semester_enrollments:
        if enrollment.program_id not in scoped_program_ids or str(enrollment.status).lower() != "active":
            continue
        if dashboard_period and (
            str(enrollment.academic_year or "").strip() != period_academic_year
            or normalize_semester(enrollment.semester) != period_semester
        ):
            continue
        period_student_ids.add(enrollment.student_id)
    class_ids = {item.id for item in raw_classes if item.id is not None}
    if class_ids:
        for enrollment in session.exec(select(ClassStudent).where(ClassStudent.class_id.in_(class_ids))).all():
            if str(enrollment.status).lower() == "active":
                period_student_ids.add(enrollment.student_id)

    all_students = list(session.exec(select(Student).order_by(Student.student_no)).all())
    if user.role == Role.STUDENT:
        students = [student for student in all_students if student.user_id == user.id]
    else:
        students = [student for student in all_students if student.id in period_student_ids]

    all_users = list(session.exec(select(User).order_by(User.name)).all())
    if user.role == Role.SUPER_ADMIN:
        users = [item for item in all_users if item.program_id == selected_program_id] if selected_program_id else all_users
    elif user.role == Role.DEAN:
        users = [item for item in all_users if item.faculty_id == user.faculty_id]
    else:
        users = [item for item in all_users if item.program_id in scoped_program_ids]

    if user.role == Role.SUPER_ADMIN and not selected_program_id:
        faculties = list(session.exec(select(Faculty)).all())
    else:
        faculty_ids = {program.faculty_id for program in programs}
        faculties = [faculty for faculty in session.exec(select(Faculty)).all() if faculty.id in faculty_ids]

    plos: list[PLO] = []
    for program in programs:
        versions = list(
            session.exec(
                select(PLOVersion)
                .where(PLOVersion.programme_id == program.id)
                .order_by(PLOVersion.id.desc())
            ).all()
        )
        current_version = next((item for item in versions if item.status == "Active"), None)
        if current_version is None:
            current_version = next((item for item in versions if item.status == "Published"), None)
        if current_version is None and versions:
            current_version = versions[0]
        program_plos = list(
            session.exec(
                select(PLO).where(
                    PLO.program_id == program.id,
                    PLO.plo_version_id == current_version.id,
                    PLO.status == "Active",
                )
            ).all()
        ) if current_version else list(
            session.exec(
                select(PLO).where(PLO.program_id == program.id, PLO.status == "Active")
            ).all()
        )
        plos.extend(program_plos)
    plos = sorted(plos, key=lambda item: (plo_sort_key(item), item.program_id, item.code))

    clos = list(session.exec(select(CLO).where(CLO.course_id.in_(course_ids))).all()) if course_ids else []
    clo_ids = {clo.id for clo in clos if clo.id is not None}
    assessments = list(
        session.exec(select(Assessment).where(Assessment.clo_id.in_(clo_ids))).all()
    ) if clo_ids else []
    assessment_ids = {assessment.id for assessment in assessments if assessment.id is not None}
    score_rows = list(
        session.exec(
            select(StudentScore).where(
                StudentScore.student_id.in_(period_student_ids),
                StudentScore.assessment_id.in_(assessment_ids),
            )
        ).all()
    ) if period_student_ids and assessment_ids else []

    default_target_record = session.exec(
        select(SystemSetting).where(SystemSetting.key == "attainment_target")
    ).first()
    try:
        default_target = float(default_target_record.value) if default_target_record else 70.0
    except (TypeError, ValueError):
        default_target = 70.0
    default_target = max(0.0, min(100.0, default_target))

    plo_ids = {plo.id for plo in plos if plo.id is not None}
    mappings = list(
        session.exec(
            select(CLOPLOMapping).where(
                CLOPLOMapping.clo_id.in_(clo_ids),
                CLOPLOMapping.plo_id.in_(plo_ids),
            )
        ).all()
    ) if clo_ids and plo_ids else []
    targets = list(
        session.exec(
            select(PLOTarget).where(
                PLOTarget.plo_id.in_(plo_ids),
                PLOTarget.academic_year == period_academic_year,
            )
        ).all()
    ) if plo_ids and period_academic_year else []

    course_by_clo = {clo.id: next((course for course in courses if course.id == clo.course_id), None) for clo in clos}
    assessments_by_clo: dict[int, list[Assessment]] = {}
    for assessment in assessments:
        assessments_by_clo.setdefault(assessment.clo_id, []).append(assessment)
    scores_by_assessment: dict[int, list[StudentScore]] = {}
    for score in score_rows:
        scores_by_assessment.setdefault(score.assessment_id, []).append(score)
    mappings_by_plo: dict[int, list[CLOPLOMapping]] = {}
    for mapping in mappings:
        mappings_by_plo.setdefault(mapping.plo_id, []).append(mapping)
    targets_by_plo: dict[int, list[float]] = {}
    for target in targets:
        targets_by_plo.setdefault(target.plo_id, []).append(float(target.target))

    programme_by_id = {program.id: program for program in programs}
    dashboard_rows_by_code: dict[str, dict] = {}
    for plo in plos:
        weighted_values: list[tuple[float, float]] = []
        for mapping in mappings_by_plo.get(plo.id, []):
            assessment_values: list[tuple[float, float]] = []
            for assessment in assessments_by_clo.get(mapping.clo_id, []):
                normalized_scores = [
                    max(0.0, min(100.0, score.score / assessment.max_score * 100))
                    for score in scores_by_assessment.get(assessment.id, [])
                    if assessment.max_score and assessment.max_score > 0
                ]
                if normalized_scores:
                    assessment_weight = float(assessment.weight or 1)
                    assessment_values.append((sum(normalized_scores) / len(normalized_scores), max(assessment_weight, 0.01)))
            if not assessment_values:
                continue
            clo_value = sum(value * weight for value, weight in assessment_values) / sum(weight for _, weight in assessment_values)
            course = course_by_clo.get(mapping.clo_id)
            mapping_weight = float(mapping.weight or 1)
            result_weight = max(mapping_weight, 0.01) * max(float(course.credits if course else 1), 0.01)
            weighted_values.append((clo_value, result_weight))

        value = None
        value_weight = 0.0
        if weighted_values:
            value_weight = sum(weight for _, weight in weighted_values)
            value = sum(item * weight for item, weight in weighted_values) / value_weight
        plo_targets = targets_by_plo.get(plo.id, [])
        target = sum(plo_targets) / len(plo_targets) if plo_targets else default_target
        target = max(0.0, min(100.0, target))
        row = dashboard_rows_by_code.setdefault(
            plo.code,
            {"code": plo.code, "values": [], "targets": [], "programmes": set()},
        )
        if value is not None:
            row["values"].append((value, value_weight or 1.0))
        row["targets"].append(target)
        program = programme_by_id.get(plo.program_id)
        if program:
            row["programmes"].add(program.code)

    dashboard_plo_rows = []
    for item in dashboard_rows_by_code.values():
        value = None
        if item["values"]:
            value = sum(value * weight for value, weight in item["values"]) / sum(weight for _, weight in item["values"])
        target = sum(item["targets"]) / len(item["targets"]) if item["targets"] else default_target
        dashboard_plo_rows.append(
            {
                "code": item["code"],
                "value": round(value, 1) if value is not None else None,
                "target": round(target, 1),
                "programmes": ", ".join(sorted(item["programmes"])),
            }
        )
    dashboard_plo_rows = sorted(
        dashboard_plo_rows,
        key=lambda item: int("".join(character for character in item["code"] if character.isdigit()) or 0),
    )[:10]
    dashboard_target = round(
        sum(item["target"] for item in dashboard_plo_rows) / len(dashboard_plo_rows), 1
    ) if dashboard_plo_rows else round(default_target, 1)
    measured_rows = [item for item in dashboard_plo_rows if item["value"] is not None]
    achieved_count = len([item for item in measured_rows if item["value"] >= item["target"]])
    not_achieved_count = len(measured_rows) - achieved_count
    no_data_count = len(dashboard_plo_rows) - len(measured_rows)
    total_indicator_count = len(dashboard_plo_rows)
    achieved_percentage = round(achieved_count / total_indicator_count * 100, 1) if total_indicator_count else 0
    measured_percentage = round(len(measured_rows) / total_indicator_count * 100, 1) if total_indicator_count else 0
    average_attainment = round(
        sum(item["value"] for item in measured_rows) / len(measured_rows), 1
    ) if measured_rows else None

    icon_by_module = {
        "User Management": ("bi-person", "blue"),
        "Programme Management": ("bi-mortarboard", "green"),
        "Academic Year": ("bi-calendar3", "green"),
        "PLO Target Setup": ("bi-bullseye", "orange"),
        "Programme Mapping": ("bi-diagram-3", "purple"),
        "Outcome Version Management": ("bi-layers", "purple"),
    }
    recent_activities = []
    for log in session.exec(select(AuditLog).order_by(AuditLog.id.desc()).limit(5)).all():
        icon, tone = icon_by_module.get(log.module, ("bi-activity", "blue"))
        recent_activities.append(
            {
                "icon": icon,
                "tone": tone,
                "text": log.description,
                "meta": f"{log.user_name} · {log.date_time}",
            }
        )

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "user": user,
            "classes": classes,
            "courses": courses,
            "students": students,
            "users": users,
            "programs": programs,
            "faculties": faculties,
            "plos": plos,
            "assessments": assessments,
            "dashboard_period": dashboard_period,
            "dashboard_program_options": dashboard_program_options,
            "dashboard_selected_program_id": selected_program_id,
            "dashboard_plo_rows": dashboard_plo_rows,
            "dashboard_target": dashboard_target,
            "dashboard_plo_summary": {
                "achieved": achieved_count,
                "not_achieved": not_achieved_count,
                "no_data": no_data_count,
                "average": average_attainment,
                "achieved_percentage": achieved_percentage,
                "measured_percentage": measured_percentage,
            },
            "recent_activities": recent_activities,
            "reports": list(session.exec(select(SystemReport).order_by(SystemReport.id.desc())).all()),
            "dean_faculty": session.get(Faculty, user.faculty_id) if user.role == Role.DEAN and user.faculty_id else None,
            "dean_overview": dean_overview,
            "dean_programmes": dean_overview_programmes(dean_overview) if dean_overview else [],
            "access_scope": scope_label(user),
            "page": build_teacher_page("courses", session, user) if user.role == Role.TEACHER else None,
            "student_page": build_student_page("dashboard", session, user) if user.role == Role.STUDENT else None,
        },
    )


def parse_int_values(values: list[str] | tuple[str, ...] | None) -> set[int]:
    parsed: set[int] = set()
    for value in values or []:
        item = optional_int(str(value))
        if item:
            parsed.add(item)
    return parsed


def sync_teacher_teaching_assignments(
    session: Session,
    teacher: Teacher,
    selected_course_ids: set[int],
    selected_class_ids: set[int],
) -> None:
    """Synchronize explicit teaching scope separately from the user's home ABAC scope."""
    if not teacher.id:
        return

    valid_class_ids: set[int] = set()
    for class_id in selected_class_ids:
        course_class = session.get(CourseClass, class_id)
        if not course_class:
            continue
        valid_class_ids.add(class_id)
        if course_class.course_id:
            selected_course_ids.add(course_class.course_id)

    valid_course_ids = {course_id for course_id in selected_course_ids if session.get(Course, course_id)}

    seen_courses: set[int] = set()
    for assignment in session.exec(select(CourseTeacher).where(CourseTeacher.teacher_id == teacher.id)).all():
        if assignment.course_id not in valid_course_ids or assignment.course_id in seen_courses:
            session.delete(assignment)
        else:
            seen_courses.add(assignment.course_id)
    for course_id in sorted(valid_course_ids - seen_courses):
        session.add(CourseTeacher(course_id=course_id, teacher_id=teacher.id))

    seen_classes: set[int] = set()
    for assignment in session.exec(select(ClassTeacher).where(ClassTeacher.teacher_id == teacher.id)).all():
        if assignment.class_id not in valid_class_ids or assignment.class_id in seen_classes:
            session.delete(assignment)
        else:
            seen_classes.add(assignment.class_id)
    for class_id in sorted(valid_class_ids - seen_classes):
        session.add(ClassTeacher(class_id=class_id, teacher_id=teacher.id))


def teaching_assignment_ids_for_programmes(
    session: Session,
    selected_faculty_ids: set[int],
    selected_program_ids: set[int],
) -> tuple[set[int], set[int]]:
    """Expand selected teaching faculties/programmes into course and class IDs."""
    if not selected_program_ids and selected_faculty_ids:
        selected_program_ids = {
            program.id
            for program in session.exec(select(Program)).all()
            if program.id and program.faculty_id in selected_faculty_ids
        }
    if not selected_program_ids:
        return set(), set()
    course_ids = {
        course.id
        for course in session.exec(select(Course).where(Course.program_id.in_(selected_program_ids))).all()
        if course.id
    }
    class_ids = {
        course_class.id
        for course_class in session.exec(select(CourseClass).where(CourseClass.course_id.in_(course_ids))).all()
        if course_class.id
    } if course_ids else set()
    return course_ids, class_ids


@app.post("/admin/users/create")
def admin_create_user(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
    email: str = Form(...),
    password: str = Form("password"),
    role: Role = Form(...),
    faculty_id: str | None = Form(None),
    program_id: str | None = Form(None),
    staff_no: str | None = Form(None),
    student_no: str | None = Form(None),
    name_kh: str | None = Form(None),
    class_id: str | None = Form(None),
    redirect_to: str = Form("/admin/users"),
    teacher_faculty_ids: list[str] = Form([]),
    teacher_program_ids: list[str] = Form([]),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    name = name.strip()
    email = email.strip().lower()
    existing = user_by_email(session, email)
    if existing:
        return redirect(f"{redirect_to}?error=email_exists#add-role-user" if redirect_to.startswith("/admin/roles") else "/admin/users?error=email_exists#add-user")
    parsed_faculty_id = optional_int(faculty_id)
    parsed_program_id = optional_int(program_id)
    if role_scope_is_missing(role, parsed_faculty_id, parsed_program_id):
        return redirect(f"{redirect_to}?error=missing_scope#add-role-user" if redirect_to.startswith("/admin/roles") else "/admin/users?error=missing_scope#add-user")

    cohort_details = None
    selected_class_id = optional_int(class_id)
    normalized_student_no = str(student_no or "").strip()
    if role == Role.STUDENT:
        if not normalized_student_no:
            return redirect("/admin/users?error=missing_student_no#add-user")
        if session.exec(select(Student).where(Student.student_no == normalized_student_no)).first():
            return redirect("/admin/users?error=student_exists#add-user")
        cohort_details = cohort_group_details(session, selected_class_id) if selected_class_id else None
        intake_parts = cohort_code_parts(cohort_details[0].name) if cohort_details else None
        if (
            not parsed_program_id
            or not cohort_details
            or cohort_details[1].id != parsed_program_id
            or not intake_parts
            or intake_parts["year"] != 1
            or intake_parts["semester"] != 1
        ):
            return redirect("/admin/users?error=invalid_student_cohort#add-user")

    faculty_id, program_id = normalize_user_scope(session, role, parsed_faculty_id, parsed_program_id)
    created = User(name=name, email=email, password_hash=hash_password(password), role=role, faculty_id=faculty_id, program_id=program_id)
    session.add(created)
    session.flush()
    if role == Role.TEACHER:
        teacher = Teacher(user_id=created.id, staff_no=staff_no or f"T-{created.id:03d}")
        session.add(teacher)
        session.flush()
        course_ids, class_ids = teaching_assignment_ids_for_programmes(
            session,
            parse_int_values(teacher_faculty_ids),
            parse_int_values(teacher_program_ids),
        )
        sync_teacher_teaching_assignments(
            session,
            teacher,
            course_ids,
            class_ids,
        )
    if role == Role.STUDENT:
        student = Student(
            user_id=created.id,
            student_no=normalized_student_no,
            name_en=name,
            name_kh=str(name_kh or "").strip() or None,
        )
        session.add(student)
        session.flush()
        try:
            enrollment, course_count, _created_period = enroll_student_in_cohort(
                session, student, selected_class_id, user.id
            )
        except ValueError:
            session.rollback()
            return redirect("/admin/users?error=invalid_student_cohort#add-user")
        add_audit_record(
            session,
            user,
            "Student Management",
            "CREATE",
            f"Created student login and assigned intake cohort {enrollment.cohort_name} from User Management.",
            student.student_no,
            request.client.host if request.client else "127.0.0.1",
        )
        session.commit()
        return redirect(f"/admin/users?created=1&student_enrolled=1&courses={course_count}#record-list")
    session.commit()
    return redirect(f"{redirect_to}?created=1#record-list")


@app.post("/admin/roles/create")
def create_admin_role(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    role_name: str = Form(...),
    role_code: str = Form(...),
    description: str = Form(""),
    status: str = Form("Active"),
    is_system_role: str = Form("false"),
    abac_scope_type: str = Form("All"),
    menu_access: str = Form(""),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    normalized_code = role_code.strip().upper().replace(" ", "_")
    if session.exec(select(RoleDefinition).where(RoleDefinition.role_code == normalized_code)).first():
        return redirect("/admin/roles?error=duplicate#add-role")
    role_definition = RoleDefinition(
        role_name=role_name.strip(),
        role_code=normalized_code,
        description=description.strip(),
        status=status,
        is_system_role=is_system_role == "true",
        abac_scope_type=abac_scope_type,
        menu_access=menu_access.strip(),
    )
    session.add(role_definition)
    session.commit()
    session.refresh(role_definition)
    for module in ROLE_PERMISSION_MODULES:
        session.add(RolePermission(role_definition_id=role_definition.id, module=module, can_view=True))
    session.commit()
    return redirect("/admin/roles?created=1#record-list")


@app.post("/admin/roles/{role_id}/update")
async def update_admin_role(
    role_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    role_definition = session.get(RoleDefinition, role_id)
    if not role_definition:
        raise HTTPException(status_code=404)
    form = await request.form()
    normalized_code = str(form.get("role_code") or "").strip().upper().replace(" ", "_")
    if session.exec(select(RoleDefinition).where(RoleDefinition.role_code == normalized_code, RoleDefinition.id != role_id)).first():
        return redirect(f"/admin/roles?error=duplicate&edit_role={role_id}#edit-role")
    role_definition.role_name = str(form.get("role_name") or role_definition.role_name).strip()
    role_definition.role_code = normalized_code
    role_definition.description = str(form.get("description") or "").strip()
    role_definition.status = str(form.get("status") or "Active")
    role_definition.is_system_role = str(form.get("is_system_role") or "false") == "true"
    role_definition.abac_scope_type = str(form.get("abac_scope_type") or "All")
    role_definition.menu_access = str(form.get("menu_access") or "").strip()
    role_definition.updated_at = datetime.utcnow()
    session.add(role_definition)
    permissions = session.exec(select(RolePermission).where(RolePermission.role_definition_id == role_id)).all()
    for permission in permissions:
        permission.can_view = form.get(f"perm_{permission.id}_view") == "true"
        permission.can_create = form.get(f"perm_{permission.id}_create") == "true"
        permission.can_edit = form.get(f"perm_{permission.id}_edit") == "true"
        permission.can_delete = form.get(f"perm_{permission.id}_delete") == "true"
        permission.can_export = form.get(f"perm_{permission.id}_export") == "true"
        permission.updated_at = datetime.utcnow()
        session.add(permission)
    session.commit()
    return redirect(f"/admin/roles?updated=1&view_role={role_id}#role-detail")


@app.post("/admin/roles/{role_id}/delete")
def delete_admin_role(
    role_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    role_definition = session.get(RoleDefinition, role_id)
    if not role_definition:
        raise HTTPException(status_code=404)
    assigned_users = []
    if role_definition.role_key:
        assigned_users = session.exec(select(User).where(User.role == Role(role_definition.role_key))).all()
    if role_definition.is_system_role and assigned_users:
        return redirect("/admin/roles?error=system_role")
    if assigned_users:
        role_definition.status = "Inactive"
        role_definition.updated_at = datetime.utcnow()
        session.add(role_definition)
    else:
        for permission in session.exec(select(RolePermission).where(RolePermission.role_definition_id == role_id)).all():
            session.delete(permission)
        session.delete(role_definition)
    session.commit()
    return redirect("/admin/roles?deleted=1")


@app.post("/admin/roles/{role}/assign-user")
def admin_assign_user_role(
    role: Role,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    target_user_id: int = Form(...),
    faculty_id: str | None = Form(None),
    program_id: str | None = Form(None),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    target = session.get(User, target_user_id)
    if not target:
        raise HTTPException(status_code=404)
    parsed_faculty_id = optional_int(faculty_id)
    parsed_program_id = optional_int(program_id)
    if role_scope_is_missing(role, parsed_faculty_id, parsed_program_id):
        return redirect(f"/admin/roles?error=missing_scope&view_role={role.value}#role-detail")
    faculty_id, program_id = normalize_user_scope(session, role, parsed_faculty_id, parsed_program_id)
    target.role = role
    target.faculty_id = faculty_id
    target.program_id = program_id
    target.is_active = True
    session.add(target)
    session.commit()
    return redirect(f"/admin/roles?assigned=1&view_role={role.value}#role-detail")


@app.post("/admin/roles/{role}/deactivate-users")
def admin_deactivate_role_users(
    role: Role,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    if role == Role.SUPER_ADMIN:
        return redirect("/admin/roles?error=system_role")
    role_users = session.exec(select(User).where(User.role == role)).all()
    for target in role_users:
        if target.id != user.id:
            target.is_active = False
            session.add(target)
    session.commit()
    return redirect(f"/admin/roles?deactivated=1&view_role={role.value}#role-detail")


@app.post("/admin/users/{target_user_id}/update")
def admin_update_user(
    target_user_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
    email: str = Form(...),
    role: Role = Form(...),
    faculty_id: str | None = Form(None),
    program_id: str | None = Form(None),
    password: str | None = Form(None),
    is_active: bool = Form(False),
    staff_no: str | None = Form(None),
    teacher_faculty_ids: list[str] = Form([]),
    teacher_program_ids: list[str] = Form([]),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    target = session.get(User, target_user_id)
    if not target:
        raise HTTPException(status_code=404)
    existing = session.exec(select(User).where(User.email == email, User.id != target_user_id)).first()
    if existing:
        return redirect(f"/admin/users?edit_user_id={target_user_id}&error=email_exists#edit-user")
    if role == Role.STUDENT and not session.exec(
        select(Student).where(Student.user_id == target_user_id)
    ).first():
        return redirect(
            f"/admin/users?edit_user_id={target_user_id}&error=student_setup_required#edit-user"
        )
    parsed_faculty_id = optional_int(faculty_id)
    parsed_program_id = optional_int(program_id)
    if role_scope_is_missing(role, parsed_faculty_id, parsed_program_id):
        return redirect(f"/admin/users?edit_user_id={target_user_id}&error=missing_scope#edit-user")
    faculty_id, program_id = normalize_user_scope(session, role, parsed_faculty_id, parsed_program_id)
    target.name = name
    target.email = email
    target.role = role
    target.faculty_id = faculty_id
    target.program_id = program_id
    target.is_active = is_active
    if password:
        target.password_hash = hash_password(password)
    session.add(target)
    teacher = session.exec(select(Teacher).where(Teacher.user_id == target.id)).first()
    if role == Role.TEACHER:
        if not teacher:
            teacher = Teacher(user_id=target.id, staff_no=staff_no or f"T-{target.id:03d}")
            session.add(teacher)
            session.flush()
        else:
            teacher.staff_no = staff_no or teacher.staff_no or f"T-{target.id:03d}"
            session.add(teacher)
        course_ids, class_ids = teaching_assignment_ids_for_programmes(
            session,
            parse_int_values(teacher_faculty_ids),
            parse_int_values(teacher_program_ids),
        )
        sync_teacher_teaching_assignments(
            session,
            teacher,
            course_ids,
            class_ids,
        )
    elif teacher:
        sync_teacher_teaching_assignments(session, teacher, set(), set())
    session.commit()
    return redirect("/admin/users?updated=1")


@app.post("/admin/users/{target_user_id}/delete")
def admin_delete_user(
    target_user_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    if target_user_id == user.id:
        return redirect("/admin/users?error=self_delete")
    target = session.get(User, target_user_id)
    if not target:
        raise HTTPException(status_code=404)
    target.is_active = False
    session.add(target)
    session.commit()
    return redirect("/admin/users?deleted=1")


def student_login_email(student_no: str, email: str | None = None) -> str:
    """Return one stable login email for a student account."""
    supplied = str(email or "").strip().lower()
    if supplied:
        return supplied
    local_part = "".join(
        character.lower()
        for character in str(student_no or "").strip()
        if character.isalnum() or character in {".", "_", "-"}
    )
    return f"{local_part or 'student'}@student.local"


def user_by_email(session: Session, email: str) -> User | None:
    normalized = str(email or "").strip().lower()
    return next(
        (item for item in session.exec(select(User)).all() if str(item.email or "").strip().lower() == normalized),
        None,
    )


def cohort_group_details(
    session: Session,
    representative_class_id: int,
) -> tuple[CourseClass, Program, list[CourseClass]] | None:
    """Resolve one displayed cohort option to all of its course classes."""
    representative = session.get(CourseClass, representative_class_id)
    if not representative:
        return None
    representative_course = session.get(Course, representative.course_id)
    if not representative_course:
        return None
    program = session.get(Program, representative_course.program_id)
    if not program:
        return None
    cohort_key = " ".join(str(representative.name or "").split()).lower()
    academic_year = str(representative.academic_year or "").strip()
    semester = normalize_semester(representative.semester)
    group: list[CourseClass] = []
    for course_class in session.exec(select(CourseClass)).all():
        course = session.get(Course, course_class.course_id)
        if not course or course.program_id != program.id:
            continue
        if " ".join(str(course_class.name or "").split()).lower() != cohort_key:
            continue
        if str(course_class.academic_year or "").strip() != academic_year:
            continue
        if normalize_semester(course_class.semester) != semester:
            continue
        group.append(course_class)
    return representative, program, group


def enroll_student_in_cohort(
    session: Session,
    student: Student,
    representative_class_id: int,
    _actor_user_id: int | None,
) -> tuple[StudentSemesterEnrollment, int, bool]:
    """Add one study-period enrollment and all course-class links for a cohort."""
    details = cohort_group_details(session, representative_class_id)
    if not details or student.id is None:
        raise ValueError("Invalid starting cohort")
    representative, program, course_classes = details
    academic_year = str(representative.academic_year or "").strip()
    semester = normalize_semester(representative.semester)
    cohort_name = str(representative.name or "").strip()
    target_family = cohort_family_key(cohort_name)
    existing_families = {
        family
        for item in session.exec(
            select(StudentSemesterEnrollment).where(
                StudentSemesterEnrollment.student_id == student.id
            )
        ).all()
        if (family := cohort_family_key(item.cohort_name))
    }
    if target_family and existing_families and target_family not in existing_families:
        raise ValueError("Student is already assigned to another four-year cohort")
    enrollment = session.exec(
        select(StudentSemesterEnrollment).where(
            StudentSemesterEnrollment.student_id == student.id,
            StudentSemesterEnrollment.program_id == program.id,
            StudentSemesterEnrollment.cohort_name == cohort_name,
            StudentSemesterEnrollment.academic_year == academic_year,
            StudentSemesterEnrollment.semester == semester,
        )
    ).first()
    created_period = enrollment is None
    if enrollment is None:
        enrollment = StudentSemesterEnrollment(
            student_id=student.id,
            program_id=program.id,
            cohort_name=cohort_name,
            academic_year=academic_year,
            semester=semester,
            status="Active",
        )
    else:
        enrollment.status = "Active"
    session.add(enrollment)

    added_course_classes = 0
    for course_class in course_classes:
        existing = session.exec(
            select(ClassStudent).where(
                ClassStudent.class_id == course_class.id,
                ClassStudent.student_id == student.id,
            )
        ).first()
        if existing:
            existing.status = "Active"
            session.add(existing)
            continue
        session.add(ClassStudent(class_id=course_class.id, student_id=student.id, status="Active"))
        added_course_classes += 1

    if student.user:
        student.user.role = Role.STUDENT
        student.user.faculty_id = program.faculty_id
        student.user.program_id = program.id
        student.user.is_active = True
        session.add(student.user)
    return enrollment, added_course_classes, created_period


def backfill_student_semester_enrollments(session: Session) -> int:
    """Convert legacy course-class links into one enrollment per study period."""
    existing_keys = {
        (
            item.student_id,
            item.program_id,
            item.cohort_name,
            str(item.academic_year or "").strip(),
            normalize_semester(item.semester),
        )
        for item in session.exec(select(StudentSemesterEnrollment)).all()
    }
    grouped_statuses: dict[tuple[int, int, str, str, str], list[str]] = {}
    for class_student in session.exec(select(ClassStudent)).all():
        course_class = session.get(CourseClass, class_student.class_id)
        course = session.get(Course, course_class.course_id) if course_class else None
        if not course_class or not course:
            continue
        key = (
            class_student.student_id,
            course.program_id,
            str(course_class.name or "").strip(),
            str(course_class.academic_year or "").strip(),
            normalize_semester(course_class.semester),
        )
        grouped_statuses.setdefault(key, []).append(str(class_student.status or "Active"))

    created = 0
    for key, statuses in grouped_statuses.items():
        if key in existing_keys:
            continue
        student_id, program_id, cohort_name, academic_year, semester = key
        status = "Active" if "Active" in statuses else statuses[0]
        session.add(
            StudentSemesterEnrollment(
                student_id=student_id,
                program_id=program_id,
                cohort_name=cohort_name,
                academic_year=academic_year,
                semester=semester,
                status=status,
            )
        )
        created += 1
    if created:
        session.commit()
    return created


@app.post("/admin/students/create")
def admin_create_student_record(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    student_no: str = Form(...),
    name_en: str = Form(...),
    name_kh: str = Form(""),
    email: str = Form(...),
    program_id: int = Form(...),
    class_id: int = Form(...),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    student_no = student_no.strip()
    name_en = name_en.strip()
    email = student_login_email(student_no, email)
    program = session.get(Program, program_id)
    cohort_details = cohort_group_details(session, class_id)
    intake_parts = cohort_code_parts(cohort_details[0].name) if cohort_details else None
    if not program:
        return redirect("/admin/students?error=invalid_program#add-student")
    if (
        not cohort_details
        or cohort_details[1].id != program.id
        or not intake_parts
        or intake_parts["year"] != 1
        or intake_parts["semester"] != 1
    ):
        return redirect("/admin/students?error=invalid_cohort#add-student")
    existing = session.exec(select(Student).where(Student.student_no == student_no)).first()
    if existing:
        return redirect("/admin/students?error=student_exists#add-student")

    linked_user = user_by_email(session, email)
    if linked_user and linked_user.role != Role.STUDENT:
        return redirect("/admin/students?error=email_exists#add-student")
    if linked_user and session.exec(select(Student).where(Student.user_id == linked_user.id)).first():
        return redirect("/admin/students?error=account_exists#add-student")
    if not linked_user:
        linked_user = User(
            name=name_en,
            email=email,
            password_hash=hash_password("password"),
            role=Role.STUDENT,
            is_active=True,
            faculty_id=program.faculty_id,
            program_id=program.id,
        )
        session.add(linked_user)
        session.flush()
    else:
        linked_user.name = name_en
        linked_user.faculty_id = program.faculty_id
        linked_user.program_id = program.id
        linked_user.is_active = True
        session.add(linked_user)

    student = Student(
        user_id=linked_user.id,
        student_no=student_no,
        name_en=name_en,
        name_kh=name_kh or None,
    )
    session.add(student)
    session.flush()
    enrollment, course_count, _created_period = enroll_student_in_cohort(
        session, student, class_id, user.id
    )
    add_audit_record(
        session,
        user,
        "Student Management",
        "CREATE",
        f"Created one student account and enrolled it in {enrollment.cohort_name}.",
        student.student_no,
        request.client.host if request.client else "127.0.0.1",
    )
    session.commit()
    return redirect(f"/admin/students?created=1&courses={course_count}#student-list")


@app.get("/admin/students/import-template")
def admin_download_student_import_template(user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Student Import Example"
    headers = ["Student ID", "Student Name", "Khmer Name", "Email"]
    examples = [
        ["S2024001", "Sothy Chomroeun", "សុធី ចំរើន", "s2024001@student.edu.kh"],
        ["S2024002", "Chan Sreyneng", "ចាន់ ស្រីណេង", "s2024002@student.edu.kh"],
        ["S2024003", "Heng Vutha", "", ""],
    ]
    sheet.append(headers)
    for row in examples:
        sheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="E8F1FF")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="061847")
        cell.fill = header_fill

    widths = {"A": 18, "B": 28, "C": 24, "D": 34}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student-import-example.xlsx"},
    )


@app.post("/admin/students/import")
async def admin_import_student_records(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    students_file: UploadFile = File(...),
    class_id: int = Form(...),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    cohort_details = cohort_group_details(session, class_id)
    intake_parts = cohort_code_parts(cohort_details[0].name) if cohort_details else None
    if (
        not cohort_details
        or not intake_parts
        or intake_parts["year"] != 1
        or intake_parts["semester"] != 1
    ):
        return redirect("/admin/students?import_error=Please%20select%20a%20valid%20starting%20cohort#import-students")
    _representative, program, _course_classes = cohort_details
    selected_family = cohort_family_key(_representative.name)
    try:
        rows = parse_student_import_rows(students_file.filename or "", await students_file.read())
    except Exception as exc:
        return redirect(f"/admin/students?import_error={quote(str(exc))}#import-students")

    if not rows:
        return redirect("/admin/students?import_error=No%20valid%20student%20rows%20found#import-students")

    created = 0
    updated = 0
    enrolled = 0
    course_links = 0
    skipped = 0

    for row in rows:
        student_no = str(row["student_no"] or "").strip()
        name_en = str(row["name_en"] or "").strip()
        login_email = student_login_email(student_no, row["email"])
        student = session.exec(select(Student).where(Student.student_no == student_no)).first()
        if student and selected_family:
            existing_families = {
                family
                for item in session.exec(
                    select(StudentSemesterEnrollment).where(
                        StudentSemesterEnrollment.student_id == student.id
                    )
                ).all()
                if (family := cohort_family_key(item.cohort_name))
            }
            if existing_families and selected_family not in existing_families:
                skipped += 1
                continue
        linked_user = user_by_email(session, login_email)
        linked_student = (
            session.exec(select(Student).where(Student.user_id == linked_user.id)).first()
            if linked_user and linked_user.id
            else None
        )
        if linked_user and linked_user.role != Role.STUDENT:
            skipped += 1
            continue
        if linked_student and (not student or linked_student.id != student.id):
            skipped += 1
            continue

        if student:
            if student.user and linked_user and student.user_id != linked_user.id:
                skipped += 1
                continue
            student.name_en = name_en
            student.name_kh = row["name_kh"] or None
            updated += 1
        else:
            student = Student(
                student_no=student_no,
                name_en=name_en,
                name_kh=row["name_kh"] or None,
            )
            session.add(student)
            session.flush()
            created += 1

        account = student.user or linked_user
        if account is None:
            account = User(
                name=name_en,
                email=login_email,
                password_hash=hash_password("password"),
                role=Role.STUDENT,
                is_active=True,
                faculty_id=program.faculty_id,
                program_id=program.id,
            )
            session.add(account)
            session.flush()
            student.user_id = account.id
        else:
            account.name = name_en
            account.email = login_email
            account.role = Role.STUDENT
            account.is_active = True
            account.faculty_id = program.faculty_id
            account.program_id = program.id
            session.add(account)
            student.user_id = account.id
        session.add(student)
        session.flush()
        _enrollment, added_courses, created_period = enroll_student_in_cohort(
            session, student, class_id, user.id
        )
        course_links += added_courses
        if created_period:
            enrolled += 1

    add_audit_record(
        session,
        user,
        "Student Management",
        "IMPORT",
        f"Imported student accounts: {created} created, {updated} updated, {skipped} skipped.",
        program.code,
        request.client.host if request.client else "127.0.0.1",
    )
    session.commit()
    return redirect(f"/admin/students?imported=1&created={created}&updated={updated}&enrolled={enrolled}&courses={course_links}&skipped={skipped}#student-list")


@app.post("/admin/students/{student_id}/assign-cohort")
def admin_assign_student_intake_cohort(
    student_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    program_id: int = Form(...),
    class_id: int = Form(...),
):
    """Repair a legacy Student account that was created without a cohort."""
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    student = session.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404)
    if session.exec(
        select(StudentSemesterEnrollment).where(
            StudentSemesterEnrollment.student_id == student.id
        )
    ).first() or session.exec(
        select(ClassStudent).where(ClassStudent.student_id == student.id)
    ).first():
        return redirect(f"/admin/students?edit_id={student_id}&error=already_enrolled#edit-record")

    cohort_details = cohort_group_details(session, class_id)
    intake_parts = cohort_code_parts(cohort_details[0].name) if cohort_details else None
    if (
        not cohort_details
        or cohort_details[1].id != program_id
        or not intake_parts
        or intake_parts["year"] != 1
        or intake_parts["semester"] != 1
    ):
        return redirect(f"/admin/students?edit_id={student_id}&error=invalid_cohort#edit-record")

    try:
        enrollment, course_count, _created_period = enroll_student_in_cohort(
            session, student, class_id, user.id
        )
    except ValueError:
        session.rollback()
        return redirect(f"/admin/students?edit_id={student_id}&error=invalid_cohort#edit-record")
    add_audit_record(
        session,
        user,
        "Student Management",
        "ASSIGN",
        f"Assigned legacy student account to intake cohort {enrollment.cohort_name}.",
        student.student_no,
        request.client.host if request.client else "127.0.0.1",
    )
    session.commit()
    return redirect(f"/admin/students?cohort_assigned=1&courses={course_count}#student-list")


@app.post("/admin/students/{student_id}/update")
def admin_update_student_record(
    student_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    student_no: str = Form(...),
    name_en: str = Form(...),
    name_kh: str = Form(""),
    email: str = Form(...),
    is_active: bool = Form(False),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    student = session.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404)
    student_no = student_no.strip()
    name_en = name_en.strip()
    email = student_login_email(student_no, email)
    existing = session.exec(select(Student).where(Student.student_no == student_no, Student.id != student_id)).first()
    if existing:
        return redirect(f"/admin/students?edit_id={student_id}&error=student_exists#edit-record")

    student.student_no = student_no
    student.name_en = name_en
    student.name_kh = name_kh.strip() or None

    linked_user = user_by_email(session, email)
    if linked_user and linked_user.id != student.user_id:
        linked_student = session.exec(select(Student).where(Student.user_id == linked_user.id)).first()
        if student.user or linked_user.role != Role.STUDENT or linked_student:
            return redirect(f"/admin/students?edit_id={student_id}&error=email_exists#edit-record")

    account = student.user or linked_user
    latest_enrollment = session.exec(
        select(StudentSemesterEnrollment)
        .where(StudentSemesterEnrollment.student_id == student.id)
        .order_by(StudentSemesterEnrollment.created_at.desc(), StudentSemesterEnrollment.id.desc())
    ).first()
    program = session.get(Program, latest_enrollment.program_id) if latest_enrollment else None
    if account is None:
        account = User(
            name=name_en,
            email=email,
            password_hash=hash_password("password"),
            role=Role.STUDENT,
            is_active=is_active,
            faculty_id=program.faculty_id if program else None,
            program_id=program.id if program else None,
        )
        session.add(account)
        session.flush()
    else:
        account.name = name_en
        account.email = email
        account.role = Role.STUDENT
        account.is_active = is_active
        if program:
            account.faculty_id = program.faculty_id
            account.program_id = program.id
        session.add(account)
    student.user_id = account.id

    session.add(student)
    add_audit_record(
        session,
        user,
        "Student Management",
        "UPDATE",
        "Updated the student identity and single login without changing progression history.",
        student.student_no,
        request.client.host if request.client else "127.0.0.1",
    )
    session.commit()
    return redirect("/admin/students?updated=1")


@app.post("/admin/students/{student_id}/delete")
def admin_delete_student_record(
    student_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    student = session.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404)
    student_no = student.student_no
    for enrollment in session.exec(
        select(StudentSemesterEnrollment)
        .where(StudentSemesterEnrollment.student_id == student.id)
        .order_by(StudentSemesterEnrollment.id.desc())
    ).all():
        session.delete(enrollment)
    for enrollment in list(student.enrollments):
        session.delete(enrollment)
    for score in list(student.scores):
        session.delete(score)
    if student.user:
        student.user.is_active = False
        session.add(student.user)
    session.delete(student)
    add_audit_record(
        session,
        user,
        "Student Management",
        "DELETE",
        "Removed the student profile and study enrollments; the login account was deactivated for audit history.",
        student_no,
        request.client.host if request.client else "127.0.0.1",
    )
    session.commit()
    return redirect("/admin/students?deleted=1")


@app.post("/admin/faculties/create")
def admin_create_faculty_record(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    existing = session.exec(select(Faculty).where(Faculty.name == name)).first()
    if existing:
        return redirect("/admin/faculties?error=faculty_exists#add-faculty")
    session.add(Faculty(name=name))
    session.commit()
    return redirect("/admin/faculties?created=1")


@app.post("/admin/faculties/{faculty_id}/update")
def admin_update_faculty_record(
    faculty_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    faculty = session.get(Faculty, faculty_id)
    if not faculty:
        raise HTTPException(status_code=404)
    existing = session.exec(select(Faculty).where(Faculty.name == name, Faculty.id != faculty_id)).first()
    if existing:
        return redirect(f"/admin/faculties?edit_id={faculty_id}&error=faculty_exists#edit-record")
    faculty.name = name
    session.add(faculty)
    session.commit()
    return redirect("/admin/faculties?updated=1")


@app.post("/admin/faculties/{faculty_id}/delete")
def admin_delete_faculty_record(
    faculty_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    faculty = session.get(Faculty, faculty_id)
    if not faculty:
        raise HTTPException(status_code=404)
    has_programmes = session.exec(select(Program).where(Program.faculty_id == faculty_id)).first()
    if has_programmes:
        return redirect("/admin/faculties?error=faculty_has_programmes")
    session.delete(faculty)
    session.commit()
    return redirect("/admin/faculties?deleted=1")


@app.post("/admin/programmes/create")
def admin_create_programme_record(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    faculty_id: int = Form(...),
    code: str = Form(...),
    name: str = Form(...),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    existing = session.exec(select(Program).where(Program.code == code)).first()
    if existing:
        return redirect("/admin/programmes?error=programme_exists#add-programme")
    session.add(Program(faculty_id=faculty_id, code=code, name=name))
    session.commit()
    return redirect("/admin/programmes?created=1")


@app.post("/admin/programmes/{program_id}/update")
def admin_update_programme_record(
    program_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    faculty_id: int = Form(...),
    code: str = Form(...),
    name: str = Form(...),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    program = session.get(Program, program_id)
    if not program:
        raise HTTPException(status_code=404)
    existing = session.exec(select(Program).where(Program.code == code, Program.id != program_id)).first()
    if existing:
        return redirect(f"/admin/programmes?edit_id={program_id}&error=programme_exists#edit-record")
    program.faculty_id = faculty_id
    program.code = code
    program.name = name
    session.add(program)
    session.commit()
    return redirect("/admin/programmes?updated=1")


@app.post("/admin/programmes/{program_id}/delete")
def admin_delete_programme_record(
    program_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    program = session.get(Program, program_id)
    if not program:
        raise HTTPException(status_code=404)
    has_dependency = (
        session.exec(select(Course).where(Course.program_id == program_id)).first()
        or session.exec(select(PLO).where(PLO.program_id == program_id)).first()
        or session.exec(select(PEO).where(PEO.program_id == program_id)).first()
        or session.exec(select(User).where(User.program_id == program_id)).first()
    )
    if has_dependency:
        return redirect("/admin/programmes?error=programme_in_use")
    session.delete(program)
    session.commit()
    return redirect("/admin/programmes?deleted=1")


@app.post("/admin/academic-years/create")
def admin_create_academic_year_record(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    is_active: bool = Form(False),
    is_default: bool = Form(False),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    existing = session.exec(select(AcademicYear).where(AcademicYear.name == name)).first()
    if existing:
        return redirect("/admin/academic-years?error=academic_year_exists#add-academic-year")
    if is_default:
        for item in session.exec(select(AcademicYear)).all():
            item.is_default = False
            session.add(item)
    session.add(AcademicYear(name=name, start_date=start_date, end_date=end_date, is_active=is_active, is_default=is_default))
    session.commit()
    sync_study_periods_from_academic_semesters(session)
    return redirect("/admin/cohorts?created=1#years")


@app.post("/admin/academic-years/{year_id}/update")
def admin_update_academic_year_record(
    year_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    is_active: bool = Form(False),
    is_default: bool = Form(False),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    year = session.get(AcademicYear, year_id)
    if not year:
        raise HTTPException(status_code=404)
    existing = session.exec(select(AcademicYear).where(AcademicYear.name == name, AcademicYear.id != year_id)).first()
    if existing:
        return redirect(f"/admin/academic-years?edit_id={year_id}&error=academic_year_exists#edit-record")
    old_name = year.name
    if is_default:
        for item in session.exec(select(AcademicYear).where(AcademicYear.id != year_id)).all():
            item.is_default = False
            session.add(item)
    year.name = name
    year.start_date = start_date
    year.end_date = end_date
    year.is_active = is_active
    year.is_default = is_default
    session.add(year)
    for semester in session.exec(select(AcademicSemester).where(AcademicSemester.academic_year == old_name)).all():
        semester.academic_year = name
        session.add(semester)
    for course_class in session.exec(select(CourseClass).where(CourseClass.academic_year == old_name)).all():
        course_class.academic_year = name
        session.add(course_class)
    for enrollment in session.exec(select(StudentSemesterEnrollment).where(StudentSemesterEnrollment.academic_year == old_name)).all():
        enrollment.academic_year = name
        session.add(enrollment)
    for target in session.exec(select(PLOTarget).where(PLOTarget.academic_year == old_name)).all():
        target.academic_year = name
        session.add(target)
    session.commit()
    sync_study_periods_from_academic_semesters(session)
    return redirect("/admin/cohorts?updated=1#years")


@app.post("/admin/academic-years/{year_id}/delete")
def admin_delete_academic_year_record(
    year_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    year = session.get(AcademicYear, year_id)
    if not year:
        raise HTTPException(status_code=404)
    has_dependency = (
        session.exec(select(CourseClass).where(CourseClass.academic_year == year.name)).first()
        or session.exec(select(AcademicSemester).where(AcademicSemester.academic_year == year.name)).first()
    )
    if has_dependency:
        return redirect("/admin/academic-years?error=academic_year_in_use")
    session.delete(year)
    session.commit()
    sync_study_periods_from_academic_semesters(session)
    return redirect("/admin/cohorts?deleted=1#years")


@app.post("/admin/semesters/create")
def admin_create_semester_record(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
    code: str = Form(...),
    academic_year: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    is_active: bool = Form(False),
    is_default: bool = Form(False),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    semester_no = semester_number(name)
    logical_duplicate = next(
        (
            item for item in session.exec(select(AcademicSemester)).all()
            if item.academic_year == academic_year and semester_number(item.name) == semester_no
        ),
        None,
    )
    existing = session.exec(select(AcademicSemester).where(AcademicSemester.code == code)).first()
    if existing or logical_duplicate:
        return redirect("/admin/cohorts?error=semester_exists#semesters")
    if is_default:
        for item in session.exec(select(AcademicSemester)).all():
            item.is_default = False
            session.add(item)
    session.add(AcademicSemester(name=name, code=code, academic_year=academic_year, start_date=start_date, end_date=end_date, is_active=is_active, is_default=is_default))
    session.commit()
    sync_study_periods_from_academic_semesters(session)
    return redirect("/admin/cohorts?created=1#semesters")


@app.post("/admin/semesters/{semester_id}/update")
def admin_update_semester_record(
    semester_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
    code: str = Form(...),
    academic_year: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    is_active: bool = Form(False),
    is_default: bool = Form(False),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    semester = session.get(AcademicSemester, semester_id)
    if not semester:
        raise HTTPException(status_code=404)
    old_semester_no = semester_number(semester.name)
    semester_no = semester_number(name)
    logical_duplicate = next(
        (
            item for item in session.exec(select(AcademicSemester)).all()
            if item.id != semester_id
            and item.academic_year == academic_year
            and semester_number(item.name) == semester_no
        ),
        None,
    )
    existing = session.exec(select(AcademicSemester).where(AcademicSemester.code == code, AcademicSemester.id != semester_id)).first()
    if existing or logical_duplicate:
        return redirect(f"/admin/cohorts?edit_semester_id={semester_id}&error=semester_exists#semesters")
    if semester.academic_year != academic_year or old_semester_no != semester_no:
        used_by_class = any(
            item.academic_year == semester.academic_year
            and semester_number(item.semester) == old_semester_no
            for item in session.exec(select(CourseClass).where(CourseClass.academic_year == semester.academic_year)).all()
        )
        used_by_enrollment = any(
            item.academic_year == semester.academic_year
            and semester_number(item.semester) == old_semester_no
            for item in session.exec(
                select(StudentSemesterEnrollment).where(
                    StudentSemesterEnrollment.academic_year == semester.academic_year
                )
            ).all()
        )
        if used_by_class or used_by_enrollment:
            return redirect(f"/admin/cohorts?edit_semester_id={semester_id}&error=semester_in_use#semesters")
    if is_default:
        for item in session.exec(select(AcademicSemester).where(AcademicSemester.id != semester_id)).all():
            item.is_default = False
            session.add(item)
    semester.name = name
    semester.code = code
    semester.academic_year = academic_year
    semester.start_date = start_date
    semester.end_date = end_date
    semester.is_active = is_active
    semester.is_default = is_default
    session.add(semester)
    session.commit()
    sync_study_periods_from_academic_semesters(session)
    return redirect("/admin/cohorts?updated=1#semesters")


@app.post("/admin/semesters/{semester_id}/delete")
def admin_delete_semester_record(
    semester_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    semester = session.get(AcademicSemester, semester_id)
    if not semester:
        raise HTTPException(status_code=404)
    semester_no = semester_number(semester.name)
    used_by_class = any(
        semester_number(item.semester) == semester_no
        for item in session.exec(
            select(CourseClass).where(CourseClass.academic_year == semester.academic_year)
        ).all()
    )
    used_by_enrollment = any(
        semester_number(item.semester) == semester_no
        for item in session.exec(
            select(StudentSemesterEnrollment).where(
                StudentSemesterEnrollment.academic_year == semester.academic_year
            )
        ).all()
    )
    if used_by_class or used_by_enrollment:
        return redirect("/admin/cohorts?error=semester_in_use#semesters")
    session.delete(semester)
    session.commit()
    sync_study_periods_from_academic_semesters(session)
    return redirect("/admin/cohorts?deleted=1#semesters")


def build_class_code_from_parts(
    generation: str | None,
    programme_code: str | None,
    study_year: str | None,
    semester_no: str | None,
    shift_code: str | None,
    degree_code: str | None,
    group_no: str | None,
) -> str:
    generation = "".join(ch for ch in str(generation or "").strip() if ch.isalnum())
    programme_code = "".join(ch for ch in str(programme_code or "").strip().upper() if ch.isalnum())
    study_year = "".join(ch for ch in str(study_year or "").strip() if ch.isdigit()) or "1"
    semester_no = "".join(ch for ch in str(semester_no or "").strip() if ch.isdigit()) or "1"
    shift_code = (str(shift_code or "M").strip()[:1] or "M").upper()
    degree_code = (str(degree_code or "b").strip()[:1] or "b")
    group_no = "".join(ch for ch in str(group_no or "").strip() if ch.isdigit()) or "1"
    return f"{generation}{programme_code}{study_year}{semester_no}{shift_code}{degree_code}{group_no}"


def clean_class_code(name: str | None) -> str:
    return "".join(str(name or "").split())


def parse_class_code(code: str | None) -> dict | None:
    text = clean_class_code(code)
    if len(text) < 8:
        return None
    generation = text[:2]
    tail = text[-5:]
    if not (generation.isdigit() and tail[0].isdigit() and tail[1].isdigit() and tail[-1].isdigit()):
        return None
    programme_code = text[2:-5].upper()
    if not programme_code:
        return None
    return {
        "generation": generation,
        "programme_code": programme_code,
        "study_year": tail[0],
        "semester_no": tail[1],
        "shift_code": tail[2].upper(),
        "degree_code": tail[3],
        "group_no": tail[4],
    }


def class_code_for_semester(base_code: str, study_year: int, semester_no: int) -> str:
    parts = parse_class_code(base_code)
    if not parts:
        return clean_class_code(base_code)
    return build_class_code_from_parts(
        parts["generation"],
        parts["programme_code"],
        str(study_year),
        str(semester_no),
        parts["shift_code"],
        parts["degree_code"],
        parts["group_no"],
    )


def academic_year_for_study_year(base_academic_year: str, study_year: int) -> str:
    try:
        start = int(str(base_academic_year).split("-")[0]) + max(study_year - 1, 0)
        return f"{start}-{start + 1}"
    except Exception:
        return base_academic_year


def get_or_create_academic_year(session: Session, year_name: str) -> AcademicYear:
    existing = session.exec(select(AcademicYear).where(AcademicYear.name == year_name)).first()
    if existing:
        return existing
    start_year = year_name.split("-")[0]
    end_year = year_name.split("-")[-1]
    item = AcademicYear(name=year_name, start_date=f"Aug 01, {start_year}", end_date=f"Jul 31, {end_year}", is_active=True, is_default=False)
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


def get_or_create_academic_semester(session: Session, academic_year: str, semester_no: str) -> AcademicSemester:
    semester_no = semester_number(semester_no)
    code = f"{academic_year}-S{semester_no}"
    existing = session.exec(select(AcademicSemester).where(AcademicSemester.code == code)).first()
    if existing:
        return existing
    start_year = academic_year.split("-")[0]
    end_year = academic_year.split("-")[-1]
    item = AcademicSemester(
        name=f"Semester {semester_no}",
        code=code,
        academic_year=academic_year,
        start_date=f"Aug 01, {start_year}" if semester_no == "1" else f"Jan 02, {end_year}",
        end_date=f"Dec 31, {start_year}" if semester_no == "1" else f"May 31, {end_year}",
        is_active=True,
        is_default=False,
    )
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


def normalize_class_semester(value: str | None) -> str:
    digits = "".join(ch for ch in str(value or "").strip() if ch.isdigit())
    return digits or "1"


def cohort_duplicate_exists(session: Session, course_id: int, name: str, academic_year: str, semester: str, ignore_id: int | None = None) -> bool:
    stmt = select(CourseClass).where(
        CourseClass.course_id == course_id,
        CourseClass.name == name,
        CourseClass.academic_year == academic_year,
        CourseClass.semester == semester,
    )
    matches = session.exec(stmt).all()
    return any(item.id != ignore_id for item in matches)


def add_audit_record(
    session: Session,
    user: User,
    module: str,
    action: str,
    description: str,
    item_record: str,
    ip_address: str = "127.0.0.1",
    status: str = "Success",
) -> None:
    """Create one audit log entry for important admin actions."""
    session.add(
        AuditLog(
            date_time=datetime.now().strftime("%b %d, %Y %I:%M %p"),
            user_name=getattr(user, "name", None) or getattr(user, "username", None) or "Admin",
            module=module,
            action=action,
            description=description,
            item_record=item_record,
            ip_address=ip_address,
            status=status,
        )
    )


@app.post("/admin/cohorts/create")
def admin_create_cohort_record(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    course_id: int | None = Form(None),
    program_id: int | None = Form(None),
    name: str | None = Form(None),
    academic_year: str = Form(...),
    semester: str | None = Form(None),
    semester_start: str | None = Form(None),
    semester_end: str | None = Form(None),
    generation: str | None = Form(None),
    programme_code: str | None = Form(None),
    study_year: str | None = Form(None),
    semester_no: str | None = Form(None),
    shift_code: str | None = Form(None),
    degree_code: str | None = Form(None),
    group_no: str | None = Form(None),
    create_full_programme: str | None = Form("1"),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)

    target_program: Program | None = None
    class_name = clean_class_code(name)
    parsed = parse_class_code(class_name)

    if parsed and not programme_code:
        programme_code = parsed["programme_code"]
    if parsed and not study_year:
        study_year = parsed["study_year"]
    if parsed and not semester_no:
        semester_no = parsed["semester_no"]
    if parsed and not shift_code:
        shift_code = parsed["shift_code"]
    if parsed and not degree_code:
        degree_code = parsed["degree_code"]
    if parsed and not group_no:
        group_no = parsed["group_no"]

    if program_id:
        target_program = session.get(Program, program_id)
    if not target_program and programme_code:
        target_program = session.exec(select(Program).where(Program.code == str(programme_code).strip().upper())).first()
    if not target_program and course_id:
        course = session.get(Course, course_id)
        target_program = course.program if course and course.program else None
    if not target_program:
        return redirect("/admin/cohorts?error=missing_program#classes")

    if not class_name:
        class_name = build_class_code_from_parts(
            generation,
            programme_code or target_program.code,
            study_year,
            semester_no,
            shift_code,
            degree_code,
            group_no,
        )
    parsed = parse_class_code(class_name)
    if not parsed:
        return redirect("/admin/cohorts?error=invalid_class_code#classes")

    # One input such as 21ME11Mb1 creates the full 4-year/8-semester class structure.
    # Course rows are still created internally because CourseClass requires course_id,
    # but the UI shows only class/cohort records, not subjects.
    year_semester_pairs = [(year, sem) for year in range(1, 5) for sem in range(1, 3)] if create_full_programme else [(int(parsed["study_year"]), int(parsed["semester_no"]))]
    created = 0
    skipped = 0
    missing_course_semesters: list[str] = []

    for year_no, sem_no in year_semester_pairs:
        semester_text = str(sem_no)
        class_code = class_code_for_semester(class_name, year_no, sem_no)
        class_academic_year = academic_year_for_study_year(academic_year, year_no)
        get_or_create_academic_year(session, class_academic_year)
        get_or_create_academic_semester(session, class_academic_year, semester_text)

        target_courses = [
            c for c in session.exec(select(Course).where(Course.program_id == target_program.id, Course.curriculum_year == year_no)).all()
            if semester_number(c.curriculum_semester) == semester_text
        ]
        if not target_courses:
            missing_course_semesters.append(f"Y{year_no}S{sem_no}")
            continue
        for course in target_courses:
            if cohort_duplicate_exists(session, course.id, class_code, class_academic_year, semester_text):
                skipped += 1
                continue
            session.add(CourseClass(course_id=course.id, name=class_code, academic_year=class_academic_year, semester=semester_text, semester_start=semester_start, semester_end=semester_end))
            created += 1
    if created > 0:
        created_class_codes = [class_code_for_semester(class_name, year, sem) for year, sem in year_semester_pairs]
        unique_created_codes = sorted(set(created_class_codes))
        add_audit_record(
            session=session,
            user=user,
            module="Cohort / Batch",
            action="CREATE",
            description=f"Created cohort structure for {target_program.name}: {len(unique_created_codes)} class codes and {created} internal course-class rows.",
            item_record=f"{class_name} → {', '.join(unique_created_codes)}",
        )
    session.commit()
    if created == 0:
        return redirect("/admin/cohorts?error=duplicate_or_no_courses#classes")
    return redirect(f"/admin/cohorts?created=1&classes=8&courses={created}&skipped={skipped}#classes")


@app.post("/admin/cohorts/group/update")
def admin_update_cohort_group(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    old_code: str = Form(...),
    old_academic_year: str = Form(...),
    old_semester: str = Form(...),
    name: str = Form(...),
    academic_year: str = Form(...),
    semester_no: str = Form(...),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    old_code = clean_class_code(old_code)
    new_code = clean_class_code(name)
    new_semester = semester_number(semester_no)
    if not parse_class_code(new_code):
        return redirect("/admin/cohorts?error=invalid_class_code#classes")
    rows = session.exec(
        select(CourseClass).where(
            CourseClass.name == old_code,
            CourseClass.academic_year == old_academic_year,
            CourseClass.semester == semester_number(old_semester),
        )
    ).all()
    updated_count = 0
    for row in rows:
        row.name = new_code
        row.academic_year = academic_year
        row.semester = new_semester
        session.add(row)
        updated_count += 1
    if updated_count:
        add_audit_record(
            session=session,
            user=user,
            module="Cohort / Batch",
            action="UPDATE",
            description=f"Updated cohort/class code group ({updated_count} internal rows).",
            item_record=f"{old_code} {old_academic_year} S{semester_number(old_semester)} → {new_code} {academic_year} S{new_semester}",
        )
    session.commit()
    return redirect("/admin/cohorts?updated=1#classes")


@app.post("/admin/cohorts/group/delete")
def admin_delete_cohort_group(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    code: str = Form(...),
    academic_year: str = Form(...),
    semester: str = Form(...),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    class_code = clean_class_code(code)
    rows = session.exec(
        select(CourseClass).where(
            CourseClass.name == class_code,
            CourseClass.academic_year == academic_year,
            CourseClass.semester == semester_number(semester),
        )
    ).all()
    for row in rows:
        has_students = session.exec(select(ClassStudent).where(ClassStudent.class_id == row.id)).first()
        if has_students:
            return redirect("/admin/cohorts?error=cohort_has_students#classes")
    deleted_count = len(rows)
    for row in rows:
        for assignment in session.exec(select(ClassTeacher).where(ClassTeacher.class_id == row.id)).all():
            session.delete(assignment)
        session.delete(row)
    if deleted_count:
        add_audit_record(
            session=session,
            user=user,
            module="Cohort / Batch",
            action="DELETE",
            description=f"Deleted cohort/class code group ({deleted_count} internal rows).",
            item_record=f"{class_code} · {academic_year} · Semester {semester_number(semester)}",
        )
    session.commit()
    return redirect("/admin/cohorts?deleted=1#classes")


@app.post("/admin/cohorts/{cohort_id}/update")
def admin_update_cohort_record(
    cohort_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    course_id: int = Form(...),
    name: str | None = Form(None),
    academic_year: str = Form(...),
    semester: str | None = Form(None),
    semester_start: str | None = Form(None),
    semester_end: str | None = Form(None),
    generation: str | None = Form(None),
    programme_code: str | None = Form(None),
    study_year: str | None = Form(None),
    semester_no: str | None = Form(None),
    shift_code: str | None = Form(None),
    degree_code: str | None = Form(None),
    group_no: str | None = Form(None),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    cohort = session.get(CourseClass, cohort_id)
    if not cohort:
        raise HTTPException(status_code=404)
    course = session.get(Course, course_id)
    if not course:
        return redirect("/admin/cohorts?error=missing_course")
    class_name = clean_class_code(name) or build_class_code_from_parts(generation, programme_code or (course.program.code if course.program else ""), study_year, semester_no, shift_code, degree_code, group_no)
    class_semester = normalize_class_semester(semester or semester_no)
    if cohort_duplicate_exists(session, course_id, class_name, academic_year, class_semester, ignore_id=cohort_id):
        return redirect(f"/admin/cohorts?edit_id={cohort_id}&error=duplicate_class#edit-record")
    cohort.course_id = course_id
    cohort.name = class_name
    cohort.academic_year = academic_year
    cohort.semester = class_semester
    cohort.semester_start = semester_start
    cohort.semester_end = semester_end
    session.add(cohort)
    session.commit()
    return redirect("/admin/cohorts?updated=1")


@app.post("/admin/cohorts/{cohort_id}/delete")
def admin_delete_cohort_record(
    cohort_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    cohort = session.get(CourseClass, cohort_id)
    if not cohort:
        raise HTTPException(status_code=404)
    has_students = session.exec(select(ClassStudent).where(ClassStudent.class_id == cohort_id)).first()
    if has_students:
        return redirect("/admin/cohorts?error=cohort_has_students")
    for assignment in session.exec(select(ClassTeacher).where(ClassTeacher.class_id == cohort_id)).all():
        session.delete(assignment)
    session.delete(cohort)
    session.commit()
    return redirect("/admin/cohorts?deleted=1")


@app.post("/admin/plos/create")
def create_admin_plo(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    program_id: int = Form(...),
    plo_version_id: int | None = Form(None),
    code: str = Form(...),
    description: str = Form(...),
    domain: str = Form("Knowledge"),
    bloom_level: str = Form("C1"),
    status: str = Form("Active"),
    remark: str = Form(""),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    version = session.get(PLOVersion, plo_version_id) if plo_version_id else active_plo_version_for_program(session, program_id)
    if not version or version.programme_id != program_id:
        return redirect(f"/admin/plos?program_id={program_id}&error=invalid_version")
    if version.is_locked:
        return redirect(f"/admin/plos?program_id={program_id}&version_id={version.id}&error=version_locked")
    version_id = version.id if version else None
    if session.exec(select(PLO).where(PLO.program_id == program_id, PLO.plo_version_id == version_id, PLO.code == code)).first():
        return redirect(f"/admin/plos?program_id={program_id}&version_id={version.id}&error=duplicate")
    session.add(
        PLO(
            program_id=program_id,
            plo_version_id=version_id,
            code=code.strip(),
            description=description.strip(),
            domain=domain,
            bloom_level=bloom_level,
            status=status,
            remark=remark.strip(),
            created_by=user.id,
        )
    )
    session.commit()
    return redirect(f"/admin/plos?program_id={program_id}&version_id={version.id}&created=1")


@app.post("/admin/plos/{plo_id}/update")
def update_admin_plo(
    plo_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    program_id: int = Form(...),
    plo_version_id: int | None = Form(None),
    code: str = Form(...),
    description: str = Form(...),
    domain: str = Form("Knowledge"),
    bloom_level: str = Form("C1"),
    status: str = Form("Active"),
    remark: str = Form(""),
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    plo = session.get(PLO, plo_id)
    if not plo:
        raise HTTPException(status_code=404)
    version = session.get(PLOVersion, plo_version_id) if plo_version_id else active_plo_version_for_program(session, program_id)
    if not version or version.programme_id != program_id:
        return redirect(f"/admin/plos?program_id={program_id}&error=invalid_version")
    if version.is_locked:
        return redirect(f"/admin/plos?program_id={program_id}&version_id={version.id}&error=version_locked")
    version_id = version.id if version else None
    if session.exec(select(PLO).where(PLO.program_id == program_id, PLO.plo_version_id == version_id, PLO.code == code, PLO.id != plo_id)).first():
        return redirect(f"/admin/plos?program_id={program_id}&version_id={version.id}&error=duplicate")
    plo.program_id = program_id
    plo.plo_version_id = version_id
    plo.code = code.strip()
    plo.description = description.strip()
    plo.domain = domain
    plo.bloom_level = bloom_level
    plo.status = status
    plo.remark = remark.strip()
    plo.updated_at = datetime.utcnow()
    session.add(plo)
    session.commit()
    return redirect(f"/admin/plos?program_id={program_id}&version_id={version.id}&updated=1")


@app.post("/admin/plos/{plo_id}/delete")
def delete_admin_plo(plo_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    plo = session.get(PLO, plo_id)
    if plo:
        version = session.get(PLOVersion, plo.plo_version_id) if plo.plo_version_id else None
        if version and version.is_locked:
            return redirect(f"/admin/plos?program_id={plo.program_id}&version_id={version.id}&error=version_locked")
        plo.status = "Inactive"
        plo.updated_at = datetime.utcnow()
        session.add(plo)
        session.commit()
    return redirect(f"/admin/plos?program_id={plo.program_id}&version_id={plo.plo_version_id}&deleted=1" if plo else "/admin/plos?deleted=1")


@app.post("/admin/plos/{plo_id}/restore")
def restore_admin_plo(plo_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    plo = session.get(PLO, plo_id)
    if plo:
        version = session.get(PLOVersion, plo.plo_version_id) if plo.plo_version_id else None
        if version and version.is_locked:
            return redirect(f"/admin/plos?program_id={plo.program_id}&version_id={version.id}&error=version_locked")
        plo.status = "Active"
        plo.updated_at = datetime.utcnow()
        session.add(plo)
        session.commit()
    return redirect(f"/admin/plos?program_id={plo.program_id}&version_id={plo.plo_version_id}&restored=1" if plo else "/admin/plos?restored=1")


@app.post("/admin/outcome-versions/create")
async def create_admin_outcome_version(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Create one Draft CQI package containing PEOs, PLOs, and their mappings."""
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = (optional_int(form.get("program_id")) or 0)
    return_section = str(form.get("return_section") or "plos")
    if return_section not in {"plos", "peos", "targets", "outcome-versions", "programme-mapping"}:
        return_section = "plos"
    return_path = f"/admin/{return_section}"
    name = str(form.get("version_name") or "").strip()
    source_id = (optional_int(form.get("source_version_id")) or 0)
    program = session.get(Program, program_id)
    if not program or not name:
        return redirect(f"{return_path}?version_error=required")
    if session.exec(select(PLOVersion).where(PLOVersion.programme_id == program.id, PLOVersion.version_name == name)).first():
        return redirect(f"{return_path}?program_id={program.id}&version_error=duplicate")
    source = session.get(PLOVersion, source_id) if source_id else None
    if source and source.programme_id != program.id:
        return redirect(f"{return_path}?program_id={program.id}&version_error=invalid_source")
    version = PLOVersion(programme_id=program.id, version_name=name, status="Draft", is_locked=False, created_by=user.id)
    session.add(version)
    session.flush()
    if source:
        plo_ids: dict[int, int] = {}
        peo_ids: dict[int, int] = {}
        for old in session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == source.id)).all():
            new = PLO(program_id=program.id, plo_version_id=version.id, code=old.code, description=old.description, domain=old.domain, bloom_level=old.bloom_level, status=old.status, remark=old.remark, created_by=user.id)
            session.add(new); session.flush(); plo_ids[old.id] = new.id
        for old in session.exec(select(PEO).where(PEO.program_id == program.id, PEO.plo_version_id == source.id)).all():
            new = PEO(program_id=program.id, plo_version_id=version.id, code=old.code, description=old.description, status=old.status, remark=old.remark, created_by=user.id)
            session.add(new); session.flush(); peo_ids[old.id] = new.id
        for old in session.exec(select(PEOPLOMapping).where(PEOPLOMapping.plo_version_id == source.id)).all():
            if old.peo_id in peo_ids and old.plo_id in plo_ids:
                session.add(PEOPLOMapping(program_id=program.id, plo_version_id=version.id, peo_id=peo_ids[old.peo_id], plo_id=plo_ids[old.plo_id], mapping_mode=old.mapping_mode, is_mapped=old.is_mapped, contribution_percentage=old.contribution_percentage, created_by=user.id))
        for old in session.exec(select(CLOPLOMapping)).all():
            if old.plo_id in plo_ids:
                session.add(CLOPLOMapping(clo_id=old.clo_id, plo_id=plo_ids[old.plo_id], weight=old.weight))
        for old in session.exec(select(PLOTarget).where(PLOTarget.program_id == program.id)).all():
            if old.plo_id in plo_ids:
                session.add(PLOTarget(program_id=program.id, plo_id=plo_ids[old.plo_id], academic_year=old.academic_year, cohort=old.cohort, target=old.target, set_by=user.name, updated_at=datetime.utcnow().strftime("%b %d, %Y %I:%M %p")))
    session.commit()
    return redirect(f"{return_path}?program_id={program.id}&version_id={version.id}&version_created=1")


@app.post("/admin/outcome-versions/publish")
async def publish_admin_outcome_version(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = (optional_int(form.get("program_id")) or 0)
    return_section = str(form.get("return_section") or "plos")
    if return_section not in {"plos", "peos", "targets", "outcome-versions", "programme-mapping"}:
        return_section = "plos"
    return_path = f"/admin/{return_section}"
    version_id = (optional_int(form.get("version_id")) or 0)
    version = session.get(PLOVersion, version_id)
    if not version or version.programme_id != program_id:
        return redirect(f"{return_path}?program_id={program_id}&version_error=invalid_version")
    plos = session.exec(select(PLO).where(PLO.program_id == program_id, PLO.plo_version_id == version.id, PLO.status == "Active")).all()
    if not plos:
        return redirect(f"{return_path}?program_id={program_id}&version_id={version.id}&version_error=incomplete_version")
    version.status = "Published"
    version.is_locked = True
    version.updated_at = datetime.utcnow()
    session.add(version)
    session.commit()
    return redirect(f"{return_path}?program_id={program_id}&version_id={version.id}&version_published=1")


@app.post("/admin/outcome-versions/unlock")
async def unlock_admin_outcome_version(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = (optional_int(form.get("program_id")) or 0)
    return_section = str(form.get("return_section") or "plos")
    if return_section not in {"plos", "peos", "targets", "outcome-versions", "programme-mapping"}:
        return_section = "plos"
    return_path = f"/admin/{return_section}"
    version_id = (optional_int(form.get("version_id")) or 0)
    version = session.get(PLOVersion, version_id)
    if not version or version.programme_id != program_id:
        return redirect(f"{return_path}?program_id={program_id}&version_error=invalid_version")
    version.status = "Draft"
    version.is_locked = False
    version.updated_at = datetime.utcnow()
    session.add(version)
    session.commit()
    return redirect(f"{return_path}?program_id={program_id}&version_id={version.id}&version_unlocked=1")


@app.post("/admin/outcome-versions/delete")
async def delete_admin_outcome_version(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Permanently remove an unused Draft outcome package and its dependants."""
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = (optional_int(form.get("program_id")) or 0)
    version_id = (optional_int(form.get("version_id")) or 0)
    return_section = str(form.get("return_section") or "plos")
    if return_section not in {"plos", "peos", "targets", "outcome-versions", "programme-mapping"}:
        return_section = "plos"
    return_path = f"/admin/{return_section}"
    program = session.get(Program, program_id)
    version = session.get(PLOVersion, version_id)
    if not program or not version or version.programme_id != program.id:
        return redirect(f"{return_path}?program_id={program_id}&version_error=invalid_version")

    versions = list(
        session.exec(
            select(PLOVersion)
            .where(PLOVersion.programme_id == program.id)
            .order_by(PLOVersion.id.desc())
        ).all()
    )
    if len(versions) <= 1:
        return redirect(
            f"{return_path}?program_id={program.id}&version_id={version.id}&version_error=last_version"
        )
    assignment = session.exec(
        select(CohortOutcomeVersion).where(CohortOutcomeVersion.outcome_version_id == version.id)
    ).first()
    if assignment:
        return redirect(
            f"{return_path}?program_id={program.id}&version_id={version.id}&version_error=assigned_version"
        )

    plos = list(
        session.exec(
            select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)
        ).all()
    )
    peos = list(
        session.exec(
            select(PEO).where(PEO.program_id == program.id, PEO.plo_version_id == version.id)
        ).all()
    )
    plo_ids = {item.id for item in plos if item.id is not None}
    peo_ids = {item.id for item in peos if item.id is not None}

    # Delete referencing rows before their PEO/PLO parents to satisfy FK rules.
    for mapping in session.exec(select(PEOPLOMapping)).all():
        if (
            mapping.plo_version_id == version.id
            or mapping.peo_id in peo_ids
            or mapping.plo_id in plo_ids
        ):
            session.delete(mapping)
    if plo_ids:
        for mapping in session.exec(
            select(CLOPLOMapping).where(CLOPLOMapping.plo_id.in_(plo_ids))
        ).all():
            session.delete(mapping)
        for mapping in session.exec(
            select(CoursePLOMapping).where(CoursePLOMapping.plo_id.in_(plo_ids))
        ).all():
            session.delete(mapping)
        for target in session.exec(
            select(PLOTarget).where(PLOTarget.plo_id.in_(plo_ids))
        ).all():
            session.delete(target)
    for peo in peos:
        session.delete(peo)
    for plo in plos:
        session.delete(plo)

    deleted_name = version.version_name
    remaining = [item for item in versions if item.id != version_id]
    session.delete(version)
    session.commit()
    add_audit_record(
        session,
        user,
        "Outcome Version Management",
        "DELETE",
        f"Deleted unassigned outcome version {deleted_name} and its dependent mappings.",
        program.code,
        request.client.host if request.client else "127.0.0.1",
    )
    fallback = next(
        (item for item in remaining if item.status in {"Active", "Published"}),
        remaining[0],
    )
    return redirect(
        f"{return_path}?program_id={program.id}&version_id={fallback.id}&version_deleted=1"
    )


@app.post("/admin/outcome-versions/assign")
async def assign_admin_outcome_version(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = (optional_int(form.get("program_id")) or 0)
    return_section = str(form.get("return_section") or "plos")
    if return_section not in {"plos", "peos", "targets", "outcome-versions", "programme-mapping"}:
        return_section = "plos"
    return_path = f"/admin/{return_section}"
    version_id = (optional_int(form.get("version_id")) or 0)
    cohort_name = outcome_cohort_key(form.get("cohort_name"))
    version = session.get(PLOVersion, version_id)
    if not cohort_name or not version or version.programme_id != program_id:
        return redirect(f"{return_path}?program_id={program_id}&version_error=invalid_assignment")
    if not version.is_locked or version.status not in {"Active", "Published"}:
        return redirect(f"{return_path}?program_id={program_id}&version_id={version.id}&version_error=publish_before_assignment")
    record = next(
        (
            item for item in session.exec(
                select(CohortOutcomeVersion).where(CohortOutcomeVersion.programme_id == program_id)
            ).all()
            if outcome_cohort_key(item.cohort_name) == cohort_name
        ),
        None,
    )
    if record:
        record.outcome_version_id = version.id
        record.cohort_name = cohort_name
        record.assigned_by = user.id
        record.assigned_at = datetime.utcnow()
        session.add(record)
        session.commit()
    elif not record:
        session.add(CohortOutcomeVersion(programme_id=program_id, cohort_name=cohort_name, outcome_version_id=version.id, assigned_by=user.id))
        session.commit()
    return redirect(f"{return_path}?program_id={program_id}&version_id={version.id}&version_assigned=1")


@app.post("/admin/peos/create")
async def create_admin_peo(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = (optional_int(form.get("program_id")) or 0)
    plo_version_id = optional_int(form.get("plo_version_id"))
    code = str(form.get("code") or "").strip().upper()
    description = str(form.get("description") or "").strip()
    status = str(form.get("status") or "Active")
    remark = str(form.get("remark") or "")
    version = session.get(PLOVersion, plo_version_id) if plo_version_id else active_plo_version_for_program(session, program_id)
    if not version or version.programme_id != program_id:
        return redirect(f"/admin/peos?program_id={program_id}&error=invalid_version")
    if version.is_locked:
        return redirect(f"/admin/peos?program_id={program_id}&version_id={version.id}&error=version_locked")
    version_id = version.id
    if session.exec(select(PEO).where(PEO.program_id == program_id, PEO.plo_version_id == version_id, PEO.code == code)).first():
        return redirect(f"/admin/peos?program_id={program_id}&version_id={version.id}&error=duplicate")
    peo = PEO(
        program_id=program_id,
        plo_version_id=version_id,
        code=code,
        description=description,
        status=status,
        remark=remark.strip(),
        created_by=user.id,
    )
    session.add(peo)
    session.flush()
    save_admin_peo_inline_weights(session, form, peo, program_id, version_id)
    session.commit()
    return redirect(f"/admin/peos?program_id={program_id}&version_id={version.id}&created=1")


@app.post("/admin/peos/{peo_id}/update")
async def update_admin_peo(
    peo_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    peo = session.get(PEO, peo_id)
    if not peo:
        raise HTTPException(status_code=404)
    form = await request.form()
    program_id = (optional_int(form.get("program_id")) or 0)
    plo_version_id = optional_int(form.get("plo_version_id"))
    code = str(form.get("code") or "").strip().upper()
    description = str(form.get("description") or "").strip()
    status = str(form.get("status") or "Active")
    remark = str(form.get("remark") or "")
    version = session.get(PLOVersion, plo_version_id) if plo_version_id else active_plo_version_for_program(session, program_id)
    if not version or version.programme_id != program_id:
        return redirect(f"/admin/peos?program_id={program_id}&error=invalid_version")
    if version.is_locked:
        return redirect(f"/admin/peos?program_id={program_id}&version_id={version.id}&error=version_locked")
    version_id = version.id
    if session.exec(select(PEO).where(PEO.program_id == program_id, PEO.plo_version_id == version_id, PEO.code == code, PEO.id != peo_id)).first():
        return redirect(f"/admin/peos?program_id={program_id}&version_id={version.id}&error=duplicate")
    peo.program_id = program_id
    peo.plo_version_id = version_id
    peo.code = code
    peo.description = description
    peo.status = status
    peo.remark = remark.strip()
    peo.updated_at = datetime.utcnow()
    session.add(peo)
    save_admin_peo_inline_weights(session, form, peo, program_id, version_id)
    session.commit()
    return redirect(f"/admin/peos?program_id={program_id}&version_id={version.id}&updated=1")


@app.post("/admin/peos/{peo_id}/delete")
def delete_admin_peo(peo_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    peo = session.get(PEO, peo_id)
    if peo:
        version = session.get(PLOVersion, peo.plo_version_id) if peo.plo_version_id else None
        if version and version.is_locked:
            return redirect(f"/admin/peos?program_id={peo.program_id}&version_id={version.id}&error=version_locked")
        peo.status = "Inactive"
        peo.updated_at = datetime.utcnow()
        session.add(peo)
        session.commit()
    return redirect(f"/admin/peos?program_id={peo.program_id}&version_id={peo.plo_version_id}&deleted=1" if peo else "/admin/peos?deleted=1")


@app.post("/admin/peos/{peo_id}/restore")
def restore_admin_peo(peo_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    peo = session.get(PEO, peo_id)
    if peo:
        version = session.get(PLOVersion, peo.plo_version_id) if peo.plo_version_id else None
        if version and version.is_locked:
            return redirect(f"/admin/peos?program_id={peo.program_id}&version_id={version.id}&error=version_locked")
        peo.status = "Active"
        peo.updated_at = datetime.utcnow()
        session.add(peo)
        session.commit()
    return redirect(f"/admin/peos?program_id={peo.program_id}&version_id={peo.plo_version_id}&restored=1" if peo else "/admin/peos?restored=1")


def save_admin_peo_inline_weights(session: Session, form, peo: PEO, program_id: int, version_id: int) -> None:
    for mapping in session.exec(select(PEOPLOMapping).where(PEOPLOMapping.peo_id == peo.id, PEOPLOMapping.plo_version_id == version_id)).all():
        session.delete(mapping)
    plos = session.exec(
        select(PLO).where(PLO.program_id == program_id, PLO.plo_version_id == version_id, PLO.status == "Active")
    ).all()
    for plo in plos:
        weight = clamp_percent(form.get(f"weight_{plo.id}"))
        if plo.id and weight > 0:
            session.add(
                PEOPLOMapping(
                    program_id=program_id,
                    plo_version_id=version_id,
                    peo_id=peo.id,
                    plo_id=plo.id,
                    weight=weight,
                    contribution_percentage=weight,
                    updated_at=datetime.utcnow(),
                )
            )


@app.post("/admin/peos/mapping")
async def update_admin_peo_mapping_matrix(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = (optional_int(form.get("program_id")) or 0)
    plo_version_id = (optional_int(form.get("plo_version_id")) or 0)
    mapping_mode = str(form.get("mapping_mode") or "percentage")
    version = session.get(PLOVersion, plo_version_id)
    return_url = f"/admin/peos?program_id={program_id}&version_id={plo_version_id}"
    if not version or version.programme_id != program_id:
        return redirect(f"{return_url}&error=invalid_version#peo-mapping")
    if version.is_locked:
        return redirect(f"{return_url}&error=version_locked#peo-mapping")
    peos = session.exec(select(PEO).where(PEO.program_id == program_id, PEO.plo_version_id == plo_version_id, PEO.status == "Active").order_by(PEO.code)).all()
    plos = session.exec(select(PLO).where(PLO.program_id == program_id, PLO.plo_version_id == plo_version_id, PLO.status == "Active").order_by(PLO.code)).all()
    if not peos or not plos:
        return redirect(f"{return_url}&error=mapping_empty#peo-mapping")

    if mapping_mode == "percentage":
        for peo in peos:
            total = sum(clamp_percent(form.get(f"map_{peo.id}_{plo.id}")) for plo in plos)
            if abs(total - 100) > 0.01:
                return redirect(f"{return_url}&error=mapping_total#peo-map-row-{peo.id}")

    for peo in peos:
        for existing in session.exec(select(PEOPLOMapping).where(PEOPLOMapping.peo_id == peo.id, PEOPLOMapping.plo_version_id == plo_version_id)).all():
            session.delete(existing)
        for plo in plos:
            if mapping_mode == "check":
                is_mapped = form.get(f"check_{peo.id}_{plo.id}") == "true"
                percentage = 100.0 if is_mapped else 0.0
            else:
                percentage = clamp_percent(form.get(f"map_{peo.id}_{plo.id}"))
                is_mapped = percentage > 0
            if is_mapped:
                session.add(
                    PEOPLOMapping(
                        program_id=program_id,
                        plo_version_id=plo_version_id,
                        peo_id=peo.id,
                        plo_id=plo.id,
                        mapping_mode=mapping_mode,
                        is_mapped=is_mapped,
                        contribution_percentage=percentage,
                        weight=percentage,
                        created_by=user.id,
                        updated_at=datetime.utcnow(),
                    )
                )
    session.commit()
    return redirect(f"{return_url}&mapping=updated#peo-mapping")


@app.post("/admin/targets/create")
def create_admin_target(session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)], program_id: int = Form(...), version_id: int = Form(...), plo_id: int = Form(...), academic_year: str = Form(...), cohort: str = Form(...), target: float = Form(...)):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    plo = session.get(PLO, plo_id)
    if not plo or plo.program_id != program_id or plo.plo_version_id != version_id:
        return redirect(f"/admin/targets?program_id={program_id}&version_id={version_id}&error=invalid_plo")
    if session.exec(select(PLOTarget).where(PLOTarget.program_id == program_id, PLOTarget.plo_id == plo_id, PLOTarget.academic_year == academic_year, PLOTarget.cohort == cohort)).first():
        return redirect(f"/admin/targets?program_id={program_id}&version_id={version_id}&error=duplicate")
    session.add(PLOTarget(program_id=program_id, plo_id=plo_id, academic_year=academic_year, cohort=cohort, target=target, set_by=user.name))
    session.commit()
    return redirect(f"/admin/targets?program_id={program_id}&version_id={version_id}&created=1")


@app.post("/admin/targets/{target_id}/update")
def update_admin_target(target_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)], program_id: int = Form(...), version_id: int = Form(...), plo_id: int = Form(...), academic_year: str = Form(...), cohort: str = Form(...), target: float = Form(...)):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    target_record = session.get(PLOTarget, target_id)
    if not target_record:
        raise HTTPException(status_code=404)
    plo = session.get(PLO, plo_id)
    if not plo or plo.program_id != program_id or plo.plo_version_id != version_id:
        return redirect(f"/admin/targets?program_id={program_id}&version_id={version_id}&error=invalid_plo")
    target_record.program_id = program_id
    target_record.plo_id = plo_id
    target_record.academic_year = academic_year
    target_record.cohort = cohort
    target_record.target = target
    target_record.set_by = user.name
    target_record.updated_at = "May 15, 2024 10:30 AM"
    session.add(target_record)
    session.commit()
    return redirect(f"/admin/targets?program_id={program_id}&version_id={version_id}&updated=1")


@app.post("/admin/targets/{target_id}/delete")
def delete_admin_target(target_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    target_record = session.get(PLOTarget, target_id)
    return_url = "/admin/targets?deleted=1"
    if target_record:
        plo = session.get(PLO, target_record.plo_id)
        version_id = plo.plo_version_id if plo else ""
        return_url = f"/admin/targets?program_id={target_record.program_id}&version_id={version_id}&deleted=1"
        session.delete(target_record)
        session.commit()
    return redirect(return_url)


@app.post("/admin/reports/create")
def create_admin_report(session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)], name: str = Form(...), category: str = Form(...), description: str = Form(...), format: str = Form("PDF"), status: str = Form("Ready")):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    session.add(SystemReport(name=name, category=category, description=description, format=format, status=status, created_by=user.name))
    session.commit()
    return redirect("/admin/reports?created=1")


@app.post("/admin/reports/{report_id}/update")
def update_admin_report(report_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)], name: str = Form(...), category: str = Form(...), description: str = Form(...), format: str = Form("PDF"), status: str = Form("Ready")):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    report = session.get(SystemReport, report_id)
    if not report:
        raise HTTPException(status_code=404)
    report.name = name
    report.category = category
    report.description = description
    report.format = format
    report.status = status
    report.created_by = user.name
    session.add(report)
    session.commit()
    return redirect("/admin/reports?updated=1")


@app.post("/admin/reports/{report_id}/delete")
def delete_admin_report(report_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    report = session.get(SystemReport, report_id)
    if report:
        session.delete(report)
        session.commit()
    return redirect("/admin/reports?deleted=1")


@app.post("/admin/audit-logs/create")
def create_admin_audit_log(session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)], date_time: str = Form(...), user_name: str = Form(...), module: str = Form(...), action: str = Form(...), description: str = Form(...), item_record: str = Form(...), ip_address: str = Form(...), status: str = Form(...)):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    session.add(AuditLog(date_time=date_time, user_name=user_name, module=module, action=action, description=description, item_record=item_record, ip_address=ip_address, status=status))
    session.commit()
    return redirect("/admin/audit-logs?created=1")


@app.post("/admin/audit-logs/{log_id}/update")
def update_admin_audit_log(log_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)], date_time: str = Form(...), user_name: str = Form(...), module: str = Form(...), action: str = Form(...), description: str = Form(...), item_record: str = Form(...), ip_address: str = Form(...), status: str = Form(...)):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    log = session.get(AuditLog, log_id)
    if not log:
        raise HTTPException(status_code=404)
    log.date_time = date_time
    log.user_name = user_name
    log.module = module
    log.action = action
    log.description = description
    log.item_record = item_record
    log.ip_address = ip_address
    log.status = status
    session.add(log)
    session.commit()
    return redirect("/admin/audit-logs?updated=1")


@app.post("/admin/audit-logs/{log_id}/delete")
def delete_admin_audit_log(log_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    log = session.get(AuditLog, log_id)
    if log:
        session.delete(log)
        session.commit()
    return redirect("/admin/audit-logs?deleted=1")


@app.get("/admin/audit-logs/export")
def export_admin_audit_logs(session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    rows = session.exec(select(AuditLog).order_by(AuditLog.id)).all()
    lines = ["Date & Time,User,Module,Action,Description,Item,IP Address,Status"]
    lines.extend(f'"{row.date_time}","{row.user_name}","{row.module}","{row.action}","{row.description}","{row.item_record}","{row.ip_address}","{row.status}"' for row in rows)
    return Response("\n".join(lines), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=audit-logs.csv"})


def setting_category(key: str) -> str:
    if key in {"system_name", "institution_name", "address", "timezone", "date_format", "time_format", "default_language", "number_format", "currency", "multilingual_support"}:
        return "general"
    if key in {"academic_year", "default_semester", "passing_score", "attainment_target", "grading_scheme", "allow_manual_data_entry"}:
        return "academic"
    if key in {"excellent_min", "good_min", "satisfactory_min", "needs_improvement_max"}:
        return "grading"
    if key in {"email_notifications", "in_app_notifications", "sms_notifications", "notify_assessments", "notify_reports"}:
        return "notifications"
    if key in {"smtp_status", "smtp_host", "smtp_port", "smtp_security", "smtp_last_test", "from_email", "email_template_footer"}:
        return "email"
    if key in {"session_timeout", "password_expiry", "two_factor_auth", "login_attempt_limit"}:
        return "security"
    if key in {"auto_backup", "backup_frequency", "last_backup", "last_backup_file"}:
        return "backup"
    if key in {"theme", "primary_color", "compact_mode"}:
        return "theme"
    if key in {"api_enabled", "webhook_url", "lms_integration"}:
        return "integrations"
    if key in {"audit_retention", "audit_export_format", "audit_tracking"}:
        return "audit"
    return "other"


def ensure_system_settings(session: Session) -> dict[str, str]:
    create_db_and_tables()
    existing = {item.key: item for item in session.exec(select(SystemSetting)).all()}
    changed = False
    for key, value in DEFAULT_SYSTEM_SETTINGS.items():
        if key not in existing:
            session.add(SystemSetting(key=key, value=value, category=setting_category(key)))
            changed = True
        elif not existing[key].category:
            existing[key].category = setting_category(key)
            session.add(existing[key])
            changed = True

    # Remove the original demo SMTP values. They made the page appear connected
    # even though no connection had ever been attempted.
    smtp_host = existing.get("smtp_host")
    if smtp_host and smtp_host.value == "smtp.university.edu.kh":
        for key, value in {
            "smtp_host": "",
            "from_email": "",
            "smtp_status": "Not tested",
            "smtp_last_test": "Never",
        }.items():
            if key in existing:
                existing[key].value = value
                session.add(existing[key])
        changed = True
    if changed:
        session.commit()
    return {item.key: item.value for item in session.exec(select(SystemSetting)).all()}


def save_setting(session: Session, key: str, value: str, category: str | None = None) -> None:
    setting = session.exec(select(SystemSetting).where(SystemSetting.key == key)).first()
    if setting:
        setting.value = value
        setting.category = category or setting.category or setting_category(key)
    else:
        setting = SystemSetting(key=key, value=value, category=category or setting_category(key))
    session.add(setting)


def system_setting_values() -> dict[str, str]:
    """Return live settings for shared template elements such as the brand."""
    try:
        with Session(engine) as session:
            return {item.key: item.value for item in session.exec(select(SystemSetting)).all()}
    except Exception:
        return DEFAULT_SYSTEM_SETTINGS.copy()


templates.env.globals["system_setting_values"] = system_setting_values


def sqlite_database_path() -> Path | None:
    """Resolve the live SQLite database without assuming the working directory."""
    if engine.url.get_backend_name() != "sqlite" or not engine.url.database or engine.url.database == ":memory:":
        return None
    path = Path(engine.url.database).expanduser()
    return (Path.cwd() / path).resolve() if not path.is_absolute() else path.resolve()


def human_file_size(size: int) -> str:
    value = float(max(size, 0))
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.0f} {unit}" if unit == "B" else f"{value:.1f} {unit}"
        value /= 1024
    return f"{value:.1f} GB"


def database_size_label() -> str:
    path = sqlite_database_path()
    if not path or not path.exists():
        return "External DB" if engine.url.get_backend_name() != "sqlite" else "Unavailable"
    return human_file_size(path.stat().st_size)


def database_backup_directory() -> Path:
    path = Path(app_config.export_dir).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    return (path / "backups").resolve()


def latest_database_backup() -> Path | None:
    directory = database_backup_directory()
    if not directory.exists():
        return None
    backups = [item for item in directory.glob("obe_plo_*.db") if item.is_file()]
    return max(backups, key=lambda item: item.stat().st_mtime) if backups else None


def create_database_backup() -> Path:
    """Create a transaction-safe copy of the live SQLite database."""
    source_path = sqlite_database_path()
    if source_path is None:
        raise RuntimeError("Database backup is currently supported for SQLite only.")
    if not source_path.exists():
        raise RuntimeError("The live database file could not be found.")

    directory = database_backup_directory()
    directory.mkdir(parents=True, exist_ok=True)
    backup_path = directory / f"obe_plo_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.db"
    with sqlite3.connect(str(source_path)) as source_connection:
        with sqlite3.connect(str(backup_path)) as backup_connection:
            source_connection.backup(backup_connection)
    return backup_path


def sync_setting_academic_defaults(session: Session) -> StudyPeriod | None:
    """Make the settings display follow the one real current Study Period."""
    periods = sync_study_periods_from_academic_semesters(session)
    current = next((period for period in periods if period.is_current), periods[0] if periods else None)
    if current:
        save_setting(session, "academic_year", current.academic_year)
        save_setting(session, "default_semester", f"Semester {current.semester}")
        session.commit()
    return current


def apply_default_academic_period(session: Session, academic_year: str, semester_value: str) -> str | None:
    """Set the admin-selected default in AcademicYear, AcademicSemester and StudyPeriod."""
    year = session.exec(
        select(AcademicYear).where(AcademicYear.name == academic_year, AcademicYear.is_active == True)  # noqa: E712
    ).first()
    if year is None:
        return "Select an active academic year created in Academic Year management."

    semester_no = int(semester_number(semester_value))
    semester = next(
        (
            item
            for item in session.exec(
                select(AcademicSemester).where(
                    AcademicSemester.academic_year == academic_year,
                    AcademicSemester.is_active == True,  # noqa: E712
                )
            ).all()
            if int(semester_number(item.name or item.code)) == semester_no
        ),
        None,
    )
    if semester is None:
        return f"Semester {semester_no} is not active for {academic_year}."

    for item in session.exec(select(AcademicYear)).all():
        item.is_default = item.id == year.id
        session.add(item)
    for item in session.exec(select(AcademicSemester)).all():
        item.is_default = item.id == semester.id
        session.add(item)
    session.commit()
    sync_study_periods_from_academic_semesters(session)
    return None


READ_ONLY_SYSTEM_SETTINGS = {
    "smtp_status",
    "smtp_last_test",
    "last_backup",
    "last_backup_file",
    "version",
}


def validate_system_setting_form(form) -> tuple[dict[str, str], str | None]:
    values = {
        key: str(form.get(key) or "").strip()
        for key in DEFAULT_SYSTEM_SETTINGS
        if key not in READ_ONLY_SYSTEM_SETTINGS and key in form
    }
    for key, label in {"system_name": "System name", "institution_name": "Institution name"}.items():
        if key in values and not values[key]:
            return {}, f"{label} is required."

    percentage_keys = ("passing_score", "attainment_target", "excellent_min", "good_min", "satisfactory_min", "needs_improvement_max")
    percentages: dict[str, float] = {}
    for key in percentage_keys:
        if key not in values:
            continue
        try:
            percentages[key] = float(values[key])
        except ValueError:
            return {}, f"{key.replace('_', ' ').title()} must be a number."
        if percentages[key] < 0 or percentages[key] > 100:
            return {}, f"{key.replace('_', ' ').title()} must be between 0 and 100."

    grade_keys = ("needs_improvement_max", "satisfactory_min", "good_min", "excellent_min")
    if all(key in percentages for key in grade_keys):
        if not (
            percentages["needs_improvement_max"]
            < percentages["satisfactory_min"]
            <= percentages["good_min"]
            <= percentages["excellent_min"]
        ):
            return {}, "Grading thresholds must increase from Needs Improvement to Excellent."

    if "smtp_port" in values:
        try:
            smtp_port = int(values["smtp_port"])
        except ValueError:
            return {}, "SMTP port must be a whole number."
        if smtp_port < 1 or smtp_port > 65535:
            return {}, "SMTP port must be between 1 and 65535."
    if "primary_color" in values and not re.fullmatch(r"#[0-9A-Fa-f]{6}", values["primary_color"]):
        return {}, "Primary color must be a valid hex color."
    return values, None


@app.get("/admin/settings", response_class=HTMLResponse)
def admin_settings_page(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    settings = ensure_system_settings(session)
    current_period = sync_setting_academic_defaults(session)
    settings = ensure_system_settings(session)
    academic_years = list(
        session.exec(
            select(AcademicYear)
            .where(AcademicYear.is_active == True)  # noqa: E712
            .order_by(AcademicYear.name.desc())
        ).all()
    )
    academic_semesters = list(
        session.exec(
            select(AcademicSemester)
            .where(AcademicSemester.is_active == True)  # noqa: E712
            .order_by(AcademicSemester.academic_year.desc(), AcademicSemester.code)
        ).all()
    )
    semester_options_by_year: dict[str, list[str]] = {}
    for semester in academic_semesters:
        label = f"Semester {semester_number(semester.name or semester.code)}"
        semester_options_by_year.setdefault(semester.academic_year, [])
        if label not in semester_options_by_year[semester.academic_year]:
            semester_options_by_year[semester.academic_year].append(label)
    latest_backup = latest_database_backup()
    active_user_count = len(session.exec(select(User).where(User.is_active == True)).all())  # noqa: E712
    page = {
        "kind": "settings",
        "title": "System Settings",
        "description": "Manage system settings stored in the database.",
        "settings": settings,
        "academic_years": academic_years,
        "academic_semesters": academic_semesters,
        "semester_options_by_year": semester_options_by_year,
        "current_period": current_period,
        "database_backend": engine.url.get_backend_name().upper(),
        "database_size": database_size_label(),
        "latest_backup_name": latest_backup.name if latest_backup else "No backup created",
        "latest_backup_exists": bool(latest_backup),
        "stats": [
            ("Database Size", database_size_label(), "bi-database", "blue"),
            ("Active Users", active_user_count, "bi-people", "green"),
            ("Current Study Period", current_period.label if current_period else "Not configured", "bi-calendar2-week", "purple"),
            ("Last Real Backup", settings.get("last_backup", "Never") if latest_backup else "Never", "bi-shield-check", "orange"),
        ],
    }
    return templates.TemplateResponse(
        "system_settings.html",
        {"request": request, "user": user, "page": page, "admin_section": "settings"},
    )


@app.post("/admin/settings/update")
async def update_admin_settings(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    current_values = ensure_system_settings(session)
    form = await request.form()
    values, error = validate_system_setting_form(form)
    if error:
        return redirect(f"/admin/settings?error={quote(error)}")

    academic_year = values.get("academic_year")
    default_semester = values.get("default_semester")
    if academic_year and default_semester:
        academic_error = apply_default_academic_period(session, academic_year, default_semester)
        if academic_error:
            return redirect(f"/admin/settings?error={quote(academic_error)}#settings-academic")

    updated_keys: list[str] = []
    for key, value in values.items():
        save_setting(session, key, value)
        updated_keys.append(key)
    if any(values.get(key, current_values.get(key, "")) != current_values.get(key, "") for key in ("smtp_host", "smtp_port", "smtp_security")):
        save_setting(session, "smtp_status", "Not tested")
        save_setting(session, "smtp_last_test", "Never")
    session.commit()
    add_audit_record(session, user, "System Settings", "UPDATE", f"Updated {len(updated_keys)} system setting(s).", "Settings", request.client.host if request.client else "127.0.0.1")
    session.commit()
    return redirect("/admin/settings?updated=1")


@app.post("/admin/settings/reset")
def reset_admin_settings(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    ensure_system_settings(session)
    for key, value in DEFAULT_SYSTEM_SETTINGS.items():
        if key not in {"academic_year", "default_semester", "last_backup", "last_backup_file"}:
            save_setting(session, key, value)
    sync_setting_academic_defaults(session)
    session.commit()
    add_audit_record(session, user, "System Settings", "RESET", "Reset system settings to default values.", "Settings", request.client.host if request.client else "127.0.0.1")
    session.commit()
    return redirect("/admin/settings?reset=1")


@app.post("/admin/settings/backup")
def backup_admin_settings(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    ensure_system_settings(session)
    session.commit()
    try:
        backup_path = create_database_backup()
    except Exception as exc:
        message = str(exc) or "Database backup failed."
        add_audit_record(session, user, "System Settings", "BACKUP", message, "Database Backup", request.client.host if request.client else "127.0.0.1", "Failed")
        session.commit()
        return redirect(f"/admin/settings?backup_error={quote(message)}#settings-backup")

    backup_time = datetime.fromtimestamp(backup_path.stat().st_mtime).strftime("%b %d, %Y %I:%M %p")
    save_setting(session, "last_backup", backup_time)
    save_setting(session, "last_backup_file", backup_path.name)
    add_audit_record(session, user, "System Settings", "BACKUP", f"Created SQLite backup {backup_path.name}.", "Database Backup", request.client.host if request.client else "127.0.0.1")
    session.commit()
    return redirect("/admin/settings?backup=1")


@app.get("/admin/settings/backup/download")
def download_admin_database_backup(session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    values = ensure_system_settings(session)
    filename = Path(values.get("last_backup_file", "")).name
    backup_path = database_backup_directory() / filename if filename else latest_database_backup()
    if not backup_path or not backup_path.is_file() or backup_path.resolve().parent != database_backup_directory():
        raise HTTPException(status_code=404, detail="No database backup is available.")
    return FileResponse(backup_path, media_type="application/vnd.sqlite3", filename=backup_path.name)


@app.post("/admin/settings/test-email")
async def test_admin_email(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    values = ensure_system_settings(session)
    form = await request.form()
    for key in ("smtp_host", "smtp_port", "smtp_security", "from_email", "email_template_footer"):
        if key in form:
            values[key] = str(form.get(key) or "").strip()
            save_setting(session, key, values[key])

    host = values.get("smtp_host", "").strip()
    if not host:
        message = "Enter an SMTP host before testing the connection."
        save_setting(session, "smtp_status", "Not configured")
        add_audit_record(session, user, "System Settings", "TEST", message, "SMTP Connection", request.client.host if request.client else "127.0.0.1", "Failed")
        session.commit()
        return redirect(f"/admin/settings?email_error={quote(message)}#settings-email")
    try:
        port = int(values.get("smtp_port", "587"))
        if port < 1 or port > 65535:
            raise ValueError("SMTP port must be between 1 and 65535.")
    except ValueError as exc:
        message = str(exc) if str(exc) else "SMTP port must be a whole number."
        save_setting(session, "smtp_status", "Failed")
        add_audit_record(session, user, "System Settings", "TEST", message, "SMTP Connection", request.client.host if request.client else "127.0.0.1", "Failed")
        session.commit()
        return redirect(f"/admin/settings?email_error={quote(message)}#settings-email")

    tested_at = datetime.now().strftime("%b %d, %Y %I:%M %p")
    security_mode = values.get("smtp_security", "STARTTLS")
    try:
        if security_mode == "SSL/TLS":
            with smtplib.SMTP_SSL(host, port, timeout=8) as smtp:
                smtp.ehlo()
                smtp.noop()
        else:
            with smtplib.SMTP(host, port, timeout=8) as smtp:
                smtp.ehlo()
                if security_mode == "STARTTLS":
                    smtp.starttls()
                    smtp.ehlo()
                smtp.noop()
    except Exception as exc:
        message = f"SMTP connection failed: {str(exc) or exc.__class__.__name__}"
        message = message[:220]
        save_setting(session, "smtp_status", "Failed")
        save_setting(session, "smtp_last_test", tested_at)
        add_audit_record(session, user, "System Settings", "TEST", message, "SMTP Connection", request.client.host if request.client else "127.0.0.1", "Failed")
        session.commit()
        return redirect(f"/admin/settings?email_error={quote(message)}#settings-email")

    save_setting(session, "smtp_status", "Connected")
    save_setting(session, "smtp_last_test", tested_at)
    add_audit_record(session, user, "System Settings", "TEST", f"Connected to SMTP server {host}:{port} using {security_mode}.", "SMTP Connection", request.client.host if request.client else "127.0.0.1")
    session.commit()
    return redirect("/admin/settings?test_email=1")


def build_admin_programme_mapping_page(
    session: Session,
    program: Program,
    version: PLOVersion,
    programs: list[Program],
    edit_course_id: int | None = None,
) -> dict:
    plo_records = sorted(
        session.exec(
            select(PLO).where(
                PLO.program_id == program.id,
                PLO.plo_version_id == version.id,
                PLO.status == "Active",
            )
        ).all(),
        key=plo_sort_key,
    )
    plos = [
        {"id": plo.id, "code": plo.code, "description": plo.description, "value": 0}
        for plo in plo_records
    ]
    courses = manager_courses(session, None, plos, program)
    edit_course = next((course for course in courses if course["id"] == edit_course_id), None)
    edit_weights = []
    if edit_course:
        edit_weights = [
            {"plo": plo, "percent": edit_course["levels"][index]["percent"]}
            for index, plo in enumerate(plos)
        ]

    fully_mapped = sum(
        1 for course in courses
        if course["levels"] and abs(sum(item["percent"] for item in course["levels"]) - 100) < 0.01
    )
    partially_mapped = sum(
        1 for course in courses
        if 0 < sum(item["percent"] for item in course["levels"]) < 99.99
        or sum(item["percent"] for item in course["levels"]) > 100.01
    )
    not_mapped = sum(1 for course in courses if not any(item["percent"] for item in course["levels"]))
    mapped_cells = sum(
        1 for course in courses for item in course["levels"] if item["percent"] > 0
    )
    total_cells = len(courses) * len(plos)
    coverage = round(mapped_cells / total_cells * 100, 1) if total_cells else 0
    return {
        "kind": "programme-mapping",
        "title": "Programme Mapping",
        "description": "Define and manage course-to-PLO contribution weights for every programme.",
        "programs": programs,
        "program": program,
        "selected_program": program,
        "plos": plos,
        "courses": courses,
        "edit_course": edit_course,
        "edit_weights": edit_weights,
        "version_data": programme_version_data(session, program, version),
        "stats": [
            ("Total Courses", len(courses), "bi-mortarboard", "blue"),
            ("Fully Mapped", fully_mapped, "bi-check-circle", "green"),
            ("Partially Mapped", partially_mapped, "bi-diagram-3", "orange"),
            ("Not Mapped", not_mapped, "bi-x-circle", "red"),
            ("PLO Coverage", f"{coverage}%", "bi-bar-chart", "purple"),
        ],
        "summary": {
            "mapped_cells": mapped_cells,
            "total_cells": total_cells,
            "coverage": coverage,
        },
    }


@app.get("/admin/programme-mapping", response_class=HTMLResponse)
def admin_programme_mapping_page(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    programs = list(session.exec(select(Program).order_by(Program.code, Program.name)).all())
    if not programs:
        raise HTTPException(status_code=404, detail="No programmes are available")
    requested_program_id = optional_int(request.query_params.get("program_id"))
    program = next((item for item in programs if item.id == requested_program_id), programs[0])
    requested_version_id = optional_int(request.query_params.get("version_id"))
    version, _versions = selected_outcome_version(session, program, requested_version_id)
    edit_course_id = optional_int(request.query_params.get("edit_course_id"))
    page = build_admin_programme_mapping_page(session, program, version, programs, edit_course_id)
    return templates.TemplateResponse(
        "admin_programme_mapping.html",
        {"request": request, "user": user, "page": page, "admin_section": "programme-mapping"},
    )


@app.post("/admin/programme-mapping/courses/create")
async def create_admin_programme_mapping_course(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = optional_int(form.get("program_id"))
    version_id = optional_int(form.get("version_id"))
    program = session.get(Program, program_id) if program_id else None
    version = session.get(PLOVersion, version_id) if version_id else None
    return_path = f"/admin/programme-mapping?program_id={program_id or ''}&version_id={version_id or ''}"
    if not program or not version or version.programme_id != program.id:
        return redirect(f"{return_path}&error=invalid_scope&show_add=1")
    if version.is_locked or version.status in {"Published", "Retired"}:
        return redirect(f"{return_path}&error=version_locked")

    code = str(form.get("code") or "").strip()
    title = str(form.get("title") or "").strip()
    semester = str(form.get("curriculum_semester") or "").strip()
    try:
        credits = float(form.get("credits") or 0)
        curriculum_year = (optional_int(form.get("curriculum_year")) or 0)
    except (TypeError, ValueError):
        return redirect(f"{return_path}&error=invalid_course&show_add=1")
    if not code or not title or curriculum_year not in {1, 2, 3, 4} or semester not in {"1", "2"} or credits < 0:
        return redirect(f"{return_path}&error=invalid_course&show_add=1")
    if session.exec(select(Course).where(Course.program_id == program.id, Course.code == code)).first():
        return redirect(f"{return_path}&error=duplicate_course&show_add=1")

    plos = sorted(
        session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)).all(),
        key=plo_sort_key,
    )
    weights = {plo.id: clamp_percent(form.get(f"weight_{plo.id}")) for plo in plos}
    total_weight = round(sum(weights.values()), 2)
    if total_weight and abs(total_weight - 100) > 0.01:
        return redirect(f"{return_path}&error=mapping_total&show_add=1")

    course = Course(
        program_id=program.id,
        code=code,
        title=title,
        credits=credits,
        curriculum_year=curriculum_year,
        curriculum_semester=semester,
    )
    session.add(course)
    session.flush()
    mapping_clo = CLO(course_id=course.id, code="CLO1", description=f"Programme mapping contribution for {code}")
    session.add(mapping_clo)
    session.flush()
    for plo in plos:
        weight = weights.get(plo.id, 0)
        if weight:
            session.add(CLOPLOMapping(clo_id=mapping_clo.id, plo_id=plo.id, weight=weight))
        session.add(CoursePLOMapping(course_id=course.id, plo_id=plo.id, level=weight, symbol=f"{weight:g}%" if weight else ""))
    session.commit()
    add_audit_record(
        session, user, "Programme Mapping", "CREATE", f"Created course {code} and its PLO mapping.",
        program.code, request.client.host if request.client else "127.0.0.1",
    )
    return redirect(f"{return_path}&created=1")


@app.post("/admin/programme-mapping/courses/{course_id}/update")
async def update_admin_programme_mapping_course(
    course_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = optional_int(form.get("program_id"))
    version_id = optional_int(form.get("version_id"))
    program = session.get(Program, program_id) if program_id else None
    version = session.get(PLOVersion, version_id) if version_id else None
    course = session.get(Course, course_id)
    return_path = f"/admin/programme-mapping?program_id={program_id or ''}&version_id={version_id or ''}"
    if not program or not version or version.programme_id != program.id or not course or course.program_id != program.id:
        return redirect(f"{return_path}&error=invalid_scope")
    if version.is_locked or version.status in {"Published", "Retired"}:
        return redirect(f"{return_path}&error=version_locked")

    plos = sorted(
        session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)).all(),
        key=plo_sort_key,
    )
    desired_weights = {plo.id: clamp_percent(form.get(f"weight_{plo.id}")) for plo in plos}
    total_weight = round(sum(desired_weights.values()), 2)
    if total_weight and abs(total_weight - 100) > 0.01:
        return redirect(f"{return_path}&edit_course_id={course.id}&error=mapping_total")

    code = str(form.get("code") or course.code).strip()
    title = str(form.get("title") or course.title).strip()
    duplicate = session.exec(
        select(Course).where(Course.program_id == program.id, Course.code == code, Course.id != course.id)
    ).first()
    if not code or not title or duplicate:
        return redirect(f"{return_path}&edit_course_id={course.id}&error=invalid_course")
    course.code = code
    course.title = title
    course.curriculum_year = optional_int(form.get("curriculum_year")) or course.curriculum_year or 1
    course.curriculum_semester = str(form.get("curriculum_semester") or course.curriculum_semester or "1")
    course.credits = float(form.get("credits") or course.credits or 0)
    session.add(course)

    clos = list(session.exec(select(CLO).where(CLO.course_id == course.id).order_by(CLO.id)).all())
    if not clos:
        mapping_clo = CLO(course_id=course.id, code="CLO1", description=f"Programme mapping contribution for {code}")
        session.add(mapping_clo)
        session.flush()
        clos = [mapping_clo]
    clo_ids = [clo.id for clo in clos if clo.id]
    for plo in plos:
        desired = desired_weights.get(plo.id, 0)
        mappings = list(
            session.exec(
                select(CLOPLOMapping).where(
                    CLOPLOMapping.clo_id.in_(clo_ids),
                    CLOPLOMapping.plo_id == plo.id,
                )
            ).all()
        )
        if desired == 0:
            for mapping in mappings:
                session.delete(mapping)
            continue
        current_total = sum(stored_percent(mapping.weight) for mapping in mappings)
        if not mappings or current_total <= 0:
            session.add(CLOPLOMapping(clo_id=clos[0].id, plo_id=plo.id, weight=desired))
            continue
        assigned = 0.0
        for index, mapping in enumerate(mappings):
            if index == len(mappings) - 1:
                new_weight = round(desired - assigned, 4)
            else:
                new_weight = round(desired * stored_percent(mapping.weight) / current_total, 4)
                assigned += new_weight
            mapping.weight = max(0, new_weight)
            session.add(mapping)
    session.commit()
    add_audit_record(
        session, user, "Programme Mapping", "UPDATE", f"Updated {course.code} PLO contribution weights.",
        program.code, request.client.host if request.client else "127.0.0.1",
    )
    return redirect(f"{return_path}&updated=1")


@app.post("/admin/programme-mapping/courses/{course_id}/delete")
async def delete_admin_programme_mapping_course(
    course_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    form = await request.form()
    program_id = optional_int(form.get("program_id"))
    version_id = optional_int(form.get("version_id"))
    program = session.get(Program, program_id) if program_id else None
    version = session.get(PLOVersion, version_id) if version_id else None
    course = session.get(Course, course_id)
    return_path = f"/admin/programme-mapping?program_id={program_id or ''}&version_id={version_id or ''}"
    if not program or not version or version.programme_id != program.id or not course or course.program_id != program.id:
        return redirect(f"{return_path}&error=invalid_scope")
    if version.is_locked or version.status in {"Published", "Retired"}:
        return redirect(f"{return_path}&error=version_locked")
    plo_ids = [
        plo.id for plo in session.exec(
            select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)
        ).all() if plo.id
    ]
    clo_ids = [clo.id for clo in session.exec(select(CLO).where(CLO.course_id == course.id)).all() if clo.id]
    if clo_ids and plo_ids:
        for mapping in session.exec(
            select(CLOPLOMapping).where(CLOPLOMapping.clo_id.in_(clo_ids), CLOPLOMapping.plo_id.in_(plo_ids))
        ).all():
            session.delete(mapping)
    if plo_ids:
        for mapping in session.exec(
            select(CoursePLOMapping).where(CoursePLOMapping.course_id == course.id, CoursePLOMapping.plo_id.in_(plo_ids))
        ).all():
            session.delete(mapping)
    session.commit()
    add_audit_record(
        session, user, "Programme Mapping", "DELETE", f"Cleared {course.code} mapping for {version.version_name}.",
        program.code, request.client.host if request.client else "127.0.0.1",
    )
    return redirect(f"{return_path}&deleted=1")


def build_student_progression_page(
    request: Request,
    session: Session,
    program: Program,
    user: User | None = None,
) -> dict:
    """Build one progression view used by both Admin and Programme Manager."""
    all_enrollments = list(
        session.exec(
            select(StudentSemesterEnrollment)
            .where(StudentSemesterEnrollment.program_id == program.id)
            .order_by(StudentSemesterEnrollment.created_at.desc(), StudentSemesterEnrollment.id.desc())
        ).all()
    )

    # The global sidebar Study Period is the single source of truth.
    global_period = selected_study_period(request, user)
    if global_period is not None:
        selected_academic_year = str(global_period.academic_year or "").strip()
        selected_semester = normalize_semester(global_period.semester)
        selected_period = f"{selected_academic_year}|{selected_semester}"
    else:
        selected_academic_year = str(request.cookies.get("global_academic_year") or "2025-2026").strip()
        selected_semester = normalize_semester(request.cookies.get("global_semester") or "1")
        selected_period = f"{selected_academic_year}|{selected_semester}"

    programme_course_ids = set(
        session.exec(select(Course.id).where(Course.program_id == program.id)).all()
    )
    programme_classes = list(
        session.exec(
            select(CourseClass)
            .where(CourseClass.course_id.in_(programme_course_ids))
            .order_by(CourseClass.academic_year.desc(), CourseClass.semester, CourseClass.name)
        ).all()
    ) if programme_course_ids else []

    period_enrollments = [
        item for item in all_enrollments
        if str(item.academic_year or "").strip() == selected_academic_year
        and normalize_semester(item.semester) == selected_semester
    ]
    latest_by_student: dict[int, StudentSemesterEnrollment] = {}
    for enrollment in period_enrollments:
        latest_by_student.setdefault(enrollment.student_id, enrollment)

    class_groups: dict[str, dict] = {}
    for course_class in programme_classes:
        if str(course_class.academic_year or "").strip() != selected_academic_year:
            continue
        if normalize_semester(course_class.semester) != selected_semester:
            continue
        class_code = str(course_class.name or "").strip()
        year, class_semester = _cohort_year_semester(class_code)
        if not class_code or not year:
            continue
        class_groups.setdefault(
            class_code,
            {
                "class_code": class_code,
                "year": year,
                "semester": class_semester or selected_semester,
                "generation": _cohort_generation(class_code),
                "students": 0,
                "active": 0,
            },
        )

    for enrollment in latest_by_student.values():
        class_code = str(enrollment.cohort_name or "Unassigned").strip() or "Unassigned"
        year, class_semester = _cohort_year_semester(class_code)
        group = class_groups.setdefault(
            class_code,
            {
                "class_code": class_code,
                "year": year,
                "semester": class_semester or selected_semester,
                "generation": _cohort_generation(class_code),
                "students": 0,
                "active": 0,
            },
        )
        group["students"] += 1
        if enrollment.status not in {"Dropped", "Withdrawn", "Graduated"}:
            group["active"] += 1

    classes_by_year = {year: [] for year in range(1, 5)}
    other_classes: list[dict] = []
    for group in class_groups.values():
        if group["year"] in classes_by_year:
            classes_by_year[group["year"]].append(group)
        else:
            other_classes.append(group)
    for groups in classes_by_year.values():
        groups.sort(key=lambda item: item["class_code"])
    other_classes.sort(key=lambda item: item["class_code"])

    requested_class = str(request.query_params.get("class_code") or "").strip()
    selected_class = requested_class if requested_class in class_groups else ""
    selected_status = str(request.query_params.get("status") or "").strip()
    search_text = str(request.query_params.get("q") or "").strip().lower()

    rows: list[dict] = []
    for enrollment in latest_by_student.values():
        if not selected_class or enrollment.cohort_name != selected_class:
            continue
        student = session.get(Student, enrollment.student_id)
        if not student or (selected_status and enrollment.status != selected_status):
            continue
        haystack = f"{student.student_no} {student.name_en} {student.name_kh or ''}".lower()
        if search_text and search_text not in haystack:
            continue
        year, semester = _cohort_year_semester(enrollment.cohort_name)
        rows.append({
            "student": student,
            "enrollment": enrollment,
            "generation": _cohort_generation(enrollment.cohort_name),
            "year": year,
            "semester": semester,
        })
    rows.sort(key=lambda item: item["student"].student_no)

    counts = {
        year: sum(group["active"] for group in groups)
        for year, groups in classes_by_year.items()
    }
    study_periods = list(
        session.exec(
            select(StudyPeriod)
            .where(StudyPeriod.is_active == True)
            .order_by(StudyPeriod.academic_year.desc(), StudyPeriod.semester)
        ).all()
    )
    return {
        "title": "Student Progression",
        "description": f"Manage {program.code} students from Year 1 to Year 4.",
        "program": program,
        "study_periods": study_periods,
        "selected_period": selected_period,
        "selected_academic_year": selected_academic_year,
        "selected_semester": selected_semester,
        "classes_by_year": classes_by_year,
        "other_classes": other_classes,
        "selected_class": selected_class,
        "selected_class_info": class_groups.get(selected_class),
        "rows": rows,
        "counts": counts,
        "total_students": len(latest_by_student),
        "filters": {"status": selected_status, "q": request.query_params.get("q") or ""},
        "updated": request.query_params.get("updated") == "1",
    }


@app.get("/admin/student-progression", response_class=HTMLResponse)
def admin_student_progression_page(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    programs = list(session.exec(select(Program).order_by(Program.code, Program.name)).all())
    if not programs:
        raise HTTPException(status_code=404, detail="No programmes are available")
    requested_program_id = optional_int(request.query_params.get("program_id"))
    program = next((item for item in programs if item.id == requested_program_id), programs[0])
    page = build_student_progression_page(request, session, program, user)
    page["programs"] = programs
    return templates.TemplateResponse(
        "student_progression.html",
        {"request": request, "user": user, "page": page, "admin_section": "student-progression"},
    )


@app.post("/admin/student-progression/enrollments/{enrollment_id}/status")
async def update_admin_student_progression_status(
    enrollment_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    enrollment = session.get(StudentSemesterEnrollment, enrollment_id)
    if not enrollment:
        raise HTTPException(status_code=404)
    form = await request.form()
    status = str(form.get("status") or "Active").strip()
    allowed = {"Active", "Probation", "Repeat", "Leave", "Withdrawn", "Dropped", "Graduated"}
    if status not in allowed:
        raise HTTPException(status_code=400, detail="Invalid student status")
    enrollment.status = status
    session.add(enrollment)
    session.commit()
    student = session.get(Student, enrollment.student_id)
    add_audit_record(
        session,
        user,
        "Student Management",
        "UPDATE",
        f"Changed {student.student_no if student else enrollment.student_id} status to {status}.",
        enrollment.cohort_name,
        request.client.host if request.client else "127.0.0.1",
    )
    program_id = optional_int(form.get("program_id")) or enrollment.program_id
    class_code = str(form.get("class_code") or enrollment.cohort_name or "").strip()
    return redirect(
        f"/admin/student-progression?program_id={program_id}&class_code={quote(class_code)}&updated=1"
    )


@app.get("/admin/student-promotion", response_class=HTMLResponse)
def admin_student_promotion_page(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "promote_students"):
        raise HTTPException(status_code=403)
    page = build_student_promotion_page(request, session, user)
    return templates.TemplateResponse(
        "student_promotion.html",
        {"request": request, "user": user, "page": page, "admin_section": "student-promotion"},
    )


@app.post("/admin/student-promotion")
async def admin_promote_students(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "promote_students"):
        raise HTTPException(status_code=403)
    form = await request.form()
    if not str(form.get("cohort") or "").strip():
        return redirect("/admin/student-promotion?error=no_cohort")
    if normalize_semester(str(form.get("from_semester") or "")) == normalize_semester(str(form.get("to_semester") or "")):
        return redirect("/admin/student-promotion?error=same_semester")
    student_ids = [int(value) for value in form.getlist("student_ids") if str(value).isdigit()]
    if not student_ids:
        return redirect("/admin/student-promotion?error=no_students")
    promoted, courses, skipped = promote_selected_students(
        session=session,
        user=user,
        academic_year=str(form.get("academic_year") or ""),
        faculty_id=optional_int(form.get("faculty_id")),
        program_id=(optional_int(form.get("program_id")) or 0),
        cohort_name=str(form.get("cohort") or ""),
        from_semester=str(form.get("from_semester") or ""),
        to_semester=str(form.get("to_semester") or ""),
        student_ids=student_ids,
    )
    if promoted == 0 and skipped:
        return redirect("/admin/student-promotion?error=duplicate")
    if promoted == 0:
        return redirect("/admin/student-promotion?error=no_target")
    return redirect(f"/admin/student-promotion?success=1&promoted={promoted}&courses={courses}&skipped={skipped}")


@app.post("/manager/enter-scores/save")
async def manager_save_course_scores(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)

    form = await request.form()
    course_id = optional_int(form.get("course_id"))
    class_id = optional_int(form.get("class_id"))

    # Verify course belongs to manager's programme
    program = manager_program(session, user)
    course = session.get(Course, course_id)
    if not course or course.program_id != program.id:
        raise HTTPException(status_code=403, detail="Course not in your programme")

    # Verify class exists for this course
    if class_id:
        cls = session.get(CourseClass, class_id)
        if not cls or cls.course_id != course.id:
            raise HTTPException(status_code=400, detail="Invalid class for this course")
    else:
        # If no class_id, try to get the first class for this course in the current study period
        study_period = selected_study_period(request, user)
        classes = session.exec(select(CourseClass).where(CourseClass.course_id == course.id)).all()
        if study_period:
            classes = [c for c in classes if class_matches_study_period(c, study_period)]
        if not classes:
            raise HTTPException(status_code=400, detail="No class found for this course in the current study period")
        class_id = classes[0].id

    # Reuse teacher's save logic but with manager permissions
    # We need to adapt the teacher_save_course_scores function to accept a user and course/class
    # For simplicity, we'll copy the core logic here (or refactor into a helper).
    # To avoid duplication, we'll create a helper function `_save_course_scores` that is called by both endpoints.
    # But for now, we'll copy the logic from teacher_save_course_scores and adjust.

    # ---------- Begin copied logic (adapted) ----------
    study_period = selected_study_period(request, user)
    selected_course = course
    selected_class = session.get(CourseClass, class_id)
    if not selected_class or selected_class.course_id != selected_course.id:
        raise HTTPException(status_code=400, detail="Invalid class")

    students = teacher_course_student_rows(session, selected_course.id, selected_class.id)
    valid_student_ids = {student["id"] for student in students}
    assessments = teacher_assessment_rows(session, selected_course.id, students)
    assessment_lookup = {
        assessment.id: assessment
        for assessment in teacher_course_assessments(session, selected_course.id)
        if assessment.id
    }

    action = str(form.get("score_action") or "draft")
    final_submit = action == "submit"
    weight_total = assessment_weight_total(assessments)
    if action != "draft" and round(weight_total, 2) != 100:
        return redirect(f"/manager/enter-scores?course_id={course_id}&class_id={class_id}&error={quote('Assessment weight total must equal 100% before saving final course scores. You can still save a draft.')}#manager-score-entry")

    if course_scores_locked(session, selected_course.id, valid_student_ids):
        return redirect(f"/manager/enter-scores?course_id={course_id}&class_id={class_id}&error={quote('Final scores are locked. Unlock before editing.')}#manager-score-entry")

    status = "Submitted" if final_submit else "Saved" if action == "save" else "Draft"
    errors = []
    saved_count = 0
    now = datetime.utcnow()

    for key, value in form.items():
        if not key.startswith("score_"):
            continue
        parts = key.split("_")
        if len(parts) != 3:
            continue
        student_id = optional_int(parts[1])
        assessment_id = optional_int(parts[2])
        if not student_id or not assessment_id or student_id not in valid_student_ids or assessment_id not in assessment_lookup:
            continue
        raw_value = str(value).strip()
        if raw_value == "":
            continue
        assessment = assessment_lookup[assessment_id]
        max_score = float(assessment.max_score or assessment_weight_percent(assessment) or 100)
        try:
            score_value = float(raw_value)
        except ValueError:
            errors.append(f"Invalid score for student {student_id}.")
            continue
        if score_value < 0 or score_value > max_score:
            errors.append(f"Score for {assessment.name} must be between 0 and {max_score:g}.")
            continue

        existing = session.exec(
            select(StudentScore).where(
                StudentScore.student_id == student_id,
                StudentScore.assessment_id == assessment_id,
            )
        ).first()
        if existing:
            if existing.locked:
                continue
            existing.score = score_value
            existing.status = status
            existing.updated_at = now
            existing.entered_by_user_id = user.id
            if final_submit:
                existing.locked = True
                existing.submitted_at = now
                existing.submitted_by_user_id = user.id
            session.add(existing)
        else:
            session.add(
                StudentScore(
                    student_id=student_id,
                    assessment_id=assessment_id,
                    score=score_value,
                    status=status,
                    locked=final_submit,
                    updated_at=now,
                    submitted_at=now if final_submit else None,
                    submitted_by_user_id=user.id if final_submit else None,
                    entered_by_user_id=user.id,
                )
            )
        saved_count += 1

    if errors:
        session.rollback()
        return redirect(f"/manager/enter-scores?course_id={course_id}&class_id={class_id}&error={quote(errors[0])}#manager-score-entry")

    session.commit()
    message = "final" if final_submit else "draft" if action == "draft" else "all"
    return redirect(f"/manager/enter-scores?course_id={course_id}&class_id={class_id}&saved=1&mode={message}#manager-score-entry")


@app.post("/manager/enter-scores/unlock")
async def manager_unlock_course_scores(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)

    form = await request.form()
    course_id = optional_int(form.get("course_id"))
    class_id = optional_int(form.get("class_id"))

    program = manager_program(session, user)
    course = session.get(Course, course_id)
    if not course or course.program_id != program.id:
        raise HTTPException(status_code=403, detail="Course not in your programme")

    if class_id:
        cls = session.get(CourseClass, class_id)
        if not cls or cls.course_id != course.id:
            raise HTTPException(status_code=400, detail="Invalid class")
    else:
        study_period = selected_study_period(request, user)
        classes = session.exec(select(CourseClass).where(CourseClass.course_id == course.id)).all()
        if study_period:
            classes = [c for c in classes if class_matches_study_period(c, study_period)]
        if not classes:
            raise HTTPException(status_code=400, detail="No class found")
        class_id = classes[0].id

    students = teacher_course_student_rows(session, course_id, class_id)
    student_ids = {student["id"] for student in students}
    assessment_ids = [assessment.id for assessment in teacher_course_assessments(session, course_id) if assessment.id]
    if not assessment_ids or not student_ids:
        return redirect(f"/manager/enter-scores?course_id={course_id}&class_id={class_id}&error={quote('No locked scores found to unlock.')}#manager-score-entry")

    for score in session.exec(
        select(StudentScore).where(
            StudentScore.assessment_id.in_(assessment_ids),
            StudentScore.student_id.in_(student_ids),
        )
    ).all():
        score.locked = False
        score.status = "Unlocked"
        score.updated_at = datetime.utcnow()
        session.add(score)

    session.commit()
    return redirect(f"/manager/enter-scores?course_id={course_id}&class_id={class_id}&unlocked=1#manager-score-entry")





@app.get("/manager/student-promotion", response_class=HTMLResponse)
@app.get("/dean/student-promotion", response_class=HTMLResponse)
def manager_student_promotion_page(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "promote_students"):
        raise HTTPException(status_code=403)
    page = build_student_promotion_page(request, session, user)
    return templates.TemplateResponse(
        "student_promotion.html",
        {"request": request, "user": user, "page": page, "manager_section": "student-promotion", "section_base": programme_section_base(user)},
    )


@app.post("/manager/student-promotion")
async def manager_promote_students(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "promote_students"):
        raise HTTPException(status_code=403)
    form = await request.form()
    if not str(form.get("cohort") or "").strip():
        return redirect("/manager/student-promotion?error=no_cohort")
    if normalize_semester(str(form.get("from_semester") or "")) == normalize_semester(str(form.get("to_semester") or "")):
        return redirect("/manager/student-promotion?error=same_semester")
    student_ids = [int(value) for value in form.getlist("student_ids") if str(value).isdigit()]
    if not student_ids:
        return redirect("/manager/student-promotion?error=no_students")
    promoted, courses, skipped = promote_selected_students(
        session=session,
        user=user,
        academic_year=str(form.get("academic_year") or ""),
        faculty_id=optional_int(form.get("faculty_id")),
        program_id=(optional_int(form.get("program_id")) or 0),
        cohort_name=str(form.get("cohort") or ""),
        from_semester=str(form.get("from_semester") or ""),
        to_semester=str(form.get("to_semester") or ""),
        student_ids=student_ids,
    )
    if promoted == 0 and skipped:
        return redirect("/manager/student-promotion?error=duplicate")
    if promoted == 0:
        return redirect("/manager/student-promotion?error=no_target")
    return redirect(f"/manager/student-promotion?success=1&promoted={promoted}&courses={courses}&skipped={skipped}")


@app.get("/admin", response_class=HTMLResponse)
def admin_home(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Admin landing page.

    The admin module uses the same management layout as Role Management.
    Redirecting to the role page avoids the previous 404 at /admin and keeps
    one consistent table/action design for administrators.
    """
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    return redirect("/admin/roles")


@app.get("/admin/{section}", response_class=HTMLResponse)
def admin_management_section(
    section: str,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_system"):
        raise HTTPException(status_code=403)
    allowed_sections = {
        "users",
        "roles",
        "students",
        "faculties",
        "programmes",
        "academic-years",
        "semesters",
        "cohorts",
        "academic-structure",
        "plos",
        "peos",
        "targets",
        "outcome-versions",
        "settings",
        "reports",
        "enter-scores",
        "audit-logs",
        "documents",
    }
    if section not in allowed_sections:
        raise HTTPException(status_code=404)
    if section == "enter-scores":
        page = score_unlock_page_data(session, user)
        return templates.TemplateResponse(
            "score_unlock_management.html",
            {"request": request, "user": user, "page": page, "admin_section": section},
        )
    if section == "documents":
        page = document_page_data(session, user, "/admin/documents")
        return templates.TemplateResponse(
            "admin_management.html",
            {"request": request, "user": user, "page": page, "admin_section": section},
        )
    if section in {"academic-years", "semesters", "cohorts", "academic-structure"}:
        page = build_admin_management_page("academic-structure", session)
        active_tab = "years" if section == "academic-years" else "semesters" if section == "semesters" else "classes"
        return templates.TemplateResponse(
            "academic_structure.html",
            {"request": request, "user": user, "page": page, "admin_section": section, "active_tab": active_tab},
        )
    raw_program_id = request.query_params.get("program_id")
    raw_version_id = request.query_params.get("version_id")
    program_id = int(raw_program_id) if raw_program_id and raw_program_id.isdigit() else None
    version_id = int(raw_version_id) if raw_version_id and raw_version_id.isdigit() else None
    page = build_admin_management_page(section, session, program_id, version_id)
    return templates.TemplateResponse(
        "admin_management.html",
        {"request": request, "user": user, "page": page, "admin_section": section},
    )


DEAN_PROGRAMME_SECTION_ALIASES = {
    "programme-plo-performance": "plo-performance",
    "programme-reports": "reports",
    "peos": "peo-management",
}

DEAN_ATTAINMENT_SECTION_ALIASES: dict[str, str] = {}

# Full attainment pages: charts, programme comparison and CQI actions.
DEAN_ATTAINMENT_DETAIL_SECTIONS = {"plo-attainment": "plo", "peo-attainment": "peo"}


@app.get("/dean/{section}", response_class=HTMLResponse)
def dean_section_page(
    section: str,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if user.role != Role.DEAN:
        raise HTTPException(status_code=403)

    if section in DEAN_PROGRAMME_SECTION_ALIASES:
        return render_programme_section(DEAN_PROGRAMME_SECTION_ALIASES[section], request, session, user)

    if section in DEAN_ATTAINMENT_DETAIL_SECTIONS:
        page = dean_attainment_detail_page(
            session, user, DEAN_ATTAINMENT_DETAIL_SECTIONS[section]
        )
        faculty = session.get(Faculty, user.faculty_id) if user.faculty_id else None
        page.update(
            {
                "faculty_name": faculty.name if faculty else "All Faculties",
                "faculty_code": faculty_code(faculty),
                "programmes": [],
                "programme_count": len(page["entries"]),
            }
        )
        return templates.TemplateResponse(
            "dean_management.html",
            {"request": request, "user": user, "page": page, "dean_section": section, "programmes": []},
        )

    faculty_section = DEAN_ATTAINMENT_SECTION_ALIASES.get(section, section)
    # Faculty-level pages that only exist for a Dean. Everything else falls
    # through to the shared programme-management sections, so a Dean gets the
    # Programme Manager functions under their own /dean URLs.
    dean_own_sections = {
        "profile",
        "members",
        "faculty-reports",
        "assessment-reports",
        "plo-performance",
        "plo-attainment",
        "peo-performance",
        "peo-attainment",
        "programmes",
        "audit-logs",
    }
    if section not in dean_own_sections:
        if section in MANAGER_SECTIONS or section in {"students", "assessments"}:
            return render_programme_section(section, request, session, user)
        raise HTTPException(status_code=404)
    page = build_dean_page(faculty_section, session, user, optional_int(request.query_params.get("program_id")))
    return templates.TemplateResponse(
        "dean_management.html",
        {"request": request, "user": user, "page": page, "dean_section": section, "programmes": page["programmes"]},
    )

@app.get("/manager/students/scores", response_class=HTMLResponse)
@app.get("/dean/students/scores", response_class=HTMLResponse)
def manager_student_scores_page(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    student_id: str | int | None = None,
):
    require_programme_scope(user)
    # Accept any value so a hand-edited ?student_id= falls back to the default
    # student instead of returning a validation error.
    student_id = optional_int(student_id)
    student_query = str(request.query_params.get("student_query") or "").strip()

    program = manager_program(session, user)

    # Four-year cohorts for this programme, using the same identity as the
    # Outcome Version Management "Assign Version to Cohort" picker, so one
    # cohort covers all of its semester classes (21ME11Mb1, 21ME12Mb1, ...).
    enrollments = list(
        session.exec(
            select(StudentSemesterEnrollment).where(
                StudentSemesterEnrollment.program_id == program.id
            )
        ).all()
    )
    student_rows = list(
        session.exec(
            select(Student)
            .join(StudentSemesterEnrollment)
            .where(StudentSemesterEnrollment.program_id == program.id)
            .order_by(Student.student_no)
        ).all()
    )
    students_by_id: dict[int, Student] = {}
    for student in student_rows:
        if student.id is not None:
            students_by_id[student.id] = student
    students = sorted(students_by_id.values(), key=lambda item: (item.student_no or "", item.name_en or ""))
    known_student_ids = set(students_by_id)

    # Only offer cohorts that still have real student records. Some enrolment
    # rows point at students that no longer exist, and listing those cohorts
    # would give an option that always shows an empty list.
    cohort_samples: dict[str, str] = {}
    for row in enrollments:
        if row.student_id not in known_student_ids:
            continue
        key = outcome_cohort_key(row.cohort_name)
        if key:
            cohort_samples.setdefault(key, str(row.cohort_name).strip())
    cohort_options = sorted(
        (
            {"key": key, "label": outcome_cohort_label(sample)}
            for key, sample in cohort_samples.items()
        ),
        key=lambda item: item["label"],
    )
    selected_class = str(request.query_params.get("cohort") or "").strip()
    if selected_class and selected_class not in cohort_samples:
        selected_class = ""

    if selected_class:
        class_student_ids = {
            row.student_id
            for row in enrollments
            if outcome_cohort_key(row.cohort_name) == selected_class
        }
        students = [item for item in students if item.id in class_student_ids]

    if student_query:
        lowered_query = student_query.lower()
        students = [
            item for item in students
            if lowered_query in str(item.student_no or "").lower()
            or lowered_query in str(item.name_en or "").lower()
            or lowered_query in str(item.name_kh or "").lower()
        ]

    student_options = [
        {"id": s.id, "student_no": s.student_no, "name": s.name_en}
        for s in students
    ]

    selected_student = None
    if student_id:
        selected_student = session.get(Student, student_id)
        if not selected_student:
            raise HTTPException(status_code=404)
        # Verify programme
        enrollment = session.exec(
            select(StudentSemesterEnrollment).where(
                StudentSemesterEnrollment.student_id == student_id,
                StudentSemesterEnrollment.program_id == program.id
            )
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Student not in your programme")
        # Keep the class picker in step with a student opened from a direct link.
        if not selected_class:
            student_class = next(
                (
                    outcome_cohort_key(row.cohort_name)
                    for row in enrollments
                    if row.student_id == student_id and row.cohort_name
                ),
                "",
            )
            selected_class = student_class if student_class in cohort_samples else ""
    else:
        # If no student selected, pick the first one in the current class.
        if students:
            selected_student = students[0]

    portal_data = None
    if selected_student:
        portal_data = student_portal_data_for_student(session, selected_student)

    # Build stats.
    stats = []
    if portal_data:
        attained = sum(1 for _, value, _ in portal_data["plo_values"] if value >= portal_data["target"])
        not_attained = max(len(portal_data["plo_values"]) - attained, 0)
        completed_assessments = sum(1 for row in portal_data["assessments"] if row["score"] is not None)
        stats = [
            ("Enrolled Courses", len(portal_data["courses"]), "bi-book", "blue"),
            ("Completed Assessments", completed_assessments, "bi-clipboard-check", "green"),
            ("Overall PLO Attainment", f"{portal_data['overall']}%", "bi-bullseye", "purple"),
            ("PLOs >= 70%", f"{attained} / {len(portal_data['plo_values'])}", "bi-graph-up-arrow", "orange"),
            ("PLOs < 70%", not_attained, "bi-graph-down-arrow", "red"),
            ("Total Assessments", len(portal_data["assessments"]), "bi-file-earmark-text", "cyan"),
        ]
    else:
        stats = [("No student selected", "—", "bi-person", "gray")]

    page = {
        "kind": "manager_student_scores",
        "title": "Student Scores & PLO Graph",
        "description": "View scores and PLO attainment for any student in your programme.",
        "students": student_options,
        "cohorts": cohort_options,
        "selected_cohort": selected_class,
        "student_query": student_query,
        "selected_student": selected_student,
        "portal_data": portal_data,
        "stats": stats,
    }

    return templates.TemplateResponse(
        "manager_management.html",
        {
            "request": request,
            "user": user,
            "page": page,
            "manager_section": "student-scores",
            "section_base": programme_section_base(user),
        },
    )

# Also support direct student ID in URL (optional)
@app.get("/manager/students/{student_id}/scores", response_class=HTMLResponse)
@app.get("/dean/students/{student_id}/scores", response_class=HTMLResponse)
def manager_student_scores_with_id(
    student_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    return manager_student_scores_page(request, session, user, student_id=student_id)








@app.get("/manager/students", response_class=HTMLResponse)
def manager_students_page(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)

    program = manager_program(session, user)
    page = build_student_progression_page(request, session, program, user)
    return templates.TemplateResponse(
        "student_progression.html",
        {"request": request, "user": user, "page": page, "manager_section": "students", "section_base": programme_section_base(user)},
    )


@app.get("/admin/students/class/{class_code}", response_class=HTMLResponse)
@app.get("/manager/students/class/{class_code}", response_class=HTMLResponse)
@app.get("/dean/students/class/{class_code}", response_class=HTMLResponse)
def manager_student_class_curriculum_page(
    class_code: str,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Show course scores or a PLO spider graph for Admin, Dean and Manager."""
    if user.role not in {Role.SUPER_ADMIN, Role.PROGRAM_MANAGER, Role.DEAN}:
        raise HTTPException(status_code=403)

    class_code = str(class_code or "").strip()
    if user.role == Role.PROGRAM_MANAGER:
        program = manager_program(session, user)
    else:
        requested_program_id = optional_int(request.query_params.get("program_id"))
        program = session.get(Program, requested_program_id) if requested_program_id else None
        if program is None:
            cohort_enrollment = session.exec(
                select(StudentSemesterEnrollment).where(
                    StudentSemesterEnrollment.cohort_name == class_code
                ).order_by(StudentSemesterEnrollment.id.desc())
            ).first()
            program = (
                session.get(Program, cohort_enrollment.program_id)
                if cohort_enrollment else None
            )
        if program is None:
            for course_class in session.exec(
                select(CourseClass).where(CourseClass.name == class_code)
            ).all():
                course = session.get(Course, course_class.course_id)
                if course:
                    program = session.get(Program, course.program_id)
                    if program:
                        break
        if program is None:
            raise HTTPException(status_code=404, detail="Class programme not found")
        if user.role == Role.DEAN and not can_access_program(user, program):
            raise HTTPException(status_code=403)

    global_period = selected_study_period(request, user)
    academic_year = (
        str(global_period.academic_year or "").strip()
        if global_period is not None
        else str(request.cookies.get("global_academic_year") or "").strip()
    )
    class_year, class_semester = _cohort_year_semester(class_code)
    if not class_year or not class_semester:
        raise HTTPException(status_code=400, detail="Invalid class code")

    # Resolve the class from both semester-enrollment and course-class links.
    # The global period is preferred when it contains this class, but a direct
    # class URL remains usable for historical cohorts and legacy enrolments.
    all_enrollments = list(session.exec(select(StudentSemesterEnrollment).where(
        StudentSemesterEnrollment.program_id == program.id,
        StudentSemesterEnrollment.cohort_name == class_code,
    ).order_by(
        StudentSemesterEnrollment.created_at.desc(),
        StudentSemesterEnrollment.id.desc(),
    )).all())

    programme_course_ids = set(
        session.exec(select(Course.id).where(Course.program_id == program.id)).all()
    )
    matching_class_rows = list(session.exec(
        select(CourseClass).where(
            CourseClass.course_id.in_(programme_course_ids or {-1}),
            CourseClass.name == class_code,
        )
    ).all())
    matching_class_rows = [
        item for item in matching_class_rows
        if normalize_semester(item.semester) == normalize_semester(class_semester)
    ]
    matching_class_ids = [item.id for item in matching_class_rows if item.id]
    linked_class_students = list(session.exec(
        select(ClassStudent).where(ClassStudent.class_id.in_(matching_class_ids or [-1]))
    ).all())
    class_year_by_id = {
        item.id: str(item.academic_year or "").strip()
        for item in matching_class_rows if item.id
    }
    participant_ids_by_year: dict[str, set[int]] = {}
    for enrollment in all_enrollments:
        year_name = str(enrollment.academic_year or "").strip()
        participant_ids_by_year.setdefault(year_name, set()).add(enrollment.student_id)
    for class_student in linked_class_students:
        year_name = class_year_by_id.get(class_student.class_id, "")
        participant_ids_by_year.setdefault(year_name, set()).add(class_student.student_id)

    available_years = {
        str(item.academic_year or "").strip() for item in matching_class_rows
    } | {
        str(item.academic_year or "").strip() for item in all_enrollments
    }
    available_years.discard("")
    if academic_year in available_years and participant_ids_by_year.get(academic_year):
        resolved_academic_year = academic_year
    elif participant_ids_by_year:
        resolved_academic_year = max(
            participant_ids_by_year,
            key=lambda year_name: (len(participant_ids_by_year[year_name]), year_name),
        )
    elif academic_year in available_years:
        resolved_academic_year = academic_year
    else:
        resolved_academic_year = max(available_years, default=academic_year)

    enrollments = [
        item for item in all_enrollments
        if str(item.academic_year or "").strip() == resolved_academic_year
        and normalize_semester(item.semester) == normalize_semester(class_semester)
    ]
    selected_class_rows = [
        item for item in matching_class_rows
        if str(item.academic_year or "").strip() == resolved_academic_year
    ]
    selected_class_ids = {item.id for item in selected_class_rows if item.id}
    selected_class_students = [
        item for item in linked_class_students if item.class_id in selected_class_ids
    ]
    if not all_enrollments and not matching_class_rows:
        raise HTTPException(status_code=404, detail="Class not found")

    latest_by_student: dict[int, StudentSemesterEnrollment] = {}
    for enrollment in enrollments:
        latest_by_student.setdefault(enrollment.student_id, enrollment)

    class_status_by_student: dict[int, str] = {}
    for class_student in selected_class_students:
        current_status = class_status_by_student.get(class_student.student_id)
        if current_status != "Active" or str(class_student.status or "") == "Active":
            class_status_by_student[class_student.student_id] = class_student.status or "Active"

    class_student_ids = set(latest_by_student) | set(class_status_by_student)
    students = [session.get(Student, sid) for sid in class_student_ids]
    students = sorted([student for student in students if student], key=lambda item: item.student_no)
    student_ids = [student.id for student in students if student.id is not None]

    courses = unique_courses_by_code(list(
        session.exec(
            select(Course)
            .where(Course.program_id == program.id)
            .order_by(Course.curriculum_year, Course.curriculum_semester, Course.code)
        ).all()
    ))
    course_ids = [course.id for course in courses if course.id is not None]

    clos = list(session.exec(select(CLO).where(CLO.course_id.in_(course_ids or [-1]))).all())
    clo_ids = [clo.id for clo in clos if clo.id is not None]
    course_by_clo = {clo.id: clo.course_id for clo in clos}
    assessments = list(
        session.exec(select(Assessment).where(Assessment.clo_id.in_(clo_ids or [-1]))).all()
    )
    assessment_ids = [assessment.id for assessment in assessments if assessment.id is not None]
    course_assessments: dict[int, list[Assessment]] = {course.id: [] for course in courses}
    assessments_by_clo: dict[int, list[Assessment]] = {clo.id: [] for clo in clos}
    for assessment in assessments:
        assessments_by_clo.setdefault(assessment.clo_id, []).append(assessment)
        course_id = course_by_clo.get(assessment.clo_id)
        if course_id is not None:
            course_assessments.setdefault(course_id, []).append(assessment)

    scores: list[StudentScore] = []
    if student_ids and assessment_ids:
        scores = list(
            session.exec(
                select(StudentScore).where(
                    StudentScore.student_id.in_(student_ids),
                    StudentScore.assessment_id.in_(assessment_ids),
                )
            ).all()
        )
    score_lookup = {(score.student_id, score.assessment_id): score for score in scores}

    # Keep the spider axes on the PLO version assigned to this four-year cohort.
    # A version assigned at intake (for example 21ME11Mb1) also applies to
    # 21ME12Mb1 through 21ME42Mb1 in the same cohort family.
    class_family = outcome_cohort_key(class_code)
    cohort_assignment = next(
        (
            item for item in session.exec(
                select(CohortOutcomeVersion).where(CohortOutcomeVersion.programme_id == program.id)
            ).all()
            if outcome_cohort_key(item.cohort_name) == class_family
        ),
        None,
    )
    outcome_version = (
        session.get(PLOVersion, cohort_assignment.outcome_version_id)
        if cohort_assignment else active_plo_version_for_program(session, program.id)
    )
    plo_stmt = select(PLO).where(PLO.program_id == program.id, PLO.status == "Active")
    if outcome_version and outcome_version.id:
        plo_stmt = plo_stmt.where(PLO.plo_version_id == outcome_version.id)
    plos = sorted(list(session.exec(plo_stmt).all()), key=plo_sort_key)
    if not plos:
        plos = sorted(
            list(session.exec(select(PLO).where(PLO.program_id == program.id)).all()),
            key=plo_sort_key,
        )
    plo_ids = [plo.id for plo in plos if plo.id is not None]
    plo_labels = [plo.code for plo in plos]
    clo_plo_mappings = list(
        session.exec(
            select(CLOPLOMapping).where(
                CLOPLOMapping.clo_id.in_(clo_ids or [-1]),
                CLOPLOMapping.plo_id.in_(plo_ids or [-1]),
            )
        ).all()
    )
    mappings_by_course_plo: dict[tuple[int, int], list[CLOPLOMapping]] = {}
    for mapping in clo_plo_mappings:
        course_id = course_by_clo.get(mapping.clo_id)
        if course_id is not None:
            mappings_by_course_plo.setdefault((course_id, mapping.plo_id), []).append(mapping)

    current_position = (int(class_year) - 1) * 2 + int(class_semester)
    course_columns: list[dict] = []
    for course in courses:
        course_year = int(course.curriculum_year or 0)
        course_sem = int(semester_number(course.curriculum_semester))
        position = (course_year - 1) * 2 + course_sem if course_year else 999
        course_columns.append({
            "course": course,
            "year": course_year,
            "semester": course_sem,
            "position": position,
            "is_future": position > current_position,
            "is_current": position == current_position,
        })

    rows: list[dict] = []
    for student in students:
        clo_attainment: dict[int, float] = {}
        for clo in clos:
            related = assessments_by_clo.get(clo.id, [])
            weighted_sum = 0.0
            total_weight = 0.0
            for assessment in related:
                score_row = score_lookup.get((student.id, assessment.id))
                weight = assessment_weight_percent(assessment)
                total_weight += weight
                if score_row is None:
                    continue
                max_score = float(assessment.max_score or 0)
                normalized = (
                    float(score_row.score or 0) / max_score * 100.0
                    if max_score > 0 else float(score_row.score or 0)
                )
                normalized = max(0.0, min(100.0, normalized))
                weighted_sum += normalized * weight
            clo_attainment[clo.id] = round(weighted_sum / total_weight, 2) if total_weight > 0 else 0.0

        course_cells: list[dict] = []
        for column in course_columns:
            course = column["course"]
            if column["is_future"]:
                course_cells.append({
                    "state": "future", "label": "Not studied yet", "score": None,
                    "completion_label": "", "radar": [0.0 for _ in plos],
                    "radar_mapped": [False for _ in plos], "mapped_count": 0,
                })
                continue

            related_assessments = course_assessments.get(course.id, [])
            weighted_earned = 0.0
            assessment_weight_total = sum(
                assessment_weight_percent(assessment)
                for assessment in related_assessments
            )
            entered_assessments = 0
            for assessment in related_assessments:
                score_row = score_lookup.get((student.id, assessment.id))
                if score_row is None:
                    continue
                entered_assessments += 1
                max_score = float(assessment.max_score or 0)
                weight = assessment_weight_percent(assessment)
                if max_score > 0:
                    weighted_earned += (float(score_row.score or 0) / max_score) * weight
                else:
                    weighted_earned += float(score_row.score or 0)

            radar_values: list[float] = []
            radar_mapped: list[bool] = []
            mapped_count = 0
            for plo in plos:
                mapped_rows = mappings_by_course_plo.get((course.id, plo.id), [])
                weighted_plo = 0.0
                mapping_total = 0.0
                for mapping in mapped_rows:
                    mapping_weight = stored_percent(mapping.weight)
                    if mapping_weight <= 0:
                        continue
                    weighted_plo += float(clo_attainment.get(mapping.clo_id, 0.0)) * mapping_weight
                    mapping_total += mapping_weight
                if mapping_total > 0:
                    mapped_count += 1
                    radar_mapped.append(True)
                    radar_values.append(round(weighted_plo / mapping_total, 2))
                else:
                    radar_mapped.append(False)
                    radar_values.append(0.0)

            if entered_assessments:
                value = round(
                    weighted_earned / assessment_weight_total * 100.0,
                    2,
                ) if assessment_weight_total > 0 else round(weighted_earned, 2)
                state, label = "score", f"{value:.2f}"
            elif column["is_current"]:
                value = None
                state, label = "progress", "In progress"
            else:
                value = None
                state, label = "missing", "No score"

            course_cells.append({
                "state": state,
                "label": label,
                "score": value,
                "completion_label": (
                    f"{entered_assessments}/{len(related_assessments)} assessments"
                    if related_assessments else "No assessment mapping"
                ),
                "radar": radar_values,
                "radar_mapped": radar_mapped,
                "mapped_count": mapped_count,
            })

        def _score_average(indexes: list[int]) -> float | None:
            values = [
                (
                    float(course_cells[index]["score"]),
                    max(float(course_columns[index]["course"].credits or 1), 0.01),
                )
                for index in indexes
                if course_cells[index].get("score") is not None
            ]
            total_credits = sum(credits for _score, credits in values)
            return (
                round(sum(score * credits for score, credits in values) / total_credits, 2)
                if values and total_credits > 0 else None
            )

        def _radar_average(indexes: list[int]) -> tuple[list[float], int]:
            cells = [
                (
                    course_cells[index],
                    max(float(course_columns[index]["course"].credits or 1), 0.01),
                )
                for index in indexes
                if course_cells[index].get("score") is not None
                and int(course_cells[index].get("mapped_count") or 0) > 0
            ]
            if not cells:
                return [0.0 for _ in plos], 0
            result: list[float] = []
            for plo_index in range(len(plos)):
                mapped_values = [
                    (float(cell["radar"][plo_index]), credits)
                    for cell, credits in cells
                    if plo_index < len(cell.get("radar", []))
                    and plo_index < len(cell.get("radar_mapped", []))
                    and bool(cell["radar_mapped"][plo_index])
                ]
                mapped_credits = sum(credits for _value, credits in mapped_values)
                result.append(
                    round(
                        sum(value * credits for value, credits in mapped_values) / mapped_credits,
                        2,
                    ) if mapped_values and mapped_credits > 0 else 0.0
                )
            return result, len(cells)

        period_summaries: dict[str, dict] = {}
        for study_year in range(1, 5):
            for study_semester in (1, 2):
                key = f"y{study_year}s{study_semester}"
                indexes = [
                    index for index, column in enumerate(course_columns)
                    if int(column["year"]) == study_year
                    and int(column["semester"]) == study_semester
                ]
                radar, radar_count = _radar_average(indexes)
                period_summaries[key] = {
                    "average": _score_average(indexes),
                    "radar": radar,
                    "radar_count": radar_count,
                }

            year_key = f"y{study_year}"
            year_indexes = [
                index for index, column in enumerate(course_columns)
                if int(column["year"]) == study_year
            ]
            year_radar, year_radar_count = _radar_average(year_indexes)
            period_summaries[year_key] = {
                "average": _score_average(year_indexes),
                "radar": year_radar,
                "radar_count": year_radar_count,
            }

        overall_indexes = list(range(len(course_columns)))
        overall_radar, overall_radar_count = _radar_average(overall_indexes)

        rows.append({
            "student": student,
            "enrollment": latest_by_student.get(student.id),
            "status": (
                latest_by_student[student.id].status
                if student.id in latest_by_student
                else class_status_by_student.get(student.id, "Active")
            ),
            "cells": course_cells,
            "period_summaries": period_summaries,
            "overall_average": _score_average(overall_indexes),
            "overall_radar": overall_radar,
            "overall_radar_count": overall_radar_count,
        })

    curriculum_groups: list[dict] = []
    for study_year in range(1, 5):
        semesters: list[dict] = []
        for study_semester in (1, 2):
            grouped_courses = [
                {**column, "cell_index": index}
                for index, column in enumerate(course_columns)
                if int(column["year"]) == study_year
                and int(column["semester"]) == study_semester
            ]
            semesters.append({
                "year": study_year,
                "semester": study_semester,
                "key": f"y{study_year}s{study_semester}",
                "courses": grouped_courses,
                "colspan": len(grouped_courses) + 1,
            })
        curriculum_groups.append({
            "year": study_year,
            "key": f"y{study_year}",
            "semesters": semesters,
            "colspan": sum(item["colspan"] for item in semesters) + 1,
        })

    page = {
        "title": f"{class_code} Curriculum Progress",
        "program": program,
        "class_code": class_code,
        "academic_year": resolved_academic_year,
        "class_year": class_year,
        "class_semester": class_semester,
        "generation": _cohort_generation(class_code),
        "courses": course_columns,
        "curriculum_groups": curriculum_groups,
        "rows": rows,
        "student_count": len(rows),
        "course_count": len(course_columns),
        "plo_labels": plo_labels,
        "outcome_version": outcome_version,
    }
    template_context = {"request": request, "user": user, "page": page}
    if user.role == Role.SUPER_ADMIN:
        template_context["admin_section"] = "student-progression"
    else:
        template_context["manager_section"] = "students"
        template_context["section_base"] = programme_section_base(user)
    return templates.TemplateResponse("student_class_curriculum.html", template_context)


@app.post("/manager/students/{enrollment_id}/status")
async def update_manager_student_status(
    enrollment_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    enrollment = session.get(StudentSemesterEnrollment, enrollment_id)
    if not enrollment or enrollment.program_id != program.id:
        raise HTTPException(status_code=404)
    form = await request.form()
    status = str(form.get("status") or "Active").strip()
    allowed = {"Active", "Probation", "Repeat", "Leave", "Withdrawn", "Dropped", "Graduated"}
    if status not in allowed:
        raise HTTPException(status_code=400, detail="Invalid student status")
    enrollment.status = status
    session.add(enrollment)
    session.commit()
    student = session.get(Student, enrollment.student_id)
    add_audit_record(
        session,
        user,
        "Student Management",
        "UPDATE",
        f"Changed {student.student_no if student else enrollment.student_id} status to {status}.",
        enrollment.cohort_name,
        request.client.host if request.client else "127.0.0.1",
    )
    class_code = str(form.get("class_code") or enrollment.cohort_name or "").strip()
    return redirect(f"/manager/students?class_code={class_code}&updated=1")


@app.post("/manager/plo-targets/create")
@app.post("/manager/plo-target-setup/create")
async def create_manager_plo_target(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    form = await request.form()
    try:
        plo_id = (optional_int(form.get("plo_id")) or 0)
        target_value = float(form.get("target") or 0)
    except (TypeError, ValueError):
        return redirect("/manager/plo-target-setup?error=invalid&show_add=1")
    academic_year = str(form.get("academic_year") or "").strip()
    cohort = str(form.get("cohort") or "").strip()
    plo = session.get(PLO, plo_id)
    if not plo or plo.program_id != program.id or not academic_year or not cohort or not 0 <= target_value <= 100:
        return redirect("/manager/plo-target-setup?error=invalid&show_add=1")
    duplicate = session.exec(
        select(PLOTarget).where(
            PLOTarget.program_id == program.id,
            PLOTarget.plo_id == plo_id,
            PLOTarget.academic_year == academic_year,
            PLOTarget.cohort == cohort,
        )
    ).first()
    if duplicate:
        return redirect("/manager/plo-target-setup?error=duplicate&show_add=1")
    session.add(PLOTarget(
        program_id=program.id,
        plo_id=plo_id,
        academic_year=academic_year,
        cohort=cohort,
        target=round(target_value, 2),
        set_by=user.name,
        updated_at=datetime.utcnow().strftime("%b %d, %Y %I:%M %p"),
    ))
    session.commit()
    return redirect("/manager/plo-target-setup?created=1")


@app.post("/manager/plo-targets/{target_id}/update")
@app.post("/manager/plo-target-setup/{target_id}/update")
async def update_manager_plo_target(
    target_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    record = session.get(PLOTarget, target_id)
    if not record or record.program_id != program.id:
        raise HTTPException(status_code=404)
    form = await request.form()
    try:
        plo_id = (optional_int(form.get("plo_id")) or 0)
        target_value = float(form.get("target") or 0)
    except (TypeError, ValueError):
        return redirect(f"/manager/plo-target-setup?target_id={target_id}&error=invalid")
    academic_year = str(form.get("academic_year") or "").strip()
    cohort = str(form.get("cohort") or "").strip()
    plo = session.get(PLO, plo_id)
    if not plo or plo.program_id != program.id or not academic_year or not cohort or not 0 <= target_value <= 100:
        return redirect(f"/manager/plo-target-setup?target_id={target_id}&error=invalid")
    duplicate = session.exec(
        select(PLOTarget).where(
            PLOTarget.program_id == program.id,
            PLOTarget.plo_id == plo_id,
            PLOTarget.academic_year == academic_year,
            PLOTarget.cohort == cohort,
            PLOTarget.id != target_id,
        )
    ).first()
    if duplicate:
        return redirect(f"/manager/plo-target-setup?target_id={target_id}&error=duplicate")
    record.plo_id = plo_id
    record.academic_year = academic_year
    record.cohort = cohort
    record.target = round(target_value, 2)
    record.set_by = user.name
    record.updated_at = datetime.utcnow().strftime("%b %d, %Y %I:%M %p")
    session.add(record)
    session.commit()
    return redirect("/manager/plo-target-setup?updated=1")


@app.post("/manager/plo-targets/{target_id}/delete")
@app.post("/manager/plo-target-setup/{target_id}/delete")
def delete_manager_plo_target(
    target_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    record = session.get(PLOTarget, target_id)
    if not record or record.program_id != program.id:
        raise HTTPException(status_code=404)
    session.delete(record)
    session.commit()
    return redirect("/manager/plo-target-setup?deleted=1")


@app.get("/manager", response_class=HTMLResponse)
def manager_home(
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    return redirect("/dashboard")

MANAGER_SECTIONS = {
    "outcome-versions",
    "plo-performance",
    "programme-mapping",
    "plo-targets",
    "plo-target-setup",
    "peos",
    "peo-management",
    "plo-management",
    "courses",
    "course-create",
    "course-mapping",
    "assign-teachers",
    "documents",
    "calendar",
    "announcements",
    "reports",
    "enter-scores",
    "users",
    "settings",
    "course-assessment-setup",
}


def programme_section_base(user: User) -> str:
    """Deans work under /dean/*, Programme Managers under /manager/*."""
    return "/dean" if user.role == Role.DEAN else "/manager"


@app.get("/manager/{section}", response_class=HTMLResponse)
def manager_section_page(
    section: str,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    if user.role == Role.DEAN:
        # Deans own these pages under their own URL space; keep old links working.
        query = f"?{request.url.query}" if request.url.query else ""
        return redirect(f"/dean/{section}{query}")
    return render_programme_section(section, request, session, user)


def render_programme_section(
    section: str,
    request: Request,
    session: Session,
    user: User,
):
    """Render one /manager or /dean programme-management section."""
    base = programme_section_base(user)
    # IMPORTANT: this dynamic route is registered before /manager/students.
    # Handle the students page here so Starlette does not return 404 first.
    if section == "students":
        return manager_students_page(request=request, session=session, user=user)
    if section == "assessments":
        return redirect(f"{base}/course-assessment-setup")
    if section not in MANAGER_SECTIONS:
        raise HTTPException(status_code=404)
    
    # --- NEW: Handle enter-scores with combined view ---
    if section == "enter-scores":
        study_period = selected_study_period(request, user)
        unlock_data = score_unlock_page_data(session, user, study_period)
        course_id = optional_int(request.query_params.get("course_id"))
        class_id = optional_int(request.query_params.get("class_id"))
        assessment_id = optional_int(request.query_params.get("assessment_id"))
        score_entry_data = build_manager_score_entry_data(
            session, user, course_id, class_id, assessment_id, study_period
        )
        page = {
            "kind": "manager_enter_scores",
            "title": "Score Entry & Unlock",
            "description": "Unlock locked scores and enter/update scores for your programme courses.",
            "unlock_data": unlock_data,
            "score_entry_data": score_entry_data,
        }
        return templates.TemplateResponse(
            "manager_management.html",
            {"request": request, "user": user, "page": page, "manager_section": section, "section_base": programme_section_base(user)},
        )
    
    # Original handling for other sections
    raw_course_id = request.query_params.get("course_id")
    if section in {"plo-targets", "plo-target-setup"}:
        raw_course_id = request.query_params.get("target_id")
    course_id = int(raw_course_id) if raw_course_id and raw_course_id.isdigit() else None
    raw_version_id = request.query_params.get("version_id")
    version_id = int(raw_version_id) if raw_version_id and raw_version_id.isdigit() else None
    if section == "documents":
        page = document_page_data(session, user, f"{base}/documents")
        return templates.TemplateResponse(
            "manager_management.html",
            {"request": request, "user": user, "page": page,
             "manager_section": section, "section_base": base},
        )
    page = build_manager_page(section, session, user, course_id, version_id, selected_study_period(request, user))
    return templates.TemplateResponse(
        "manager_management.html",
        {"request": request, "user": user, "page": page, "manager_section": section, "section_base": programme_section_base(user)},
    )




@app.post("/manager/programme-mapping/courses/create")
async def create_manager_mapping_course(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)

    form = await request.form()
    program = manager_program(session, user)
    version_id = optional_int(form.get("version_id"))
    version, _versions = selected_outcome_version(session, program, version_id)
    if version.is_locked or version.status in {"Published", "Retired"}:
        return redirect(f"/manager/programme-mapping?version_id={version.id}&error=version_locked")
    code = str(form.get("code", "")).strip()
    title = str(form.get("title", "")).strip()
    semester = str(form.get("curriculum_semester", "")).strip()

    try:
        credits = float(form.get("credits", 0) or 0)
        curriculum_year = (optional_int(form.get("curriculum_year")) or 0)
    except (TypeError, ValueError):
        return redirect(f"/manager/programme-mapping?version_id={version.id}&error=invalid_course&show_add=1")

    if not code or not title or curriculum_year not in {1, 2, 3, 4} or semester not in {"1", "2"} or credits < 0:
        return redirect(f"/manager/programme-mapping?version_id={version.id}&error=invalid_course&show_add=1")

    duplicate = session.exec(
        select(Course).where(Course.program_id == program.id, Course.code == code)
    ).first()
    if duplicate:
        return redirect(f"/manager/programme-mapping?version_id={version.id}&error=duplicate_course&show_add=1")

    course = Course(
        program_id=program.id,
        code=code,
        title=title,
        credits=credits,
        curriculum_year=curriculum_year,
        curriculum_semester=semester,
    )
    session.add(course)
    session.commit()
    session.refresh(course)

    plos = session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)).all()
    for plo in plos:
        raw_weight = form.get(f"weight_{plo.id}", 0)
        try:
            weight = max(0.0, min(100.0, float(raw_weight or 0)))
        except (TypeError, ValueError):
            weight = 0.0
        session.add(
            CoursePLOMapping(
                course_id=course.id,
                plo_id=plo.id,
                level=weight,
                symbol="",
            )
        )
    session.commit()
    return redirect(f"/manager/programme-mapping?version_id={version.id}&course=created&course_id={course.id}")


@app.post("/manager/programme-mapping/courses/{course_id}/delete")
def delete_manager_course_mapping(course_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    require_programme_scope(user)
    program = manager_program(session, user)
    course = session.get(Course, course_id)
    if not course or course.program_id != program.id:
        raise HTTPException(status_code=404)
    for mapping in session.exec(select(CoursePLOMapping).where(CoursePLOMapping.course_id == course_id)).all():
        session.delete(mapping)
    official_codes = {code for _year, _semester, code, _title, _credits in ME_CURRICULUM} if program.code == "ME" else set()
    has_class = session.exec(select(CourseClass).where(CourseClass.course_id == course_id)).first()
    has_clo = session.exec(select(CLO).where(CLO.course_id == course_id)).first()
    if course.code not in official_codes and not has_class and not has_clo:
        session.delete(course)
    session.commit()
    return redirect("/manager/programme-mapping?mapping=deleted")


@app.post("/manager/courses/create")
def create_manager_course(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    code: str = Form(...),
    title: str = Form(...),
    credits: float = Form(...),
    curriculum_year: int = Form(...),
    curriculum_semester: str = Form(...),
):
    require_programme_scope(user)
    program = manager_program(session, user)
    clean_code = code.strip()
    if session.exec(select(Course).where(Course.program_id == program.id, Course.code == clean_code)).first():
        return redirect("/manager/courses?error=duplicate_course")
    course = Course(
        program_id=program.id,
        code=clean_code,
        title=title.strip(),
        credits=credits,
        curriculum_year=curriculum_year,
        curriculum_semester=curriculum_semester.strip(),
    )
    session.add(course)
    session.commit()
    session.refresh(course)
    # Course CRUD is intentionally separate from PLO mapping. Mapping records
    # are created later from /manager/programme-mapping or /manager/course-mapping.
    return redirect(f"/manager/courses?created=1&new_course_id={course.id}")


@app.post("/manager/courses/{course_id}/update")
def update_manager_course(
    course_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    code: str = Form(...),
    title: str = Form(...),
    credits: float = Form(...),
    curriculum_year: int = Form(...),
    curriculum_semester: str = Form(...),
):
    require_programme_scope(user)
    program = manager_program(session, user)
    course = session.get(Course, course_id)
    if not course or course.program_id != program.id:
        raise HTTPException(status_code=404)
    clean_code = code.strip()
    duplicate = session.exec(select(Course).where(Course.program_id == program.id, Course.code == clean_code, Course.id != course_id)).first()
    if duplicate:
        return redirect(f"/manager/courses?course_id={course_id}&error=duplicate_course")
    course.code = clean_code
    course.title = title.strip()
    course.credits = credits
    course.curriculum_year = curriculum_year
    course.curriculum_semester = curriculum_semester.strip()
    session.add(course)
    session.commit()
    return redirect(f"/manager/courses?course_id={course_id}&updated=1")


@app.post("/manager/courses/{course_id}/delete")
def delete_manager_course(course_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    require_programme_scope(user)
    program = manager_program(session, user)
    course = session.get(Course, course_id)
    if not course or course.program_id != program.id:
        raise HTTPException(status_code=404)
    official_codes = {code for _year, _semester, code, _title, _credits in ME_CURRICULUM} if program.code == "ME" else set()
    has_class = session.exec(select(CourseClass).where(CourseClass.course_id == course_id)).first()
    for mapping in session.exec(select(CoursePLOMapping).where(CoursePLOMapping.course_id == course_id)).all():
        session.delete(mapping)
    if course.code in official_codes or has_class:
        session.commit()
        return redirect("/manager/courses?mapping_cleared=1")
    for clo in session.exec(select(CLO).where(CLO.course_id == course_id)).all():
        for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all():
            session.delete(mapping)
        for assessment in session.exec(select(Assessment).where(Assessment.clo_id == clo.id)).all():
            delete_assessment_with_scores(session, assessment)
        session.delete(clo)
    session.delete(course)
    session.commit()
    return redirect("/manager/courses?deleted=1")




def ensure_course_mapping_records(session: Session, course: Course, plos: list[PLO]) -> None:
    """Make sure every course has a Course-PLO mapping row so the UI is always connected to DB rows."""
    for plo in plos:
        existing = session.exec(
            select(CoursePLOMapping).where(CoursePLOMapping.course_id == course.id, CoursePLOMapping.plo_id == plo.id)
        ).first()
        if not existing:
            session.add(CoursePLOMapping(course_id=course.id, plo_id=plo.id, level=0, symbol=""))


def default_clo_code(session: Session, course_id: int) -> str:
    total = len(session.exec(select(CLO).where(CLO.course_id == course_id)).all()) + 1
    return f"CLO{total}"


def create_standard_assessment_plan(session: Session, course: Course, plos: list[PLO]) -> int:
    """Create a simple real database assessment plan for testing and first setup."""
    existing = session.exec(select(CLO).where(CLO.course_id == course.id)).first()
    if existing:
        return 0
    clo_templates = [
        ("CLO1", f"Explain fundamental concepts of {course.title}."),
        ("CLO2", f"Apply methods and tools related to {course.title}."),
        ("CLO3", f"Analyze problems and propose appropriate solutions in {course.title}."),
        ("CLO4", f"Communicate results and demonstrate professional responsibility in {course.title}."),
    ]
    created = 0
    for idx, (code, desc) in enumerate(clo_templates):
        clo = CLO(course_id=course.id, code=code, description=desc)
        session.add(clo)
        session.commit()
        session.refresh(clo)
        # Map each CLO to one or two PLOs in the real CLO-PLO mapping table.
        if plos:
            primary = plos[idx % len(plos)]
            session.add(CLOPLOMapping(clo_id=clo.id, plo_id=primary.id, weight=70))
            if len(plos) > 1:
                secondary = plos[(idx + 1) % len(plos)]
                session.add(CLOPLOMapping(clo_id=clo.id, plo_id=secondary.id, weight=30))
        created += 1
    assessment_templates = [
        ("CLO1", "Attendance", 10),
        ("CLO1", "Quiz", 10),
        ("CLO2", "Assignment", 20),
        ("CLO3", "Midterm Exam", 20),
        ("CLO4", "Final Exam", 40),
    ]
    clos_by_code = {clo.code: clo for clo in session.exec(select(CLO).where(CLO.course_id == course.id)).all()}
    for clo_code, name, weight in assessment_templates:
        clo = clos_by_code.get(clo_code)
        if clo:
            session.add(Assessment(clo_id=clo.id, name=name, max_score=weight, weight=weight, description=f"{name} assessment for {clo.code}"))
            created += 1
    ensure_course_mapping_records(session, course, plos)
    session.commit()
    return created


@app.post("/manager/course-mapping/{course_id}/update")
async def update_manager_course_mapping(course_id: int, request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    require_programme_scope(user)
    program = manager_program(session, user)
    course = session.get(Course, course_id)
    if not course or course.program_id != program.id:
        raise HTTPException(status_code=404)
    form = await request.form()
    version_id = (optional_int(form.get("version_id")) or 0)
    version = editable_outcome_version(session, program, version_id)
    plos = session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id).order_by(PLO.code)).all()
    ensure_course_mapping_records(session, course, plos)
    for plo in plos:
        level = clamp_percent(form.get(f"weight_{plo.id}", form.get(f"level_{plo.id}", "0")))
        mapping = session.exec(select(CoursePLOMapping).where(CoursePLOMapping.course_id == course.id, CoursePLOMapping.plo_id == plo.id)).first()
        symbol = f"{level}%" if level else ""
        if mapping:
            mapping.level = level
            mapping.symbol = symbol
            session.add(mapping)
        else:
            session.add(CoursePLOMapping(course_id=course.id, plo_id=plo.id, level=level, symbol=symbol))
    session.commit()
    return redirect(f"/manager/course-mapping?course_id={course_id}&version_id={version.id}&updated=1")


@app.post("/manager/course-assessment-setup/{course_id}/update")
async def update_manager_course_assessment_setup(
    course_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    course = session.get(Course, course_id)
    if not course or course.program_id != program.id:
        raise HTTPException(status_code=404)

    form = await request.form()
    version_id = (optional_int(form.get("version_id")) or 0)
    version = editable_outcome_version(session, program, version_id)
    submit_action = str(form.get("submit_action", "save_all")).strip()
    plos = session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)).all()
    clos = session.exec(select(CLO).where(CLO.course_id == course.id).order_by(CLO.code)).all()

    delete_assessment_id = str(form.get("delete_assessment_id", "")).strip()
    if delete_assessment_id:
        assessment = session.get(Assessment, int(delete_assessment_id))
        if assessment and assessment.clo and assessment.clo.course_id == course.id:
            delete_assessment_with_scores(session, assessment)
            session.commit()
            return redirect(f"/manager/course-mapping?course_id={course_id}&deleted=1")

    delete_clo_id = str(form.get("delete_clo_id", "")).strip()
    if delete_clo_id:
        clo = session.get(CLO, int(delete_clo_id))
        if clo and clo.course_id == course.id:
            for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all():
                session.delete(mapping)
            for assessment in session.exec(select(Assessment).where(Assessment.clo_id == clo.id)).all():
                delete_assessment_with_scores(session, assessment)
            session.delete(clo)
            session.commit()
            return redirect(f"/manager/course-mapping?course_id={course_id}&deleted=1")

    # Assessment capacity rule:
    # A CLO's assessments may total only up to the highest PLO percentage mapped to that CLO.
    # Example: CLO1 -> PLO1 20%, PLO2 50% means CLO1 assessment total must be 50%, not 100%.
    target_by_clo: dict[int, float] = {}
    proposed_total_by_clo: dict[int, float] = {}
    for clo in clos:
        posted_weights = [
            clamp_percent(form.get(f"mapping_{clo.id}_{plo.id}"))
            for plo in plos
            if f"mapping_{clo.id}_{plo.id}" in form
        ]
        if posted_weights:
            target_by_clo[clo.id] = max(posted_weights, default=0.0)
        else:
            stored = session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all()
            target_by_clo[clo.id] = max((float(item.weight or 0) for item in stored), default=0.0)

        proposed_total = 0.0
        for assessment in session.exec(select(Assessment).where(Assessment.clo_id == clo.id)).all():
            proposed_total += clamp_percent(form.get(f"assessment_weight_{assessment.id}", assessment.weight))
        proposed_total_by_clo[clo.id] = proposed_total

    bulk_clo_ids_for_validation = form.getlist("bulk_assessment_clo_id")
    bulk_names_for_validation = form.getlist("bulk_assessment_name")
    bulk_weights_for_validation = form.getlist("bulk_assessment_weight")
    for index, raw_name in enumerate(bulk_names_for_validation):
        if not str(raw_name).strip():
            continue
        raw_clo_id = str(bulk_clo_ids_for_validation[index] if index < len(bulk_clo_ids_for_validation) else "").strip()
        try:
            validation_clo_id = int(raw_clo_id)
        except (TypeError, ValueError):
            continue
        if validation_clo_id in proposed_total_by_clo:
            raw_weight = bulk_weights_for_validation[index] if index < len(bulk_weights_for_validation) else 0
            proposed_total_by_clo[validation_clo_id] += clamp_percent(raw_weight)

    for clo in clos:
        target = round(target_by_clo.get(clo.id, 0.0), 6)
        proposed = round(proposed_total_by_clo.get(clo.id, 0.0), 6)
        if proposed > target:
            return redirect(
                f"/manager/course-mapping?course_id={course_id}"
                f"&mapping_error=assessment_over_target&clo_code={clo.code}&target={target:g}#preview"
            )

    for clo in clos:
        clo.code = str(form.get(f"clo_code_{clo.id}", clo.code)).strip() or clo.code
        clo.description = str(form.get(f"clo_description_{clo.id}", clo.description)).strip() or clo.description
        session.add(clo)
        for plo in plos:
            key = f"mapping_{clo.id}_{plo.id}"
            if key not in form:
                continue
            weight = clamp_percent(form.get(key))
            mapping = session.exec(
                select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id, CLOPLOMapping.plo_id == plo.id)
            ).first()
            if weight == 0:
                if mapping:
                    session.delete(mapping)
                continue
            if mapping:
                mapping.weight = weight
                session.add(mapping)
            else:
                session.add(CLOPLOMapping(clo_id=clo.id, plo_id=plo.id, weight=weight))

    for assessment in session.exec(select(Assessment).join(CLO).where(CLO.course_id == course.id)).all():
        assessment.name = str(form.get(f"assessment_name_{assessment.id}", assessment.name)).strip() or assessment.name
        assessment.weight = clamp_percent(form.get(f"assessment_weight_{assessment.id}", assessment.weight))
        assessment.max_score = assessment.weight
        assessment.description = str(form.get(f"assessment_description_{assessment.id}", assessment.description or "")).strip() or None
        session.add(assessment)

    new_clo_code = str(form.get("new_clo_code", "")).strip()
    new_clo_description = str(form.get("new_clo_description", "")).strip()
    new_assessment_name = str(form.get("new_assessment_name", "")).strip()
    new_assessment_weight = clamp_percent(form.get("new_assessment_weight"))
    new_assessment_score = clamp_percent(form.get("new_assessment_score"))
    selected_clo_id = str(form.get("new_assessment_clo_id", "")).strip()
    target_clo = session.get(CLO, int(selected_clo_id)) if selected_clo_id else None
    if target_clo and target_clo.course_id != course.id:
        target_clo = None
    if not target_clo and new_clo_code and new_clo_description:
        target_clo = CLO(course_id=course.id, code=new_clo_code, description=new_clo_description)
        session.add(target_clo)
        session.commit()
        session.refresh(target_clo)
        for plo in plos:
            weight = clamp_percent(form.get(f"new_mapping_{plo.id}"))
            if weight:
                session.add(CLOPLOMapping(clo_id=target_clo.id, plo_id=plo.id, weight=weight))
    if target_clo and new_assessment_name:
        session.add(
            Assessment(
                clo_id=target_clo.id,
                name=new_assessment_name,
                max_score=new_assessment_weight,
                weight=new_assessment_weight,
            )
        )

    bulk_clo_ids = form.getlist("bulk_assessment_clo_id")
    bulk_names = form.getlist("bulk_assessment_name")
    bulk_weights = form.getlist("bulk_assessment_weight")
    for index, bulk_name in enumerate(bulk_names):
        assessment_name = str(bulk_name).strip()
        clo_id_value = str(bulk_clo_ids[index] if index < len(bulk_clo_ids) else "").strip()
        if not assessment_name or not clo_id_value:
            continue
        try:
            bulk_clo_id = int(clo_id_value)
        except ValueError:
            continue
        bulk_clo = session.get(CLO, bulk_clo_id)
        if not bulk_clo or bulk_clo.course_id != course.id:
            continue
        weight_value = bulk_weights[index] if index < len(bulk_weights) else 0
        assessment_weight = clamp_percent(weight_value)
        session.add(
            Assessment(
                clo_id=bulk_clo.id,
                name=assessment_name,
                max_score=assessment_weight,
                weight=assessment_weight,
            )
        )

    session.commit()
    anchor = "assessmentMappingPanel" if submit_action in {"add_assessment", "update_assessments"} else "clo-section"
    return redirect(f"/manager/course-mapping?course_id={course_id}&version_id={version.id}&updated=1#{anchor}")




@app.post("/manager/course-mapping/{course_id}/assessment-update")
async def update_manager_course_mapping_assessments_alias(
    course_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Alias used by the /manager/course-mapping page so Assessment Mapping saves to the same real DB function."""
    return await update_manager_course_assessment_setup(course_id, request, session, user)


@app.post("/manager/course-mapping/{course_id}/create-standard-assessments")
def create_manager_standard_assessments(
    course_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    course = session.get(Course, course_id)
    if not course or course.program_id != program.id:
        raise HTTPException(status_code=404)
    plos = session.exec(select(PLO).where(PLO.program_id == program.id).order_by(PLO.code)).all()
    created = create_standard_assessment_plan(session, course, plos)
    return redirect(f"/manager/course-mapping?course_id={course_id}&standard_created={created}")


@app.post("/manager/assign-teachers")
def assign_manager_course_teacher(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    course_id: int = Form(...),
    class_id: int = Form(...),
    teacher_id: int = Form(...),
):
    require_programme_scope(user)
    program = manager_program(session, user)
    course = session.get(Course, course_id)
    course_class = session.get(CourseClass, class_id)
    teacher = session.get(Teacher, teacher_id)
    study_period = selected_study_period(request, user)
    if (
        not course
        or course.program_id != program.id
        or not course_class
        or course_class.course_id != course.id
        or not class_matches_study_period(course_class, study_period)
        or not teacher
    ):
        raise HTTPException(status_code=404)

    class_assignment = session.exec(
        select(ClassTeacher).where(
            ClassTeacher.class_id == class_id,
            ClassTeacher.teacher_id == teacher_id,
        )
    ).first()
    if not class_assignment:
        session.add(ClassTeacher(class_id=class_id, teacher_id=teacher_id))
    course_assignment = session.exec(
        select(CourseTeacher).where(
            CourseTeacher.course_id == course_id,
            CourseTeacher.teacher_id == teacher_id,
        )
    ).first()
    if not course_assignment:
        session.add(CourseTeacher(course_id=course_id, teacher_id=teacher_id))
    session.commit()
    return redirect(f"/manager/course-mapping?course_id={course_id}&teacher_assigned=1#teacher-assignment")


@app.post("/manager/class-teacher-assignments/{assignment_id}/delete")
def delete_manager_class_teacher(
    assignment_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    assignment = session.get(ClassTeacher, assignment_id)
    if (
        not assignment
        or not assignment.course_class
        or not assignment.course_class.course
        or assignment.course_class.course.program_id != program.id
    ):
        raise HTTPException(status_code=404)
    course_id = assignment.course_class.course_id
    teacher_id = assignment.teacher_id
    session.delete(assignment)
    session.flush()

    remaining_class_assignment = next(
        (
            row
            for row in session.exec(select(ClassTeacher).where(ClassTeacher.teacher_id == teacher_id)).all()
            if row.course_class and row.course_class.course_id == course_id
        ),
        None,
    )
    if remaining_class_assignment is None:
        for row in session.exec(
            select(CourseTeacher).where(
                CourseTeacher.course_id == course_id,
                CourseTeacher.teacher_id == teacher_id,
            )
        ).all():
            session.delete(row)
    session.commit()
    return redirect(f"/manager/course-mapping?course_id={course_id}&teacher_removed=1#teacher-assignment")


@app.post("/manager/assign-teachers/{assignment_id}/delete")
def delete_manager_course_teacher(
    assignment_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    assignment = session.get(CourseTeacher, assignment_id)
    if not assignment or not assignment.course or assignment.course.program_id != program.id:
        raise HTTPException(status_code=404)
    course_id = assignment.course_id
    session.delete(assignment)
    session.commit()
    return redirect(f"/manager/course-mapping?course_id={course_id}&teacher_removed=1#teacher-assignment")


MANAGER_OUTCOME_RETURN_SECTIONS = {
    "outcome-versions",
    "programme-mapping",
    "plo-management",
    "peos",
    "peo-management",
    "plo-targets",
    "plo-target-setup",
}


def manager_version_return_path(value: object, base: str = "/manager") -> str:
    """Keep outcome-version redirects on the pages that host the panel.

    Accepts either URL space so a Dean posting from /dean/* lands back there.
    """
    path = str(value or "").split("?")[0].strip()
    for prefix in ("/manager/", "/dean/"):
        if path.startswith(prefix):
            section = path[len(prefix):]
            if section in MANAGER_OUTCOME_RETURN_SECTIONS:
                return f"{prefix}{section}"
            break
    return f"{base}/programme-mapping"


@app.post("/manager/select-programme")
async def select_manager_programme(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Switch which faculty programme a Dean is working on."""
    require_programme_scope(user)
    form = await request.form()
    return_to = safe_return_path(form.get("return_to"))
    if not return_to.startswith(("/manager/", "/dean/")):
        return_to = f"{programme_section_base(user)}/programme-mapping"
    if user.role != Role.DEAN:
        # Programme Managers own exactly one programme; nothing to switch.
        return redirect(return_to)
    program = session.get(Program, optional_int(form.get("program_id")) or 0)
    if not program or not can_access_program(user, program):
        return redirect(f"{return_to}{'&' if '?' in return_to else '?'}programme_error=1")
    set_active_programme(session, user, program)
    return redirect(return_to)


@app.post("/manager/outcome-versions/create")
async def create_manager_outcome_version(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    require_programme_scope(user)
    program = manager_program(session, user)
    form = await request.form()
    name = str(form.get("version_name") or "").strip()
    return_to = manager_version_return_path(form.get("return_to"), programme_section_base(user))
    if not name:
        return redirect(f"{return_to}?version_error=required")
    if session.exec(select(PLOVersion).where(PLOVersion.programme_id == program.id, PLOVersion.version_name == name)).first():
        return redirect(f"{return_to}?version_error=duplicate")
    source_id = (optional_int(form.get("source_version_id")) or 0)
    source = session.get(PLOVersion, source_id) if source_id else None
    version = PLOVersion(programme_id=program.id, version_name=name, status="Draft", created_by=user.id)
    session.add(version); session.flush()
    if source and source.programme_id == program.id:
        old_plos = session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == source.id)).all()
        plo_map = {}
        for old in old_plos:
            new = PLO(program_id=program.id, plo_version_id=version.id, code=old.code, description=old.description, domain=old.domain, bloom_level=old.bloom_level, status=old.status, remark=old.remark, created_by=user.id)
            session.add(new); session.flush(); plo_map[old.id] = new.id
        old_peos = session.exec(select(PEO).where(PEO.program_id == program.id, PEO.plo_version_id == source.id)).all()
        peo_map = {}
        for old in old_peos:
            new = PEO(program_id=program.id, plo_version_id=version.id, code=old.code, description=old.description, status=old.status, remark=old.remark, created_by=user.id)
            session.add(new); session.flush(); peo_map[old.id] = new.id
        for old in session.exec(select(PEOPLOMapping).where(PEOPLOMapping.plo_version_id == source.id)).all():
            if old.peo_id in peo_map and old.plo_id in plo_map:
                session.add(PEOPLOMapping(program_id=program.id, plo_version_id=version.id, peo_id=peo_map[old.peo_id], plo_id=plo_map[old.plo_id], mapping_mode=old.mapping_mode, is_mapped=old.is_mapped, contribution_percentage=old.contribution_percentage, created_by=user.id))
        for old in session.exec(select(CLOPLOMapping)).all():
            if old.plo_id in plo_map:
                session.add(CLOPLOMapping(clo_id=old.clo_id, plo_id=plo_map[old.plo_id], weight=old.weight))
        for old in session.exec(select(PLOTarget).where(PLOTarget.program_id == program.id)).all():
            if old.plo_id in plo_map:
                session.add(PLOTarget(program_id=program.id, plo_id=plo_map[old.plo_id], academic_year=old.academic_year, cohort=old.cohort, target=old.target, set_by=user.name, updated_at=datetime.utcnow().strftime("%b %d, %Y %I:%M %p")))
    session.commit()
    sep = "&" if "?" in return_to else "?"
    return redirect(f"{return_to}{sep}version_id={version.id}&version_created=1")


@app.post("/manager/outcome-versions/assign")
async def assign_manager_outcome_version(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    require_programme_scope(user)
    program = manager_program(session, user)
    form = await request.form()
    cohort_name = outcome_cohort_key(form.get("cohort_name"))
    version_id = (optional_int(form.get("version_id")) or 0)
    return_to = manager_version_return_path(form.get("return_to"), programme_section_base(user))
    version = session.get(PLOVersion, version_id)
    if not cohort_name or not version or version.programme_id != program.id:
        return redirect(f"{return_to}?version_error=invalid_assignment")
    if version.status not in {"Active", "Published"} or not version.is_locked:
        return redirect(f"{return_to}?version_error=publish_before_assignment")
    record = next(
        (
            item for item in session.exec(
                select(CohortOutcomeVersion).where(CohortOutcomeVersion.programme_id == program.id)
            ).all()
            if outcome_cohort_key(item.cohort_name) == cohort_name
        ),
        None,
    )
    if record:
        if record.outcome_version_id != version.id:
            return redirect(f"{return_to}?version_error=cohort_already_assigned")
        if record.cohort_name != cohort_name:
            record.cohort_name = cohort_name
            session.add(record)
    else:
        session.add(CohortOutcomeVersion(programme_id=program.id, cohort_name=cohort_name, outcome_version_id=version.id, assigned_by=user.id))
    session.commit()
    sep = "&" if "?" in return_to else "?"
    return redirect(f"{return_to}{sep}version_id={version.id}&version_assigned=1")


@app.post("/manager/outcome-versions/publish")
async def publish_manager_outcome_version(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    """Freeze an approved CQI version so reports remain reproducible."""
    require_programme_scope(user)
    program = manager_program(session, user)
    form = await request.form()
    version_id = (optional_int(form.get("version_id")) or 0)
    return_to = manager_version_return_path(form.get("return_to"), programme_section_base(user))
    version = session.get(PLOVersion, version_id)
    if not version or version.programme_id != program.id:
        return redirect(f"{return_to}?version_error=invalid_version")
    peos = session.exec(select(PEO).where(PEO.program_id == program.id, PEO.plo_version_id == version.id, PEO.status == "Active")).all()
    plos = session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id, PLO.status == "Active")).all()
    if not peos or not plos:
        return redirect(f"{return_to}?version_error=incomplete_version")
    version.status = "Published"
    version.is_locked = True
    version.updated_at = datetime.utcnow()
    session.add(version)
    session.commit()
    sep = "&" if "?" in return_to else "?"
    return redirect(f"{return_to}{sep}version_id={version.id}&version_published=1")


@app.post("/manager/outcome-versions/unlock")
async def unlock_manager_outcome_version(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    """Reopen a published version so the programme can continue CQI editing."""
    require_programme_scope(user)
    program = manager_program(session, user)
    form = await request.form()
    version_id = (optional_int(form.get("version_id")) or 0)
    return_to = manager_version_return_path(form.get("return_to"), programme_section_base(user))
    version = session.get(PLOVersion, version_id)
    if not version or version.programme_id != program.id:
        return redirect(f"{return_to}?version_error=invalid_version")
    version.status = "Draft"
    version.is_locked = False
    version.updated_at = datetime.utcnow()
    session.add(version)
    session.commit()
    return redirect(f"{return_to}?version_id={version.id}&version_unlocked=1")


@app.post("/manager/outcome-versions/delete")
async def delete_manager_outcome_version(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    """Permanently remove an unused outcome package and its dependants."""
    require_programme_scope(user)
    program = manager_program(session, user)
    form = await request.form()
    version_id = (optional_int(form.get("version_id")) or 0)
    return_to = manager_version_return_path(form.get("return_to"), programme_section_base(user))
    version = session.get(PLOVersion, version_id)
    if not version or version.programme_id != program.id:
        return redirect(f"{return_to}?version_error=invalid_version")

    versions = list(
        session.exec(
            select(PLOVersion)
            .where(PLOVersion.programme_id == program.id)
            .order_by(PLOVersion.id.desc())
        ).all()
    )
    if len(versions) <= 1:
        return redirect(f"{return_to}?version_id={version.id}&version_error=last_version")
    assignment = session.exec(
        select(CohortOutcomeVersion).where(CohortOutcomeVersion.outcome_version_id == version.id)
    ).first()
    if assignment:
        return redirect(f"{return_to}?version_id={version.id}&version_error=assigned_version")

    plos = list(
        session.exec(
            select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)
        ).all()
    )
    peos = list(
        session.exec(
            select(PEO).where(PEO.program_id == program.id, PEO.plo_version_id == version.id)
        ).all()
    )
    plo_ids = {item.id for item in plos if item.id is not None}
    peo_ids = {item.id for item in peos if item.id is not None}

    # Delete referencing rows before their PEO/PLO parents to satisfy FK rules.
    for mapping in session.exec(select(PEOPLOMapping)).all():
        if (
            mapping.plo_version_id == version.id
            or mapping.peo_id in peo_ids
            or mapping.plo_id in plo_ids
        ):
            session.delete(mapping)
    if plo_ids:
        for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.plo_id.in_(plo_ids))).all():
            session.delete(mapping)
        for mapping in session.exec(select(CoursePLOMapping).where(CoursePLOMapping.plo_id.in_(plo_ids))).all():
            session.delete(mapping)
        for target in session.exec(select(PLOTarget).where(PLOTarget.plo_id.in_(plo_ids))).all():
            session.delete(target)
    for peo in peos:
        session.delete(peo)
    for plo in plos:
        session.delete(plo)

    deleted_name = version.version_name
    remaining = [item for item in versions if item.id != version_id]
    session.delete(version)
    session.commit()
    add_audit_record(
        session,
        user,
        "Outcome Version Management",
        "DELETE",
        f"Deleted unassigned outcome version {deleted_name} and its dependent mappings.",
        program.code,
        request.client.host if request.client else "127.0.0.1",
    )
    fallback = next((item for item in remaining if item.status in {"Active", "Published"}), remaining[0])
    return redirect(f"{return_to}?version_id={fallback.id}&version_deleted=1")


@app.post("/manager/peos/create")
@app.post("/manager/peo-management/create")
async def create_manager_peo(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    form = await request.form()
    code = str(form.get("code") or "").strip().upper()
    description = str(form.get("description") or "").strip()
    version_id = (optional_int(form.get("version_id")) or 0)
    version = editable_outcome_version(session, program, version_id)
    if not code or not description:
        return redirect(f"/manager/peo-management?version_id={version.id}&error=required&show_add=1")
    if session.exec(select(PEO).where(PEO.program_id == program.id, PEO.plo_version_id == version.id, PEO.code == code)).first():
        return redirect(f"/manager/peo-management?version_id={version.id}&error=duplicate&show_add=1")

    peo = PEO(program_id=program.id, plo_version_id=version.id, code=code, description=description)
    session.add(peo)
    session.flush()
    plos = session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)).all()
    for plo in plos:
        weight = clamp_percent(form.get(f"weight_{plo.id}"))
        if plo.id and weight > 0:
            session.add(PEOPLOMapping(program_id=program.id, plo_version_id=version.id, peo_id=peo.id, plo_id=plo.id, contribution_percentage=weight, is_mapped=True))
    session.commit()
    return redirect(f"/manager/peo-management?version_id={version.id}&created=1")


@app.post("/manager/peos/{peo_id}/update")
@app.post("/manager/peo-management/{peo_id}/update")
async def update_manager_peo(
    peo_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    peo = session.get(PEO, peo_id)
    if not peo or peo.program_id != program.id:
        raise HTTPException(status_code=404)
    form = await request.form()
    code = str(form.get("code") or "").strip().upper()
    description = str(form.get("description") or "").strip()
    version = editable_outcome_version(session, program, peo.plo_version_id)
    if not code or not description:
        return redirect(f"/manager/peo-management?version_id={version.id}&peo_id={peo_id}&error=required")
    duplicate = session.exec(
        select(PEO).where(PEO.program_id == program.id, PEO.plo_version_id == version.id, PEO.code == code, PEO.id != peo_id)
    ).first()
    if duplicate:
        return redirect(f"/manager/peo-management?version_id={version.id}&peo_id={peo_id}&error=duplicate")

    peo.code = code
    peo.description = description
    session.add(peo)
    for mapping in session.exec(select(PEOPLOMapping).where(PEOPLOMapping.peo_id == peo_id)).all():
        session.delete(mapping)
    plos = session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)).all()
    for plo in plos:
        weight = clamp_percent(form.get(f"weight_{plo.id}"))
        if plo.id and weight > 0:
            session.add(PEOPLOMapping(program_id=program.id, plo_version_id=version.id, peo_id=peo_id, plo_id=plo.id, contribution_percentage=weight, is_mapped=True))
    session.commit()
    return redirect(f"/manager/peo-management?version_id={version.id}&updated=1")


@app.post("/manager/peos/{peo_id}/delete")
@app.post("/manager/peo-management/{peo_id}/delete")
def delete_manager_peo(
    peo_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    peo = session.get(PEO, peo_id)
    if not peo or peo.program_id != program.id:
        raise HTTPException(status_code=404)
    version = editable_outcome_version(session, program, peo.plo_version_id)
    for mapping in session.exec(select(PEOPLOMapping).where(PEOPLOMapping.peo_id == peo_id)).all():
        session.delete(mapping)
    session.delete(peo)
    session.commit()
    return redirect(f"/manager/peo-management?version_id={version.id}&deleted=1")


@app.post("/manager/peos/{peo_id}/mapping")
@app.post("/manager/peo-management/{peo_id}/mapping")
async def update_manager_peo_mapping(
    peo_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    # Backward-compatible endpoint. It updates the same real mapping rows.
    require_programme_scope(user)
    program = manager_program(session, user)
    peo = session.get(PEO, peo_id)
    if not peo or peo.program_id != program.id:
        raise HTTPException(status_code=404)
    version = editable_outcome_version(session, program, peo.plo_version_id)
    form = await request.form()
    for mapping in session.exec(select(PEOPLOMapping).where(PEOPLOMapping.peo_id == peo_id)).all():
        session.delete(mapping)
    for plo in session.exec(select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id)).all():
        weight = clamp_percent(form.get(f"weight_{plo.id}"))
        if plo.id and weight > 0:
            session.add(PEOPLOMapping(program_id=program.id, plo_version_id=version.id, peo_id=peo_id, plo_id=plo.id, contribution_percentage=weight, is_mapped=True))
    session.commit()
    return redirect(f"/manager/peo-management?version_id={version.id}&mapping=updated")


@app.post("/manager/plos/create")
@app.post("/manager/plo-management/create")
async def create_manager_plo(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    form = await request.form()
    version = editable_outcome_version(session, program, optional_int(form.get("version_id")))
    code = str(form.get("code") or "").strip().upper()
    description = str(form.get("description") or "").strip()
    domain = str(form.get("domain") or "Knowledge").strip()
    bloom_level = str(form.get("bloom_level") or "C1").strip()
    status = str(form.get("status") or "Active").strip()
    remark = str(form.get("remark") or "").strip()
    if not code or not description:
        return redirect(f"/manager/plo-management?version_id={version.id}&error=required&show_add=1")
    duplicate = session.exec(
        select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id, PLO.code == code)
    ).first()
    if duplicate:
        return redirect(f"/manager/plo-management?version_id={version.id}&error=duplicate&show_add=1")
    session.add(
        PLO(
            program_id=program.id,
            plo_version_id=version.id,
            code=code,
            description=description,
            domain=domain,
            bloom_level=bloom_level,
            status=status,
            remark=remark,
            created_by=user.id,
        )
    )
    session.commit()
    return redirect(f"/manager/plo-management?version_id={version.id}&created=1")


@app.post("/manager/plos/{plo_id}/update")
@app.post("/manager/plo-management/{plo_id}/update")
async def update_manager_plo(
    plo_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    plo = session.get(PLO, plo_id)
    if not plo or plo.program_id != program.id:
        raise HTTPException(status_code=404)
    version = editable_outcome_version(session, program, plo.plo_version_id)
    form = await request.form()
    code = str(form.get("code") or "").strip().upper()
    description = str(form.get("description") or "").strip()
    if not code or not description:
        return redirect(f"/manager/plo-management?version_id={version.id}&plo_id={plo_id}&error=required")
    duplicate = session.exec(
        select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == version.id, PLO.code == code, PLO.id != plo_id)
    ).first()
    if duplicate:
        return redirect(f"/manager/plo-management?version_id={version.id}&plo_id={plo_id}&error=duplicate")
    plo.code = code
    plo.description = description
    plo.domain = str(form.get("domain") or plo.domain or "Knowledge").strip()
    plo.bloom_level = str(form.get("bloom_level") or plo.bloom_level or "C1").strip()
    plo.status = str(form.get("status") or plo.status or "Active").strip()
    plo.remark = str(form.get("remark") or "").strip()
    plo.updated_at = datetime.utcnow()
    session.add(plo)
    session.commit()
    return redirect(f"/manager/plo-management?version_id={version.id}&updated=1")


@app.post("/manager/plos/{plo_id}/delete")
@app.post("/manager/plo-management/{plo_id}/delete")
def delete_manager_plo(
    plo_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    plo = session.get(PLO, plo_id)
    if not plo or plo.program_id != program.id:
        raise HTTPException(status_code=404)
    version = editable_outcome_version(session, program, plo.plo_version_id)
    plo.status = "Inactive"
    plo.updated_at = datetime.utcnow()
    session.add(plo)
    session.commit()
    return redirect(f"/manager/plo-management?version_id={version.id}&deleted=1")


@app.post("/manager/plos/{plo_id}/restore")
@app.post("/manager/plo-management/{plo_id}/restore")
def restore_manager_plo(
    plo_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    require_programme_scope(user)
    program = manager_program(session, user)
    plo = session.get(PLO, plo_id)
    if not plo or plo.program_id != program.id:
        raise HTTPException(status_code=404)
    version = editable_outcome_version(session, program, plo.plo_version_id)
    plo.status = "Active"
    plo.updated_at = datetime.utcnow()
    session.add(plo)
    session.commit()
    return redirect(f"/manager/plo-management?version_id={version.id}&restored=1")



def _cohort_year_semester(cohort_name: str) -> tuple[int | None, int | None]:
    """Read year/semester from class codes such as 21ME11Mb1."""
    import re
    match = re.match(r"^\d{2}[A-Za-z]+([1-4])([1-2])", (cohort_name or "").strip())
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _cohort_generation(cohort_name: str) -> str:
    value = (cohort_name or "").strip()
    return value[:2] if len(value) >= 2 and value[:2].isdigit() else "-"




@app.get("/teacher/{section}", response_class=HTMLResponse)
def teacher_section_page(
    section: str,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if user.role != Role.TEACHER:
        raise HTTPException(status_code=403)
    allowed_sections = {
        "outcome-versions",
        "course-mapping",
        "plo-target-view",
        "plo-target-setup",
        "courses",
        "assessments",
        "enter-scores",
        "students",
        "plo-reports",
        "course-reports",
        "announcements",
        "calendar",
        "documents",
    }
    if section not in allowed_sections:
        raise HTTPException(status_code=404)
    study_period = selected_study_period(request, user)
    if section == "documents":
        page = document_page_data(session, user, "/teacher/documents")
        return templates.TemplateResponse(
            "teacher_management.html",
            {"request": request, "user": user, "page": page, "teacher_section": section},
        )
    requested_course_id = optional_int(request.query_params.get("course_id"))
    if requested_course_id and not selected_teacher_course(session, user, requested_course_id, study_period):
        raise HTTPException(status_code=403)
    page = build_teacher_page(
        section,
        session,
        user,
        requested_course_id,
        optional_int(request.query_params.get("assessment_id")),
        optional_int(request.query_params.get("class_id")),
        study_period,
        optional_int(request.query_params.get("version_id")),
    )
    page["score_saved"] = request.query_params.get("saved") == "1"
    page["score_error"] = request.query_params.get("error") or ""
    return templates.TemplateResponse(
        "teacher_management.html",
        {"request": request, "user": user, "page": page, "teacher_section": section},
    )


@app.post("/teacher/course-mapping/{course_id}/update")
async def update_teacher_course_mapping(
    course_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    class_id: int | None = Form(None),
    version_id: int | None = Form(None),
):
    # CLO-PLO mapping is view only for teachers; the programme coordinator
    # owns these weights. The teacher page renders no form for this.
    raise HTTPException(status_code=403, detail="Course mapping is view only for teachers")
    study_period = selected_study_period(request, user)
    course = selected_teacher_course(session, user, course_id, study_period)
    if not course or course.id != course_id:
        raise HTTPException(status_code=403)
    program = course.program
    if not program:
        raise HTTPException(status_code=404)
    selected_version, _versions = selected_outcome_version(session, program, version_id)
    if selected_version.is_locked or selected_version.status in {"Published", "Retired"}:
        return redirect(f"/teacher/course-mapping?course_id={course_id}&class_id={class_id or ''}&version_id={selected_version.id}&error=readonly")
    form = await request.form()
    plos = session.exec(
        select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == selected_version.id)
    ).all()
    plos_by_id = {plo.id: plo for plo in plos if plo.id is not None}
    for clo in course.clos:
        if clo.id is None:
            continue
        existing = {
            mapping.plo_id: mapping
            for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all()
            if mapping.plo_id in plos_by_id
        }
        for plo_id in plos_by_id:
            raw = str(form.get(f"mapping_{clo.id}_{plo_id}") or "").strip()
            weight = 0.0
            if raw:
                try:
                    weight = min(max(float(raw), 0), 100)
                except ValueError:
                    weight = 0.0
            mapping = existing.get(plo_id)
            if weight <= 0:
                if mapping:
                    session.delete(mapping)
            elif mapping:
                mapping.weight = weight
            else:
                session.add(CLOPLOMapping(clo_id=clo.id, plo_id=plo_id, weight=weight))
    session.commit()
    return redirect(f"/teacher/course-mapping?course_id={course_id}&class_id={class_id or ''}&version_id={selected_version.id}&saved=1")


@app.post("/teacher/course-mapping/{course_id}/copy-previous")
async def copy_teacher_course_mapping(
    course_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    class_id: int | None = Form(None),
    version_id: int | None = Form(None),
    source_version_id: int | None = Form(None),
):
    # CLO-PLO mapping is view only for teachers; the programme coordinator
    # owns these weights. The teacher page renders no form for this.
    raise HTTPException(status_code=403, detail="Course mapping is view only for teachers")
    study_period = selected_study_period(request, user)
    course = selected_teacher_course(session, user, course_id, study_period)
    if not course or course.id != course_id:
        raise HTTPException(status_code=403)
    program = course.program
    if not program:
        raise HTTPException(status_code=404)
    target_version, versions = selected_outcome_version(session, program, version_id)
    if target_version.is_locked or target_version.status in {"Published", "Retired"}:
        return redirect(
            f"/teacher/course-mapping?course_id={course_id}&class_id={class_id or ''}"
            f"&version_id={target_version.id}&error=readonly"
        )
    source_version = session.get(PLOVersion, source_version_id) if source_version_id else None
    if not source_version or source_version.programme_id != program.id or source_version.id == target_version.id:
        source_version = next(
            (
                item
                for item in versions
                if item.id != target_version.id and item.status in {"Published", "Active", "Current"}
            ),
            None,
        ) or next((item for item in versions if item.id != target_version.id), None)
    if not source_version:
        return redirect(
            f"/teacher/course-mapping?course_id={course_id}&class_id={class_id or ''}"
            f"&version_id={target_version.id}&error=no-source-version"
        )

    source_plos = {
        plo.code: plo
        for plo in session.exec(
            select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == source_version.id)
        ).all()
        if plo.id is not None
    }
    target_plos = {
        plo.code: plo
        for plo in session.exec(
            select(PLO).where(PLO.program_id == program.id, PLO.plo_version_id == target_version.id)
        ).all()
        if plo.id is not None
    }
    source_id_to_code = {plo.id: code for code, plo in source_plos.items() if plo.id is not None}
    target_ids = {plo.id for plo in target_plos.values() if plo.id is not None}

    copied = 0
    for clo in course.clos:
        if clo.id is None:
            continue
        existing = list(session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all())
        source_rows = [mapping for mapping in existing if mapping.plo_id in source_id_to_code]
        for mapping in existing:
            if mapping.plo_id in target_ids:
                session.delete(mapping)
        for source_row in source_rows:
            target_plo = target_plos.get(source_id_to_code.get(source_row.plo_id))
            if target_plo and target_plo.id is not None:
                session.add(
                    CLOPLOMapping(
                        clo_id=clo.id,
                        plo_id=target_plo.id,
                        weight=stored_percent(source_row.weight),
                    )
                )
                copied += 1
    session.commit()
    return redirect(
        f"/teacher/course-mapping?course_id={course_id}&class_id={class_id or ''}"
        f"&version_id={target_version.id}&saved=1&copied={copied}"
    )


@app.post("/teacher/scores/{assessment_id}")
async def teacher_save_scores(
    assessment_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if user.role != Role.TEACHER:
        raise HTTPException(status_code=403)
    assessment = session.get(Assessment, assessment_id)
    if not assessment or not assessment.clo or not assessment.clo.course_id:
        raise HTTPException(status_code=404)
    study_period = selected_study_period(request, user)
    selected_course = selected_teacher_course(session, user, assessment.clo.course_id, study_period)
    if not selected_course or selected_course.id != assessment.clo.course_id:
        raise HTTPException(status_code=403)
    assigned_classes = teacher_course_classes(session, selected_course.id, user, study_period)
    valid_student_ids = {
        student["id"]
        for course_class in assigned_classes
        for student in teacher_course_student_rows(session, selected_course.id, course_class.id)
    }
    form = await request.form()
    max_score = float(assessment.max_score or assessment_weight_percent(assessment) or 100)
    for key, value in form.items():
        if not key.startswith("score_") or value in ("", None):
            continue
        student_id = optional_int(key.replace("score_", "", 1))
        if not student_id or student_id not in valid_student_ids:
            continue
        try:
            score_value = float(value)
        except (TypeError, ValueError):
            continue
        score_value = max(0, min(max_score, score_value))
        existing = session.exec(
            select(StudentScore).where(
                StudentScore.student_id == student_id,
                StudentScore.assessment_id == assessment_id,
            )
        ).first()
        now = datetime.utcnow()
        if existing:
            if existing.locked:
                continue
            existing.score = score_value
            existing.status = "Draft"
            existing.updated_at = now
            existing.entered_by_user_id = user.id
            session.add(existing)
        else:
            session.add(StudentScore(student_id=student_id, assessment_id=assessment_id, score=score_value, status="Draft", updated_at=now, entered_by_user_id=user.id))
    session.commit()
    return redirect(f"/teacher/enter-scores?course_id={selected_course.id}&assessment_id={assessment_id}&saved=1")

##############################












@app.post("/teacher/course-scores/{course_id}")
async def teacher_save_course_scores(
    course_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if user.role != Role.TEACHER:
        raise HTTPException(status_code=403)
    study_period = selected_study_period(request, user)
    selected_course = selected_teacher_course(session, user, course_id, study_period)
    if not selected_course or selected_course.id != course_id:
        raise HTTPException(status_code=403)
    form = await request.form()
    class_id = optional_int(form.get("class_id"))
    assigned_classes = teacher_course_classes(session, course_id, user, study_period)
    allowed_class_ids = {item.id for item in assigned_classes if item.id}
    if class_id not in allowed_class_ids:
        if len(allowed_class_ids) == 1:
            class_id = next(iter(allowed_class_ids))
        else:
            raise HTTPException(status_code=403, detail="Select one of your assigned classes in the current Study Period.")
    return_path = f"/teacher/enter-scores?course_id={course_id}"
    if class_id:
        return_path += f"&class_id={class_id}"
    students = teacher_course_student_rows(session, course_id, class_id)
    valid_student_ids = {student["id"] for student in students}
    assessments = teacher_assessment_rows(session, course_id, students)
    assessment_lookup = {
        assessment.id: assessment
        for assessment in teacher_course_assessments(session, course_id)
        if assessment.id
    }
    action = str(form.get("score_action") or "draft")
    final_submit = action == "submit"
    weight_total = assessment_weight_total(assessments)
    if action != "draft" and round(weight_total, 2) != 100:
        return redirect(f"{return_path}&error={quote('Assessment weight total must equal 100% before saving final course scores. You can still save a draft.')}")
    if course_scores_locked(session, course_id, valid_student_ids):
        return redirect(f"{return_path}&error={quote('Final scores are locked. Ask Admin or Programme Manager to unlock before editing.')}")

    status = "Submitted" if final_submit else "Saved" if action == "save" else "Draft"
    errors: list[str] = []
    saved_count = 0
    now = datetime.utcnow()
    for key, value in form.items():
        if not key.startswith("score_"):
            continue
        parts = key.split("_")
        if len(parts) != 3:
            continue
        student_id = optional_int(parts[1])
        assessment_id = optional_int(parts[2])
        if not student_id or not assessment_id or student_id not in valid_student_ids or assessment_id not in assessment_lookup:
            continue
        raw_value = str(value).strip()
        if raw_value == "":
            continue
        assessment = assessment_lookup[assessment_id]
        max_score = float(assessment.max_score or assessment_weight_percent(assessment) or 100)
        try:
            score_value = float(raw_value)
        except ValueError:
            errors.append(f"Invalid score for student {student_id}.")
            continue
        if score_value < 0 or score_value > max_score:
            errors.append(f"Score for {assessment.name} must be between 0 and {max_score:g}.")
            continue
        existing = session.exec(
            select(StudentScore).where(
                StudentScore.student_id == student_id,
                StudentScore.assessment_id == assessment_id,
            )
        ).first()
        if existing:
            if existing.locked:
                continue
            existing.score = score_value
            existing.status = status
            existing.updated_at = now
            existing.entered_by_user_id = user.id
            if final_submit:
                existing.locked = True
                existing.submitted_at = now
                existing.submitted_by_user_id = user.id
            session.add(existing)
        else:
            session.add(
                StudentScore(
                    student_id=student_id,
                    assessment_id=assessment_id,
                    score=score_value,
                    status=status,
                    locked=final_submit,
                    updated_at=now,
                    submitted_at=now if final_submit else None,
                    submitted_by_user_id=user.id if final_submit else None,
                    entered_by_user_id=user.id,
                )
            )
        saved_count += 1
    if errors:
        session.rollback()
        return redirect(f"{return_path}&error={quote(errors[0])}")
    session.commit()
    message = "final" if final_submit else "draft" if action == "draft" else "all"
    return redirect(f"{return_path}&saved=1&mode={message}")


@app.post("/teacher/course-scores/{course_id}/unlock")
async def unlock_course_scores(
    course_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if user.role not in {Role.SUPER_ADMIN, Role.PROGRAM_MANAGER, Role.DEAN}:
        raise HTTPException(status_code=403)
    form = await request.form()
    class_id = optional_int(form.get("class_id"))
    return_to = str(form.get("return_to") or "").strip()
    course = session.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404)
    if user.role == Role.PROGRAM_MANAGER and course.program_id != user.program_id:
        raise HTTPException(status_code=403)
    if user.role == Role.DEAN:
        program = session.get(Program, course.program_id) if course.program_id else None
        if not program or program.faculty_id != user.faculty_id:
            raise HTTPException(status_code=403)
    students = teacher_course_student_rows(session, course_id, class_id)
    student_ids = {student["id"] for student in students}
    assessment_ids = [assessment.id for assessment in teacher_course_assessments(session, course_id) if assessment.id]
    if not assessment_ids or not student_ids:
        fallback = return_to or f"/teacher/enter-scores?course_id={course_id}"
        sep = "&" if "?" in fallback else "?"
        return redirect(f"{fallback}{sep}error={quote('No locked scores found to unlock.')}")
    for score in session.exec(select(StudentScore).where(StudentScore.assessment_id.in_(assessment_ids), StudentScore.student_id.in_(student_ids))).all():
        score.locked = False
        score.status = "Unlocked"
        score.updated_at = datetime.utcnow()
        session.add(score)
    session.commit()
    suffix = f"&class_id={class_id}" if class_id else ""
    if return_to:
        sep = "&" if "?" in return_to else "?"
        return redirect(f"{return_to}{sep}unlocked=1")
    return redirect(f"/teacher/enter-scores?course_id={course_id}{suffix}&saved=1")


@app.get("/setup", response_class=HTMLResponse)
def setup_page(request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    selected_cohort_id = request.query_params.get("cohort_id")
    selected_cohort_id_int = int(selected_cohort_id) if selected_cohort_id and selected_cohort_id.isdigit() else None
    programs = list(session.exec(select(Program)))
    all_courses = sorted(
        session.exec(select(Course)).all(),
        key=lambda course: (course.curriculum_year or 99, course.curriculum_semester or "", course.code),
    )
    raw_classes = list(session.exec(select(CourseClass)))
    classes_by_generation: dict[str, CourseClass] = {}
    for course_class in raw_classes:
        key = " ".join(course_class.name.split()).lower()
        classes_by_generation.setdefault(key, course_class)
    classes = list(classes_by_generation.values())
    selected_cohort = session.get(CourseClass, selected_cohort_id_int) if selected_cohort_id_int else None
    selected_courses = [course for course in all_courses if course.cohort_id == (selected_cohort.id if selected_cohort else None)]
    peos = list(session.exec(select(PEO).order_by(PEO.code)))
    plos = list(session.exec(select(PLO).order_by(PLO.code)))
    teachers = list(session.exec(select(Teacher)))
    students = list(session.exec(select(Student).order_by(Student.student_no)))
    return templates.TemplateResponse(
        "setup.html",
        {
            "request": request,
            "user": user,
            "programs": programs,
            "courses": selected_courses,
            "all_courses": all_courses,
            "peos": peos,
            "plos": plos,
            "classes": classes,
            "selected_cohort": selected_cohort,
            "curriculum_years": build_curriculum_years(selected_courses, classes),
            "teachers": teachers,
            "students": students,
        },
    )


@app.post("/setup/peo")
def create_peo(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    program_id: int = Form(...),
    code: str = Form(...),
    description: str = Form(...),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    peo = session.exec(select(PEO).where(PEO.program_id == program_id, PEO.code == code)).first()
    if peo:
        peo.description = description
        session.add(peo)
    else:
        session.add(PEO(program_id=program_id, code=code, description=description))
    session.commit()
    return redirect("/setup")


@app.post("/setup/peo/{peo_id}")
def update_peo(
    peo_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    code: str = Form(...),
    description: str = Form(...),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    peo = session.get(PEO, peo_id)
    if not peo:
        raise HTTPException(status_code=404)
    peo.code = code
    peo.description = description
    session.add(peo)
    session.commit()
    return redirect("/setup")


@app.post("/setup/peo/{peo_id}/delete")
def delete_peo(peo_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    peo = session.get(PEO, peo_id)
    if peo:
        session.delete(peo)
        session.commit()
    return redirect("/setup")


@app.post("/setup/plo")
def create_plo(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    program_id: int = Form(...),
    code: str = Form(...),
    description: str = Form(...),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    plo = session.exec(select(PLO).where(PLO.program_id == program_id, PLO.code == code)).first()
    if plo:
        plo.description = description
        session.add(plo)
    else:
        session.add(PLO(program_id=program_id, code=code, description=description))
    session.commit()
    return redirect("/setup")


@app.post("/setup/plo/{plo_id}")
def update_plo(
    plo_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    code: str = Form(...),
    description: str = Form(...),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    plo = session.get(PLO, plo_id)
    if not plo:
        raise HTTPException(status_code=404)
    plo.code = code
    plo.description = description
    session.add(plo)
    session.commit()
    return redirect("/setup")


@app.post("/setup/plo/{plo_id}/delete")
def delete_plo(plo_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    plo = session.get(PLO, plo_id)
    if plo:
        for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.plo_id == plo_id)).all():
            session.delete(mapping)
        session.delete(plo)
        session.commit()
    return redirect("/setup")


@app.post("/setup/class")
def create_class(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
    academic_year: str = Form(...),
    semester: str = Form("Full Program"),
    semester_start: str | None = Form(None),
    semester_end: str | None = Form(None),
    course_id: int | None = Form(None),
    return_to: str | None = Form(None),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    if course_id is None:
        course = session.exec(select(Course).order_by(Course.curriculum_year, Course.curriculum_semester, Course.code)).first()
        if not course:
            raise HTTPException(status_code=400, detail="Create a subject before creating a class/cohort.")
        course_id = course.id
    name = " ".join(name.split())
    session.add(
        CourseClass(
            course_id=course_id,
            name=name,
            academic_year=academic_year,
            semester=semester,
            semester_start=semester_start,
            semester_end=semester_end,
        )
    )
    session.commit()
    if return_to:
        return redirect(return_to)
    return redirect("/setup#classes")


@app.post("/setup/course")
def create_course(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    program_id: int = Form(...),
    cohort_id: int | None = Form(None),
    curriculum_year: int = Form(...),
    curriculum_semester: str = Form(...),
    code: str = Form(...),
    title: str = Form(...),
    credits: float = Form(...),
    return_to: str | None = Form(None),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    course = session.exec(
        select(Course).where(Course.program_id == program_id, Course.cohort_id == cohort_id, Course.code == code)
    ).first()
    if course:
        course.title = title
        course.credits = credits
        course.curriculum_year = curriculum_year
        course.curriculum_semester = curriculum_semester
        session.add(course)
    else:
        session.add(
            Course(
                program_id=program_id,
                cohort_id=cohort_id,
                curriculum_year=curriculum_year,
                curriculum_semester=curriculum_semester,
                code=code,
                title=title,
                credits=credits,
            )
        )
    session.commit()
    if return_to:
        return redirect(return_to)
    return redirect(f"/setup?cohort_id={cohort_id}#subjects" if cohort_id else "/setup#subjects")


@app.post("/setup/course/{course_id}")
def update_course(
    course_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    cohort_id: int | None = Form(None),
    curriculum_year: int = Form(...),
    curriculum_semester: str = Form(...),
    code: str = Form(...),
    title: str = Form(...),
    credits: float = Form(...),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    course = session.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404)
    course.cohort_id = cohort_id
    course.curriculum_year = curriculum_year
    course.curriculum_semester = curriculum_semester
    course.code = code
    course.title = title
    course.credits = credits
    session.add(course)
    session.commit()
    return redirect(f"/setup?cohort_id={cohort_id}#subjects" if cohort_id else "/setup#subjects")


@app.post("/setup/course/{course_id}/delete")
def delete_course(course_id: int, request: Request, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    course = session.get(Course, course_id)
    if course:
        has_class = session.exec(select(CourseClass).where(CourseClass.course_id == course_id)).first()
        if not has_class:
            for clo in session.exec(select(CLO).where(CLO.course_id == course_id)).all():
                for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all():
                    session.delete(mapping)
                for assessment in session.exec(select(Assessment).where(Assessment.clo_id == clo.id)).all():
                    delete_assessment_with_scores(session, assessment)
                session.delete(clo)
            session.delete(course)
            session.commit()
    cohort_id = request.query_params.get("cohort_id")
    return redirect(f"/setup?cohort_id={cohort_id}#subjects" if cohort_id else "/setup#subjects")


@app.post("/setup/cohort/{cohort_id}/sync-courses")
def sync_cohort_courses(cohort_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    cohort = session.get(CourseClass, cohort_id)
    if not cohort:
        raise HTTPException(status_code=404)
    base_courses = session.exec(select(Course).where(Course.cohort_id == None)).all()
    cohort_courses = session.exec(select(Course).where(Course.cohort_id == cohort_id)).all()
    courses_by_code = {course.code: course for course in cohort_courses}
    for course in base_courses:
        copied_course = courses_by_code.get(course.code)
        if copied_course is None:
            copied_course = Course(
                program_id=course.program_id,
                cohort_id=cohort_id,
                code=course.code,
                title=course.title,
                credits=course.credits,
                curriculum_year=course.curriculum_year,
                curriculum_semester=course.curriculum_semester,
            )
            session.add(copied_course)
            session.commit()
            session.refresh(copied_course)
            courses_by_code[course.code] = copied_course

        existing_clo_codes = {
            clo.code
            for clo in session.exec(select(CLO).where(CLO.course_id == copied_course.id)).all()
        }
        for base_clo in session.exec(select(CLO).where(CLO.course_id == course.id)).all():
            if base_clo.code in existing_clo_codes:
                continue
            copied_clo = CLO(
                course_id=copied_course.id,
                code=base_clo.code,
                description=base_clo.description,
                domain=base_clo.domain,
                pass_threshold=base_clo.pass_threshold,
            )
            session.add(copied_clo)
            session.commit()
            session.refresh(copied_clo)
            for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == base_clo.id)).all():
                session.add(CLOPLOMapping(clo_id=copied_clo.id, plo_id=mapping.plo_id, weight=mapping.weight))
            for assessment in session.exec(select(Assessment).where(Assessment.clo_id == base_clo.id)).all():
                session.add(
                    Assessment(
                        clo_id=copied_clo.id,
                        name=assessment.name,
                        max_score=assessment.max_score,
                        weight=assessment.weight,
                    )
                )
    session.commit()
    return redirect(f"/setup?cohort_id={cohort_id}#subjects")


@app.post("/setup/class/{class_id}")
def update_class(
    class_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
    academic_year: str = Form(...),
    semester: str = Form("Full Program"),
    semester_start: str | None = Form(None),
    semester_end: str | None = Form(None),
    course_id: int | None = Form(None),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    course_class = session.get(CourseClass, class_id)
    if not course_class:
        raise HTTPException(status_code=404)
    if course_id is not None:
        course_class.course_id = course_id
    course_class.name = " ".join(name.split())
    course_class.academic_year = academic_year
    course_class.semester = semester
    course_class.semester_start = semester_start
    course_class.semester_end = semester_end
    session.add(course_class)
    session.commit()
    return redirect("/setup#classes")


@app.post("/setup/class/{class_id}/delete")
def delete_class(class_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    course_class = session.get(CourseClass, class_id)
    if course_class:
        for assignment in session.exec(select(ClassTeacher).where(ClassTeacher.class_id == class_id)).all():
            session.delete(assignment)
        for enrollment in session.exec(select(ClassStudent).where(ClassStudent.class_id == class_id)).all():
            session.delete(enrollment)
        session.delete(course_class)
        session.commit()
    return redirect("/setup#classes")


@app.post("/setup/class/{class_id}/copy")
def copy_class(
    class_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    name: str = Form(...),
    academic_year: str = Form(...),
    semester: str = Form("Full Program"),
    semester_start: str | None = Form(None),
    semester_end: str | None = Form(None),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    source = session.get(CourseClass, class_id)
    if not source:
        raise HTTPException(status_code=404)
    copied = CourseClass(
        course_id=source.course_id,
        name=name,
        academic_year=academic_year,
        semester=semester,
        semester_start=semester_start,
        semester_end=semester_end,
    )
    session.add(copied)
    session.commit()
    session.refresh(copied)
    for assignment in source.teachers:
        session.add(ClassTeacher(class_id=copied.id, teacher_id=assignment.teacher_id))
    for enrollment in source.students:
        session.add(ClassStudent(class_id=copied.id, student_id=enrollment.student_id))
    session.commit()
    return redirect("/setup#classes")


@app.post("/setup/student")
def create_student(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    student_no: str = Form(...),
    name_en: str = Form(...),
    name_kh: str = Form(""),
    class_id: int = Form(...),
    return_to: str | None = Form(None),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    student = session.exec(select(Student).where(Student.student_no == student_no)).first()
    if not student:
        student = Student(student_no=student_no, name_en=name_en, name_kh=name_kh or None)
        session.add(student)
        session.commit()
        session.refresh(student)
    existing = session.exec(select(ClassStudent).where(ClassStudent.class_id == class_id, ClassStudent.student_id == student.id)).first()
    if not existing:
        session.add(ClassStudent(class_id=class_id, student_id=student.id))
        session.commit()
    if return_to:
        return redirect(return_to)
    return redirect("/setup#students")


@app.post("/setup/students/import")
async def import_students(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    class_id: int = Form(...),
    file: UploadFile = File(...),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    workbook = load_workbook(BytesIO(await file.read()), data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        student_no = str(row[0]).strip()
        name_en = str(row[1] or "").strip() or student_no
        name_kh = str(row[2] or "").strip() if len(row) > 2 else ""
        student = session.exec(select(Student).where(Student.student_no == student_no)).first()
        if not student:
            student = Student(student_no=student_no, name_en=name_en, name_kh=name_kh or None)
            session.add(student)
            session.commit()
            session.refresh(student)
        existing = session.exec(select(ClassStudent).where(ClassStudent.class_id == class_id, ClassStudent.student_id == student.id)).first()
        if not existing:
            session.add(ClassStudent(class_id=class_id, student_id=student.id))
    session.commit()
    return redirect("/setup#students")


@app.post("/setup/teacher-assignment")
def assign_teacher(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    class_id: int = Form(...),
    teacher_id: int = Form(...),
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    existing = session.exec(select(ClassTeacher).where(ClassTeacher.class_id == class_id, ClassTeacher.teacher_id == teacher_id)).first()
    if not existing:
        session.add(ClassTeacher(class_id=class_id, teacher_id=teacher_id))
        session.commit()
    return redirect("/setup")


@app.post("/setup/class-teacher/{assignment_id}/delete")
def delete_teacher_assignment(
    assignment_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_program"):
        raise HTTPException(status_code=403)
    assignment = session.get(ClassTeacher, assignment_id)
    if assignment:
        session.delete(assignment)
        session.commit()
    return redirect("/setup#classes")


@app.post("/setup/clo")
def create_clo(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    course_id: int = Form(...),
    code: str = Form(...),
    domain: str = Form(...),
    description: str = Form(...),
    pass_threshold: float = Form(0.5),
):
    if not can(user, "manage_clo_assessment"):
        raise HTTPException(status_code=403)
    session.add(CLO(course_id=course_id, code=code, domain=domain, description=description, pass_threshold=pass_threshold))
    session.commit()
    return redirect("/setup")


@app.post("/setup/assessment")
def create_assessment(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    clo_id: int = Form(...),
    name: str = Form(...),
    max_score: float = Form(...),
    description: str = Form(""),
):
    if not can(user, "manage_clo_assessment"):
        raise HTTPException(status_code=403)
    session.add(Assessment(clo_id=clo_id, name=name, max_score=max_score, description=description))
    session.commit()
    return redirect("/setup")


@app.post("/setup/mapping")
def create_mapping(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    clo_id: int = Form(...),
    plo_id: int = Form(...),
    weight: float = Form(...),
):
    if not can(user, "manage_clo_assessment"):
        raise HTTPException(status_code=403)
    mapping = session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo_id, CLOPLOMapping.plo_id == plo_id)).first()
    if mapping:
        mapping.weight = weight
        session.add(mapping)
    else:
        session.add(CLOPLOMapping(clo_id=clo_id, plo_id=plo_id, weight=weight))
    session.commit()
    return redirect("/setup")


def delete_assessment_with_scores(session: Session, assessment: Assessment) -> None:
    for score in session.exec(select(StudentScore).where(StudentScore.assessment_id == assessment.id)).all():
        session.delete(score)
    session.delete(assessment)


@app.post("/setup/clo/{clo_id}/delete")
def delete_clo(clo_id: int, session: Annotated[Session, Depends(get_session)], user: Annotated[User, Depends(require_user)]):
    if not can(user, "manage_clo_assessment"):
        raise HTTPException(status_code=403)
    clo = session.get(CLO, clo_id)
    if clo:
        for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo_id)).all():
            session.delete(mapping)
        for assessment in session.exec(select(Assessment).where(Assessment.clo_id == clo_id)).all():
            delete_assessment_with_scores(session, assessment)
        session.delete(clo)
        session.commit()
    return redirect("/setup")


@app.post("/setup/assessment-detail")
async def update_assessment_detail(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_clo_assessment"):
        raise HTTPException(status_code=403)
    form = await request.form()
    clos = session.exec(select(CLO)).all()
    plos = session.exec(select(PLO)).all()

    for clo in clos:
        if form.get(f"delete_clo_{clo.id}"):
            for mapping in session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id)).all():
                session.delete(mapping)
            for assessment in session.exec(select(Assessment).where(Assessment.clo_id == clo.id)).all():
                delete_assessment_with_scores(session, assessment)
            session.delete(clo)
            continue

        code = str(form.get(f"clo_code_{clo.id}", clo.code)).strip()
        domain = str(form.get(f"clo_domain_{clo.id}", clo.domain)).strip()
        description = str(form.get(f"clo_description_{clo.id}", clo.description)).strip()
        threshold_raw = str(form.get(f"clo_threshold_{clo.id}", clo.pass_threshold)).strip()
        clo.code = code or clo.code
        clo.domain = domain
        clo.description = description
        clo.pass_threshold = float(threshold_raw or 0.5)
        session.add(clo)

        for plo in plos:
            key = f"mapping_{clo.id}_{plo.id}"
            if key not in form:
                continue
            raw_value = str(form.get(key, "")).strip()
            mapping = session.exec(select(CLOPLOMapping).where(CLOPLOMapping.clo_id == clo.id, CLOPLOMapping.plo_id == plo.id)).first()
            if raw_value == "":
                if mapping:
                    session.delete(mapping)
                continue
            weight = float(raw_value)
            if mapping:
                mapping.weight = weight
                session.add(mapping)
            else:
                session.add(CLOPLOMapping(clo_id=clo.id, plo_id=plo.id, weight=weight))

        new_assessment_name = str(form.get(f"new_assessment_name_{clo.id}", "")).strip()
        new_assessment_score = str(form.get(f"new_assessment_score_{clo.id}", "")).strip()
        if new_assessment_name and new_assessment_score:
            session.add(Assessment(clo_id=clo.id, name=new_assessment_name, max_score=float(new_assessment_score)))

    assessments = session.exec(select(Assessment)).all()
    for assessment in assessments:
        if form.get(f"delete_assessment_{assessment.id}"):
            delete_assessment_with_scores(session, assessment)
            continue
        name = str(form.get(f"assessment_name_{assessment.id}", assessment.name)).strip()
        score_raw = str(form.get(f"assessment_score_{assessment.id}", assessment.max_score)).strip()
        description = str(form.get(f"assessment_description_{assessment.id}", assessment.description or "")).strip()
        if not name:
            delete_assessment_with_scores(session, assessment)
            continue
        assessment.name = name
        assessment.max_score = float(score_raw or 0)
        assessment.description = description or None
        session.add(assessment)

    session.commit()
    return redirect("/setup")


@app.get("/classes/{class_id}/marks", response_class=HTMLResponse)
def marks_page(
    class_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "input_marks"):
        raise HTTPException(status_code=403)
    course_class = session.get(CourseClass, class_id)
    clos = list(session.exec(select(CLO).where(CLO.course_id == course_class.course_id)))
    plos = sorted(course_class.course.program.plos, key=plo_sort_key)
    assessments = list(session.exec(select(Assessment).join(CLO).where(CLO.course_id == course_class.course_id)))
    scores = list(session.exec(select(StudentScore)))
    score_map = {(score.student_id, score.assessment_id): score for score in scores}
    return templates.TemplateResponse(
        "marks.html",
        {
            "request": request,
            "user": user,
            "class": course_class,
            "clos": clos,
            "plos": plos,
            "assessments": assessments,
            "score_map": score_map,
        },
    )


@app.post("/classes/{class_id}/marks")
async def save_marks(
    class_id: int,
    session: Annotated[Session, Depends(get_session)],
    request: Request,
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "input_marks"):
        raise HTTPException(status_code=403)

    form = await request.form()
    for key, value in form.items():
        if not key.startswith("score_") or value == "":
            continue
        _, student_id, assessment_id = key.split("_")
        existing = session.exec(
            select(StudentScore).where(
                StudentScore.student_id == int(student_id),
                StudentScore.assessment_id == int(assessment_id),
            )
        ).first()
        if existing:
            existing.score = float(value)
            session.add(existing)
        else:
            session.add(StudentScore(student_id=int(student_id), assessment_id=int(assessment_id), score=float(value)))
    session.commit()
    return redirect(f"/classes/{class_id}/marks")


@app.get("/classes/{class_id}/reports", response_class=HTMLResponse)
def reports_page(
    class_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "view_clo_report") and not can(user, "view_all_reports"):
        raise HTTPException(status_code=403)
    report = get_course_report(session, class_id)
    plo_summary = get_plo_summary(session, class_id)
    return templates.TemplateResponse(
        "reports.html",
        {"request": request, "user": user, "report": report, "plo_summary": plo_summary},
    )


@app.get("/programs/{program_id}/reports", response_class=HTMLResponse)
def program_reports_page(
    program_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "manage_program") and not can(user, "view_all_reports"):
        raise HTTPException(status_code=403)
    report = get_program_report(session, program_id)
    return templates.TemplateResponse("program_reports.html", {"request": request, "user": user, "report": report})


@app.get("/classes/{class_id}/clo/{clo_id}", response_class=HTMLResponse)
def clo_report_page(
    class_id: int,
    clo_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "view_clo_report"):
        raise HTTPException(status_code=403)
    report = get_clo_report(session, class_id, clo_id)
    return templates.TemplateResponse("clo_report.html", {"request": request, "user": user, "report": report})


@app.get("/student/report", response_class=HTMLResponse)
def student_report(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if user.role != Role.STUDENT:
        raise HTTPException(status_code=403)
    return redirect("/student/course-reports")


@app.get("/student/{section}", response_class=HTMLResponse)
def student_section_page(
    section: str,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if user.role != Role.STUDENT:
        raise HTTPException(status_code=403)
    allowed_sections = {
    "courses",
    "assessments",
    "clo-attainment",
    "plo-overview",
    "my-scores",          # <-- add this
    "course-reports",
    "announcements",
    "calendar",
    "documents",
    }
    if section == "dashboard":
        return redirect("/dashboard")
    if section not in allowed_sections:
        raise HTTPException(status_code=404)
    return templates.TemplateResponse(
        "student_management.html",
        {"request": request, "user": user, "page": build_student_page(section, session, user), "student_section": section},
    )


@app.get("/classes/{class_id}/assessment/{assessment_id}/template")
def download_mark_template(
    class_id: int,
    assessment_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "import_marks"):
        raise HTTPException(status_code=403)
    content = build_mark_template(session, class_id, assessment_id)
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mark-template.xlsx"},
    )


@app.post("/classes/{class_id}/assessment/{assessment_id}/import")
async def import_marks(
    class_id: int,
    assessment_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
    file: UploadFile = File(...),
):
    if not can(user, "import_marks"):
        raise HTTPException(status_code=403)
    content = await file.read()
    import_marks_from_excel(session, class_id, assessment_id, content)
    return redirect(f"/classes/{class_id}/marks")


@app.get("/classes/{class_id}/export")
def export_course_report(
    class_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if not can(user, "view_clo_report") and not can(user, "view_all_reports"):
        raise HTTPException(status_code=403)
    content = build_course_report_excel(session, class_id)
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=course-report.xlsx"},
    )


# --- Account profile (all roles) --------------------------------------------------

ROLE_TITLES_UI = {
    Role.SUPER_ADMIN: "System Administrator",
    Role.DEAN: "Dean",
    Role.PROGRAM_MANAGER: "Programme Coordinator",
    Role.TEACHER: "Lecturer",
    Role.STUDENT: "Student",
}


def role_title(user: User | None) -> str:
    """Display title for the topbar and profile page."""
    if not user:
        return ""
    return ROLE_TITLES_UI.get(user.role, ROLE_LABELS.get(user.role, "User"))


templates.env.globals["role_title"] = role_title


def profile_page_data(session: Session, user: User) -> dict:
    """Real account details for the signed-in user, whatever their role."""
    faculty = session.get(Faculty, user.faculty_id) if user.faculty_id else None
    program = session.get(Program, user.program_id) if user.program_id else None
    teacher = session.exec(select(Teacher).where(Teacher.user_id == user.id)).first()
    student = session.exec(select(Student).where(Student.user_id == user.id)).first()

    details = [("Role", role_title(user)), ("Access scope", scope_label(user))]
    if faculty:
        details.append(("Faculty", faculty.name))
    if program:
        details.append(("Programme", f"{program.code} · {program.name}"))
    if teacher:
        details.append(("Staff number", teacher.staff_no or "-"))
    if student:
        details.append(("Student number", student.student_no or "-"))
        if student.name_kh:
            details.append(("Name (Khmer)", student.name_kh))
    details.append(("Account status", "Active" if user.is_active else "Inactive"))

    stats: list[dict[str, str]] = [
        {"label": "Role", "value": role_title(user), "icon": "bi-shield-check", "tone": "blue"},
        {"label": "Status", "value": "Active" if user.is_active else "Inactive", "icon": "bi-check-circle", "tone": "green" if user.is_active else "orange"},
    ]
    if faculty:
        stats.append({"label": "Faculty", "value": faculty.name, "icon": "bi-building", "tone": "cyan"})
    if program:
        stats.append({"label": "Programme", "value": program.code, "icon": "bi-mortarboard", "tone": "purple"})
    if teacher and teacher.id:
        assigned_courses = session.exec(
            select(Course)
            .join(CourseTeacher, CourseTeacher.course_id == Course.id)
            .where(CourseTeacher.teacher_id == teacher.id)
            .order_by(Course.code)
        ).all()
        details.append(("Assigned courses", str(len(assigned_courses))))
        if assigned_courses:
            details.append(("Course scope", ", ".join(course.code for course in assigned_courses[:6]) + (" ..." if len(assigned_courses) > 6 else "")))
        stats.append({"label": "Assigned Courses", "value": str(len(assigned_courses)), "icon": "bi-journal-bookmark", "tone": "orange"})
    if student and student.id:
        semester_enrollments = session.exec(
            select(StudentSemesterEnrollment)
            .where(StudentSemesterEnrollment.student_id == student.id)
            .order_by(StudentSemesterEnrollment.academic_year, StudentSemesterEnrollment.semester)
        ).all()
        details.append(("Semester enrollments", str(len(semester_enrollments))))
        if semester_enrollments:
            latest_enrollment = semester_enrollments[-1]
            details.append(("Current cohort", latest_enrollment.cohort_name))
            details.append(("Current semester", f"{latest_enrollment.academic_year} · Semester {latest_enrollment.semester}"))
        stats.append({"label": "Enrollments", "value": str(len(semester_enrollments)), "icon": "bi-stack", "tone": "orange"})

    return {
        "kind": "profile",
        "title": "My Profile",
        "description": "Manage your account information and password.",
        "account": user,
        "role_title": role_title(user),
        "stats": stats,
        "details": details,
        "faculty": faculty,
        "program": program,
    }


@app.get("/profile", response_class=HTMLResponse)
def account_profile_page(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    page = profile_page_data(session, user)
    return templates.TemplateResponse(
        "profile.html", {"request": request, "user": user, "page": page}
    )


@app.post("/profile/update")
async def update_account_profile(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Update the signed-in user's own name and email."""
    form = await request.form()
    name = str(form.get("name") or "").strip()
    email = str(form.get("email") or "").strip().lower()
    if not name or not email:
        return redirect("/profile?profile_error=required")
    if "@" not in email or "." not in email.split("@")[-1]:
        return redirect("/profile?profile_error=invalid_email")

    account = session.get(User, user.id)
    if not account:
        raise HTTPException(status_code=404)
    taken = session.exec(select(User).where(User.email == email)).first()
    if taken and taken.id != account.id:
        return redirect("/profile?profile_error=email_taken")

    account.name = name
    account.email = email
    session.add(account)
    session.commit()
    return redirect("/profile?profile_updated=1")


@app.post("/profile/password")
async def change_account_password(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Change the signed-in user's own password."""
    form = await request.form()
    current = str(form.get("current_password") or "")
    new_password = str(form.get("new_password") or "")
    confirm = str(form.get("confirm_password") or "")

    account = session.get(User, user.id)
    if not account:
        raise HTTPException(status_code=404)
    if not current or not new_password or not confirm:
        return redirect("/profile?password_error=required#password")
    if not verify_password(current, account.password_hash):
        return redirect("/profile?password_error=wrong_current#password")
    if len(new_password) < 8:
        return redirect("/profile?password_error=too_short#password")
    if new_password != confirm:
        return redirect("/profile?password_error=mismatch#password")
    if verify_password(new_password, account.password_hash):
        return redirect("/profile?password_error=same#password")

    account.password_hash = hash_password(new_password)
    session.add(account)
    session.commit()
    return redirect("/profile?password_updated=1#password")


# --- Documents (all roles) ---------------------------------------------------------

DOCUMENT_ROOT = Path(app_config.upload_dir) / "documents"
DOCUMENT_CATEGORIES = [
    "General",
    "Curriculum",
    "Assessment",
    "Accreditation",
    "Policy",
    "Report",
    "Template",
]
DOCUMENT_STATUSES = ["Published", "Draft", "Archived"]
MAX_DOCUMENT_BYTES = 25 * 1024 * 1024


def document_scope(session: Session, user: User) -> tuple[list[int], int | None]:
    """Programme ids and faculty id a user may see documents for."""
    if user.role == Role.SUPER_ADMIN:
        return [], None
    if user.role == Role.DEAN:
        return [item.id for item in faculty_programmes(session, user) if item.id], user.faculty_id
    program = session.get(Program, user.program_id) if user.program_id else None
    if program is None and user.role in PROGRAMME_SCOPE_ROLES:
        program = manager_program(session, user)
    return ([program.id] if program else []), (program.faculty_id if program else user.faculty_id)


def visible_documents(session: Session, user: User) -> list[Document]:
    """Documents this user may read, newest first.

    A document with no faculty and no programme is system-wide and visible to
    everyone; otherwise it must match the user's faculty or programme.
    """
    rows = list(session.exec(select(Document).order_by(Document.id.desc())).all())
    if user.role == Role.SUPER_ADMIN:
        return rows
    programme_ids, faculty_id = document_scope(session, user)
    visible = []
    for row in rows:
        if row.program_id is None and row.faculty_id is None:
            visible.append(row)
        elif row.program_id is not None and row.program_id in programme_ids:
            visible.append(row)
        elif row.program_id is None and faculty_id is not None and row.faculty_id == faculty_id:
            visible.append(row)
    return visible


def can_edit_document(session: Session, user: User, document: Document) -> bool:
    """Admins manage everything; others manage what falls inside their scope."""
    if user.role == Role.SUPER_ADMIN:
        return True
    if document.uploaded_by == user.id:
        return True
    programme_ids, faculty_id = document_scope(session, user)
    if user.role == Role.DEAN:
        return document.program_id in programme_ids or (
            faculty_id is not None and document.faculty_id == faculty_id
        )
    if user.role == Role.PROGRAM_MANAGER:
        return document.program_id in programme_ids
    return False


def document_row(session: Session, user: User, document: Document) -> dict:
    program = session.get(Program, document.program_id) if document.program_id else None
    faculty = session.get(Faculty, document.faculty_id) if document.faculty_id else None
    if program:
        scope = f"{program.code} · {program.name}"
    elif faculty:
        scope = faculty.name
    else:
        scope = "All faculties"
    size = document.size_bytes or 0
    if size >= 1024 * 1024:
        size_label = f"{size / 1024 / 1024:.1f} MB"
    elif size >= 1024:
        size_label = f"{size / 1024:.0f} KB"
    else:
        size_label = f"{size} B"
    return {
        "document": document,
        "id": document.id,
        "title": document.title,
        "description": document.description,
        "category": document.category,
        "status": document.status,
        "scope": scope,
        "file_name": document.original_name,
        "size_label": size_label,
        "has_file": bool(document.stored_name),
        "uploaded_by": document.uploaded_by_name or "-",
        "created_at": format_datetimeish(document.created_at),
        "can_manage": can_edit_document(session, user, document),
    }


def document_page_data(session: Session, user: User, base: str) -> dict:
    """Everything the documents page needs, scoped to the signed-in user."""
    documents = visible_documents(session, user)
    rows = [document_row(session, user, item) for item in documents]

    if user.role == Role.SUPER_ADMIN:
        programmes = list(session.exec(select(Program).order_by(Program.code)).all())
        faculties = list(session.exec(select(Faculty).order_by(Faculty.name)).all())
    elif user.role == Role.DEAN:
        programmes = faculty_programmes(session, user)
        faculties = [session.get(Faculty, user.faculty_id)] if user.faculty_id else []
    else:
        programme_ids, faculty_id = document_scope(session, user)
        programmes = [session.get(Program, pid) for pid in programme_ids]
        faculties = [session.get(Faculty, faculty_id)] if faculty_id else []
    programmes = [item for item in programmes if item]
    faculties = [item for item in faculties if item]

    by_status = defaultdict(int)
    for row in rows:
        by_status[row["status"]] += 1
    total_bytes = sum(item.size_bytes or 0 for item in documents)

    return {
        "kind": "documents",
        "title": "Documents",
        "description": "Upload, organise and download documents for your scope.",
        "rows": rows,
        "categories": DOCUMENT_CATEGORIES,
        "statuses": DOCUMENT_STATUSES,
        "programmes": programmes,
        "faculties": faculties,
        "base": base,
        "can_upload": user.role != Role.STUDENT,
        "stats": [
            ("Total Documents", len(rows), "bi-folder", "blue"),
            ("Published", by_status.get("Published", 0), "bi-file-earmark-check", "green"),
            ("Draft", by_status.get("Draft", 0), "bi-pencil", "orange"),
            ("Archived", by_status.get("Archived", 0), "bi-archive", "purple"),
            ("Total Size", f"{total_bytes / 1024 / 1024:.1f} MB" if total_bytes else "0 MB", "bi-hdd", "cyan"),
        ],
    }


def document_return_path(value: object, user: User) -> str:
    """Send the user back to the documents page they came from."""
    path = safe_return_path(value)
    allowed = {"/admin/documents", "/dean/documents", "/manager/documents", "/teacher/documents"}
    if path in allowed:
        return path
    default = {
        Role.SUPER_ADMIN: "/admin/documents",
        Role.DEAN: "/dean/documents",
        Role.PROGRAM_MANAGER: "/manager/documents",
        Role.TEACHER: "/teacher/documents",
    }
    return default.get(user.role, "/dashboard")


def resolve_document_scope(session: Session, user: User, form) -> tuple[int | None, int | None]:
    """Work out which faculty/programme an upload belongs to, within the user's rights."""
    program_id = optional_int(form.get("program_id"))
    program = session.get(Program, program_id) if program_id else None
    if user.role == Role.SUPER_ADMIN:
        faculty_id = optional_int(form.get("faculty_id"))
        if program:
            return program.faculty_id, program.id
        return faculty_id, None
    programme_ids, faculty_id = document_scope(session, user)
    if program and program.id in programme_ids:
        return program.faculty_id, program.id
    if user.role == Role.DEAN:
        return faculty_id, None
    return faculty_id, (programme_ids[0] if programme_ids else None)


@app.post("/documents/upload")
async def upload_document(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    if user.role == Role.STUDENT:
        raise HTTPException(status_code=403)
    form = await request.form()
    return_to = document_return_path(form.get("return_to"), user)
    title = str(form.get("title") or "").strip()
    if not title:
        return redirect(f"{return_to}?doc_error=required")

    upload = form.get("file")
    payload = b""
    original_name = ""
    content_type = ""
    if upload is not None and getattr(upload, "filename", ""):
        original_name = Path(str(upload.filename)).name
        content_type = str(getattr(upload, "content_type", "") or "")
        payload = await upload.read()
        if len(payload) > MAX_DOCUMENT_BYTES:
            return redirect(f"{return_to}?doc_error=too_large")

    category = str(form.get("category") or "General").strip()
    if category not in DOCUMENT_CATEGORIES:
        category = "General"
    status = str(form.get("status") or "Published").strip()
    if status not in DOCUMENT_STATUSES:
        status = "Published"
    faculty_id, program_id = resolve_document_scope(session, user, form)

    document = Document(
        title=title,
        description=str(form.get("description") or "").strip(),
        category=category,
        status=status,
        original_name=original_name,
        content_type=content_type,
        size_bytes=len(payload),
        faculty_id=faculty_id,
        program_id=program_id,
        uploaded_by=user.id,
        uploaded_by_name=user.name,
    )
    session.add(document)
    session.commit()
    session.refresh(document)

    if payload:
        DOCUMENT_ROOT.mkdir(parents=True, exist_ok=True)
        suffix = Path(original_name).suffix[:16]
        document.stored_name = f"{document.id}{suffix}"
        (DOCUMENT_ROOT / document.stored_name).write_bytes(payload)
        session.add(document)
        session.commit()
    return redirect(f"{return_to}?doc_created=1")


@app.post("/documents/{document_id}/update")
async def update_document(
    document_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    document = session.get(Document, document_id)
    form = await request.form()
    return_to = document_return_path(form.get("return_to"), user)
    if not document:
        raise HTTPException(status_code=404)
    if not can_edit_document(session, user, document):
        raise HTTPException(status_code=403)

    title = str(form.get("title") or "").strip()
    if not title:
        return redirect(f"{return_to}?doc_error=required")
    category = str(form.get("category") or document.category).strip()
    status = str(form.get("status") or document.status).strip()
    document.title = title
    document.description = str(form.get("description") or "").strip()
    document.category = category if category in DOCUMENT_CATEGORIES else document.category
    document.status = status if status in DOCUMENT_STATUSES else document.status
    document.updated_at = datetime.utcnow()
    session.add(document)
    session.commit()
    return redirect(f"{return_to}?doc_updated=1")


@app.post("/documents/{document_id}/delete")
async def delete_document(
    document_id: int,
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    document = session.get(Document, document_id)
    form = await request.form()
    return_to = document_return_path(form.get("return_to"), user)
    if not document:
        raise HTTPException(status_code=404)
    if not can_edit_document(session, user, document):
        raise HTTPException(status_code=403)
    stored = document.stored_name
    session.delete(document)
    session.commit()
    if stored:
        target = DOCUMENT_ROOT / stored
        if target.exists():
            target.unlink()
    return redirect(f"{return_to}?doc_deleted=1")


@app.get("/documents/{document_id}/download")
def download_document(
    document_id: int,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    document = session.get(Document, document_id)
    if not document:
        raise HTTPException(status_code=404)
    if document.id not in {item.id for item in visible_documents(session, user)}:
        raise HTTPException(status_code=403)
    if not document.stored_name:
        raise HTTPException(status_code=404, detail="This record has no attached file")
    target = DOCUMENT_ROOT / document.stored_name
    if not target.exists():
        raise HTTPException(status_code=404, detail="The stored file is missing")
    return FileResponse(
        target,
        media_type=document.content_type or "application/octet-stream",
        filename=document.original_name or f"document-{document.id}",
    )


# --- Profile picture ---------------------------------------------------------------

AVATAR_ROOT = Path(app_config.upload_dir) / "avatars"
MAX_AVATAR_BYTES = 3 * 1024 * 1024
# Only raster formats. SVG is deliberately excluded because it can carry script.
AVATAR_SIGNATURES = [
    (b"\x89PNG\r\n\x1a\n", ".png", "image/png"),
    (b"\xff\xd8\xff", ".jpg", "image/jpeg"),
    (b"GIF87a", ".gif", "image/gif"),
    (b"GIF89a", ".gif", "image/gif"),
]


def sniff_avatar(payload: bytes) -> tuple[str, str] | None:
    """Identify an image by its magic bytes, ignoring the declared content type."""
    for signature, suffix, media_type in AVATAR_SIGNATURES:
        if payload.startswith(signature):
            return suffix, media_type
    if payload[:4] == b"RIFF" and payload[8:12] == b"WEBP":
        return ".webp", "image/webp"
    return None


def avatar_media_type(stored_name: str) -> str:
    suffix = Path(stored_name).suffix.lower()
    return {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".webp": "image/webp",
    }.get(suffix, "application/octet-stream")


def avatar_url(user: User | None) -> str:
    """URL for a user's picture, or empty when they have none."""
    if not user or not getattr(user, "avatar_name", "") or not user.id:
        return ""
    return f"/avatars/{user.id}?v={Path(user.avatar_name).stem}"


templates.env.globals["avatar_url"] = avatar_url


@app.get("/avatars/{user_id}")
def serve_avatar(
    user_id: int,
    session: Annotated[Session, Depends(get_session)],
    _user: Annotated[User, Depends(require_user)],
):
    """Serve a profile picture to any signed-in user."""
    account = session.get(User, user_id)
    if not account or not account.avatar_name:
        raise HTTPException(status_code=404)
    target = AVATAR_ROOT / account.avatar_name
    if not target.exists():
        raise HTTPException(status_code=404)
    return FileResponse(
        target,
        media_type=avatar_media_type(account.avatar_name),
        headers={"X-Content-Type-Options": "nosniff", "Cache-Control": "private, max-age=300"},
    )


@app.post("/profile/avatar")
async def update_profile_avatar(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Upload or replace the signed-in user's own profile picture."""
    form = await request.form()
    upload = form.get("avatar")
    if upload is None or not getattr(upload, "filename", ""):
        return redirect("/profile?avatar_error=required")
    payload = await upload.read()
    if not payload:
        return redirect("/profile?avatar_error=required")
    if len(payload) > MAX_AVATAR_BYTES:
        return redirect("/profile?avatar_error=too_large")
    sniffed = sniff_avatar(payload)
    if not sniffed:
        return redirect("/profile?avatar_error=unsupported")
    suffix, _media_type = sniffed

    account = session.get(User, user.id)
    if not account:
        raise HTTPException(status_code=404)
    previous = account.avatar_name

    AVATAR_ROOT.mkdir(parents=True, exist_ok=True)
    # The filename changes on every upload so browsers refetch the new picture.
    stored_name = f"{account.id}-{uuid4().hex[:8]}{suffix}"
    (AVATAR_ROOT / stored_name).write_bytes(payload)

    account.avatar_name = stored_name
    session.add(account)
    session.commit()

    if previous and previous != stored_name:
        old = AVATAR_ROOT / previous
        if old.exists():
            old.unlink()
    return redirect("/profile?avatar_updated=1")


@app.post("/profile/avatar/remove")
def remove_profile_avatar(
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Drop the signed-in user's picture and fall back to the initials badge."""
    account = session.get(User, user.id)
    if not account:
        raise HTTPException(status_code=404)
    stored = account.avatar_name
    account.avatar_name = ""
    session.add(account)
    session.commit()
    if stored:
        target = AVATAR_ROOT / stored
        if target.exists():
            target.unlink()
    return redirect("/profile?avatar_removed=1")


@app.get("/teacher/students/export")
def export_teacher_students(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    user: Annotated[User, Depends(require_user)],
):
    """Export the student list with per-PLO averages as an .xlsx workbook."""
    if user.role != Role.TEACHER:
        raise HTTPException(status_code=403)
    study_period = selected_study_period(request, user)
    course_id = optional_int(request.query_params.get("course_id"))
    class_id = optional_int(request.query_params.get("class_id"))
    course = selected_teacher_course(session, user, course_id, study_period)
    if not course:
        raise HTTPException(status_code=404, detail="No assigned course selected")

    page = build_teacher_page(
        "students", session, user, course.id, None, class_id, study_period, None
    )
    students = page["students"]
    plos = page["plos"]          # dicts with id/code/description
    sp = page["student_plo"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students PLO Averages"

    headers = ["#", "Student ID", "Student Name", "Class", "Status"]
    headers += [f"{plo['code']} score" for plo in plos]
    headers += ["Total Score"]
    sheet.append(headers)

    for index, student in enumerate(students, 1):
        row = [
            index,
            student["student_no"],
            student["name"],
            student.get("class_name", ""),
            student.get("status", ""),
        ]
        row += [sp["per_student"].get(student["id"], {}).get(plo["id"]) for plo in plos]
        row += [sp["student_total"].get(student["id"])]
        sheet.append(row)

    if students:
        average_row = ["", "Class average", "", "", ""]
        average_row += [sp["plo_average"].get(plo["id"]) for plo in plos]
        average_row += [sp["total"]]
        sheet.append(average_row)
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True, color="061847")
            cell.fill = PatternFill(fill_type="solid", fgColor="EEF6FF")

    header_fill = PatternFill(fill_type="solid", fgColor="E8F1FF")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="061847")
        cell.fill = header_fill

    widths = {"A": 6, "B": 16, "C": 28, "D": 14, "E": 12}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "C2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"{course.code}-students-plo.xlsx".replace(" ", "-")
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
